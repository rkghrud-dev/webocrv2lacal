import csv
import re
from openpyxl import load_workbook


def _cell_to_text(value):
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value)


def _dedupe_header(headers, header):
    if header not in headers:
        return header

    idx = 2
    while f'{header}_{idx}' in headers:
        idx += 1
    return f'{header}_{idx}'


def _read_csv_rows(file_path):
    last_error = None
    for encoding in ('utf-8-sig', 'cp949', 'euc-kr', 'utf-8'):
        try:
            with open(file_path, newline='', encoding=encoding) as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                return list(csv.reader(f, dialect))
        except UnicodeDecodeError as e:
            last_error = e

    raise ValueError('CSV 인코딩을 읽을 수 없습니다. UTF-8 또는 CP949 형식으로 저장해주세요') from last_error


def _table_from_rows(rows, sheet_name):
    header_row = None
    header_index = 0
    for i, row in enumerate(rows):
        if row and any(value is not None and str(value).strip() for value in row):
            header_row = row
            header_index = i
            break

    if not header_row:
        return {
            'sheet_name': sheet_name,
            'headers': [],
            'rows': [],
            'row_count': 0
        }

    headers = []
    for i, value in enumerate(header_row):
        header = _cell_to_text(value).strip() or f'col_{i + 1}'
        headers.append(_dedupe_header(headers, header))

    table_rows = []
    for row in rows[header_index + 1:]:
        if not row or not any(value is not None and str(value).strip() for value in row):
            continue

        item = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            item[header] = _cell_to_text(value)
        table_rows.append(item)

    return {
        'sheet_name': sheet_name,
        'headers': headers,
        'rows': table_rows,
        'row_count': len(table_rows)
    }


def parse_table_csv(file_path):
    return _table_from_rows(_read_csv_rows(file_path), 'CSV')


def parse_table_excel(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)

    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header_row = None

            for row in rows:
                if row and any(value is not None and str(value).strip() for value in row):
                    header_row = row
                    break

            if not header_row:
                continue

            headers = []
            for i, value in enumerate(header_row):
                header = _cell_to_text(value).strip() or f'col_{i + 1}'
                headers.append(_dedupe_header(headers, header))

            table_rows = []
            for row in rows:
                if not row or not any(value is not None and str(value).strip() for value in row):
                    continue

                item = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    item[header] = _cell_to_text(value)
                table_rows.append(item)

            return {
                'sheet_name': sheet_name,
                'headers': headers,
                'rows': table_rows,
                'row_count': len(table_rows)
            }

        return {
            'sheet_name': wb.sheetnames[0] if wb.sheetnames else '',
            'headers': [],
            'rows': [],
            'row_count': 0
        }
    finally:
        wb.close()


def parse_gs_code(code):
    if not code or not isinstance(code, str):
        raw = str(code).strip() if code else ''
        return '기타', raw, '', raw

    code = code.strip()

    match = re.match(r'^(GS\d{2})(\d{5})([A-Z])(_.+)?$', code)
    if match:
        supplier = match.group(1)
        seq = match.group(2)
        option = match.group(3)
        suffix = match.group(4) or ''
        sku_group = supplier + seq
        return supplier, seq, option + suffix, sku_group

    if code.upper().startswith('GS'):
        match2 = re.match(r'^(GS\d{2})', code)
        if match2:
            supplier = match2.group(1)
            last_alpha = re.search(r'([A-Z])(_.+)?$', code)
            if last_alpha:
                sku_group = code[:last_alpha.start()]
                option = code[last_alpha.start():]
            else:
                sku_group = code
                option = ''
            return supplier, code[4:].rstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ_0123456789') or code[4:], option, sku_group

    return '기타', code, '', code


def _price_to_float(value):
    try:
        return float(str(value or '0').replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _is_naver_duplicate_status(status):
    normalized = str(status or '').strip().lower()
    return normalized not in ('', '신규', 'n', 'no', '0', 'false', '없음')


def _product_from_row(row_dict):
    product_code = str(row_dict.get('자체 상품코드', '')).strip()
    supplier, seq, option, sku_group = parse_gs_code(product_code)

    return {
        'product_code': product_code,
        'cafe24_code': row_dict.get('상품코드', ''),
        'supplier_code': supplier,
        'product_seq': seq,
        'option_code': option,
        'sku_group': sku_group,
        'product_name': row_dict.get('상품명', ''),
        'price': _price_to_float(row_dict.get('판매가', '0')),
        'image_url': row_dict.get('이미지등록(목록)', ''),
        'display_status': row_dict.get('진열상태', ''),
        'sale_status': row_dict.get('판매상태', ''),
        'naver_status': row_dict.get('네이버중복여부', '신규'),
        'naver_product_id': row_dict.get('네이버상품번호', ''),
        'raw_data': row_dict
    }


def parse_csv(file_path):
    table = parse_table_csv(file_path)
    result = {
        'products': [],
        'naver_duplicates': [],
        'naver_listed_codes': [],
        'headers': table['headers'],
        'sheet_info': {'main': 'CSV'}
    }

    for row_dict in table['rows']:
        product_code = str(row_dict.get('자체 상품코드', '')).strip()
        if not product_code:
            continue

        product = _product_from_row(row_dict)
        result['products'].append(product)

        if _is_naver_duplicate_status(product['naver_status']):
            result['naver_duplicates'].append(product_code)

    return result


def parse_excel(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)

    result = {
        'products': [],
        'naver_duplicates': [],
        'naver_listed_codes': [],
        'headers': [],
        'sheet_info': {}
    }

    main_sheet = None
    for name in wb.sheetnames:
        if '전체' in name and '상품' in name:
            main_sheet = wb[name]
            result['sheet_info']['main'] = name
            break
    if not main_sheet:
        main_sheet = wb[wb.sheetnames[0]]
        result['sheet_info']['main'] = wb.sheetnames[0]

    headers = []
    first_row = next(main_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for i, val in enumerate(first_row):
        headers.append(str(val) if val else f'col_{i+1}')
    result['headers'] = headers

    col_indices = {}
    target_cols = {
        '자체 상품코드': None, '상품코드': None, '상품명': None,
        '판매가': None, '이미지등록(목록)': None, '진열상태': None,
        '판매상태': None, '네이버중복여부': None, '네이버상품번호': None
    }
    for i, h in enumerate(headers):
        if h in target_cols:
            col_indices[h] = i

    for row in main_sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        idx = col_indices.get('자체 상품코드')
        if idx is None or idx >= len(row) or not row[idx]:
            continue

        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                if val is not None:
                    row_dict[headers[i]] = str(val)
                else:
                    row_dict[headers[i]] = ''

        product_code = str(row[col_indices['자체 상품코드']]).strip()
        supplier, seq, option, sku_group = parse_gs_code(product_code)

        result['products'].append(_product_from_row(row_dict))

    for name in wb.sheetnames:
        if '이미' in name and '올린' in name:
            listed_sheet = wb[name]
            listed_headers = []
            first_listed = next(listed_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            for val in first_listed:
                listed_headers.append(str(val) if val else '')

            gs_idx = None
            for i, h in enumerate(listed_headers):
                if '자체' in h and '상품' in h:
                    gs_idx = i
                    break

            if gs_idx is not None:
                for row in listed_sheet.iter_rows(min_row=2, values_only=True):
                    if row and gs_idx < len(row) and row[gs_idx]:
                        result['naver_listed_codes'].append(str(row[gs_idx]).strip())

            result['sheet_info']['naver_listed'] = name
            break

    for name in wb.sheetnames:
        if '중복' in name:
            dup_sheet = wb[name]
            dup_headers = []
            first = next(dup_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            for val in first:
                dup_headers.append(str(val) if val else '')

            gs_idx = None
            for i, h in enumerate(dup_headers):
                if 'GS' in h and '코드' in h:
                    gs_idx = i
                    break

            if gs_idx is not None:
                for row in dup_sheet.iter_rows(min_row=2, values_only=True):
                    if row and gs_idx < len(row) and row[gs_idx]:
                        result['naver_duplicates'].append(str(row[gs_idx]).strip())

            result['sheet_info']['duplicates'] = name
            break

    wb.close()
    return result
