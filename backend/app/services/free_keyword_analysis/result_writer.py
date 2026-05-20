from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


MARKET_A_TITLE_COL = "추천키워드1"
MARKET_A_SEARCH_COL = "추천키워드2"
MARKET_B_TITLE_COL = "추천키워드4"
MARKET_B_SEARCH_COL = "추천키워드5"


def write_keyword_result(
    upload_file: str,
    analysis_excel: str,
    output_dir: str = "",
) -> str:
    upload_path = Path(upload_file).resolve()
    analysis_path = Path(analysis_excel).resolve()
    if not upload_path.is_file():
        raise FileNotFoundError(f"upload file not found: {upload_path}")
    if not analysis_path.is_file():
        raise FileNotFoundError(f"analysis excel not found: {analysis_path}")

    out_dir = Path(output_dir).resolve() if output_dir else upload_path.parent / "llm_result_v4_local"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{upload_path.stem}_llm_v4_local.xlsx"

    lookup = _load_analysis_lookup(str(analysis_path))
    wb = load_workbook(upload_path)
    for sheet_name in wb.sheetnames:
        if sheet_name not in {"분리추출후", "B마켓"}:
            continue
        ws = wb[sheet_name]
        headers = _headers(ws)
        is_b_market = sheet_name == "B마켓"
        _ensure_columns(ws, headers, ["검색어설정", "검색키워드", "쿠팡검색태그", "네이버태그", "OCR요약"])
        headers = _headers(ws)

        for row_idx in range(2, ws.max_row + 1):
            gs_code = _extract_row_gs(ws, headers, row_idx)
            if not gs_code:
                continue
            analysis = _find_analysis_row(lookup, gs_code)
            if not analysis:
                continue
            _apply_row(ws, headers, row_idx, analysis, is_b_market=is_b_market)

    wb.save(out_path)
    return str(out_path)


def _load_analysis_lookup(path: str) -> dict[str, dict[str, Any]]:
    df = pd.read_excel(path, sheet_name="상품요약").fillna("")
    lookup: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        payload = {str(k): row[k] for k in df.columns}
        gs = str(payload.get("GS코드", "")).strip().upper()
        if gs:
            lookup[gs] = payload
            base = _base_gs(gs)
            if base:
                lookup.setdefault(base, payload)
                lookup.setdefault(base + "A", payload)
    return lookup


def _headers(ws) -> dict[str, int]:
    out = {}
    for cell in ws[1]:
        name = str(cell.value or "").strip()
        if name:
            out[name] = int(cell.column)
    return out


def _ensure_columns(ws, headers: dict[str, int], names: list[str]) -> None:
    next_col = ws.max_column + 1
    for name in names:
        if name in headers:
            continue
        ws.cell(row=1, column=next_col, value=name)
        headers[name] = next_col
        next_col += 1


def _extract_row_gs(ws, headers: dict[str, int], row_idx: int) -> str:
    preferred = ["자체 상품코드", "자체상품코드", "GS상품코드", "상품명", "상품명(관리용)", "모델명"]
    for col_name in preferred:
        col = headers.get(col_name)
        if not col:
            continue
        value = str(ws.cell(row=row_idx, column=col).value or "")
        gs = _extract_gs(value)
        if gs:
            return gs
    values = [str(ws.cell(row=row_idx, column=col).value or "") for col in range(1, ws.max_column + 1)]
    return _extract_gs(" ".join(values))


def _extract_gs(value: str) -> str:
    match = re.search(r"(GS\d{7})([A-Z])?", str(value or ""), flags=re.IGNORECASE)
    if not match:
        return ""
    return (match.group(1) + (match.group(2) or "A")).upper()


def _base_gs(gs_code: str) -> str:
    match = re.search(r"(GS\d{7})", str(gs_code or ""), flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def _find_analysis_row(lookup: dict[str, dict[str, Any]], gs_code: str) -> dict[str, Any] | None:
    gs = gs_code.upper()
    return lookup.get(gs) or lookup.get(_base_gs(gs)) or lookup.get(_base_gs(gs) + "A")


def _apply_row(ws, headers: dict[str, int], row_idx: int, analysis: dict[str, Any], is_b_market: bool) -> None:
    title_col = MARKET_B_TITLE_COL if is_b_market else MARKET_A_TITLE_COL
    search_col = MARKET_B_SEARCH_COL if is_b_market else MARKET_A_SEARCH_COL
    title = _clean_line(str(analysis.get(title_col, ""))) or _clean_line(str(analysis.get(MARKET_A_TITLE_COL, "")))
    search_keywords = _search_keywords(str(analysis.get(search_col, "")) or title)
    tags = _tags_from_analysis(analysis, is_b_market=is_b_market)
    ocr_summary = _clean_line(str(analysis.get("키워드소스문장", ""))) or _clean_line(str(analysis.get("OCR정제문", "")))

    _set(ws, headers, row_idx, "상품명", _limit_title(title))
    _set(ws, headers, row_idx, "검색어설정", ",".join(tags))
    _set(ws, headers, row_idx, "검색키워드", search_keywords)
    _set(ws, headers, row_idx, "쿠팡검색태그", ",".join(tags))
    _set(ws, headers, row_idx, "네이버태그", "|".join(tags[:10]))
    _set(ws, headers, row_idx, "OCR요약", ocr_summary[:500])
    _set(ws, headers, row_idx, "1차키워드", title)
    _set(ws, headers, row_idx, "최종키워드2차", title)


def _set(ws, headers: dict[str, int], row_idx: int, col_name: str, value: str) -> None:
    col = headers.get(col_name)
    if col:
        ws.cell(row=row_idx, column=col, value=value)


def _clean_line(value: str) -> str:
    value = re.sub(r"\s+", " ", str(value or "")).strip()
    value = re.sub(r"\bnan\b", "", value, flags=re.IGNORECASE).strip()
    return value


def _limit_title(title: str, max_len: int = 100) -> str:
    title = _clean_line(title)
    if len(title) <= max_len:
        return title
    tokens = title.split()
    out = []
    for token in tokens:
        candidate = " ".join(out + [token])
        if len(candidate) > max_len:
            break
        out.append(token)
    return " ".join(out) if out else title[:max_len]


def _split_field(value: Any) -> list[str]:
    raw = str(value or "")
    if not raw:
        return []
    return [x.strip() for x in re.split(r"[,;|]", raw) if x.strip()]


def _tags_from_analysis(analysis: dict[str, Any], is_b_market: bool) -> list[str]:
    fields = ["상품정체성", "대표검색어", "다른명칭"]
    fields += ["사용처", "사용자대상", "핵심특징", "규격", "옵션"] if is_b_market else ["규격", "옵션", "소재", "핵심특징", "사용처"]
    raw_terms: list[str] = []
    for field in fields:
        raw_terms.extend(_split_field(analysis.get(field, "")))

    tags: list[str] = []
    seen: set[str] = set()
    for term in raw_terms:
        tag = _tagify(term)
        if not tag or tag in seen:
            continue
        seen.add(tag)
        tags.append(tag)
        if len(tags) >= (14 if is_b_market else 20):
            break
    return tags


def _tagify(term: str) -> str:
    term = _clean_line(term)
    if not term:
        return ""
    term = re.sub(r"\s+", "", term)
    term = re.sub(r"[^0-9A-Za-z가-힣]", "", term)
    if re.search(r"[A-Za-z]", re.sub(r"A4", "", term, flags=re.IGNORECASE)):
        return ""
    if len(term) < 2 or len(term) > 20:
        return ""
    return term


def _search_keywords(line: str, max_count: int = 18) -> str:
    tokens = []
    seen = set()
    for token in _clean_line(line).split():
        token = re.sub(r"[^0-9A-Za-z가-힣]", "", token)
        if not token or len(token) < 2:
            continue
        if re.search(r"[A-Za-z]", re.sub(r"A4", "", token, flags=re.IGNORECASE)):
            continue
        key = token.lower()
        if key in seen:
            continue
        seen.add(key)
        tokens.append(token)
        if len(tokens) >= max_count:
            break
    return " ".join(tokens)
