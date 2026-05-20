"""
쿠팡 WING API 클라이언트 + 업로드 파이프라인
- 키: ~/Desktop/key/ 폴더에서 로드
- GUI 연동을 위한 콜백 기반 인터페이스
"""
from __future__ import annotations

import hmac
import hashlib
import json
import os
import re
import ssl
import time
import urllib.request
import urllib.parse
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Callable

# ─── 키/설정 로드 ───────────────────────────────
_KEY_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "key")

def _load_key(filename: str, fallback: str = "") -> str:
    path = os.path.join(_KEY_DIR, filename)
    if os.path.isfile(path):
        return open(path, encoding="utf-8").read().strip()
    return fallback

def _load_kv_file(filename: str) -> dict[str, str]:
    """KEY=VALUE 형식 파일 로드"""
    path = os.path.join(_KEY_DIR, filename)
    kv = {}
    if os.path.isfile(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                kv[k.strip()] = v.strip()
    return kv

_COUPANG_KEYS = _load_kv_file("coupang_wing_api.txt")
ACCESS_KEY = _COUPANG_KEYS.get("access_key", "")
SECRET_KEY = _COUPANG_KEYS.get("secret_key", "")
VENDOR_ID = _COUPANG_KEYS.get("vendor_id", "")

OUTBOUND_CODE = 23273329
RETURN_CENTER_CODE = 1002256451


# ─── API 클라이언트 ─────────────────────────────

def _auth(method: str, path: str, query: str | None = None) -> str:
    if not ACCESS_KEY or not SECRET_KEY:
        raise RuntimeError(f"쿠팡 API 키를 찾을 수 없습니다: {os.path.join(_KEY_DIR, 'coupang_wing_api.txt')}")

    now = datetime.now(timezone.utc)
    dt = now.strftime("%y%m%dT%H%M%SZ")
    message = dt + method + path + (query or "")
    sig = hmac.new(SECRET_KEY.encode(), message.encode(), hashlib.sha256).hexdigest()
    return f"CEA algorithm=HmacSHA256, access-key={ACCESS_KEY}, signed-date={dt}, signature={sig}"


def api_call(method: str, path: str, query: str | None = None, body: bytes | None = None) -> dict:
    import gzip as _gzip
    url = f"https://api-gateway.coupang.com{path}"
    if query:
        url += f"?{query}"
    req = urllib.request.Request(url)
    req.add_header("Content-type", "application/json;charset=UTF-8")
    req.add_header("Authorization", _auth(method, path, query))
    req.add_header("X-EXTENDED-TIMEOUT", "90000")
    req.add_header("Accept-Encoding", "gzip, identity")
    req.get_method = lambda: method

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    def _decode_resp(resp_obj) -> dict:
        raw = resp_obj.read()
        if raw[:2] == b"\x1f\x8b":
            raw = _gzip.decompress(raw)
        return json.loads(raw.decode("utf-8"))

    try:
        resp = urllib.request.urlopen(req, body, context=ctx, timeout=30) if body else urllib.request.urlopen(req, context=ctx, timeout=30)
        return _decode_resp(resp)
    except urllib.request.HTTPError as e:
        raw = e.read()
        if raw[:2] == b"\x1f\x8b":
            try:
                raw = _gzip.decompress(raw)
            except Exception:
                pass
        msg = raw.decode("utf-8", errors="replace")
        return {"_error": e.code, "_msg": msg[:500]}


def predict_category(product_name: str) -> dict:
    body = json.dumps({"productName": product_name}).encode("utf-8")
    return api_call("POST", "/v2/providers/openapi/apis/api/v1/categorization/predict", body=body)


def get_category_meta(category_code: int) -> dict:
    path = f"/v2/providers/seller_api/apis/api/v1/marketplace/meta/category-related-metas/display-category-codes/{category_code}"
    return api_call("GET", path)


def create_product(product_json: dict) -> dict:
    body = json.dumps(product_json).encode("utf-8")
    return api_call("POST", "/v2/providers/seller_api/apis/api/v1/marketplace/seller-products", body=body)


# ─── 엑셀 → 쿠팡 JSON 변환 ────────────────────

_IMAGE_SELECTION_CACHE: dict[str, dict[str, dict[str, object]]] = {}
_COUPANG_PUBLIC_IMAGE_CACHE: dict[str, str] = {}


def _extract_gs9(code: str | None) -> str:
    match = re.search(r"GS\d{7,9}", str(code or ""), re.IGNORECASE)
    return match.group(0).upper() if match else ""


def _resolve_export_root(source_file_path: str) -> str:
    from pathlib import Path

    path = Path(source_file_path).resolve()
    parent = path.parent
    parent_name = parent.name.lower()
    grand_name = parent.parent.name.lower() if parent.parent else ""

    if parent_name == "llm_result" and grand_name == "llm_chunks":
        return str(parent.parent.parent)
    if parent_name in ("llm_result", "llm_result_v5_cli", "llm_result_v4_cli"):
        return str(parent.parent)
    return str(parent)


def _load_image_selections(export_root: str) -> dict[str, dict[str, object]]:
    cached = _IMAGE_SELECTION_CACHE.get(export_root)
    if cached is not None:
        return cached

    path = os.path.join(export_root, "image_selections.json")
    if not os.path.isfile(path):
        _IMAGE_SELECTION_CACHE[export_root] = {}
        return _IMAGE_SELECTION_CACHE[export_root]

    try:
        with open(path, encoding="utf-8-sig") as f:
            raw = json.load(f)
        parsed: dict[str, dict[str, object]] = {}
        for key, value in raw.items():
            if not isinstance(value, dict):
                continue
            parsed[key.upper()] = {
                "main": value.get("main"),
                "mainB": value.get("mainB"),
                "additional": value.get("additional") if isinstance(value.get("additional"), list) else [],
            }
        _IMAGE_SELECTION_CACHE[export_root] = parsed
    except Exception:
        _IMAGE_SELECTION_CACHE[export_root] = {}

    return _IMAGE_SELECTION_CACHE[export_root]


def _find_listing_folder(export_root: str, gs9: str) -> str | None:
    from pathlib import Path

    for root_name in ("listing_images",):
        base = Path(export_root) / root_name
        if not base.is_dir():
            continue
        matches = sorted((p for p in base.rglob(gs9) if p.is_dir()), key=lambda p: str(p))
        if matches:
            return str(matches[0])
    return None


def _pick_local_listing_images(row: dict) -> list[str]:
    source_file_path = str(row.get("_source_file_path") or "")
    gs9 = _extract_gs9(row.get("자체 상품코드"))
    if not source_file_path or not gs9:
        return []

    export_root = str(row.get("_export_root") or _resolve_export_root(source_file_path))
    folder_path = _find_listing_folder(export_root, gs9)
    if not folder_path:
        return []

    files = sorted(
        [
            os.path.join(folder_path, name)
            for name in os.listdir(folder_path)
            if os.path.splitext(name)[1].lower() in (".jpg", ".jpeg", ".png", ".bmp", ".webp")
        ],
        key=lambda p: os.path.basename(p).lower(),
    )
    if not files:
        return []

    selection = _load_image_selections(export_root).get(gs9, {})
    ordered: list[str] = []
    seen: set[int] = set()
    has_explicit_selection = False

    main_index = selection.get("main")
    if isinstance(main_index, int) and 0 <= main_index < len(files):
        ordered.append(files[main_index])
        seen.add(main_index)
        has_explicit_selection = True

    additional = selection.get("additional") if isinstance(selection.get("additional"), list) else []
    for idx in additional:
        if isinstance(idx, int) and 0 <= idx < len(files) and idx not in seen:
            ordered.append(files[idx])
            seen.add(idx)
            has_explicit_selection = True

    if has_explicit_selection:
        return ordered[:10]

    default_main = 1 if len(files) > 1 else 0
    ordered.append(files[default_main])
    seen.add(default_main)

    for idx, file_path in enumerate(files):
        if idx in seen:
            continue
        ordered.append(file_path)

    return ordered[:10]

def _resolve_public_image_url(image_path: str) -> str:
    cached = _COUPANG_PUBLIC_IMAGE_CACHE.get(image_path)
    if cached:
        return cached

    from app.services.naver_commerce import upload_image_url

    public_url = upload_image_url(image_path)
    _COUPANG_PUBLIC_IMAGE_CACHE[image_path] = public_url
    return public_url


def _build_fallback_image_urls(row: dict, detail_html: str) -> list[str]:
    import re as _re

    product_img_urls = []
    representative_candidates = []
    for value in (row.get("이미지등록(목록)"), row.get("이미지등록(추가)")):
        if not value:
            continue
        representative_candidates.extend(
            [u.strip() for u in _re.split(r"[|\n]", str(value)) if u and str(u).strip()]
        )

    seen_candidates = set()
    for candidate in representative_candidates:
        candidate = str(candidate).strip()
        if not candidate or candidate in seen_candidates:
            continue
        seen_candidates.add(candidate)
        product_img_urls.append(candidate)

    return product_img_urls[:10]

def _build_coupang_image_urls(row: dict, detail_html: str) -> list[str]:
    local_images = _pick_local_listing_images(row)
    if local_images:
        public_urls = []
        first_error = None
        for image_path in local_images[:10]:
            try:
                public_urls.append(_resolve_public_image_url(image_path))
            except Exception as ex:
                if first_error is None:
                    first_error = ex
        if public_urls:
            return public_urls[:10]
        if first_error:
            raise ValueError(f"쿠팡 이미지 공개 URL 준비 실패: {first_error}") from first_error
        raise ValueError("쿠팡 이미지 공개 URL 준비 실패")

    return _build_fallback_image_urls(row, detail_html)


def read_source_file(file_path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    # 쿠팡은 B마켓 → "B마켓" 시트 우선, 없으면 "분리추출후", 최종 fallback active
    if "B마켓" in wb.sheetnames:
        ws = wb["B마켓"]
    elif "분리추출후" in wb.sheetnames:
        ws = wb["분리추출후"]
    else:
        ws = wb.active
    headers = {c.column: c.value for c in ws[1] if c.value}
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for col, name in headers.items():
            row[name] = ws.cell(row=r, column=col).value
        row["_row_num"] = r
        row["_source_file_path"] = file_path
        row["_export_root"] = _resolve_export_root(file_path)
        rows.append(row)
    return rows

def parse_options(option_str: str | None, extra_price_str: str | None = None) -> list[dict]:
    if not option_str:
        return []

    option_text = str(option_str).strip()
    brace_match = re.match(r"^옵션\{(.*)\}$", option_text)
    if brace_match:
        option_text = brace_match.group(1)

    prices = []
    if extra_price_str:
        # 구분자: , 또는 |
        for p in re.split(r"[,|]", str(extra_price_str)):
            p = p.strip()
            if p:
                try:
                    prices.append(int(float(p)))
                except ValueError:
                    prices.append(0)

    options = []
    parts = [part.strip() for part in re.split(r"[,|]", option_text) if part and part.strip()]
    for i, part in enumerate(parts):
        match = re.match(r"^([A-Z])\s*(.+)$", part)
        value = f"{match.group(1)} {match.group(2).strip()}" if match else part
        price = prices[i] if i < len(prices) else 0
        options.append({"name": value.strip(), "price": price})
    return options

def build_notice_content(category_meta: dict, row: dict | None = None, product_name: str = "", options: list[dict] | None = None) -> list[dict]:
    notices = category_meta.get("data", {}).get("noticeCategories", [])
    if not notices:
        return []
    row = row or {}
    options = options or []
    source_text = " ".join(
        str(x or "")
        for x in (product_name, row.get("OCR요약"), row.get("상품 상세설명"), row.get("상세설명"))
    )
    notice = _select_notice_category(notices, source_text)
    notice_name = notice.get("noticeCategoryName", "")
    details = notice.get("noticeCategoryDetailNames", [])
    fields = _build_notice_fields(row, product_name, options, source_text)
    return [
        {
            "noticeCategoryName": notice_name,
            "noticeCategoryDetailName": d.get("noticeCategoryDetailName", ""),
            "content": _build_notice_detail_content(d.get("noticeCategoryDetailName", ""), fields),
        }
        for d in details
    ]


def _select_notice_category(notices: list[dict], source_text: str) -> dict:
    best = notices[0]
    best_score = -999
    for notice in notices:
        notice_name = str(notice.get("noticeCategoryName", ""))
        details_text = " ".join(str(item.get("noticeCategoryDetailName", "")) for item in notice.get("noticeCategoryDetailNames", []))
        score = 0
        if re.search(r"깔창|인솔|신발|구두|운동화|슬리퍼|부츠", source_text) and re.search(r"신발|구두", notice_name):
            score += 80
        if re.search(r"가방|백팩|파우치|숄더백|토트백", source_text) and "가방" in notice_name:
            score += 80
        if re.search(r"의류|티셔츠|셔츠|바지|자켓|재킷|점퍼|원피스|스커트", source_text) and "의류" in notice_name:
            score += 80
        if "기타" in notice_name:
            score -= 10
        if any(token in details_text for token in ("재질", "색상", "크기")):
            score += 5
        if score > best_score:
            best = notice
            best_score = score
    return best


def _build_notice_fields(row: dict, product_name: str, options: list[dict], source_text: str) -> dict:
    json_fields = _extract_notice_json_fields(
        row.get("네이버상품정보고시") or row.get("상품정보제공고시") or row.get("naverProvidedNotice") or ""
    )
    colors = _extract_notice_colors(options)
    return {
        "product_name": product_name,
        "model_name": _first_notice_value(json_fields.get("modelName"), _extract_gs_from_text(source_text), product_name),
        "material": _first_notice_value(json_fields.get("material"), _extract_labeled_notice_value(source_text, ["소재", "재질", "재료", "원단"]), _extract_notice_material(source_text)),
        "color": _first_notice_value(json_fields.get("color"), _extract_labeled_notice_value(source_text, ["색상", "컬러", "색깔"]), colors),
        "size": _first_notice_value(json_fields.get("size"), _extract_labeled_notice_value(source_text, ["사이즈", "크기", "규격", "치수", "중량"]), _extract_notice_size(source_text)),
        "quantity": _first_notice_value(json_fields.get("quantity"), _extract_labeled_notice_value(source_text, ["수량", "구성수량"]), "1개"),
        "manufacturer": _first_notice_value(json_fields.get("manufacturer"), _extract_labeled_notice_value(source_text, ["제조사", "제조자", "제조원"])),
        "importer": _first_notice_value(json_fields.get("importer"), _extract_labeled_notice_value(source_text, ["수입사", "수입원", "수입자"])),
        "origin": _first_notice_value(json_fields.get("origin"), _extract_labeled_notice_value(source_text, ["제조국", "원산지"]), "중국"),
        "customer_service": _first_notice_value(json_fields.get("customerServicePhoneNumber"), _extract_labeled_notice_value(source_text, ["A/S", "AS", "고객센터", "소비자상담"]), "홈런마켓 / 010-2324-8352"),
        "components": _first_notice_value(_extract_labeled_notice_value(source_text, ["제품 구성", "제품구성", "구성품", "구성"]), "본품 1개"),
    }


def _build_notice_detail_content(detail_name: str, fields: dict) -> str:
    if "품명" in detail_name or "모델명" in detail_name:
        return f"{fields['product_name']} / {fields['model_name']}" if fields.get("model_name") else fields["product_name"]
    if "KC" in detail_name or "인증" in detail_name or "허가" in detail_name:
        return "해당없음"
    if "크기" in detail_name or "중량" in detail_name:
        return fields.get("size") or "상세페이지 참조"
    if "색상" in detail_name:
        return fields.get("color") or "상세페이지 참조"
    if "재질" in detail_name or "소재" in detail_name or "원단" in detail_name:
        return fields.get("material") or "상세페이지 참조"
    if "제품 구성" in detail_name or "제품구성" in detail_name:
        return fields.get("components") or "본품 1개"
    if "수량" in detail_name:
        return fields.get("quantity") or "1개"
    if "출시" in detail_name:
        return "상세페이지 참조"
    if "제조자" in detail_name or "수입자" in detail_name:
        maker = fields.get("manufacturer") or "상세페이지 참조"
        importer = fields.get("importer") or "홈런마켓"
        return f"제조자: {maker} / 수입자: {importer}"
    if "제조국" in detail_name or "원산지" in detail_name:
        return fields.get("origin") or "중국"
    if "세부 사양" in detail_name:
        return "상세페이지 참조"
    if "품질보증" in detail_name:
        return "소비자분쟁해결기준에 따름"
    if "A/S" in detail_name or "소비자상담" in detail_name:
        return fields.get("customer_service") or "홈런마켓 / 010-2324-8352"
    return "상세페이지 참조"


def _extract_notice_json_fields(raw: object) -> dict:
    if isinstance(raw, dict):
        payload = raw
    else:
        text = str(raw or "").strip()
        if not text:
            return {}
        try:
            payload = json.loads(text)
        except Exception:
            return {}
    if isinstance(payload.get("productInfoProvidedNotice"), dict):
        payload = payload["productInfoProvidedNotice"]
    out: dict[str, str] = {}
    for value in payload.values():
        if not isinstance(value, dict):
            continue
        for key, inner in value.items():
            cleaned = _clean_notice_value(inner)
            if cleaned and cleaned not in {"상품상세 참조", "상세페이지 참조"}:
                out[key] = cleaned
    return out


def _extract_labeled_notice_value(source_text: str, labels: list[str]) -> str:
    if not source_text or not labels:
        return ""
    label_pattern = "|".join(re.escape(label) for label in labels)
    stop_pattern = r"소재|재질|재료|원단|수량|구성수량|색상|컬러|색깔|사이즈|크기|규격|치수|중량|제조사|제조자|제조원|수입사|수입원|수입자|제조국|원산지|A/S|AS|고객센터|소비자상담|제품 구성|제품구성|구성품|구성"
    joined = re.sub(r"<[^>]+>", " ", source_text)
    joined = re.sub(r"\s+", " ", joined).strip()
    match = re.search(
        rf"(?:^|\s)(?:{label_pattern})\s*[:：]?\s*(.+?)(?=\s+(?:{stop_pattern})\s*[:：]?|\s+홈런마켓\b|\s+급배송\b|$)",
        joined,
        re.IGNORECASE,
    )
    if not match:
        return ""
    value = _clean_notice_value(match.group(1))
    return "" if _notice_is_noise(value) else value


def _clean_notice_value(value: object) -> str:
    text = re.sub(r"[{}\"`]+", " ", str(value or ""))
    text = re.sub(r"\s+", " ", text).strip(" ,./|·:-")
    text = re.split(
        r"\s+(?:홈런마켓|급배송|평일|택배사|구매대행|국내|모든\s*제품|상품/대량구매|대량구매|퀵서비스|방문수령)\b",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    return text.strip(" ,./|·:-")


def _notice_is_noise(value: str) -> bool:
    return not value or bool(re.search(r"^(?:high|bullet|advantage|product\s*profile|size|in/mm|on|off|zero|\d{1,4}|\d{1,4}\s*mm)$", value, re.IGNORECASE))


def _first_notice_value(*values: object) -> str:
    for value in values:
        cleaned = _clean_notice_value(value)
        if cleaned and not _notice_is_noise(cleaned):
            return cleaned
    return ""


def _extract_gs_from_text(value: str) -> str:
    match = re.search(r"GS\d{7}[A-Z0-9]*", value or "", re.IGNORECASE)
    return match.group(0).upper() if match else ""


def _extract_notice_colors(options: list[dict]) -> str:
    values: list[str] = []
    for option in options:
        value = str(option.get("name", "")).strip()
        value = re.sub(r"^[A-Z]\s+", "", value).strip()
        if value and value not in values:
            values.append(value)
    return ", ".join(values[:12])


def _extract_notice_material(source_text: str) -> str:
    found: list[str] = []
    for label, pattern in (
        ("ABS", r"\bABS\b"),
        ("스테인리스", r"스테인리스|스텐"),
        ("플라스틱", r"플라스틱|PP|PVC|PE"),
        ("아크릴", r"아크릴"),
        ("실리콘", r"실리콘"),
        ("알루미늄", r"알루미늄"),
        ("철제", r"철제|스틸"),
    ):
        if re.search(pattern, source_text or "", re.IGNORECASE) and label not in found:
            found.append(label)
    return ", ".join(found[:3])


def _extract_notice_size(source_text: str) -> str:
    values: list[str] = []
    for match in re.finditer(r"\d+(?:\.\d+)?\s*(?:mm|cm|m|g|kg|ml|L|리터|개|매|장)", source_text or "", re.IGNORECASE):
        value = re.sub(r"\s+", "", match.group(0))
        if value not in values:
            values.append(value)
    return ", ".join(values[:4])


def build_attributes(category_meta: dict) -> list[dict]:
    attrs = category_meta.get("data", {}).get("attributes", [])
    result = []
    for a in attrs:
        if a["required"] != "MANDATORY":
            continue
        if a.get("exposed") == "EXPOSED":
            continue
        attr_name = a["attributeTypeName"]
        if a["inputType"] == "SELECT" and a.get("inputValues"):
            values = a["inputValues"]
            val = values[0].get("inputValueName", str(values[0])) if isinstance(values[0], dict) else str(values[0])
        elif attr_name in ("수량", "총 수량"):
            val = "1"
        elif attr_name == "색상":
            val = "기타"
        elif a["dataType"] == "NUMBER":
            val = "1"
        else:
            val = "상세페이지 참조"

        entry = {"attributeTypeName": attr_name, "attributeValueName": val}
        if a.get("basicUnit") and a["basicUnit"] != "없음":
            entry["unitCodeName"] = a["basicUnit"]
        result.append(entry)
    return result


def build_option_attributes(option_name: str | None, category_meta: dict) -> list[dict]:
    option_text = str(option_name or "").strip()
    if not option_text:
        return []

    color_value = None
    for label, pattern in (
        ("투명", r"투명"),
        ("스텐", r"스텐|스테인리스"),
        ("실버", r"실버|은색"),
        ("화이트", r"화이트|흰색"),
        ("블랙", r"블랙|검정"),
    ):
        if re.search(pattern, option_text, re.IGNORECASE):
            color_value = label
            break

    quantity_value = None
    count_match = re.search(r"(\d+)\s*(개|ea|EA|p|P)", option_text)
    if count_match:
        quantity_value = f"{count_match.group(1)}개"
    elif re.search(r"싱글|single", option_text, re.IGNORECASE):
        quantity_value = "1개"
    elif re.search(r"더블|double", option_text, re.IGNORECASE):
        quantity_value = "2개"

    result = []
    for attr in category_meta.get("data", {}).get("attributes", []):
        if attr.get("exposed") != "EXPOSED":
            continue

        attr_name = str(attr.get("attributeTypeName", "")).strip()
        if attr_name == "색상" and color_value:
            result.append({"attributeTypeName": attr_name, "attributeValueName": color_value})
        elif attr_name in ("수량", "총 수량") and quantity_value:
            result.append({"attributeTypeName": attr_name, "attributeValueName": quantity_value})

    return result


def _merge_attributes(base_attributes: list[dict], extra_attributes: list[dict]) -> list[dict]:
    merged = [dict(attr) for attr in base_attributes]
    name_to_index = {
        str(attr.get("attributeTypeName", "")).strip(): index
        for index, attr in enumerate(merged)
    }

    for attr in extra_attributes:
        attr_name = str(attr.get("attributeTypeName", "")).strip()
        if not attr_name:
            continue
        if attr_name in name_to_index:
            merged[name_to_index[attr_name]] = dict(attr)
        else:
            name_to_index[attr_name] = len(merged)
            merged.append(dict(attr))

    return merged


def parse_search_tags(raw: str | None, max_tags: int = 20) -> list[str]:
    source = str(raw or "")
    has_explicit_delimiter = bool(re.search(r"[,|\n;/]+", source))
    parts = re.split(r"[,|\n;/]+", source) if has_explicit_delimiter else [source]
    tags: list[str] = []
    seen: set[str] = set()

    for part in parts:
        cleaned = re.sub(r"[^0-9A-Za-z가-힣\s]", " ", str(part or ""))
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        if not cleaned:
            continue

        candidates = [cleaned]
        if len(cleaned) > 20 and not has_explicit_delimiter:
            candidates = [token for token in cleaned.split() if 2 <= len(token) <= 20]

        for candidate in candidates:
            if len(candidate) > 20:
                continue
            key = candidate.lower()
            if key in seen:
                continue
            seen.add(key)
            tags.append(candidate)
            if len(tags) >= max_tags:
                return tags

    return tags


def normalize_brand(value: str | None) -> str:
    brand = str(value or "").strip()
    if (
        not brand
        or brand.upper() == "B0000000"
        or "브랜드없음" in brand
        or "자체브랜드" in brand
        or brand == "없음"
    ):
        return "샤플라이"
    return brand


def build_coupang_product(row: dict, category_code: int, category_meta: dict) -> dict:
    # 상품명: 키워드 가공된 이름 우선 사용
    product_name = (
        row.get("홈런_쿠팡상품명")
        or row.get("쿠팡상품명")
        or row.get("홈런_공통마켓상품명")
        or row.get("상품명")
        or row.get("최종키워드2차")
        or row.get("1차키워드")
        or ""
    )
    product_name = str(product_name).strip()
    display_name = product_name[:100]
    sale_price = max(int(row.get("판매가", 0) or 0), 1000)
    original_price = int(row.get("소비자가", 0) or 0)
    if original_price < sale_price:
        original_price = sale_price

    # 이미지: 로컬 가공본이 있으면 public URL로 변환해서 우선 사용
    detail_html = row.get("상품 상세설명") or row.get("상세설명") or ""
    product_img_urls = _build_coupang_image_urls(row, detail_html)

    images = []
    for idx, u in enumerate(product_img_urls[:10]):
        img_type = "REPRESENTATION" if idx == 0 else "DETAIL"
        images.append({"imageOrder": idx, "imageType": img_type, "vendorPath": u})

    options = parse_options(row.get("옵션입력"), row.get("옵션추가금"))
    notice_content = build_notice_content(category_meta, row, display_name, options)
    attributes = build_attributes(category_meta)

    search_tags = (
        row.get("홈런_쿠팡검색태그")
        or row.get("홈런_쿠팡검색키워드")
        or row.get("쿠팡검색태그")
        or row.get("홈런_공통마켓검색키워드")
        or row.get("공통마켓검색키워드")
        or row.get("검색어설정")
        or row.get("검색키워드")
        or ""
    )
    tag_list = parse_search_tags(search_tags, max_tags=20)

    mgmt_name = display_name
    detail_html = row.get("상품 상세설명") or row.get("상세설명") or ""
    ext_sku = row.get("자체 상품코드") or ""

    def _make_item(item_name: str, s_price: int, o_price: int, sku: str) -> dict:
        return {
            "itemName": item_name,
            "originalPrice": o_price,
            "salePrice": s_price,
            "maximumBuyCount": 9999,
            "maximumBuyForPerson": 9999,
            "outboundShippingTimeDay": 2,
            "maximumBuyForPersonPeriod": 1,
            "unitCount": 1,
            "adultOnly": "EVERYONE",
            "taxType": "TAX",
            "parallelImported": "NOT_PARALLEL_IMPORTED",
            "overseasPurchased": "NOT_OVERSEAS_PURCHASED",
            "pccNeeded": False,
            "externalVendorSku": sku,
            "barcode": "",
            "emptyBarcode": True,
            "emptyBarcodeReason": "",
            "notices": notice_content,
            "attributes": attributes,
            "contents": [{"contentsType": "TEXT", "contentDetails": [{"content": detail_html, "detailType": "TEXT"}]}],
            "images": images,
            "searchTags": tag_list,
        }

    items = []
    if options:
        for i, opt in enumerate(options):
            item = _make_item(
                opt["name"],
                sale_price + opt["price"],
                original_price + opt["price"],
                f"{ext_sku}_{i+1}" if ext_sku else "",
            )
            item["attributes"] = _merge_attributes(attributes, build_option_attributes(opt["name"], category_meta))
            items.append(item)
    else:
        items.append(_make_item(display_name, sale_price, original_price, ext_sku))

    return {
        "displayCategoryCode": category_code,
        "sellerProductName": mgmt_name,
        "vendorId": VENDOR_ID,
        "saleStartedAt": "2020-01-01T00:00:00",
        "saleEndedAt": "2099-12-31T00:00:00",
        "displayProductName": display_name,
        "brand": normalize_brand(row.get("브랜드")),
        "generalProductName": "",
        "productGroup": "",
        "deliveryMethod": "SEQUENCIAL",
        "deliveryCompanyCode": "CJGLS",
        "deliveryChargeType": "FREE",
        "deliveryCharge": 0,
        "freeShipOverAmount": 0,
        "deliveryChargeOnReturn": 3000,
        "returnCharge": 3000,
        "outboundShippingPlaceCode": OUTBOUND_CODE,
        "returnCenterCode": RETURN_CENTER_CODE,
        "returnChargeName": "명일우진반품",
        "companyContactNumber": "010-2324-8352",
        "returnZipCode": "05287",
        "returnAddress": "서울특별시 강동구 상일로 74",
        "returnAddressDetail": "고덕리엔파크3단지아파트 고덕리엔파크 321동 CJ대한통운 명일우진대리점",
        "remoteAreaDeliverable": "Y",
        "unionDeliveryType": "UNION_DELIVERY",
        "vendorUserId": "rkghrud",
        "afterServiceInformation": "010-2324-8352",
        "afterServiceContactNumber": "010-2324-8352",
        "requested": True,
        "items": items,
        "requiredDocuments": [],
        "extraInfoMessage": "",
        "manufacture": "",
    }


# ─── 업로드 파이프라인 (GUI 콜백 버전) ─────────

@dataclass
class CoupangUploadConfig:
    file_path: str = ""
    row_start: int = 0          # 0이면 전체
    row_end: int = 0
    dry_run: bool = True

@dataclass
class CoupangUploadResult:
    row: int = 0
    name: str = ""
    status: str = ""
    category: str = ""
    seller_product_id: str = ""
    error: str = ""


def run_coupang_upload(
    config: CoupangUploadConfig,
    status_cb: Callable[[str], None] | None = None,
    progress_cb: Callable[[int], None] | None = None,
) -> list[CoupangUploadResult]:
    """GUI에서 호출하는 쿠팡 업로드 메인 함수"""

    def _status(msg: str) -> None:
        if status_cb:
            status_cb(msg)

    def _progress(pct: int) -> None:
        if progress_cb:
            progress_cb(pct)

    _status("가공파일 읽는 중...")
    rows = read_source_file(config.file_path)
    total = len(rows)
    _status(f"{total}개 상품 로드 완료")

    # 행 필터
    if config.row_start > 0:
        end = config.row_end if config.row_end > 0 else config.row_start
        target_rows = [r for r in rows if config.row_start <= (r["_row_num"] - 1) <= end]
    else:
        target_rows = rows

    target_count = len(target_rows)
    _status(f"처리 대상: {target_count}개")
    results: list[CoupangUploadResult] = []
    category_cache: dict[int, dict] = {}

    # 카테고리 추천 (배치 8건 + 1초 대기, 쿠팡 API 초당 10건 제한)
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    # 엑셀에 쿠팡카테고리코드가 있는 행은 API 호출 스킵
    rows_need_predict = []
    cat_results_ordered: list[tuple[dict, dict]] = [({}, {})] * target_count
    for i, row in enumerate(target_rows):
        preset_cat = row.get("쿠팡카테고리코드") or row.get("쿠팡카테고리")
        if preset_cat:
            cat_code = int(float(str(preset_cat)))
            cat_results_ordered[i] = (row, {"data": {"predictedCategoryId": cat_code, "predictedCategoryName": f"엑셀지정({cat_code})", "autoCategorizationPredictionResultType": "SUCCESS"}, "_preset": True})
            _status(f"  행{row['_row_num']}: 엑셀 카테고리 사용 ({cat_code})")
        else:
            rows_need_predict.append((i, row))

    if rows_need_predict:
        _status(f"카테고리 추천 중... ({len(rows_need_predict)}건 API 호출)")
        cat_lock = threading.Lock()

        BATCH_SIZE = 8
        for batch_start in range(0, len(rows_need_predict), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(rows_need_predict))
            batch = rows_need_predict[batch_start:batch_end]
            batch_t0 = time.time()

            def _predict_one(idx: int, row: dict) -> tuple[int, dict, dict]:
                product_name = row.get("상품명", "")
                cat_result = predict_category(product_name)
                return idx, row, cat_result

            with ThreadPoolExecutor(max_workers=BATCH_SIZE) as pool:
                futures = [pool.submit(_predict_one, i, row) for i, row in batch]
                for fut in as_completed(futures):
                    idx, row, cat_result = fut.result()
                    cat_results_ordered[idx] = (row, cat_result)

        _status(f"[{batch_end}/{target_count}] 카테고리 추천 중...")
        _progress(int(batch_end * 40 / target_count))

        # 초당 10건 제한: 배치 처리 후 최소 1초 대기
        elapsed = time.time() - batch_t0
        if elapsed < 1.0:
            time.sleep(1.0 - elapsed)

    # 결과 처리 (순서대로)
    for i in range(target_count):
        row, cat_result = cat_results_ordered[i]
        product_name = row.get("상품명", "")
        short_name = product_name[:50]
        row_num = row["_row_num"]

        if "_error" in cat_result:
            results.append(CoupangUploadResult(row=row_num, name=short_name, status="CATEGORY_FAIL", error=cat_result.get("_msg", "")[:200]))
            continue

        data = cat_result.get("data", {})
        cat_code = data.get("predictedCategoryId")
        cat_name = data.get("predictedCategoryName", "")
        result_type = data.get("autoCategorizationPredictionResultType", "")

        if not cat_code or result_type != "SUCCESS":
            results.append(CoupangUploadResult(row=row_num, name=short_name, status="CATEGORY_UNCERTAIN", category=cat_name))
            continue

        if cat_code not in category_cache:
            meta = get_category_meta(cat_code)
            category_cache[cat_code] = meta if "_error" not in meta else {"data": {"attributes": [], "noticeCategories": []}}

        row["_category_code"] = cat_code
        row["_category_name"] = cat_name
        row["_category_meta"] = category_cache[cat_code]

    _status(f"카테고리 추천 완료")

    # JSON 생성
    _status("상품 JSON 생성 중...")
    products = []
    for row in target_rows:
        if "_category_code" not in row:
            continue
        product_json = build_coupang_product(row, row["_category_code"], row["_category_meta"])
        products.append({"row": row["_row_num"], "name": row.get("상품명", "")[:50], "sku": row.get("자체 상품코드", ""), "category": f"[{row['_category_code']}] {row['_category_name']}", "json": product_json})
    _progress(50)
    _status(f"JSON 생성 완료: {len(products)}개")

    # 등록 (배치 5건 + 1초 대기)
    if not config.dry_run:
        prod_count = len(products)
        _status(f"쿠팡 등록 시작 ({prod_count}개)...")
        REG_BATCH = 5
        reg_results_ordered: list[CoupangUploadResult] = [CoupangUploadResult()] * prod_count

        def _register_one(idx: int, p: dict) -> tuple[int, CoupangUploadResult]:
            resp = create_product(p["json"])
            if "_error" in resp:
                return idx, CoupangUploadResult(row=p["row"], name=p["name"], status="REGISTER_FAIL", error=resp["_msg"][:200])
            code = resp.get("code", "")
            if code == "SUCCESS":
                spid = str(resp.get("data", ""))
                return idx, CoupangUploadResult(row=p["row"], name=p["name"], status="SUCCESS", seller_product_id=spid)
            return idx, CoupangUploadResult(row=p["row"], name=p["name"], status=f"FAIL_{code}", error=resp.get("message", "")[:200])

        for batch_start in range(0, prod_count, REG_BATCH):
            batch_end = min(batch_start + REG_BATCH, prod_count)
            batch_t0 = time.time()
            with ThreadPoolExecutor(max_workers=REG_BATCH) as pool:
                futures = [pool.submit(_register_one, i, products[i]) for i in range(batch_start, batch_end)]
                for fut in as_completed(futures):
                    idx, r = fut.result()
                    reg_results_ordered[idx] = r
            _status(f"[{batch_end}/{prod_count}] 등록 중...")
            _progress(50 + int(batch_end * 50 / prod_count))
            elapsed = time.time() - batch_t0
            if elapsed < 1.0:
                time.sleep(1.0 - elapsed)

        for i in range(prod_count):
            r = reg_results_ordered[i]
            results.append(r)
            _status(f"[{i+1}/{prod_count}] {r.status} - {r.name}")
    else:
        _status("DRY RUN 완료 - 등록하지 않음")
        for p in products:
            results.append(CoupangUploadResult(row=p["row"], name=p["name"], status="DRY_RUN", category=p["category"]))
        _progress(100)

    _progress(100)
    return results
