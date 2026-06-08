from __future__ import annotations



from dataclasses import dataclass

from datetime import datetime

from concurrent.futures import ThreadPoolExecutor, as_completed

import base64

import io

import json

import math

import os

import re

import time



import numpy as np

import pandas as pd



from . import legacy_core as core

from .keyword_builder import build_keyword_string

from .market_keywords import MARKET_KEYWORD_COLUMNS_10, generate_market_keyword_packages

from .env_loader import ensure_env_loaded, get_env





@dataclass

class PipelineConfig:

    file_path: str

    img_tag: str = ""

    tesseract_path: str = ""



    model_keyword: str = "gpt-4.1"        # 키워드 생성용

    model_longtail: str = "gpt-4.1-mini"      # 롱테일/R열 키워드용

    model_keyword_stage2: str = "" # 비우면 1차 모델 재사용

    keyword_version: str = "3.0"



    max_words: int = 24

    max_len: int = 140

    min_len: int = 90



    use_html_ocr: bool = False

    use_local_ocr: bool = True

    merge_ocr_with_name: bool = True



    max_imgs: int = 999

    threads: int = 6

    max_depth: int = -1

    local_img_dir: str = ""

    allow_folder_match: bool = True



    korean_only: bool = True

    drop_digits: bool = True

    psm: int = 11

    oem: int = 3



    ocr_excel_path: str = ""             # 미리 처리된 OCR 결과 Excel 경로



    write_to_r: bool = True



    debug: bool = True



    naver_enabled: bool = False

    naver_dry_run: bool = False

    naver_retry: bool = False

    naver_retry_count: int = 2

    naver_retry_delay: float = 0.8



    naver_autocomplete: bool = False

    google_autocomplete: bool = True



    make_listing: bool = True

    listing_size: int = 1000

    listing_pad: int = 20

    listing_max: int = 20



    logo_path: str = ""

    logo_ratio: int = 14

    logo_opacity: int = 65

    logo_pos: str = "tr"



    use_auto_contrast: bool = True

    use_sharpen: bool = True

    use_small_rotate: bool = True

    rotate_zoom: float = 1.04



    ultra_angle_deg: float = 0.35

    ultra_translate_px: float = 0.6

    ultra_scale_pct: float = 0.25



    trim_tol: int = 8

    jpeg_q_min: int = 88

    jpeg_q_max: int = 92

    do_flip_lr: bool = True

    phase: str = "full"          # "full" | "images" | "analysis"
    export_root_override: str = ""  # phase=analysis 시 Phase1의 export_root 재사용
    chunk_size: int = 10             # ocr_only 모드에서 분할 엑셀 개수 (0이면 미분할)

    # B마켓 설정
    enable_b_market: bool = True     # B마켓 시트 동시 생성 여부
    logo_path_b: str = ""            # B마켓 로고 경로
    img_tag_b: str = ""              # B마켓 상세 이미지 태그

    # 상품명/태그 길이 설정
    a_name_min: int = 80
    a_name_max: int = 100
    b_name_min: int = 63
    b_name_max: int = 98
    a_tag_count: int = 20
    b_tag_count: int = 14



def _normalize_keyword_version(version: str) -> str:

    version = str(version or "3.0").strip()

    if version == "1.0":
        return "1.0"
    if version == "2.0":
        return "2.0"
    if version == "3.0":
        return "3.0"
    return "3.0"



def _keyword_version_slug(version: str) -> str:

    return f"v{_normalize_keyword_version(version).replace('.', '_')}"


def _default_export_base() -> str:
    """V4 기본 출력 루트: 바탕화면 EXPORT."""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(desktop):
        desktop = os.path.join(os.environ.get("USERPROFILE", r"C:\code"), "Desktop")
    return os.path.join(desktop, "EXPORT")


def _make_export_root(date_tag: str, csv_base: str) -> str:
    return os.path.join(_default_export_base(), f"{date_tag}_{csv_base}")



def _prepare_chunk_session_dir(export_root: str, keyword_version: str, date_tag: str) -> tuple[str, str]:

    version_slug = _keyword_version_slug(keyword_version)

    chunks_root = os.path.join(export_root, "llm_chunks")

    os.makedirs(chunks_root, exist_ok=True)

    session_stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

    session_name = f"session_{date_tag}_{version_slug}_{session_stamp}"

    session_dir = os.path.join(chunks_root, session_name)

    os.makedirs(session_dir, exist_ok=True)

    marker_path = os.path.join(chunks_root, f"_active_{version_slug}.txt")

    with open(marker_path, "w", encoding="utf-8") as f:

        f.write(session_dir)

    return session_dir, version_slug



def _status(cb, msg: str) -> None:

    if cb:

        cb(msg)



def _progress(cb, value: int) -> None:

    if cb:

        cb(int(value))


def _ceil_price(value: float) -> int:

    """10원 단위 가격은 1원 자리에서 올림."""

    try:

        v = float(value)

    except Exception:

        return 0

    if v <= 0:

        return 0

    return int(math.ceil(v / 10) * 10)


def _calc_sale_price(supply_price: float) -> int:

    try:

        supply = float(supply_price)

    except Exception:

        return 0

    return _ceil_price(supply * core.get_multiplier(supply))


def _calc_consumer_price(sale_price: float) -> int:

    return _ceil_price(float(sale_price or 0) * 1.2)





def _format_naver_keyword_table(items: list, limit: int = 15) -> str:

    if not items:

        return ""

    rows = []

    for it in items:

        kw = str(it.get("relKeyword") or "").strip()

        if not kw:

            continue

        pc = int(it.get("monthlyPcQcCnt") or 0)

        mo = int(it.get("monthlyMobileQcCnt") or 0)

        total = pc + mo

        rows.append((kw, pc, mo, total))

    rows.sort(key=lambda x: x[3], reverse=True)

    lines = ["키워드|PC|MO|합계"]

    for kw, pc, mo, total in rows[: max(1, int(limit))]:

        lines.append(f"{kw}|{pc}|{mo}|{total}")

    return "\n".join(lines)

def _split_upload_excel(upload_path: str, export_root: str, chunk_size: int, date_tag: str, status_cb=None,
                        llm_chunks_dir: str | None = None):
    """업로드용 엑셀을 chunk_size 상품씩 분할하여 지정 폴더에 저장."""
    import openpyxl

    llm_chunks_dir = llm_chunks_dir or os.path.join(export_root, "llm_chunks")
    os.makedirs(llm_chunks_dir, exist_ok=True)

    wb = openpyxl.load_workbook(upload_path)
    main_ws = wb["분리추출후"]
    ocr_ws = wb["OCR결과"] if "OCR결과" in wb.sheetnames else None

    # 헤더 행
    main_headers = [cell.value for cell in main_ws[1]]
    ocr_headers = [cell.value for cell in ocr_ws[1]] if ocr_ws else []

    # 데이터 행
    main_rows = list(main_ws.iter_rows(min_row=2, values_only=True))
    ocr_rows = list(ocr_ws.iter_rows(min_row=2, values_only=True)) if ocr_ws else []

    # GS코드로 OCR 매핑
    ocr_map = {}
    if ocr_rows and ocr_headers:
        gs_col = 0
        for i, h in enumerate(ocr_headers):
            if str(h or '').strip() == 'GS코드':
                gs_col = i
                break
        for row in ocr_rows:
            gs = str(row[gs_col] or '').strip()
            if gs:
                ocr_map[gs] = row

    # 상품명에서 GS코드 추출
    name_col = 0
    for i, h in enumerate(main_headers):
        if str(h or '').strip() == '상품명':
            name_col = i
            break

    chunk_files = []
    total_chunks = (len(main_rows) + chunk_size - 1) // chunk_size

    for ci in range(total_chunks):
        start = ci * chunk_size
        end = min(start + chunk_size, len(main_rows))
        chunk_rows = main_rows[start:end]

        chunk_wb = openpyxl.Workbook()
        # 분리추출후 시트
        ws1 = chunk_wb.active
        ws1.title = "분리추출후"
        ws1.append(main_headers)
        for row in chunk_rows:
            ws1.append(list(row))

        # OCR결과 시트 (해당 상품만)
        if ocr_headers:
            ws2 = chunk_wb.create_sheet("OCR결과")
            ws2.append(ocr_headers)
            for row in chunk_rows:
                name_val = str(row[name_col] or '')
                import re as _re
                gs_match = _re.search(r'GS\d{7,9}', name_val)
                gs_key = gs_match.group()[:9] if gs_match else ''
                if gs_key and gs_key in ocr_map:
                    ws2.append(list(ocr_map[gs_key]))

        chunk_name = f"chunk_{ci+1:02d}_{date_tag}.xlsx"
        chunk_path = os.path.join(llm_chunks_dir, chunk_name)
        chunk_wb.save(chunk_path)
        chunk_files.append(chunk_path)

    _status(status_cb, f"엑셀 분할 완료: {total_chunks}개 파일 (각 {chunk_size}개 상품) → {llm_chunks_dir}")
    return chunk_files


def _find_market_category_processed_dir() -> str | None:
    """Return the shared market-category-matcher processed data directory, if available."""
    candidates: list[str] = []
    env_dir = os.environ.get("MARKET_CATEGORY_PROCESSED_DIR") or os.environ.get("MARKET_CATEGORY_DATA_DIR")
    if env_dir:
        candidates.append(env_dir)

    user_profile = os.environ.get("USERPROFILE") or os.path.expanduser("~")
    if user_profile:
        candidates.append(os.path.join(
            user_profile,
            "Desktop",
            "프로젝트",
            "market-category-matcher",
            "data",
            "processed",
        ))

    # pipeline.py -> services -> app -> backend -> keywordocr-v3. The matcher is a sibling project.
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    parent_project_dir = os.path.dirname(repo_root)
    candidates.append(os.path.join(parent_project_dir, "market-category-matcher", "data", "processed"))

    for candidate in candidates:
        if candidate and os.path.isdir(candidate):
            return candidate
    return None


def _copy_market_category_references(target_root: str, status_cb=None) -> int:
    """Copy collected market category CSV files into target_root/category_reference."""
    import shutil

    src_dir = _find_market_category_processed_dir()
    if not src_dir:
        _status(status_cb, "마켓 카테고리 기준표 폴더를 찾지 못함: category_reference 복사 생략")
        return 0

    files = [
        "naver_categories.csv",
        "coupang_categories.csv",
        "11st_categories.csv",
        "lotteon_categories.csv",
        "lotteon_standard_categories.csv",
        "lotteon_display_categories.csv",
        "auction_categories.csv",
        "esm_auction_gmarket_category_matching.csv",
    ]
    ref_dir = os.path.join(target_root, "category_reference")
    os.makedirs(ref_dir, exist_ok=True)

    copied = 0
    for file_name in files:
        src = os.path.join(src_dir, file_name)
        if os.path.isfile(src):
            shutil.copy2(src, os.path.join(ref_dir, file_name))
            copied += 1

    if copied:
        _status(status_cb, f"마켓 카테고리 기준표 {copied}개 → {ref_dir} 복사")
    else:
        _status(status_cb, f"마켓 카테고리 기준표 파일 없음: {src_dir}")
    return copied


def _generate_keyword_skill_md(export_root: str, upload_path: str, date_tag: str, chunk_size: int = 0, status_cb=None,
                               a_name_min: int = 80, a_name_max: int = 100,
                               b_name_min: int = 63, b_name_max: int = 98,
                               a_tag_count: int = 20, b_tag_count: int = 14,
                               keyword_version: str = "3.0"):
    """ocr_only 모드에서 LLM 키워드 생성용 지시서(keyword_skill.md)를 생성."""
    keyword_version = _normalize_keyword_version(keyword_version)
    md_path = os.path.join(export_root, "keyword_skill.md")

    upload_basename = os.path.basename(upload_path)
    version_slug = _keyword_version_slug(keyword_version)
    llm_result_rel = f"llm_result_{version_slug}"
    llm_result_dir = os.path.join(export_root, llm_result_rel)
    os.makedirs(llm_result_dir, exist_ok=True)
    result_filename = f"업로드용_{date_tag}_{version_slug}_llm.xlsx"
    result_rel_path = f"{llm_result_rel}/{result_filename}"
    if keyword_version == "1.0":
        version_label = "1.0 레거시 확장형"
    elif keyword_version == "3.0":
        version_label = "3.0 통합 검색형 (구매자 검색어 중심 + A/B 분리)"
    else:
        version_label = "2.0/3.0 통합 검색형"

    if keyword_version == "1.0":
        rules_block = f"""## 키워드 생성 규칙

### 키워드 우선순위 (이 순서대로 채우세요)
1. **핵심 키워드** (상품 유형의 표준명) — 검색창에 사용자가 실제로 입력하는 대표 표현
2. **세부 속성** (규격/소재/용량/호환/기능) — 롱테일 검색 매칭의 핵심
3. **용도/상황** (사무실용, 차량용, 캠핑, 욕실 등) — 구매 의도와 연결
4. **형태/별칭** (스틱형, 롱타입, 미니 등) — 다른 표현으로 검색하는 사용자 커버

### 1단계: 카테고리 확인 (필수)
- 각 상품의 키워드를 작성하기 전에 반드시 **상품이 실제로 어떤 카테고리인지** 확인하세요.
- OCR 텍스트와 원본 상품명을 기반으로 상품의 **정확한 용도/카테고리**를 판단하세요.
- ❌ 카테고리명 자체(`생활철물`, `DIY수리`, `운동용품`)를 제목에 넣지 마세요.
- ✅ 대신 상품의 **구체적 속성/용도** 키워드를 사용하세요.

### 2단계: 상품명 작성
- 공백으로 구분된 키워드 나열입니다.
- **{a_name_min}~{a_name_max}자** 범위로 최대한 채우세요. 너무 짧으면 검색 노출이 줄어듭니다.
- **구조: [핵심상품명(표준명)] + [대표옵션/규격] + [핵심속성] + [용도/사용처] + [별칭/동의어]**
- **동의어 활용**: 같은 단어를 반복하지 말고, 두 번째 언급부터는 **동의어/별칭**을 사용하세요.
- **사용처/용도를 충분히 넣으세요**: 이 상품을 어디서/어떻게 쓰는지 (차량, 자동차, 중장비, 산업현장, 가정, 사무실, 욕실 등).
- 단, OCR이나 원본에 없는 재질/대상/호환 규격을 추정해서 넣지는 마세요.

### 띄어쓰기 규칙 (중요)
- **자연스러운 띄어쓰기 1회만** 사용 — 의미 단위로 띄어쓰기
- ❌ 붙여쓰기/띄어쓰기 둘 다 넣지 마세요 (`무선청소기, 무선 청소기` → `무선 청소기`만)
- ❌ 띄어쓰기 과다/과소 모두 비권장
- ✅ 수식어-상품명 띄어쓰기: `강력 흡입`, `대용량 배터리`
- ✅ 브랜드-제품명 띄어쓰기: `노루 페인트`

### 중복 제거 규칙 (필수)
아래 4종 중복을 반드시 제거하세요. **슬롯 낭비이자 어뷰징 위험**입니다.
1. **완전 중복**: 동일 단어 반복 (`기모, 기모` → `기모`)
2. **공백 변형 중복**: 붙/띄만 다른 표현 (`무선청소기` vs `무선 청소기` → 띄어쓰기 형태 1개만)
3. **재조합 중복**: 같은 단어 순서만 바꾼 것 (`가을 패딩` vs `패딩 가을` → 1개만)
4. **동의어 중복**: 같은 의미 단어 나열 (`핸드폰/휴대폰/스마트폰` → 검색량 높은 1개만)

### 형태 정규화
- **형용사/동사 → 명사형**: `강력한` → `강력`, `사용하는` → `사용`, `튼튼한` → `내구성`
- **연속 공백 제거**: 공백은 항상 1칸
- **합성어 분리형 우선**: 필요시 합성어 1개만 허용 (`유리롤러 레일롤러` → `유리 레일 롤러`)

### 핵심 원칙
1. **핵심상품명 맨 앞**: 원본 상품명의 핵심 단어(GS코드 제외)를 상품명 맨 앞에 배치하세요.
2. **동의어 활용**: 같은 단어를 2번 쓰지 말고, 두 번째부터는 동의어/별칭을 사용하세요.
3. **속성/용도 중심**: 카테고리 대신 구체적 속성(소재, 규격, 기능)과 용도(차량용, 사무실, 캠핑)로 채우세요.
4. **길이 확장 허용**: 짧으면 검색 노출 손해가 크므로, **근거 있는 사용처/별칭/속성**으로 폭을 넓히세요.
5. **근거 없는 확장 금지**: OCR이나 원본에 없는 재질, 호환 규격, 대상층, 사용처는 추정해서 넣지 마세요.
6. **색상 제외**: 색상은 옵션이므로 키워드에서 제외합니다.
7. **카테고리명 금지**: 카테고리 일반어로 제목을 부풀리지 마세요.

### ❌ 반드시 제외할 노이즈
아래 항목은 OCR 텍스트에 포함되어 있더라도 **절대 키워드에 사용하지 마세요**:
- 상세페이지 템플릿: Product Profile, SIZE, Advantage, Features, Description, Specification
- 영어 마케팅 문구: Premium Quality, Best Seller, Hot Item, Free Shipping
- 배송/정책: 배송, 반품, 교환, AS, 원산지, 연락처
- 판매조건: 무료배송, 할인, 프로모션, 이벤트, 특가, 최저가
- 검증 곤란 수식어(단독 사용 금지): 최고급, 프리미엄, 고품질, 인기상품, 베스트
- 엑셀 컬럼명: 옵션입력, 판매가, 재고수, 이미지URL 등 데이터 필드명
- **다른 상품의 키워드를 섞지 마세요** — 각 행은 독립된 상품입니다

### 3단계: 검색어설정 (Cafe24 태그)
- **띄어쓰기 없이 붙여쓰기**, 쉼표로 구분합니다.
- 목표 수량은 **{a_tag_count}개**이며, 가능한 범위에서 넓게 채우세요.
- 상품명에서 다 넣지 못한 **용도/상황/별칭/동의어**를 여기에 배치하세요.
- ❌ 띄어쓰기 금지 — `구리스 호스`가 아니라 `구리스호스`
- ❌ 상품명과 완전히 동일한 조합을 반복하지 마세요.
- ✅ 다양한 조합으로 검색 커버리지를 넓히되, 다른 상품군 단어는 금지합니다.

### 4단계: 검색키워드
- 네이버 쇼핑 검색용 상위 키워드
- 공백 구분, 최대 18개
- 상품명 + 검색어설정에서 핵심만 추려서 배치하세요.
- 가격 분리 상품은 **핵심상품명은 유지하고 뒤쪽 키워드 조합만 조금씩 다르게** 하세요.
"""
    elif keyword_version == "3.0":
        rules_block = f"""## 키워드 생성 규칙 (3.0 타겟형)

### 0단계: 허용 근거 소스 (엄격)
- 사용 가능한 근거: **① 원본 상품명, ② OCR결과 시트의 OCR텍스트** — 이 두 가지만.
- ❌ 네이버/쿠팡/구글 등 **외부 검색 데이터, 추천어, 연관 검색어, 자동완성**은 근거로 쓰지 마세요. (이 파이프라인은 해당 단계를 스킵합니다.)
- ❌ 일반 상식/트렌드 키워드로 확장하지 마세요. 근거에 없으면 비워두는 편이 낫습니다.
- 두 소스에 **동시에 등장**하거나 둘 중 하나에 **명확히 근거**가 있는 토큰만 씁니다.

### 1단계: identity 확정
- 원본 상품명과 OCR텍스트를 보고 **이 상품이 정확히 무엇인지** 먼저 확정하세요.
- 핵심상품명(표준명), 핵심 규격, 핵심 재질, 핵심 기능만 우선 추립니다.
- 카테고리 일반어(`생활철물`, `운동용품`, `주방용품`, `DIY`, `자동차용품`)는 제목에 넣지 마세요.

### 2단계: OCR 수치 필터 (3.0 신규)
OCR 텍스트에는 의미 없는 숫자가 섞여 있습니다. **다음 기준으로 수치 토큰을 걸러내세요**:
- ✅ **유지**: 단위가 붙은 규격 — `35mm`, `2M`, `500ml`, `12V`, `100A`, `3평`, `6인치`, `1/2인치`, `M8`, `Ø15`, `800W`
- ✅ **유지**: 명확한 사이즈/수량 — `2인용`, `4구`, `3단`, `8자루`, `1박스`, `10매`
- ❌ **제외**: 단위 없는 순수 숫자 — `801`, `3457`, `2024`, `12345`, `00123` (제품코드/바코드/넘버링/연도/페이지 번호)
- ❌ **제외**: 가격/원 — `9900원`, `19,900`, `₩15000`
- ❌ **제외**: OCR 깨진 숫자 파편 — `2O2`, `I23`, `O01` 같이 영문/숫자가 혼재되어 의미 불명한 토큰
- 판단 어려우면 **빼는 쪽**으로. 숫자 토큰은 구매 의도와 거의 무관합니다.

### 3단계: A마켓(홈런마켓) 상품명 — **기능/규격 중심 확장형**
A마켓은 홈런마켓 상위노출용 긴 제목입니다. 공백 구분 1줄.
- 목표 길이: **{a_name_min}~{a_name_max}자** (근거 부족 시 더 짧아도 됨)
- **A마켓 구조**: `[핵심상품명 + 규격]` → `[근거 있는 세부 기능 2~4개]` → `[재질/소재]` → `[호환/적용 규격 1~2개]` → `[별칭/동의어 0~1개]`
- **A마켓 포커스**: 기능·규격·재질·호환성. "무엇을 할 수 있는가, 어떤 스펙인가"에 집중.
- 사용처는 A마켓에서는 **최소한만** (0~1개). 사용처는 B마켓의 영역입니다.
- 동의어는 **안전한 실무 유사어 0~1개만** 허용.

### 4단계: A마켓 검색어설정 (Cafe24 태그)
- 붙여쓰기 + 쉼표 구분, 목표 **{a_tag_count}개** (근거 있을 때까지만. 빈 슬롯 허용)
- A 태그는 **identity·규격·기능·재질·호환성** 중심.
- 상품명에 이미 들어간 단어도 **조합 변형**으로 넣어도 됩니다 (예: 상품명 `호스 컷팅밴드 2M`, 태그에 `호스컷팅밴드`, `컷팅밴드2M`).
- ❌ 오타/맞춤법 변형, 무관 카테고리, 과장 문구 금지.

### 5단계: A마켓 검색키워드 (쇼핑몰 상위 키워드)
- 공백 구분, 최대 18개
- 상품명 + A 태그에서 **핵심만** 추립니다.
- 가격 분리 상품군은 **핵심상품명은 동일**하게 유지.

### 6단계: B마켓(준비몰) 상품명 — **용도/사용처 중심 독립 패키징**
⚠️ 3.0부터 B마켓은 A의 단순 축약이 아닙니다. **다른 각도의 패키징**입니다.
- 목표 길이: **{b_name_min}~{b_name_max}자** (짧아도 됨)
- **B마켓 구조**: `[핵심상품명 + 규격]` → `[사용처/용도 2~3개]` → `[대상/상황 1~2개]` → `[대표 기능 1개]` → `[옵션 0~1개]`
- **B마켓 포커스**: "어디서·누가·언제 쓰는가". 욕실/차량/캠핑/사무실/주방 같은 구체적 사용처 중심.
- 핵심상품명·규격은 A와 **공유**하되, 뒷부분 토큰은 A와 **겹치지 않게** 고르세요.
- 단, 근거가 부족한 사용처·대상은 추가 금지. OCR/상품명에 흔적이 있어야 함.
- **A에 이미 들어간 기능/재질 토큰은 B 제목에서 생략**하고, 그 자리를 근거 있는 용도로 채우세요.
- ❌ A마켓에 없는 **새 상품군/새 재질/새 호환 규격**은 창작 금지. 반면 사용처/용도는 근거가 있으면 B에서 새로 도입 허용.

#### 예시 (가구 연결 볼트)
- A마켓(기능/규격): `가구 연결 볼트 35mm 편심 체결 니켈 도금 캠록 M8 나사산 조립 철물 커넥터`
- B마켓(용도/사용처): `가구 연결 볼트 35mm 옷장 서랍장 붙박이장 책상 DIY 조립 가구부속 체결`
- 핵심상품명(`가구 연결 볼트 35mm`)은 공유, A는 기능/재질(편심/니켈/M8/나사산), B는 사용처(옷장/서랍장/붙박이장/책상)로 분리.

### 7단계: B마켓 검색어설정 (준비몰 태그)
- 최대 **{b_tag_count}개**, 붙여쓰기 + 쉼표 구분
- B 태그는 **용도/상황/대상/별칭** 중심. A 태그와의 중복은 최소화.
- A 태그에 이미 있는 identity/규격 중 1~2개는 유지하고, 나머지는 B 고유 용도 토큰으로.

### 8단계: B마켓 검색키워드 (쇼핑몰 상위 키워드)
- 공백 구분, 최대 **{b_tag_count // 2}개**
- B마켓 상품명 + B 태그에서 핵심만. A 검색키워드와 **조합이 달라야** 합니다.

### 9단계: 카테고리 매핑
- 네이버/쿠팡 카테고리는 상품의 실제 정체성 기준.
- A/B 같은 카테고리 코드 사용.

### 띄어쓰기 규칙
- **자연스러운 띄어쓰기 1회만** — 의미 단위로.
- ❌ 붙여쓰기/띄어쓰기 둘 다 넣지 마세요
- ✅ 수식어-상품명: `강력 흡입`, `대용량 배터리`

### 중복 제거 (필수)
1. **완전 중복**: 동일 단어 반복 금지
2. **공백 변형 중복**: `무선청소기` vs `무선 청소기` → 1개만
3. **재조합 중복**: `가을 패딩` vs `패딩 가을` → 1개만
4. **동의어 중복**: 같은 뜻 단어 나열 → 검색량 높은 1개만
5. **A↔B 교차 중복 (3.0 신규)**: A 제목과 B 제목의 뒷부분 토큰이 **50% 이상 겹치면 B를 다시 작성**하세요. 핵심상품명·규격만 공유하고 뒷부분은 각도를 달리합니다.

### 최종 품질 게이트 (저장 전 행별 자체검사)
각 행을 저장하기 전에 아래 기준을 통과하는지 확인하세요. 하나라도 걸리면 해당 행만 다시 고칩니다.
- 상품명에 GS코드, 가격, 배송/할인/이벤트 문구가 남아 있지 않음
- 단위 없는 숫자, 바코드/제품코드 같은 OCR 숫자 파편이 없음
- A마켓은 기능/규격/재질 중심, B마켓은 사용처/용도 중심으로 각도가 분리됨
- A/B 제목의 핵심상품명·규격 이후 토큰이 50% 이상 겹치지 않음
- 검색어설정은 붙여쓰기이며, 같은 뜻/공백변형/재조합 중복이 없음
- 검색어설정을 개수 맞추기용 무관 태그로 채우지 않음. 근거가 약하면 목표 개수보다 적어도 됨
- 다른 행의 상품군/사용처/규격이 섞이지 않음

### 형태 정규화
- 형용사/동사 → 명사형: `강력한` → `강력`, `사용하는` → `사용`
- 연속 공백 제거: 공백 1칸
- 합성어 분리형 우선

### 핵심 원칙 (요약)
1. **핵심상품명 맨 앞** (GS코드 제외)
2. **근거는 상품명+OCR만** — 외부 검색/연관어 쓰지 마세요
3. **OCR 수치 필터 적용** — 단위 없는 순수 숫자 제거
4. **A는 기능/규격, B는 용도/사용처** — 독립 패키징
5. **색상 제외** (옵션)
6. **카테고리명 금지**
7. **오타/표기변형 생성 금지**

### ❌ 반드시 제외할 노이즈
- 상세페이지 템플릿: Product Profile, SIZE, Advantage, Features, Description, Specification
- 영어 마케팅 문구: Premium Quality, Best Seller, Hot Item, Free Shipping
- 배송/정책: 배송, 반품, 교환, AS, 원산지, 연락처
- 판매조건: 무료배송, 할인, 프로모션, 이벤트, 특가, 최저가
- 검증 곤란 수식어(단독 사용 금지): 최고급, 프리미엄, 고품질, 인기상품, 베스트
- 엑셀 컬럼명: 옵션입력, 판매가, 재고수, 이미지URL 등
- **단위 없는 OCR 숫자 토큰** (제품코드·연도·페이지번호·바코드)
- **다른 상품의 키워드 혼입 금지**
"""
    else:
        rules_block = f"""## 키워드 생성 규칙

### 전체 원칙
- **evidence-first**: 상품명 → OCR/Vision → 검색데이터 순으로만 근거를 사용하세요.
- **짧아도 됩니다.** 길이를 채우기 위해 무관 키워드를 넣지 마세요.
- **오타/맞춤법/띄어쓰기 변형을 검색 커버리지 목적으로 생성하지 마세요.**
- **같은 뜻 중복 금지**: 비슷한 표현을 여러 개 늘어놓지 마세요.
- **새 카테고리/새 상품군/새 사용처 도입 금지**: 입력 근거가 없으면 비워두세요.

### 1단계: identity 확정
- 원본 상품명과 OCR/Vision을 보고 **이 상품이 정확히 무엇인지** 먼저 확정하세요.
- 핵심상품명, 핵심 규격, 핵심 재질, 핵심 기능만 우선 추립니다.
- 카테고리 일반어(`생활철물`, `운동용품`, `주방용품`)는 제목에 넣지 마세요.

### 2단계: A 제목 생성
- 공백 구분 키워드 1줄로 작성합니다.
- 목표 길이는 **{a_name_min}~{a_name_max}자**지만 강제하지 않습니다. 근거가 부족하면 더 짧아도 됩니다.
- 구조: **[핵심상품명 + 규격] → [근거 있는 기능 1~3개] → [근거 있는 사용처 0~2개] → [재질/옵션]**
- A마켓도 **근거 있는 범위 안에서만** 확장하세요. 길이 채우기용 보강 금지.
- 동의어는 **안전한 실무 유사어 0~1개만** 허용합니다.

### 3단계: A 태그 생성
- 검색어설정은 붙여쓰기 + 쉼표 구분입니다.
- 목표 수량은 **{a_tag_count}개**지만, **반드시 채우려고 하지 마세요.** 근거 있는 후보만 사용하세요.
- 오타/표기변형/무관 카테고리/과장 문구 금지.
- 상품명에서 다 담지 못한 **온토픽 identity/기능/규격** 중심으로 배치하세요.

### 4단계: B 제목 생성
- B마켓 제목은 **A 제목의 부분집합 + 최소한의 표기 정리**입니다.
- 새 키워드를 창작하지 말고, A 제목보다 **짧고 선명하게** 만드세요.
- 구조: **[핵심상품명 + 규격] → [기능 0~2개] → [사용처 0~1개] → [옵션]**

### 5단계: B 태그 생성
- B 태그는 A 태그에서 **핵심 {b_tag_count}개 이내**만 고릅니다.
- identity/규격/핵심 기능 우선. usage/material/synonym 중복은 줄이세요.

### 6단계: 카테고리 매핑
- 마지막에 네이버/쿠팡 카테고리를 매핑합니다.
- 카테고리는 상품의 실제 정체성을 기준으로 고르고, 제목 길이를 늘리기 위한 카테고리 단어는 넣지 마세요.

### 띄어쓰기 규칙 (중요)
- **자연스러운 띄어쓰기 1회만** 사용 — 의미 단위로 띄어쓰기
- ❌ 붙여쓰기/띄어쓰기 둘 다 넣지 마세요 (`무선청소기, 무선 청소기` → `무선 청소기`만)
- ❌ 띄어쓰기 과다/과소 모두 비권장
- ✅ 수식어-상품명 띄어쓰기: `강력 흡입`, `대용량 배터리`
- ✅ 브랜드-제품명 띄어쓰기: `노루 페인트`

### 중복 제거 규칙 (필수)
아래 4종 중복을 반드시 제거하세요. **슬롯 낭비이자 어뷰징 위험**입니다.
1. **완전 중복**: 동일 단어 반복 (`기모, 기모` → `기모`)
2. **공백 변형 중복**: 붙/띄만 다른 표현 (`무선청소기` vs `무선 청소기` → 띄어쓰기 형태 1개만)
3. **재조합 중복**: 같은 단어 순서만 바꾼 것 (`가을 패딩` vs `패딩 가을` → 1개만)
4. **동의어 중복**: 같은 의미 단어 나열 (`핸드폰/휴대폰/스마트폰` → 검색량 높은 1개만)

### 형태 정규화
- **형용사/동사 → 명사형**: `강력한` → `강력`, `사용하는` → `사용`, `튼튼한` → `내구성`
- **연속 공백 제거**: 공백은 항상 1칸
- **합성어 분리형 우선**: 필요시 합성어 1개만 허용 (`유리롤러 레일롤러` → `유리 레일 롤러`)

### 핵심 원칙
1. **핵심상품명 맨 앞**: 원본 상품명의 핵심 단어(GS코드 제외)를 상품명 맨 앞에 배치하세요.
2. **evidence-first**: 근거 없는 사용처/재질/기능/대상 생성 금지.
3. **짧아도 허용**: 길이보다 온토픽 여부가 우선입니다.
4. **오타/표기변형 생성 금지**: 검색 커버리지 목적의 typo-expansion 금지.
5. **색상 제외**: 색상은 옵션이므로 키워드에서 제외합니다.
6. **카테고리명 금지**: 카테고리 일반어로 제목을 부풀리지 마세요.

### ❌ 반드시 제외할 노이즈
아래 항목은 OCR 텍스트에 포함되어 있더라도 **절대 키워드에 사용하지 마세요**:
- 상세페이지 템플릿: Product Profile, SIZE, Advantage, Features, Description, Specification
- 영어 마케팅 문구: Premium Quality, Best Seller, Hot Item, Free Shipping
- 배송/정책: 배송, 반품, 교환, AS, 원산지, 연락처
- 판매조건: 무료배송, 할인, 프로모션, 이벤트, 특가, 최저가
- 검증 곤란 수식어(단독 사용 금지): 최고급, 프리미엄, 고품질, 인기상품, 베스트
- 엑셀 컬럼명: 옵션입력, 판매가, 재고수, 이미지URL 등 데이터 필드명
- **다른 상품의 키워드를 섞지 마세요** — 각 행은 독립된 상품입니다

### 3단계: 검색어설정 (Cafe24 태그)
- **띄어쓰기 없이 붙여쓰기**, 쉼표로 구분합니다.
- 목표는 **{a_tag_count}개**지만, 개수를 맞추려고 무관 태그를 넣지 마세요.
- 상품명에서 다 못 넣은 **온토픽 identity/기능/규격** 위주로 배치하세요.
- ❌ 오타/표기변형/맞춤법변형 생성 금지
- ❌ 상품명과 완전히 동일한 조합 반복 금지
- ❌ 새 카테고리/새 사용처 도입 금지

### 4단계: 검색키워드
- 네이버 쇼핑 검색용 상위 키워드
- 공백 구분, 최대 18개
- 상품명 + 검색어설정에서 핵심만 추려서 배치
"""

    if keyword_version == "3.0":
        b_market_block = f"""## B마켓 시트 (필수 — 반드시 생성)
결과 엑셀에 **`B마켓`** 시트를 **반드시** 추가하세요. `분리추출후` 시트를 복사한 뒤, 위 **3.0 타겟형 규칙 6~8단계**에 따라 상품명/검색어설정/검색키워드를 **용도/사용처 각도로 독립 작성**합니다.
⚠️ B마켓 시트가 없거나, A 제목 뒷부분과 50% 이상 겹치면 작업 미완료로 간주합니다.

### B마켓 핵심 원칙 (3.0)
- B마켓은 **A의 부분집합이 아닙니다.** 핵심상품명·규격만 공유하고, 뒷부분 토큰은 **용도/사용처/대상** 중심으로 **독립 패키징**하세요.
- 단, 근거(상품명+OCR)에 흔적이 있는 사용처만 도입 가능. 상식에 기대어 창작 금지.
- A마켓에 없는 **새 상품군/새 재질/새 호환 규격**은 B에서도 금지. 새로 도입 허용은 **사용처/용도 축**뿐.

### B마켓 상품명
- **글자수 목표**: {b_name_min}~{b_name_max}자 (짧아도 됨)
- **구조**: [핵심상품명 + 규격] → [사용처 2~3개] → [대상/상황 1~2개] → [대표 기능 1개]
- **예시**:
  - A마켓: `가구 연결 볼트 35mm 편심 체결 니켈 도금 캠록 M8 나사산 조립 철물 커넥터`
  - B마켓: `가구 연결 볼트 35mm 옷장 서랍장 붙박이장 책상 DIY 조립 가구부속`

### B마켓 검색어설정
- 최대 **{b_tag_count}개**, 붙여쓰기 + 쉼표 구분
- B 태그는 **용도/상황/대상/별칭** 중심. A 태그와의 중복은 최소화.

### B마켓 검색키워드
- 공백 구분, 최대 **{b_tag_count // 2}개**
- A 검색키워드와 **조합이 달라야** 합니다."""
    else:
        b_market_block = f"""## B마켓 시트 (필수 — 반드시 생성)
결과 엑셀에 **`B마켓`** 시트를 **반드시** 추가하세요. `분리추출후` 시트를 복사한 뒤, 아래 "코어 버전" 규칙으로 상품명/검색어설정/검색키워드를 **별도로** 작성합니다.
⚠️ B마켓 시트가 없으면 작업 미완료로 간주합니다.

### 코어 버전 핵심 원칙
코어 버전은 **새로 생성이 아니라, A마켓(풀 버전)의 선택/축약**입니다.
- 풀 버전 토큰의 **부분집합** + 최소한의 표기 정리로 만드세요
- 근거 없는 새 키워드를 추가하지 마세요 (풀 버전에 없는 단어 금지)
- 제품 정체성(핵심상품명)을 반드시 유지하세요

### B마켓 상품명 작성 규칙
- **글자수 목표**: {b_name_min}~{b_name_max}자. 하지만 근거가 부족하면 더 짧아도 됩니다.
- **구조 (의미 순서 더 엄격)**: [핵심상품명 + 규격] → [기능 0~2개] → [사용처 0~1개]
- B마켓은 **압축형**: A마켓에서 검증된 토큰만 남겨 짧고 선명하게 만드세요.
- A마켓에 없는 새 토큰, 새 사용처, 새 재질을 추가하지 마세요.
- **옵션**: 있으면 맨 끝에 1개만
- **예시**:
  - A마켓: `가구 연결 볼트 35mm 편심 체결 고정 조립 니켈 도금 캠록 801 커넥터 캐비닛 서랍장 붙박이장 조립가구 부속`
  - B마켓: `가구 연결 볼트 35mm 결합 캠 잠금 옷장 선반 가구부속` (A에서 안 쓴 토큰 위주)

### B마켓 검색어설정 (쿠팡 태그)
- 최대 **{b_tag_count}개**
- A마켓 검색어에서 **핵심 identity/기능/규격만 선별**하세요.
- 수량을 채우기 위한 synonym/usage/material 중복 보강 금지

### B마켓 검색키워드 (네이버태그)
- 최대 **{b_tag_count // 2}개** (A마켓보다 적음)
- A마켓 검색키워드에서 핵심만 추려서 배치"""

    if keyword_version != "1.0":
        rules_block = f"""## 키워드 생성 규칙 (2.0/3.0 통합 검색형)

### 전체 목적
- 상품을 설명하는 문장을 쓰는 것이 아니라, 실제 구매자가 검색창에 입력할 가능성이 높은 단어 중심으로 상품명/태그/검색키워드를 재구성합니다.
- 모든 단어는 **"사람이 이 단어를 검색창에 직접 칠까?"** 기준을 통과해야 합니다.
- 상품명은 문장이 아니라 검색어 조합 구조입니다.
- 목표 글자 수를 채우려고 단어를 늘리지 마세요. 길이보다 검색어 밀도, 정확성, 구매자 검색 가능성이 우선입니다.
- 근거 있고 검색 가능성이 높은 후보가 충분하면 기존보다 약 5개 더 많은 단어를 담습니다.
- 결과물은 한국어 중심으로 작성합니다. 영어/로마자/중문 표현은 제거하고, 한국어 대체어가 명확한 경우만 한국어 검색어로 바꿉니다. 단, `35mm`, `2M`, `500ml`, `12V`, `M8`처럼 상품 규격으로 확실한 숫자+단위는 유지할 수 있습니다.

### 근거 우선순위
1. 원본 상품명
2. OCR/Vision 텍스트
3. 검색데이터/자동완성은 우선순위 참고용만

- 검색데이터는 키워드 순서를 조정하는 참고 자료일 뿐입니다.
- 상품명/OCR/Vision에 없는 새 상품군, 재질, 규격, 사용처, 호환성을 검색데이터만 보고 만들지 마세요.
- 근거가 약하면 짧게 끝내는 것이 맞습니다.

### 상품 정체성 확정
- GS코드를 제거한 뒤 핵심상품명, 규격, 재질, 기능을 먼저 뽑습니다.
- 핵심상품명은 구매자가 가장 많이 검색할 법한 일반 검색어로 정리합니다.
- 원본에 있더라도 검색어 가치가 낮은 내부 표현, 제조사식 표현, 번역투 표현은 우선순위를 낮춥니다.
- 카테고리 일반어(`생활철물`, `DIY`, `자동차용품`, `주방용품`)로 제목을 부풀리지 마세요.
- 근거가 약하면 핵심상품명 + 확인 가능한 스펙만 사용합니다.

### A마켓 상품명
- 목적: 기능, 규격, 재질, 호환성 중심의 검색형 상품명.
- A마켓은 **무엇이고 어떤 스펙인가**에 집중합니다.
- 구조: `핵심상품명 + 규격` → `대표 기능 1~4개` → `재질/소재` → `호환/적용 규격` → `안전한 별칭 0~1개`
- 핵심상품명과 대표 검색어는 반드시 앞쪽에 둡니다.
- 사용처는 최소화합니다.
- 설명형 문장 금지.
- 목표 길이는 **{a_name_min}~{a_name_max}자**지만, 근거가 부족하거나 검색어 가치가 낮으면 짧아도 됩니다.
- 단어 수는 기존보다 약 5개 더 담되, 검색 가능성이 낮은 단어로 억지 보강하지 않습니다.

### B마켓 상품명
- 목적: 용도, 사용처, 대상 중심의 검색형 상품명.
- B마켓은 **누가 어디에 쓰는가**에 집중합니다.
- 구조: `핵심상품명 + 규격` → `사용처/용도` → `대상/상황` → `대표 기능 1개` → `옵션`
- A와 핵심상품명, 규격은 공유합니다.
- B는 A보다 사용 장면과 구매자 검색어에 더 가깝게 작성합니다.
- 단, 근거 없는 새 상품군, 재질, 호환 규격은 만들지 않습니다.
- 새로 허용되는 것은 OCR/Vision/상품명에 흔적이 있는 사용처/용도뿐입니다.
- A 뒤쪽 토큰과 B 뒤쪽 토큰이 50% 이상 겹치면 B를 다시 작성합니다.
- 목표 길이는 **{b_name_min}~{b_name_max}자**지만, 근거가 부족하거나 검색어 가치가 낮으면 짧아도 됩니다.
- 단어 수는 기존보다 약 5개 더 담되, A와 중복되는 뒷부분 토큰으로 채우지 않습니다.

### 검색어설정 / 태그
- 태그는 붙여쓰기 + 쉼표 구분으로 작성합니다.
- 태그도 실제 검색 가능성이 있는 단어만 사용합니다.
- 단순 설명어, 내부 용어, 번역투 표현, 과한 세부 표현은 제외합니다.
- A태그: identity, 규격, 기능, 재질, 호환성 중심으로 최대 **{a_tag_count}개**.
- B태그: 용도, 상황, 대상, 별칭 중심으로 최대 **{b_tag_count}개**.
- 목표 개수를 억지로 채우지 마세요.
- 오타, 맞춤법 변형, 공백 변형, 무관 카테고리 태그는 금지합니다.
- 같은 의미의 태그는 하나만 남깁니다.
- 상품명과 너무 동떨어진 태그는 금지합니다.

### 검색키워드
- 검색키워드는 공백 구분으로 작성합니다.
- A검색키워드는 A상품명 + A태그에서 핵심 검색어만 최대 18개 추출합니다.
- B검색키워드는 B상품명 + B태그에서 핵심 검색어만 최대 **{b_tag_count // 2}개** 추출합니다.
- 상품 정체성, 핵심 기능, 규격, 재질, 사용처 중심으로 구성합니다.
- 설명형 단어보다 실제 검색어로 쓰일 가능성이 높은 단어를 우선합니다.
- 가격 분리 상품은 핵심상품명은 유지하고 뒤쪽 조합만 다르게 합니다.
- 같은 키워드 반복 금지.
- 검색 가능성이 낮은 단어는 제외합니다.

### OCR 숫자 필터
- 유지: `35mm`, `2M`, `500ml`, `12V`, `M8`, `3단`, `10매`, `4구`처럼 상품 규격으로 볼 수 있는 숫자.
- 제외: `801`, `3457`, `2024`, `12345`, 가격, 바코드, 페이지 번호, 깨진 숫자 파편.
- 판단이 애매하면 제외합니다.
- 숫자가 상품 규격인지 확실하지 않으면 상품명에 넣지 않습니다.

### 무조건 제외
- GS코드
- 배송/반품/교환/AS/무료배송/할인/이벤트/특가
- `프리미엄`, `고품질`, `베스트`, `인기`, `추천` 같은 검증 곤란 수식어
- 상세페이지 템플릿 문구
- 엑셀 컬럼명
- 다른 행 상품의 키워드
- 상품과 직접 관련 없는 카테고리 일반어
- 영어, 로마자, 중문 표현
- 구매자가 검색하지 않을 가능성이 높은 내부 관리 표현

### 최종 품질 게이트
- A마켓상품명은 기능/규격/재질/호환성 중심인가?
- B마켓상품명은 용도/사용처/대상 중심인가?
- 핵심상품명과 대표 검색어가 앞쪽에 있는가?
- 상품명, OCR/Vision에 없는 새 근거를 만들지 않았는가?
- 검색어설정은 붙여쓰기이며 중복/공백변형/오타변형이 없는가?
- 검색키워드는 실제 검색 가능성이 높은 핵심어만 남겼는가?
- 제외한 단어와 판단근거를 검토해도 설명 가능한가?
"""

        b_market_block = f"""## B마켓 시트 (필수 — 반드시 생성)
결과 엑셀에 **`B마켓`** 시트를 **반드시** 추가하세요. `분리추출후` 시트를 복사한 뒤, 위 통합 검색형 규칙에 따라 상품명/검색어설정/검색키워드를 **용도·사용처·대상 각도**로 별도 작성합니다.
⚠️ B마켓 시트가 없거나, A 제목 뒷부분과 50% 이상 겹치면 작업 미완료로 간주합니다.

### B마켓 작성 필드
- 상품명: 핵심상품명 + 규격을 유지하고, 뒷부분은 용도/사용처/대상 중심으로 작성
- 검색어설정: 최대 **{b_tag_count}개**, 붙여쓰기 + 쉼표 구분
- 검색키워드: 공백 구분, 최대 **{b_tag_count // 2}개**
- 상품 상세설명: A마켓 전용 상세태그만 제거하고 원본 이미지 태그는 유지

### B마켓 금지
- A마켓에 없는 새 상품군/새 재질/새 호환 규격 창작 금지
- 근거 없는 사용처/대상 추가 금지
- A태그와 같은 조합을 개수 맞추기용으로 반복 금지"""

    content = f"""# 키워드 생성 지시서

## 실행 모드 (필수)
- **확인 질문 없이 끝까지 자동 실행하세요.** 절대 중간에 멈추지 마세요.
- "진행할까요?", "맞나요?", "이렇게 할까요?" 등의 질문을 하지 마세요.
- 엑셀 읽기 → 모든 행 키워드 채우기 → 결과 파일 저장까지 한 번에 완료하세요.
- 완료 후 처리 건수와 결과 파일 경로만 출력하세요.
- **모든 파일 경로는 이 지시서가 있는 폴더 기준 상대 경로입니다.** 절대 경로를 사용하지 마세요.

## 작업 요약
같은 폴더에 있는 엑셀 파일(`{upload_basename}`)을 읽고, 각 상품의 **상품명** 과 **검색어설정** 컬럼을 채워서 새 엑셀로 저장하세요.

## 현재 버전
- **{version_label}**
- 결과 저장 경로: `{result_rel_path}`

## 일관성 고정 규칙 (Codex 재실행 안정화)
같은 입력 파일을 Codex에서 3회 이상 다시 실행해도 마켓별 결과의 느낌과 핵심 토큰이 거의 같아야 합니다.

### 고정 프로필
- 쿠팡A: 노련한 쿠팡형. 핵심상품명 정석 시작, 가장 풍부하게 작성합니다. 정체성, 기능, 규격, 소재, 구성품, 사용처를 넓게 담습니다.
- 쿠팡B: 쿠팡 용도형. 사용처/문제해결/상황으로 시작을 다르게 하되, 쿠팡A의 80~100% 길이를 유지합니다.
- 네이버A: 프리미엄 검수형. 핵심상품명과 대표 검색어를 앞에 두고 중복을 강하게 줄입니다.
- 네이버B: 실사용자 검색형. 구매자가 검색창에 입력할 법한 용도/상황 검색어를 앞쪽에 둡니다.
- ESMA: 전환율 중시형. 구매 이유, 문제해결, 기능 중심으로 작성합니다.
- ESMB: 테마 확장형. 사용 상황, 작업 테마, 사용자 맥락 중심으로 작성합니다.
- 롯데ON A: 마켓 균형형. 상품명 중심으로 시작하되 네이버보다 풍부하게 작성합니다.
- 롯데ON B: 용도 확장형. 사용처/상황으로 시작을 바꾸고 롯데ON A와 뒷부분 중복을 줄입니다.
- 11번가A: 균형형. 상품명, 기능, 용도를 고르게 배치합니다.
- 11번가B: 최소 중복 엄격형. 합성어 안에 들어간 반복 헤드 단어를 가장 강하게 줄입니다.

### 변동폭 제한
- 같은 상품을 다시 처리해도 각 마켓별 핵심상품명, 대표 기능, 대표 용도는 유지합니다.
- 재실행 시 바뀔 수 있는 것은 뒤쪽 보조 토큰의 순서 1~3개 정도입니다.
- 새 후보를 떠올리기보다 동일한 후보 풀에서 같은 프로필 순서로 다시 선택하세요.
- 쿠팡B는 짧은 보조형이 아닙니다. 쿠팡A와 비슷한 길이로 작성하되 시작 각도만 다르게 합니다.
- 제목 시작은 전부 동일하게 만들지 마세요. 단, 핵심상품명 범위를 벗어나면 안 됩니다.

### 3회 자체검산
각 행 저장 전 내부적으로 같은 상품을 3회 다시 만든다고 가정하고 아래를 확인하세요.
1. 세 결과의 핵심상품명과 대표 토큰이 같은가?
2. 쿠팡A/B, 네이버A/B, ESM A/B, 롯데ON A/B, 11번가 A/B의 시작 각도가 서로 다른가?
3. 색상, 배송, 가격, 리뷰, 판매자 문구가 제거되었는가?
4. 같은 헤드 단어(경첩, 깔망, 손잡이, 버클, 브라켓, 스티커, 거치대 등)가 과반복되지 않았는가?
5. 결과가 크게 흔들리면 새로 만들지 말고 1회차 후보 풀 기준으로 재정렬만 하세요.

## 대량 무검수 생성 규칙
이번 작업은 수백 개 상품을 한 번에 돌리는 전제입니다. 사용자가 행별로 검수하지 않아도 바로 쓸 수 있는 수준으로 스스로 판단하세요.

### OCR 기반 시드 확장
- OCR에서 상품을 유추할 때 먼저 후보를 5개 버킷으로 분류합니다: `정체성`, `기능`, `용도/사용처`, `소재/규격`, `문제해결/구매상황`.
- OCR에 반복 노출되는 단어, 상세페이지 제목/스펙 표에 가까운 단어, 상품명과 OCR 양쪽에 모두 보이는 단어를 우선합니다.
- `상세 정보`, `주의 사항`, `리뷰`, `배송`, `문의`, `사용 흔적`, `교환 반품`, `원산지`, `수입원`, `색상`, `옵션 선택`은 후보 풀에서 제외합니다.
- 상품 정체성이 애매하면 가장 구체적인 물건명 1개를 고르고, 같은 행 안에서는 그 정체성을 끝까지 유지합니다.

### OCR 상품정체성 확정
- 가장 중요한 1단계는 OCR로 `이게 정확히 어떤 물건인지` 확정하는 것입니다. 키워드 확장보다 상품정체성 판단을 먼저 끝내세요.
- OCR에서 `브라켓`, `마운트`, `니플`, `유니온`, `패드`, `테이프`, `커넥터`, `가스켓`, `고리`, `볼트`, `너트`, `밑창`처럼 제품군을 가리키는 단어가 보이면 정체성 후보로 우선 승격합니다.
- 상품명과 OCR이 충돌하면 원본 상품명을 그대로 믿지 말고, OCR 반복어와 이미지/스펙 문맥을 기준으로 실제 상품군을 재판단합니다.
- 최종 후보는 반드시 `정체성 1개 + 기능 1개 + 용도/사용처 1개` 이상으로 설명 가능해야 합니다. 이 3개가 맞지 않으면 키워드를 늘리지 말고 정체성부터 다시 고릅니다.
- 정체성 후보가 2개 이상이면 더 구체적인 조합을 선택합니다. 예: `브라켓`보다 `차량 조명 브라켓`, `테이프`보다 `신발 밑창 테이프`, `너트`보다 `브레이크 니플 유니온`.
- OCR 단어가 많아도 실제 물건명과 관련 없는 설명문, 주의문, 배송문, 옵션명은 정체성 후보로 쓰지 않습니다.

### 검색가능성 점수
후보를 고를 때 아래 순서로 점수를 높게 봅니다.
1. 구매자가 검색창에 직접 칠 법한 표준 상품명
2. 상품명과 OCR에 동시에 등장하는 핵심어
3. 설치/교체/수리/정리/보관/고정/방지처럼 구매 의도가 드러나는 기능어
4. 사용처가 명확한 단어
5. 규격/소재가 상품 선택에 직접 영향을 주는 단어

점수를 낮게 봅니다: 형용사, 감성어, 장식어, 설명문 조각, 너무 넓은 카테고리명, 색상, 배송/판매 문구, 리뷰 문구, 단위 없는 숫자.

### 쿠팡 확장 우선 규칙
- 쿠팡A/B는 전체 마켓 중 가장 풍부하게 작성합니다. 네이버보다 길고, ESM/롯데ON/11번가보다도 OCR 기능 단서를 더 적극 반영합니다.
- OCR에서 상품정체성이 확정된 뒤 `각도 조절`, `볼트 고정`, `라인 연결`, `배관`, `정밀 나사산`, `셀프수선 접착`, `마모 방지`처럼 구매 의도가 있는 기능 조합이 보이면 쿠팡에 우선 포함합니다.
- 쿠팡A는 `용도/상황/문제해결` 시작, 쿠팡B는 `상품군/기능/소재` 시작처럼 서로 다른 순서로 배치합니다.
- 단, 쿠팡을 늘리더라도 실제 상품군과 맞지 않는 무관 확장어는 넣지 않습니다. 예: 트럭 고정고리에 무관한 외부조명, 볼트고리, 나사고리는 제외합니다.

### 10개 결과 차별화
- 한 상품에서 만들어지는 10개 계열은 모두 같은 키워드 풀을 공유하지만, 시작 각도와 뒤쪽 토큰 조합은 달라야 합니다.
- 10개 중 7개 이상이 같은 첫 2단어로 시작하면 실패입니다. 같은 상품명 계열이 필요해도 첫 4단어 안에서 기능/용도/규격 위치를 바꿉니다.
- 쿠팡A/B는 둘 다 풍부해야 하며, B는 A의 축약본이 아니라 용도/문제해결 시작의 별도 장문입니다.
- 네이버A/B는 짧지만 서로 다른 검색 의도를 가져야 합니다. A는 정체성, B는 실사용 검색어입니다.
- ESM/롯데ON/11번가는 네이버보다 풍부하고 쿠팡보다 절제합니다. 각각 시작어와 강조축을 다르게 둡니다.
- 결과가 서로 비슷하면 새 단어를 창작하지 말고 같은 후보 풀 안에서 `정체성 시작`, `기능 시작`, `용도 시작`, `문제해결 시작`, `규격 시작`으로 재배열합니다.

## ⚠️ 가장 중요: 상품명 = 핵심상품명이 맨 앞
상품명 컬럼이 최종 쇼핑몰에 노출되는 핵심 필드입니다.
반드시 **원본 상품명에서 추출한 핵심 상품명을 맨 앞에** 배치하고, 나머지 키워드를 뒤에 붙이세요.

## 입력 파일 구조
### 시트1: `분리추출후` (상품 데이터)
- **상품명**: 원본 상품명 (GS코드9자리 포함) — 이걸 키워드가 포함된 상품명으로 **교체**
  - 원본 예시: `GS0700704 끝검 스테인리스 호스 컷팅밴드 2M`
  - ✅ 변경 예시: `스테인리스 호스 컷팅밴드 2M 파이프 배관 고정 연결 롤형 자유절단 호스밴드`
  - ❌ 잘못된 예: `호스 파이프 배관 자동차 오토바이 스테인리스 호스클램프 컷팅밴드 생활철물 DIY수리 교체용 보수용` (핵심상품명이 묻히고 카테고리 나열)
  - ❌ GS코드(GS0700704)는 상품명에서 **제거**
- **검색어설정**: 빈 컬럼 — Cafe24 검색어용 (쉼표 구분)
- **검색키워드**: 빈 컬럼 — 네이버 검색용 (공백 구분)

### 시트2: `OCR결과` (OCR 텍스트 — 반드시 참조)
- **GS코드**: 상품 매칭 키 (분리추출후 시트의 상품명에서 GS코드를 추출하여 매칭)
- **OCR텍스트**: 상품 이미지에서 OCR로 추출한 원본 텍스트
- **대표이미지수**: 이미지 수

OCR결과 시트의 텍스트를 반드시 참고하여 상품의 실제 특성(재질, 규격, 용도 등)을 파악하세요.

{rules_block}

## 행별 최종 출력 개념
엑셀에는 아래 개념을 각 시트/컬럼에 대응해 저장하세요.

A마켓상품명: `분리추출후` 시트의 `상품명`
B마켓상품명: `B마켓` 시트의 `상품명`
A태그: `분리추출후` 시트의 `검색어설정`
B태그: `B마켓` 시트의 `검색어설정`
A검색키워드: `분리추출후` 시트의 `검색키워드`
B검색키워드: `B마켓` 시트의 `검색키워드`
제외한단어: 내부 점검용으로 사용하되, 결과 컬럼이 있으면 기록
판단근거: 내부 점검용으로 사용하되, 결과 컬럼이 있으면 기록

## 병렬 처리 안내
이 파일이 여러 개로 분할되어 있을 수 있습니다 (chunk_01, chunk_02 등).
각 파일은 독립적으로 처리 가능하며, **별도의 Codex/ChatGPT 세션에서 병렬 실행**하면 빠릅니다.

## 출력
- **원본 엑셀과 동일한 구조** 유지, 키워드 관련 컬럼만 채워서 저장
- 저장 경로: `{llm_result_rel}/` (이 지시서와 같은 폴더 안의 하위 폴더)
- **파일명 규칙**: 입력 파일명에서 `.xlsx`를 `_llm.xlsx`로 바꿔서 저장
  - 예: `chunk_01_20260323.xlsx` → `{llm_result_rel}/chunk_01_20260323_llm.xlsx`
  - 예: `chunk_05_20260323.xlsx` → `{llm_result_rel}/chunk_05_20260323_llm.xlsx`
- ⚠️ 다른 청크 파일의 결과를 덮어쓰지 마세요. 반드시 입력 파일명 기준으로 저장하세요.

{b_market_block}

### B마켓 상품 상세설명 (중요)
- `분리추출후` 시트의 상품 상세설명에 `<img src='https://gi.esmplus.com/rkghrud/1.jpg' />` 같은 **상세태그**가 삽입되어 있을 수 있습니다
- B마켓 시트에서는 이 상세태그를 **반드시 제거**하세요
- 상세태그는 보통 `<center>` 바로 뒤 또는 HTML 맨 앞에 위치합니다
- 원본 상품 이미지 태그(`esmplus.com/...goodsellers/...`)는 그대로 유지하고, 추가 삽입된 상세태그만 제거

### B마켓 기타 컬럼
- 상품명/검색어설정/검색키워드/상품 상세설명 외의 모든 컬럼은 `분리추출후` 시트와 **동일하게 유지**

## 5단계: 마켓별 카테고리 매칭 파일 생성 (필수)
키워드 결과 엑셀을 저장한 뒤, 같은 입력 파일 기준으로 **상품코드/상품명별 마켓 카테고리 매칭 파일**도 추가 생성하세요.

### 출력 파일
- 기본 결과 폴더: `{llm_result_rel}/`
- 파일명 규칙: 입력 파일명 기준으로 `_category_match_{version_slug}.xlsx`를 붙입니다.
  - 예: `상품전처리GPT_v3_0_20260427_124037.xlsx` → `{llm_result_rel}/상품전처리GPT_v3_0_20260427_124037_category_match_{version_slug}.xlsx`
  - 예: `chunk_01_20260427.xlsx` → `{llm_result_rel}/chunk_01_20260427_category_match_{version_slug}.xlsx`
- 이 카테고리 파일은 키워드 결과 엑셀과 별도 파일입니다. 키워드 결과 파일을 덮어쓰지 마세요.

### 카테고리 기준표 위치
우선순위대로 사용하세요.
1. `category_reference/` 폴더의 CSV 파일
2. 같은 폴더의 CSV 파일
3. 기존 호환 파일 `naver_category_tree.txt`, `coupang_category_tree.txt`

사용 가능한 기준표:
- `category_reference/naver_categories.csv`
- `category_reference/coupang_categories.csv`
- `category_reference/11st_categories.csv`
- `category_reference/lotteon_categories.csv`
- `category_reference/lotteon_standard_categories.csv`
- `category_reference/lotteon_display_categories.csv`
- `category_reference/auction_categories.csv`
- `category_reference/esm_auction_gmarket_category_matching.csv`

### 카테고리 매칭 파일 컬럼
카테고리 매칭 엑셀에는 최소 아래 컬럼을 포함하세요.
- `상품코드`
- `상품명`
- `네이버카테고리코드`
- `네이버카테고리경로`
- `쿠팡카테고리코드`
- `쿠팡카테고리경로`
- `11번가카테고리코드`
- `11번가카테고리경로`
- `롯데ON표준카테고리코드`
- `롯데ON표준카테고리경로`
- `롯데ON전시카테고리코드`
- `롯데ON전시카테고리경로`
- `옥션카테고리코드`
- `옥션카테고리경로`
- `G마켓카테고리경로`
- `ESM카테고리경로`
- `확신도`
- `검수필요`
- `매칭근거`
- `마켓플러스검증상태`
- `마켓플러스차단마켓`
- `마켓플러스검증메모`
- `G마켓옵션위험`
- `롯데ON옵션위험`
- `옵션검수필요`

### 결과 엑셀에도 직접 쓸 컬럼
키워드 결과 엑셀의 `분리추출후` 시트와 `B마켓` 시트에는 업로드에서 바로 쓰는 아래 컬럼도 추가/갱신하세요.
- `네이버카테고리코드`
- `쿠팡카테고리코드`

다른 마켓(11번가/롯데ON/옥션/ESM)은 우선 별도 카테고리 매칭 파일에만 기록합니다.

### 매칭 기준
1. **외부 검색 금지**: 상품명, OCR텍스트, 생성한 키워드, 제공된 카테고리 기준표만 사용하세요.
2. **상품의 실제 정체성 기준**: 제목에 들어간 마케팅 단어보다 핵심 상품 유형/기능/용도를 우선합니다.
3. **가장 구체적인 leaf 카테고리**를 선택하세요. 상위 카테고리로 뭉뚱그리지 마세요.
4. A마켓과 B마켓은 같은 상품이므로 같은 카테고리 매칭값을 사용하세요.
5. 확신도가 낮으면 `검수필요`를 `Y`로 표시하고 `매칭근거`에 애매한 이유를 적으세요.
6. 롯데ON은 `category_type=standard` 행에서 표준카테고리, `category_type=display` 행에서 전시카테고리를 각각 고르세요.
7. ESM 매칭표는 `사이트=G마켓` 행의 `G/A 카테고리명`을 `G마켓카테고리경로`에 기록하고, 원본 `ESM 카테고리명`은 `ESM카테고리경로`에 함께 기록하세요.

### 마켓플러스 전송 사전검증
카테고리 매칭 파일 생성 시 마켓플러스 일괄보내기에서 실패할 가능성이 높은 항목을 함께 표시하세요.

- `마켓플러스검증상태`: `PASS`, `WARN`, `BLOCK` 중 하나를 기록합니다.
- `마켓플러스차단마켓`: 자동 전송을 막아야 할 마켓명을 쉼표로 기록합니다. 없으면 빈칸으로 둡니다.
- `마켓플러스검증메모`: 경고/차단 이유를 짧게 기록합니다.
- `G마켓옵션위험`, `롯데ON옵션위험`, `옵션검수필요`: `Y` 또는 `N`으로 기록합니다.

검증 기준:
1. 상품명이 100자를 초과하거나 초과 위험이 크면 `WARN` 또는 `BLOCK`으로 표시하세요.
2. 옵션명/옵션값이 한글 25자 또는 50바이트를 넘을 가능성이 있으면 `옵션검수필요=Y`로 표시하세요.
3. 옵션 구조가 자유명명형이고 G마켓/옥션 카테고리별 권장 옵션명과 맞지 않을 가능성이 있으면 `G마켓옵션위험=Y`로 표시하세요.
4. 롯데ON 표준카테고리 또는 전시카테고리가 비어 있으면 `롯데ON옵션위험=Y`, `마켓플러스검증상태=BLOCK`, `마켓플러스차단마켓=롯데ON`으로 표시하세요.
5. 롯데ON 카테고리별 옵션명 매칭이나 직접입력/선택입력 판단이 필요해 보이면 `롯데ON옵션위험=Y`와 `옵션검수필요=Y`로 표시하세요.
6. 확실히 자동 전송하면 실패할 가능성이 높은 상품은 `마켓플러스검증상태=BLOCK`으로 표시하고 차단 마켓을 기록하세요.
7. 판단 근거는 상품명, OCR텍스트, 생성 키워드, 제공된 카테고리 기준표만 사용하세요. 외부 검색은 금지합니다.
"""

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(content)

    _status(status_cb, f"keyword_skill.md 생성: {md_path}")
    _status(status_cb, f"LLM 결과 저장 폴더: {llm_result_dir}")
    _copy_market_category_references(export_root, status_cb)

    import shutil
    _services_dir = os.path.dirname(os.path.abspath(__file__))
    for cat_file in ("naver_category_tree.txt", "coupang_category_tree.txt"):
        src = os.path.join(_services_dir, cat_file)
        if os.path.isfile(src):
            dst = os.path.join(export_root, cat_file)
            shutil.copy2(src, dst)
            _status(status_cb, f"{cat_file} → export 폴더 복사")

    if chunk_size and chunk_size > 0:
        chunks_dir, _ = _prepare_chunk_session_dir(export_root, keyword_version, date_tag)
        _split_upload_excel(upload_path, export_root, chunk_size, date_tag, status_cb, llm_chunks_dir=chunks_dir)
        if os.path.isdir(chunks_dir):
            shutil.copy2(md_path, os.path.join(chunks_dir, "keyword_skill.md"))
            _copy_market_category_references(chunks_dir, status_cb)
            for cat_file in ("naver_category_tree.txt", "coupang_category_tree.txt"):
                src = os.path.join(_services_dir, cat_file)
                if os.path.isfile(src):
                    shutil.copy2(src, os.path.join(chunks_dir, cat_file))
            os.makedirs(os.path.join(chunks_dir, llm_result_rel), exist_ok=True)
            _status(status_cb, f"keyword_skill.md + 카테고리 트리 → {chunks_dir} 복사 완료 ({llm_result_rel})")

def run_pipeline(cfg: PipelineConfig, status_cb=None, progress_cb=None) -> tuple[str, str]:

    _status(status_cb, "🚀 run_pipeline 함수 시작!")



    if not cfg.file_path:

        raise ValueError("CSV/Excel 파일을 선택해 주세요.")



    csv_base = os.path.splitext(os.path.basename(cfg.file_path))[0]

    date_tag = datetime.now().strftime("%Y%m%d")

    # phase=analysis 시 Phase1의 export_root 재사용
    if cfg.export_root_override and os.path.isdir(cfg.export_root_override):
        export_root = cfg.export_root_override
    else:
        export_root = _make_export_root(date_tag, csv_base)

    os.makedirs(export_root, exist_ok=True)

    _status(status_cb, f"📁 작업 폴더: {export_root} (phase={cfg.phase})")

    keyword_version = _normalize_keyword_version(getattr(cfg, "keyword_version", "2.0"))

    use_keyword_v2 = keyword_version == "2.0"

    keyword_version_slug = f"v{keyword_version.replace('.', '_')}"

    min_kw_tokens = 5 if use_keyword_v2 else 4

    kw_quality_gate = 12 if use_keyword_v2 else 6

    _status(status_cb, f"키워드 버전: {keyword_version}")



    # Google Cloud Vision API 인증 설정 (통합 파이프라인에서 OCR 사용 시)

    _app_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    ensure_env_loaded(os.path.join(_app_root, ".env"))

    _gv_cred_path = get_env("GOOGLE_APPLICATION_CREDENTIALS", "GOOGLE_VISION_CREDENTIALS")

    if not _gv_cred_path:

        _gv_cred_path = os.path.join(_app_root, "google_vision_key.json")

    if _gv_cred_path and os.path.isfile(_gv_cred_path):

        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = _gv_cred_path

        _status(status_cb, f"Google Vision API 인증 설정 완료: {_gv_cred_path}")

    else:

        _status(status_cb, f"Google Vision 키 파일 없음: {_gv_cred_path}")



    if cfg.use_html_ocr or cfg.use_local_ocr:

        detected = core.setup_tesseract(cfg.tesseract_path or None)

        if not detected:

            _status(status_cb, "Tesseract 경로를 찾지 못했습니다. Google Vision API로 OCR 진행합니다.")



    # 네이버 검색광고 API/자동완성은 현재 완전 비활성화

    cfg.naver_enabled = False

    cfg.naver_autocomplete = False

    # 통합 검색형 규칙: 자동완성은 새 키워드 근거로 사용하지 않는다.
    cfg.google_autocomplete = False

    core.DRY_RUN = True

    naver_keys = {"ACCESS_LICENSE": "", "SECRET_KEY": "", "CUSTOMER_ID": ""}

    _status(status_cb, "외부 검색/자동완성 비활성화: 상품명/OCR/Vision 근거 중심으로 키워드 생성")



    _status(status_cb, "처리중... (1/2) 전처리 + OCR + 키워드 생성")

    _progress(progress_cb, 10)

    if core.client is None:

        if core.refresh_openai_client():

            _status(status_cb, "AI client 재로드 완료")

        else:

            _status(status_cb, "AI client 없음: .env에 ANTHROPIC_API_KEY 또는 OPENAI_API_KEY 확인 필요")



    # 사용자 제외 단어 리로드 (실행 사이에 추가된 것 반영)

    core.merge_user_stopwords()



    # 통합 검색형 규칙: 기존보다 약 5개 더 많은 유효 검색어를 담을 수 있게 한다.
    max_words = max(10, int(cfg.max_words) + 5)

    max_len = max(30, int(cfg.max_len))

    min_len = max(0, int(cfg.min_len))

    if min_len > max_len:

        min_len = max_len // 2



    # min_len 달성에 필요한 최소 단어수 보정 (한국어 평균 ~3.8자/단어)

    _min_words_for_len = int(min_len / 3.5) + 1 if min_len > 0 else max_words

    if max_words < _min_words_for_len:

        max_words = min(_min_words_for_len, 50)  # 최대 50단어 상한



    max_imgs = max(0, int(cfg.max_imgs))

    threads = min(16, max(1, int(cfg.threads)))

    max_depth = int(cfg.max_depth)



    listing_size = max(200, int(cfg.listing_size))

    listing_pad = max(0, int(cfg.listing_pad))

    listing_max = max(0, int(cfg.listing_max))

    logo_ratio = max(1, min(60, int(cfg.logo_ratio)))

    logo_opacity = max(0, min(100, int(cfg.logo_opacity)))

    logo_pos = cfg.logo_pos or "tr"

    use_auto_contrast = bool(cfg.use_auto_contrast)

    use_sharpen = bool(cfg.use_sharpen)

    use_small_rotate = bool(cfg.use_small_rotate)

    rotate_zoom = float(cfg.rotate_zoom)

    logo_rgba = core._load_logo(cfg.logo_path.strip())



    ultra_angle_deg = float(cfg.ultra_angle_deg)

    ultra_translate_px = float(cfg.ultra_translate_px)

    ultra_scale_pct = float(cfg.ultra_scale_pct)

    trim_tol = int(cfg.trim_tol)

    jpeg_q_min = max(70, min(99, int(cfg.jpeg_q_min)))

    jpeg_q_max = max(jpeg_q_min, min(99, int(cfg.jpeg_q_max)))



    do_flip_lr = bool(cfg.do_flip_lr)



    psm = int(cfg.psm or 3)

    oem = int(cfg.oem or 3)

    korean_only = bool(cfg.korean_only)

    drop_digits = bool(cfg.drop_digits)

    tess_lang = "kor" if korean_only else "kor+eng"



    df = core.safe_read_csv(cfg.file_path)

    if df.empty:

        raise ValueError("CSV/Excel 내용이 비어 있습니다.")

    input_cols = list(df.columns)  # 업로드용 저장 시 원본 컬럼만 유지



    # 불필요한 컬럼 제거 (원본 파일에서 유입되는 CM, Unnamed 등)

    _drop_cols = [c for c in df.columns

                  if str(c).strip().upper() == "CM"

                  or str(c).startswith("Unnamed")]

    if _drop_cols:

        df.drop(columns=_drop_cols, inplace=True, errors="ignore")



    name_col = "상품명"

    if name_col not in df.columns:

        raise ValueError("'상품명' 컬럼이 없습니다.")



    code_col = None

    for c in df.columns:

        if str(c).strip() in ["자체상품코드", "자체 상품코드", "상품코드B", "코드", "코드B"]:

            code_col = c

            break



    if str(df.columns[0]) == "상품코드" and len(df) > 0:

        # 상품코드 열의 자동값 제거 (A2 Pxxxxxx 제거 포함)

        df[df.columns[0]] = df[df.columns[0]].astype(object)
        df.iloc[:, 0] = ""

    if df.shape[1] >= 5:

        df.iloc[:, 4] = 23



    detail_col = next((col for col in df.columns if "상세" in str(col)), None)

    if detail_col and cfg.img_tag:

        df[detail_col] = df[detail_col].apply(lambda x: core.insert_img_tag(x, cfg.img_tag))



    # AU열 (이미지등록(상세)) 컬럼 확인 - 대표이미지 순차 다운로드용

    listing_img_col = None

    for c in df.columns:

        if "이미지등록" in str(c) or str(c).strip().upper() == "AU":

            listing_img_col = c

            break



    공급가_col = next((col for col in df.columns if "공급가" in str(col)), None)

    판매가_col = next((col for col in df.columns if "판매가" in str(col)), None)

    소비자가_col = next((col for col in df.columns if "소비자가" in str(col)), None)

    if not all([공급가_col, 판매가_col, 소비자가_col]):

        raise ValueError("공급가/판매가/소비자가 컬럼명을 찾을 수 없습니다.")

    df[공급가_col] = pd.to_numeric(df[공급가_col], errors="coerce").fillna(0)

    df[판매가_col] = df[공급가_col].apply(_calc_sale_price)

    df[소비자가_col] = df[판매가_col].apply(_calc_consumer_price)



    df["_code9_from_name"] = df[name_col].astype(str).str.extract(r"(GS\d{7})")

    df["_opt_from_name"] = (

        df[name_col].astype(str)

        .str.replace(r".*GS\d{7}[A-Z0-9]+", "", regex=True)

        .str.strip()

    )



    # ── 옵션 가격 분리: 추가금 > 기본판매가 시 상품 자동 분리 ──

    _price_split_log = []

    for _gs9, _grp in df.groupby("_code9_from_name", dropna=True):

        if _gs9 is None or pd.isna(_gs9) or len(_grp) <= 1:

            continue

        _sells = _grp[판매가_col].values

        _base = float(_sells.min())

        if _base <= 0 or float(_sells.max()) <= _base * 2:

            continue

        _sorted_idx = _grp[판매가_col].sort_values().index.tolist()

        _bands = []

        _current_band = [_sorted_idx[0]]

        _band_base = float(df.at[_sorted_idx[0], 판매가_col])

        for _si in _sorted_idx[1:]:

            _sell = float(df.at[_si, 판매가_col])

            if _sell <= _band_base * 2:

                _current_band.append(_si)

            else:

                _bands.append(_current_band)

                _current_band = [_si]

                _band_base = _sell

        _bands.append(_current_band)

        if len(_bands) <= 1:

            continue

        _letter_offset = 0  # 밴드 간 알파벳 오프셋 누적

        for _bi, _band in enumerate(_bands):

            _band_sells = [float(df.at[_idx, 판매가_col]) for _idx in _band]

            _min_p = int(min(_band_sells))

            _max_p = int(max(_band_sells))

            _price_tag = f"({_min_p:,}~{_max_p:,}원)" if _min_p != _max_p else f"({_min_p:,}원)"

            for _ri, _idx in enumerate(_band):

                _new_letter = chr(65 + _letter_offset + _ri)

                df.at[_idx, "_code9_from_name"] = f"{_gs9}-{_bi+1}"

                _orig_name = str(df.at[_idx, name_col])

                _orig_name = re.sub(r'(GS\d{7})[A-Z0-9]+', rf'\g<1>{_new_letter}', _orig_name, count=1)

                if _price_tag not in _orig_name:

                    _orig_name = f"{_orig_name} {_price_tag}"

                df.at[_idx, name_col] = _orig_name

                if code_col and code_col in df.columns:

                    _old_code = str(df.at[_idx, code_col])

                    _new_code = re.sub(r'(GS\d{7})[A-Z0-9]+', rf'\g<1>{_new_letter}', _old_code, count=1)

                    df.at[_idx, code_col] = _new_code

            _letter_offset += len(_band)  # 다음 밴드는 이어서 알파벳 부여

        _price_split_log.append(f"{_gs9} → {len(_bands)}개 상품으로 분리")

    if _price_split_log:

        _status(status_cb, f"⚠️ 가격 분리: {len(_price_split_log)}건")

        for _msg in _price_split_log:

            _status(status_cb, f"  {_msg}")



    def _resolve_columns(frame, candidate_names):

        matched = [c for c in frame.columns if str(c).strip() in candidate_names]

        if not matched:

            primary = candidate_names[0]

            frame[primary] = pd.Series([""] * len(frame), index=frame.index, dtype="string")

            matched = [primary]

        return matched



    def _set_row_values(frame, row_idx, candidate_names, value):

        for c in frame.columns:

            if str(c).strip() in candidate_names:

                frame.at[row_idx, c] = value



    option_name_map = {

        "옵션사용": ["옵션사용"],

        "옵션 구성방식": ["옵션 구성방식", "옵션구성방식"],

        "옵션 표시방식": ["옵션 표시방식", "옵션표시방식"],

        "옵션입력": ["옵션입력"],

    }



    for key, names in option_name_map.items():

        cols = _resolve_columns(df, names)

        for col in cols:

            df[col] = df[col].astype("string").fillna("")

            df[col] = "N" if key == "옵션사용" else ""



    # 품목구성방식/품목 구성방식 모두 유지, 기본 공란 (옵션 상품만 T)

    def _apply_item_comp_fix(frame):

        item_cols = _resolve_columns(frame, ["품목구성방식", "품목 구성방식"])

        for col in item_cols:

            frame[col] = ""

        return frame



    df = _apply_item_comp_fix(df)



    # 가격 분리된 밴드(GS...-1, GS...-2)에서 알파벳 오프셋 계산
    _band_letter_offset: dict[str, int] = {}
    grouped = df.groupby("_code9_from_name", dropna=True)
    for code in sorted(grouped.groups.keys(), key=lambda c: str(c)):
        code_str = str(code)
        # GS1234567-2 형태인 경우 같은 원본의 이전 밴드 옵션 수를 누적
        if "-" in code_str:
            base_gs = code_str.rsplit("-", 1)[0]
            _band_letter_offset.setdefault(base_gs, 0)
            _band_letter_offset[code_str] = _band_letter_offset[base_gs]
            _band_letter_offset[base_gs] += len(grouped.get_group(code))
        else:
            _band_letter_offset[code_str] = 0

    for code, group in grouped:

        if code is None or pd.isna(code):

            continue

        opts = []

        _opt_offset = _band_letter_offset.get(str(code), 0)

        for i, (idx, row) in enumerate(group.iterrows()):

            option_code = chr(65 + _opt_offset + i)

            option_val = str(row["_opt_from_name"]).strip()

            if option_val:

                opts.append(f"{option_code} {option_val}")

        if opts:

            ak_val = "옵션{" + "|".join(opts) + "}"

            aidx = group.index[0]

            _set_row_values(df, aidx, ["품목구성방식", "품목 구성방식"], "T")

            _set_row_values(df, aidx, ["옵션사용"], "Y")

            _set_row_values(df, aidx, ["옵션 구성방식", "옵션구성방식"], "T")

            _set_row_values(df, aidx, ["옵션 표시방식", "옵션표시방식"], "S")

            _set_row_values(df, aidx, ["옵션입력"], ak_val)

            # 옵션 추가금액 계산: A옵션(기본가) 대비 차액
            _base_sell = float(df.at[aidx, 판매가_col])
            _additionals = []
            for _oi, (_oidx, _orow) in enumerate(group.iterrows()):
                _opt_sell = float(df.at[_oidx, 판매가_col])
                _additionals.append(str(int(_opt_sell - _base_sell)))
            if "옵션추가금" not in df.columns:
                df["옵션추가금"] = ""
            df.at[aidx, "옵션추가금"] = "|".join(_additionals)





    df.drop(columns=["_code9_from_name", "_opt_from_name"], inplace=True, errors="ignore")



    name_s = df[name_col].astype(str)

    code_s = df[code_col].astype(str) if (code_col and code_col in df.columns) else pd.Series("", index=df.index, dtype="string")

    # 대표행 판별: GS코드A로 끝나거나, 가격 분리로 옵션입력이 설정된 행
    mask1 = code_s.str.contains(r"GS\d{7}A$", na=False, regex=True)

    mask2 = name_s.str.contains(r"GS\d{7}A\b", na=False, regex=True)

    # 가격 분리된 밴드의 대표행 (옵션입력이 설정된 행)
    _opt_input_col = [c for c in df.columns if str(c).strip() == "옵션입력"]
    mask3 = pd.Series(False, index=df.index)
    if _opt_input_col:
        mask3 = df[_opt_input_col[0]].astype(str).str.startswith("옵션{", na=False)

    rep_mask = (mask1 | mask2 | mask3)

    df_after = df.loc[rep_mask].copy()

    if df_after.empty:

        fallback_mask = name_s.str.contains(r"GS\d{7}", na=False, regex=True) | code_s.str.contains(r"GS\d{7}", na=False, regex=True)

        df_after = df.loc[fallback_mask].copy()



    # NOTE:

    # df 단계에서 옵션상품의 품목구성방식=T를 세팅했으므로,

    # 필터링된 df_after에서 다시 초기화하면 T가 사라진다.

    # 여기서는 컬럼 존재만 보장하고 기존 값은 유지한다.

    _resolve_columns(df_after, ["품목구성방식", "품목 구성방식"])



    if "검색어설정" not in df_after.columns:

        df_after["검색어설정"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["검색어설정"] = df_after["검색어설정"].astype("string").fillna("")

    if "쿠팡검색태그" not in df_after.columns:

        df_after["쿠팡검색태그"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["쿠팡검색태그"] = df_after["쿠팡검색태그"].astype("string").fillna("")

    if "네이버태그" not in df_after.columns:

        df_after["네이버태그"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["네이버태그"] = df_after["네이버태그"].astype("string").fillna("")



    if "검색키워드" not in df_after.columns:

        df_after["검색키워드"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["검색키워드"] = df_after["검색키워드"].astype("string").fillna("")

    for _market_kw_col in MARKET_KEYWORD_COLUMNS_10 + ("홈런_네이버태그",):
        if _market_kw_col not in df_after.columns:
            df_after[_market_kw_col] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")
        else:
            df_after[_market_kw_col] = df_after[_market_kw_col].astype("string").fillna("")

    if "OCR요약" not in df_after.columns:

        df_after["OCR요약"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["OCR요약"] = df_after["OCR요약"].astype("string").fillna("")

    if "네이버검색광고데이터" not in df_after.columns:

        df_after["네이버검색광고데이터"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["네이버검색광고데이터"] = df_after["네이버검색광고데이터"].astype("string").fillna("")

    if "1차키워드" not in df_after.columns:

        df_after["1차키워드"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["1차키워드"] = df_after["1차키워드"].astype("string").fillna("")

    if "최종키워드2차" not in df_after.columns:

        df_after["최종키워드2차"] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

    else:

        df_after["최종키워드2차"] = df_after["최종키워드2차"].astype("string").fillna("")

    _vision_cols = [

        "Vision힌트",

        "Vision분석JSON",

        "Vision_core_identity",

        "Vision_installation_and_physical",

        "Vision_usage_context",

        "Vision_market_expansion",

        "Vision_compatibility",

        "Vision_functional_inference",

        "Vision_search_boost_elements",

    ]

    for _c in _vision_cols:

        if _c not in df_after.columns:

            df_after[_c] = pd.Series([""] * len(df_after), index=df_after.index, dtype="string")

        else:

            df_after[_c] = df_after[_c].astype("string").fillna("")



    debug_rows = []

    debug_on = bool(cfg.debug)



    local_root = cfg.local_img_dir

    allow_folder_match = bool(cfg.allow_folder_match)

    ocr_temp_root = os.path.join(export_root, "_ocr_tmp")

    os.makedirs(ocr_temp_root, exist_ok=True)

    use_local = cfg.use_local_ocr and bool(detail_col)



    def ocr_paths(paths):

        """Google Cloud Vision API로 이미지 OCR 처리"""

        texts, raw_pairs = [], []

        if not paths:

            return texts, raw_pairs



        # Google Cloud Vision OCR 사용

        from app.services.ocr_pipeline import _ocr_google_vision



        with ThreadPoolExecutor(max_workers=threads) as ex:

            futs = {ex.submit(_ocr_google_vision, p): p for p in paths}

            for fut in as_completed(futs):

                src = futs[fut]

                try:

                    t = fut.result()

                    if t:

                        texts.append(t)

                        raw_pairs.append((os.path.basename(src), t[:200]))

                except Exception as e:

                    # Google Vision 실패 시 에러 로그

                    _status(status_cb, f"[OCR 실패] {os.path.basename(src)}: {str(e)[:50]}")

        return texts, raw_pairs



    def analyze_product_images_local(image_paths, product_name, model_name, min_fill_ratio=0.5):

        """대표이미지 Vision 분석(JSON). 실패 시 {} 반환."""

        try:

            if not core.client or not image_paths:

                return {}



            _target_paths = [

                "core_identity.category",

                "core_identity.product_type_correction",

                "core_identity.structure",

                "core_identity.material_visual",

                "core_identity.color",

                "core_identity.size_context",

                "installation_and_physical.mount_type",

                "installation_and_physical.installation_method",

                "installation_and_physical.environment_resistance",

                "installation_and_physical.durability_hint",

                "installation_and_physical.weight_feel",

                "usage_context.usage_location",

                "usage_context.usage_purpose",

                "usage_context.target_user",

                "usage_context.usage_scenario",

                "usage_context.indoor_outdoor",

                "market_expansion.emotion_tone",

                "market_expansion.design_style",

                "market_expansion.shape_motif",

                "market_expansion.seasonal_context",

                "market_expansion.trend_alignment",

                "compatibility.compatible_with",

                "compatibility.size_compatibility",

                "compatibility.device_fit",

                "functional_inference.primary_function",

                "functional_inference.secondary_function",

                "functional_inference.problem_solving_keyword",

                "functional_inference.convenience_feature",

                "search_boost_elements.installation_keywords",

                "search_boost_elements.space_keywords",

                "search_boost_elements.benefit_keywords",

                "search_boost_elements.longtail_candidates",

            ]



            valid_paths = [p for p in image_paths if os.path.isfile(p)]

            if not valid_paths:

                return {}



            def _resize_image_for_vision(img_path, max_side=768):
                """Vision API 토큰 절약을 위해 이미지를 리사이즈하여 base64 반환."""
                from PIL import Image
                img = Image.open(img_path)
                w, h = img.size
                if max(w, h) > max_side:
                    ratio = max_side / max(w, h)
                    img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format="JPEG", quality=75)
                return base64.b64encode(buf.getvalue()).decode("utf-8"), "image/jpeg"

            def _build_image_contents(paths):

                out = []

                for img_path in paths:

                    try:
                        b64, mime = _resize_image_for_vision(img_path)
                    except Exception:
                        with open(img_path, "rb") as f:
                            b64 = base64.b64encode(f.read()).decode("utf-8")
                        ext = os.path.splitext(img_path)[1].lower()
                        mime = "image/png" if ext == ".png" else "image/jpeg"

                    out.append({

                        "type": "image_url",

                        "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "low"},

                    })

                return out



            def _is_filled(v):

                if isinstance(v, str):

                    return bool(v.strip())

                if isinstance(v, list):

                    return any(bool(str(x).strip()) for x in v)

                if isinstance(v, dict):

                    return any(_is_filled(x) for x in v.values())

                return bool(v)



            def _get_path(d, path):

                cur = d

                for k in path.split("."):

                    if not isinstance(cur, dict) or k not in cur:

                        return None

                    cur = cur[k]

                return cur



            def _fill_ratio(d):

                if not isinstance(d, dict):

                    return 0.0

                filled = 0

                for p in _target_paths:

                    if _is_filled(_get_path(d, p)):

                        filled += 1

                return filled / max(1, len(_target_paths))



            def _merge_analysis(base, extra):

                if not isinstance(base, dict):

                    return extra if isinstance(extra, dict) else {}

                if not isinstance(extra, dict):

                    return base

                merged = dict(base)

                for k, v in extra.items():

                    if k not in merged:

                        merged[k] = v

                        continue

                    bv = merged[k]

                    if isinstance(bv, dict) and isinstance(v, dict):

                        merged[k] = _merge_analysis(bv, v)

                    elif isinstance(bv, list) and isinstance(v, list):

                        merged_list = []

                        seen = set()

                        for item in bv + v:

                            s = str(item).strip()

                            if not s:

                                continue

                            lk = s.lower()

                            if lk in seen:

                                continue

                            seen.add(lk)

                            merged_list.append(s)

                        merged[k] = merged_list

                    else:

                        if not _is_filled(bv) and _is_filled(v):

                            merged[k] = v

                return merged



            cleaned_name = re.sub(r"GS\d{7}[A-Z0-9]*\s*", "", str(product_name)).strip()

            system_prompt = (

                "너는 이커머스 상품 이미지 분석가다. JSON만 출력. 설명문/마크다운/코드블록 금지.\n"

                "모든 값은 검색 키워드로 쓸 수 있는 짧은 명사구(1~4단어)로 작성.\n"

                "문장형 금지. 추론 불가면 빈값(\"\") 또는 빈배열([])."

            )

            user_text = (

                f"상품명: {cleaned_name}\n\n"

                "이미지를 보고 아래 JSON을 채워라.\n"

                "규칙: 모든 값은 짧은 명사구/키워드만. 문장 금지. 모르면 비워라.\n\n"

                "{"

                "\"core_identity\":{\"category\":\"\",\"product_type_correction\":\"\",\"structure\":\"\",\"material_visual\":\"\",\"color\":\"\",\"size_context\":\"\"},"

                "\"installation_and_physical\":{\"mount_type\":\"\",\"installation_method\":\"\",\"environment_resistance\":[],\"durability_hint\":\"\",\"weight_feel\":\"\"},"

                "\"usage_context\":{\"usage_location\":[],\"usage_purpose\":[],\"target_user\":[],\"usage_scenario\":[],\"indoor_outdoor\":\"\"},"

                "\"market_expansion\":{\"emotion_tone\":[],\"design_style\":[],\"shape_motif\":\"\",\"seasonal_context\":\"\",\"trend_alignment\":\"\"},"

                "\"compatibility\":{\"compatible_with\":[],\"size_compatibility\":[],\"device_fit\":\"\"},"

                "\"functional_inference\":{\"primary_function\":\"\",\"secondary_function\":[],\"problem_solving_keyword\":[],\"convenience_feature\":[]},"

                "\"search_boost_elements\":{\"installation_keywords\":[],\"space_keywords\":[],\"benefit_keywords\":[],\"longtail_candidates\":[]}"

                "}\n\n"

                "예시(참고용):\n"

                "category→\"브라켓\" / product_type_correction→\"차량조명 마운트 브라켓\"\n"

                "mount_type→\"무타공 고정형\" / installation_method→\"클램프 체결\"\n"

                "usage_location→[\"차량 본넷\",\"트렁크\"] / usage_purpose→[\"조명 설치\",\"작업등 장착\"]\n"

                "primary_function→\"무타공 조명 고정 각도조절\" / problem_solving_keyword→[\"무타공\",\"간편설치\"]\n"

                "longtail_candidates→[\"무타공 차량조명 브라켓\",\"본넷 작업등 거치대\"]"

            )



            # 1차: 최대 3장, 2차: 부족하면 최대 5장으로 재분석 후 병합

            stages = []

            first_n = min(3, len(valid_paths))

            if first_n > 0:

                stages.append(first_n)

            if len(valid_paths) > first_n:

                second_n = min(5, len(valid_paths))

                if second_n > first_n:

                    stages.append(second_n)



            best = {}

            vision_client = core._create_client(model_name) or core.client

            if vision_client is None:

                return {}

            for n in stages:

                image_contents = _build_image_contents(valid_paths[:n])

                user_msg = [{"type": "text", "text": user_text}] + image_contents

                resp = vision_client.chat.completions.create(

                    model=model_name,

                    messages=[

                        {"role": "system", "content": system_prompt},

                        {"role": "user", "content": user_msg},

                    ],

                    temperature=0.1,

                    top_p=0.9,

                    max_tokens=2048,

                    response_format={"type": "json_object"},

                )

                raw = (resp.choices[0].message.content or "").strip()

                try:

                    cur = json.loads(raw) if raw else {}

                except json.JSONDecodeError:

                    cur = {}

                    for _closer in ['"}}', '"]}}', '"]}}}', '"}}}', '"]}}}}}']:

                        try:

                            cur = json.loads(raw + _closer)

                            break

                        except json.JSONDecodeError:

                            continue

                best = _merge_analysis(best, cur)

                if _fill_ratio(best) >= float(min_fill_ratio):

                    break



            return best if isinstance(best, dict) else {}

        except Exception as _ve:

            _status(status_cb, f"[Vision 오류] {type(_ve).__name__}: {str(_ve)[:200]}")

            return {}



    def _extract_vision_hints(analysis):

        """구/신 Vision 스키마에서 키워드 힌트 문자열만 안전하게 추출."""

        hints, filled = [], []



        def _add_hint(path, text):

            s = re.sub(r"\s+", " ", str(text or "")).strip()

            if s:

                hints.append(s)

                filled.append(path)



        def _walk(node, path):

            if isinstance(node, dict):

                # 구 스키마: {"value": "...", "confidence": 0~1, "evidence": "..."}

                if "value" in node and any(k in node for k in ("confidence", "evidence")):

                    conf_ok = True

                    try:

                        conf_ok = float(node.get("confidence", 1.0)) >= 0.5

                    except Exception:

                        conf_ok = True

                    if conf_ok:

                        _add_hint(path, node.get("value", ""))

                    return

                for k, v in node.items():

                    _walk(v, f"{path}.{k}" if path else k)

                return

            if isinstance(node, list):

                had = False

                for x in node:

                    sx = re.sub(r"\s+", " ", str(x or "")).strip()

                    if sx:

                        hints.append(sx)

                        had = True

                if had:

                    filled.append(path)

                return

            if isinstance(node, str):

                _add_hint(path, node)



        _walk(analysis or {}, "")

        uniq, seen = [], set()

        for h in hints:

            key = h.lower()

            if key in seen:

                continue

            seen.add(key)

            uniq.append(h)

        uniq_filled = list(dict.fromkeys([p for p in filled if p]))

        return uniq[:40], uniq_filled



    def _vision_excel_payload(analysis, hint_parts):

        """Vision 분석 결과를 엑셀 저장용 문자열 컬럼으로 변환."""

        def _as_json(v):

            if v is None:

                return ""

            try:

                return json.dumps(v, ensure_ascii=False, separators=(",", ":"))

            except Exception:

                return str(v)



        a = analysis or {}

        return {

            "Vision힌트": " | ".join([x for x in (hint_parts or []) if str(x).strip()])[:1000],

            "Vision분석JSON": _as_json(a),

            "Vision_core_identity": _as_json(a.get("core_identity", "")),

            "Vision_installation_and_physical": _as_json(a.get("installation_and_physical", "")),

            "Vision_usage_context": _as_json(a.get("usage_context", "")),

            "Vision_market_expansion": _as_json(a.get("market_expansion", "")),

            "Vision_compatibility": _as_json(a.get("compatibility", "")),

            "Vision_functional_inference": _as_json(a.get("functional_inference", "")),

            "Vision_search_boost_elements": _as_json(a.get("search_boost_elements", "")),

        }



    def _stamp_vision_to_ocr_results(gs_code9, payload):

        """같은 GS코드의 가장 최근 OCR 결과 레코드에 Vision 컬럼을 부착."""

        if not gs_code9 or not ocr_results_list or not payload:

            return

        for i in range(len(ocr_results_list) - 1, -1, -1):

            row = ocr_results_list[i]

            if str(row.get("GS코드", "")).strip() == str(gs_code9).strip():

                row.update(payload)

                return





    # B마켓 데이터 수집용
    b_market_rows: dict[int, dict] = {}  # idx → {상품명, 검색어설정, 검색키워드, 네이버태그, ...}

    def _apply_market_keyword_packages(

        row_idx,

        product_name: str,

        source_text: str,

        naver_keyword_table: str,

        model_name: str,

        anchors,

        baseline,

        market: str = "A",
        avoid_terms: str = "",

    ):

        result = generate_market_keyword_packages(

            product_name=product_name,

            source_text=source_text,

            model_name=model_name,

            anchors=anchors,

            baseline=baseline,

            naver_keyword_table=naver_keyword_table,

            market=market,
            avoid_terms=avoid_terms,

        )

        if market.upper() == "B":
            # B마켓 결과는 별도 컬럼에 저장
            if result.search_keywords:
                df_after.at[row_idx, "B_검색키워드"] = result.search_keywords
            if result.coupang_tags:
                b_coupang_line = ",".join(result.coupang_tags)
                df_after.at[row_idx, "B_검색어설정"] = b_coupang_line
                df_after.at[row_idx, "B_쿠팡검색태그"] = b_coupang_line
            if result.naver_tags:
                df_after.at[row_idx, "B_네이버태그"] = "|".join(result.naver_tags)
        else:
            for _col, _value in (result.market_keywords or {}).items():
                if _col in df_after.columns and _value:
                    df_after.at[row_idx, _col] = _value

            _smartstore_tags = (result.market_keywords or {}).get("홈런_스마트스토어태그", "")
            if _smartstore_tags and "홈런_네이버태그" in df_after.columns:
                df_after.at[row_idx, "홈런_네이버태그"] = _smartstore_tags

            if result.search_keywords:

                df_after.at[row_idx, "검색키워드"] = result.search_keywords

                _status(status_cb, f"검색키워드 생성: {str(product_name)[:20]} → {len(result.search_keywords)}자")

            if result.coupang_tags:

                coupang_line = ",".join(result.coupang_tags)

                df_after.at[row_idx, "검색어설정"] = coupang_line

                df_after.at[row_idx, "쿠팡검색태그"] = coupang_line

            if result.naver_tags:

                df_after.at[row_idx, "네이버태그"] = "|".join(result.naver_tags)

        # B마켓 키워드도 동시 생성
        if cfg.enable_b_market:
            b_result = generate_market_keyword_packages(
                product_name=product_name,
                source_text=source_text,
                model_name=model_name,
                anchors=anchors,
                baseline=baseline,
                naver_keyword_table=naver_keyword_table,
                market="B",
                avoid_terms=avoid_terms,
            )
            b_data = {}
            if b_result.coupang_tags:
                b_data["검색어설정"] = ",".join(b_result.coupang_tags)
            if b_result.search_keywords:
                b_data["검색키워드"] = b_result.search_keywords
            if b_result.naver_tags:
                b_data["네이버태그"] = "|".join(b_result.naver_tags)
            b_market_rows[row_idx] = b_data

        return result
    def _has_meaningful_title_evidence(text: str) -> bool:
        cleaned = core.normalize_space(str(text or "")).replace("(OCR 텍스트 없음)", "").strip()
        return bool(cleaned)


    def _compose_b_market_title(
        base_name: str,
        kw_line: str,
        option_tokens: set,
        option_text: str = "",
        ocr_text: str = "",
        b_kw_line: str = "",
        prefer_seed_only: bool = False,
        a_final_name: str = "",
    ) -> str:

        if use_keyword_v2:

            return _compose_v2_title(
                base_name_value=base_name,
                kw_line_value=b_kw_line or kw_line or base_name,
                option_text_value=option_text,
                max_len_value=int(cfg.b_name_max),
                market_value="B",
                a_final_name_value=a_final_name,
            )

        def _split_title_tokens(text: str) -> list[str]:
            return [tok for tok in core.normalize_space(str(text or "")).split() if tok]

        def _reorder_away_from_a(tokens: list[str], a_head: set[str]) -> list[str]:
            """A마켓 앞부분과 겹치는 토큰을 뒤로 밀어 B마켓 시작을 차별화"""
            if not a_head or len(tokens) < 3:
                return tokens
            front = [t for t in tokens if t not in a_head]
            back = [t for t in tokens if t in a_head]
            if not front:
                mid = max(1, len(back) // 2)
                front = back[mid:]
                back = back[:mid]
            return front + back

        def _identity_head_tokens() -> list[str]:
            head: list[str] = []
            seen: set[str] = set()
            for tok in _split_title_tokens(base_name):
                if tok in option_tokens or tok in seen:
                    continue
                seen.add(tok)
                head.append(tok)
                if len(head) >= 2:
                    break
            return head

        desired_min = min(int(cfg.b_name_max), max(int(cfg.b_name_min), min(78, int(cfg.b_name_max))))
        b_seed = core.normalize_space(str(b_kw_line or "")).strip()
        a_seed = core.normalize_space(str(kw_line or "")).strip()
        a_ref = a_final_name or a_seed
        a_head_tokens = set(_split_title_tokens(a_ref)[:3])
        identity_head = _identity_head_tokens()
        identity_head_set = set(identity_head)

        def _apply_version_order(tokens: list[str]) -> list[str]:
            if not tokens:
                return tokens
            deduped: list[str] = []
            seen: set[str] = set()
            for tok in tokens:
                if tok in option_tokens or tok in seen:
                    continue
                seen.add(tok)
                deduped.append(tok)
            if not use_keyword_v2 or not identity_head:
                return _reorder_away_from_a(deduped, a_head_tokens)
            tail = [tok for tok in deduped if tok not in identity_head_set]
            tail = _reorder_away_from_a(tail, {t for t in a_head_tokens if t not in identity_head_set})
            return identity_head + tail

        if prefer_seed_only and not _has_meaningful_title_evidence(ocr_text):
            concise_tokens = _apply_version_order(_split_title_tokens(b_seed or a_seed or base_name))
            concise = " ".join(concise_tokens).strip()
            if concise:
                return concise[: int(cfg.b_name_max)].rstrip() if len(concise) > int(cfg.b_name_max) else concise

        b_tokens = _split_title_tokens(b_seed)
        strong_b_seed = len(b_tokens) >= 4 or len(b_seed) >= max(18, desired_min - 22)

        seed_tokens: list[str] = []
        seed_seen: set[str] = set()
        for source in ([b_seed] if strong_b_seed else [b_seed, a_seed]):
            for tok in _split_title_tokens(source):
                if tok in option_tokens or tok in seed_seen:
                    continue
                seed_seen.add(tok)
                seed_tokens.append(tok)
        if not seed_tokens:
            for tok in _split_title_tokens(base_name):
                if tok in option_tokens or tok in seed_seen:
                    continue
                seed_seen.add(tok)
                seed_tokens.append(tok)

        seed_tokens = _apply_version_order(seed_tokens)
        seed_line = " ".join(seed_tokens).strip()
        out = core.merge_base_name_with_keywords(
            seed_line,
            seed_line,
            max_words,
            int(cfg.b_name_max),
            option_tokens=option_tokens,
            ocr_text=ocr_text,
        )

        out_tokens = _apply_version_order(_split_title_tokens(out))
        seen = set(out_tokens)
        fill_sources = [b_seed, a_seed, ocr_text, base_name]
        for source in fill_sources:
            for tok in _split_title_tokens(source):
                if tok in option_tokens or tok in seen:
                    continue
                candidate_tokens = _apply_version_order(out_tokens + [tok])
                candidate = " ".join(candidate_tokens).strip()
                if len(candidate) > int(cfg.b_name_max):
                    continue
                out_tokens = candidate_tokens
                seen = set(out_tokens)
                if len(candidate) >= desired_min:
                    return candidate

        final_out = " ".join(out_tokens).strip()
        return final_out[: int(cfg.b_name_max)].rstrip() if len(final_out) > int(cfg.b_name_max) else final_out


    # ── OCR 결과 Excel 로드 (미리 처리된 경우) ──

    ocr_lookup: dict = {}

    if cfg.ocr_excel_path and os.path.isfile(cfg.ocr_excel_path):

        from app.services.ocr_excel import read_ocr_results

        ocr_lookup, _ocr_meta = read_ocr_results(cfg.ocr_excel_path)

        _status(status_cb, f"OCR 결과 로드: {len(ocr_lookup)}개 상품 ({os.path.basename(cfg.ocr_excel_path)})")



    global_listing_sources = []

    naver_cache = {}

    no_detail_indices = []  # 상세 없는 상품 인덱스 수집

    ocr_results_list = []  # OCR 결과 수집 (저장용)



    def clean_base_name_for_naver(s: str) -> str:

        if not s:

            return ""

        s = re.sub(r"(GS\d{7}[A-Z0-9]*)", "", s)

        s = re.sub(r"[\[\]\(\)\-_|]+", " ", s)

        s = re.sub(r"\s+", " ", s).strip()

        return s



    CTR_THR = 0.05

    naver_stage_emitted = False

    fatal_gpt_error = ""



    def _resolve_topic_refs(base_name_value: str, expanded_name_value: str = ""):

        source = core.normalize_space(str(base_name_value or "")).strip()

        if not use_keyword_v2:

            expanded = core.normalize_space(str(expanded_name_value or "")).strip()

            if expanded:

                source = expanded

        return core.build_anchors_from_name(source), core.build_baseline_tokens_from_name(source)



    def _keyword_noise_score(text: str) -> int:

        weak_tokens = {
            "그리고", "없이", "필요", "붙이기만", "하면", "간단하게", "해보세요",
            "강력한", "특수", "제작된", "쉽게", "떨어지지", "않습니다", "끝",
        }

        score = 0

        for tok in [t for t in re.split(r"\s+", core.normalize_space(str(text or ""))) if t]:

            low = tok.lower()

            if re.fullmatch(r"[A-Za-z]", tok) or low in {"a", "b", "c", "d"}:

                score += 3

                continue

            if tok in weak_tokens:

                score += 2

            if tok.endswith(("로", "에", "을", "를", "이", "가", "은", "는")) and len(tok) <= 5:

                score += 1

        return score



    def _prune_v2_tokens(tokens: list[str], base_name_value: str = "", option_token_set=None) -> list[str]:

        if not use_keyword_v2:

            return [t for t in tokens if t]

        weak_tokens = {
            "그리고", "없이", "필요", "필요한", "붙이기만", "붙이기", "붙이면", "하면", "간단하게", "해보세요",
            "강력한", "특수", "제작된", "쉽게", "떨어지지", "않습니다", "끝",
            "공구", "구멍", "설치", "후크로", "벽에", "뚫", "뚫을", "와이어",
        }

        base_refs = set(core.tokenize_korean_words(base_name_value or ""))

        if option_token_set:

            base_refs |= {str(t) for t in option_token_set if str(t).strip()}

        out: list[str] = []

        seen: set[str] = set()

        for raw in tokens:

            tok = core.normalize_space(str(raw or "")).strip()

            if not tok:

                continue

            if re.fullmatch(r"[A-Za-z]", tok) or tok.lower() in {"a", "b", "c", "d"}:

                continue

            tok = re.sub(r"(으로|에서|에게|까지|부터|처럼|로|에|을|를|은|는)$", "", tok)

            tok = core.normalize_space(tok).strip()

            if not tok or tok in weak_tokens:

                continue

            if any(tok.endswith(suffix) for suffix in ("하기", "하게", "하면", "으로", "에서")):

                continue

            if base_refs and tok not in base_refs and any(mark in tok for mark in ("싱글", "더블", "투명", "스텐")):

                continue

            if tok in seen:

                continue

            seen.add(tok)

            out.append(tok)

        return out


    def _compose_v2_title(
        base_name_value: str,
        kw_line_value: str,
        option_text_value: str = "",
        max_len_value: int = 100,
        market_value: str = "A",
        a_final_name_value: str = "",
    ) -> str:

        title_base = core.normalize_space(str(base_name_value or "")).strip()

        if not use_keyword_v2:

            seed = core.normalize_space(str(kw_line_value or title_base)).strip()

            return seed[:max_len_value].rstrip() if len(seed) > max_len_value else seed

        base_tokens = [tok for tok in core.tokenize_korean_words(title_base) if tok]

        option_tokens_ordered = [tok for tok in core.tokenize_korean_words(option_text_value or "") if tok]

        anchors, baseline = _resolve_topic_refs(title_base)

        usage_hints = {"주방", "욕실", "거실", "사무실", "베란다", "현관", "세탁실"}

        install_hints = {"무타공", "접착", "벽부착", "부착", "장착"}

        function_hints = {"거치", "선반거치", "고정", "연결", "체결", "잠금"}

        identity_hints = {
            "후크", "걸이", "홀더", "거치대", "브라켓", "브래킷", "클립", "고리", "링", "D링",
            "와이어선반", "랙선반", "선반", "커넥터", "조인트", "캐치", "래치", "가스켓", "가스킷",
        }

        option_markers = ("싱글", "더블", "투명", "스텐", "화이트", "블랙", "실버", "304")

        def _sem_key(value: str) -> str:

            return re.sub(r"[^0-9A-Za-z가-힣]", "", str(value or "")).lower()

        base_keys = [_sem_key(tok) for tok in base_tokens if _sem_key(tok)]

        def _contained_in_base(token: str) -> bool:

            key = _sem_key(token)

            if not key:

                return False

            for base_key in base_keys:

                if key == base_key:

                    return True

                if len(key) <= len(base_key) and key in base_key:

                    return True

            return False

        candidate_tokens = _prune_v2_tokens(
            core.tokenize_korean_words(kw_line_value or ""),
            base_name_value=title_base,
            option_token_set=set(option_tokens_ordered),
        )

        buckets = {
            "identity": [],
            "install": [],
            "function": [],
            "usage": [],
            "option": [],
            "extra": [],
        }

        seen_keys = {_sem_key(tok) for tok in base_tokens if _sem_key(tok)}

        def _push(bucket: str, token: str):

            key = _sem_key(token)

            if not key or key in seen_keys:

                return

            seen_keys.add(key)

            buckets[bucket].append(token)

        for tok in option_tokens_ordered:

            if tok and not _contained_in_base(tok):

                _push("option", tok)

        for tok in candidate_tokens:

            if not tok or _contained_in_base(tok):

                continue

            if tok in option_tokens_ordered or any(marker in tok for marker in option_markers):

                _push("option", tok)

                continue

            if tok in usage_hints:

                _push("usage", tok)

                continue

            if tok in install_hints:

                _push("install", tok)

                continue

            if tok in function_hints or tok.endswith("거치"):

                _push("function", tok)

                continue

            if any(hint in tok or tok in hint for hint in identity_hints):

                _push("identity", tok)

                continue

            if anchors and baseline and core.is_on_topic(tok, anchors, baseline):

                if len(tok) >= 4 or re.search(r"\d", tok):

                    _push("extra", tok)

        a_front = set(core.tokenize_korean_words(a_final_name_value or "")[:4])

        def _reorder_for_b(tokens: list[str]) -> list[str]:

            if not a_front:

                return tokens

            front = [tok for tok in tokens if tok not in a_front]

            back = [tok for tok in tokens if tok in a_front]

            return front + back

        if market_value.upper() == "B":

            ordered_groups = [
                base_tokens,
                _reorder_for_b(buckets["install"]),
                _reorder_for_b(buckets["function"]),
                _reorder_for_b(buckets["identity"]),
                _reorder_for_b(buckets["usage"]),
                _reorder_for_b(buckets["option"]),
                _reorder_for_b(buckets["extra"]),
            ]

        else:

            ordered_groups = [
                base_tokens,
                buckets["identity"],
                buckets["install"],
                buckets["function"],
                buckets["usage"],
                buckets["option"],
                buckets["extra"],
            ]

        out_tokens: list[str] = []

        out_seen: set[str] = set()

        word_budget = max_words if market_value.upper() != "B" else max(6, min(max_words, 23))

        for group in ordered_groups:

            for tok in group:

                key = _sem_key(tok)

                if not key or key in out_seen:

                    continue

                candidate = " ".join(out_tokens + [tok]).strip()

                if out_tokens and len(candidate) > int(max_len_value):

                    continue

                out_tokens.append(tok)

                out_seen.add(key)

                if len(out_tokens) >= word_budget:

                    break

            if len(out_tokens) >= word_budget:

                break

        final_out = " ".join(out_tokens).strip() or title_base

        return final_out[:int(max_len_value)].rstrip() if len(final_out) > int(max_len_value) else final_out



    def _infer_v2_builder_analysis(base_name_value: str, option_text_value: str, ocr_text_value: str) -> dict:

        if not use_keyword_v2:

            return {}

        text_value = core.normalize_space(str(ocr_text_value or "")).strip()

        full_name = core.normalize_space(f"{base_name_value} {option_text_value}".strip())

        if not text_value and not full_name:

            return {}

        def _uniq(items):

            out = []

            seen = set()

            for item in items:

                val = core.normalize_space(str(item or "")).strip()

                if not val or val in seen:

                    continue

                seen.add(val)

                out.append(val)

            return out

        anchor_candidates = list(core.build_anchors_from_name(full_name) or [])
        category = _uniq(anchor_candidates[:3] or core.tokenize_korean_words(full_name)[:3])

        product_type = []

        for term in ("후크", "걸이", "거치대", "홀더", "브라켓", "브래킷", "클립", "고리", "클램프", "커넥터", "조인트"):

            if term in text_value and term not in category:

                product_type.append(term)

        structure = []

        if re.search(r"와이어\s*선반", text_value):

            structure.append("와이어선반")

        if ("랙" in text_value and "선반" in text_value) or re.search(r"랙\s*/\s*와이어\s*선반", text_value):

            structure.append("랙선반")

        if "선반" in text_value and not any("선반" in tok for tok in category + product_type + structure):

            structure.append("선반")

        material_visual = []

        option_clean = core.normalize_space(str(option_text_value or "")).strip()

        if option_clean:

            material_visual.append(option_clean)

        for label, pattern in (("투명", r"투명"), ("스텐", r"스텐|스테인리스"), ("실버", r"실버|은색"), ("화이트", r"화이트|흰색"), ("블랙", r"블랙|검정")):

            if re.search(pattern, f"{full_name} {text_value}"):

                material_visual.append(label)

        installation_method = []

        if re.search(r"무타공|타공\s*없|구멍\s*뚫", text_value):

            installation_method.append("무타공")

        if re.search(r"접착|테이프|붙이기", text_value):

            installation_method.append("접착")

        mount_type = []

        if re.search(r"벽(면)?", text_value):

            mount_type.append("벽부착")

        usage_location = [loc for loc in ("주방", "욕실", "거실", "사무실", "베란다", "현관", "세탁실") if loc in text_value]

        usage_purpose = []

        primary_function = []

        if "선반" in text_value and any(term in f"{full_name} {text_value}" for term in ("후크", "걸이", "거치", "홀더")):

            usage_purpose.append("선반거치")

            primary_function.append("거치")

        return {

            "core_identity": {

                "category": _uniq(category),

                "product_type_correction": _uniq(product_type),

                "structure": _uniq(structure),

                "material_visual": _uniq(material_visual),

            },

            "installation_and_physical": {

                "mount_type": _uniq(mount_type),

                "installation_method": _uniq(installation_method),

            },

            "usage_context": {

                "usage_location": _uniq(usage_location),

                "usage_purpose": _uniq(usage_purpose),

            },

            "functional_inference": {

                "primary_function": _uniq(primary_function),

                "problem_solving_keyword": _uniq(installation_method),

            },

            "search_boost_elements": {

                "installation_keywords": _uniq(installation_method + mount_type),

                "space_keywords": _uniq(structure),

                "benefit_keywords": [],

            },

        }



    def _build_builder_line(base_name_value: str, option_text_value: str, ocr_text_value: str, vision_analysis_value, market_value: str) -> str:

        fallback_name = core.normalize_space(f"{base_name_value} {option_text_value}".strip()) or core.normalize_space(str(base_name_value or "")).strip()

        analysis_value = vision_analysis_value if isinstance(vision_analysis_value, dict) and vision_analysis_value else {}

        if use_keyword_v2 and not analysis_value:

            analysis_value = _infer_v2_builder_analysis(base_name_value, option_text_value, ocr_text_value)

        builder_line = build_keyword_string(

            ocr_text=ocr_text_value,

            vision_analysis=analysis_value,

            target_count=25,

            fallback_text=fallback_name,

            market=market_value,

        )

        if not use_keyword_v2:

            return builder_line

        builder_tokens = _prune_v2_tokens(
            core.tokenize_korean_words(builder_line or ""),
            base_name_value=base_name_value,
            option_token_set=set(core.tokenize_korean_words(option_text_value or "")),
        )

        if len(builder_tokens) >= 4:

            return " ".join(builder_tokens)

        recovery_order = [
            ("core_identity", "category"),
            ("core_identity", "product_type_correction"),
            ("core_identity", "structure"),
            ("installation_and_physical", "installation_method"),
            ("installation_and_physical", "mount_type"),
            ("functional_inference", "primary_function"),
            ("usage_context", "usage_location"),
            ("core_identity", "material_visual"),
        ]

        recovered: list[str] = []

        for section, key in recovery_order:

            section_value = analysis_value.get(section, {}) if isinstance(analysis_value, dict) else {}

            values = section_value.get(key, []) if isinstance(section_value, dict) else []

            if isinstance(values, str):

                values = [values]

            for value in values or []:

                tok = core.normalize_space(str(value or "")).strip()

                if tok:

                    recovered.append(tok)

        merged_tokens = core.tokenize_korean_words(base_name_value or "") + recovered + core.tokenize_korean_words(option_text_value or "")
        merged_tokens = _prune_v2_tokens(
            merged_tokens,
            base_name_value=base_name_value,
            option_token_set=set(core.tokenize_korean_words(option_text_value or "")),
        )
        return " ".join(merged_tokens[: max_words])



    def _is_fatal_gpt_error(msg: str) -> bool:

        m = str(msg or "").lower()

        return (

            "model_not_found" in m

            or "error code: 404" in m

        )



    def _fetch_naver_items_with_retry(hint_keywords: str):

        nonlocal naver_stage_emitted

        naver_err_local = ""

        if not hint_keywords:

            return [], naver_err_local

        if hint_keywords in naver_cache:

            return naver_cache.get(hint_keywords, []), naver_err_local

        try:

            if not naver_stage_emitted:

                _status(status_cb, "네이버 키워드 조회 중")

                _progress(progress_cb, 50)

                naver_stage_emitted = True

            retries = max(0, int(cfg.naver_retry_count)) if cfg.naver_retry else 0

            attempts = retries + 1

            items_local = []

            for attempt in range(attempts):

                try:

                    items_local = core.naver_keyword_tool(naver_keys, hint_keywords, debug=False)

                    naver_cache[hint_keywords] = items_local

                    if not core.DRY_RUN:

                        time.sleep(core.SLEEP_BETWEEN_CALLS)

                    break

                except Exception as e:

                    naver_err_local = str(e)

                    if attempt < attempts - 1:

                        _status(status_cb, f"네이버 재시도 {attempt + 1}/{attempts - 1}")

                        time.sleep(float(cfg.naver_retry_delay))

                    else:

                        naver_cache[hint_keywords] = []

        except Exception as e:

            naver_cache[hint_keywords] = []

            naver_err_local = str(e)

        return naver_cache.get(hint_keywords, []), naver_err_local



    def _query_naver_two_pass(final_line: str, base_name: str):

        return [], "", "네이버 API 비활성화"



    total_rows = max(1, len(df_after))

    for row_i, idx in enumerate(df_after.index, start=1):

        search_keywords = ""  # 검색 키워드 초기화

        try:

            full_pname = str(df_after.at[idx, name_col])

            base_name, option_text = core.extract_base_and_option(full_pname)

            option_tokens = set(core.tokenize_korean_words(option_text))

            prompt_product_name = base_name



            gs_code9 = None

            if code_col and code_col in df_after.columns:

                m1 = re.search(r"(GS\d{7})", str(df_after.at[idx, code_col]) or "")

                gs_code9 = m1.group(1) if m1 else None

            if not gs_code9:

                m2 = re.search(r"(GS\d{7})", full_pname)

                gs_code9 = m2.group(1) if m2 else None



            local_texts, matched_count = [], 0

            local_pairs = []



            # OCR 결과 Excel 에서 미리 처리된 데이터 사용

            # OCR Excel에 GSxxxxxxxA 형태로 저장되어 있으므로 두 가지 키로 매칭

            _ocr_match_key = None

            if ocr_lookup and gs_code9:

                for _mk in [gs_code9, f"{gs_code9}A"]:

                    if _mk in ocr_lookup:

                        _ocr_match_key = _mk

                        break

            # OCR Excel에서 매칭 시도 (raw 텍스트가 있는 경우만 사용)

            _used_ocr_excel = False

            if _ocr_match_key:

                _ocr_data = ocr_lookup[_ocr_match_key]

                if _ocr_data["raw"] and _ocr_data["raw"].strip():

                    local_texts = [_ocr_data["raw"]]

                    matched_count = _ocr_data["count"]

                    _used_ocr_excel = True



                    # OCR 결과 수집 (OCR Excel에서 불러온 경우)

                    ocr_results_list.append({

                        "GS코드": gs_code9,

                        "상품명": base_name,

                        "처리된이미지수": 0,

                        "전체이미지수": matched_count,

                        "OCR텍스트": _ocr_data["raw"][:500],

                        "이미지경로": "OCR Excel에서 로드"

                    })



                # OCR Excel의 이미지 경로로 대표이미지 소스 수집

                _gs_low = gs_code9.lower()

                _img_list = _ocr_data.get("images", [])

                _valid_imgs = [p for p in _img_list if p.strip() and p.strip().lower() != "nan" and os.path.isfile(p) and _gs_low in os.path.basename(p).lower()]

                if _valid_imgs:

                    global_listing_sources.extend(_valid_imgs)



            # phase=analysis: 기존 _ocr_tmp에서 이미지 재사용
            if cfg.phase == "analysis" and gs_code9:
                _gs_dir = os.path.join(ocr_temp_root, f"{gs_code9}A")
                if os.path.isdir(_gs_dir):
                    _existing = sorted([
                        os.path.join(_gs_dir, f) for f in os.listdir(_gs_dir)
                        if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))
                    ])
                    if _existing:
                        global_listing_sources.extend(_existing)
                        if not _used_ocr_excel:
                            all_hits_raw_reuse = _existing
                            ocr_hits_reuse = [p for p in all_hits_raw_reuse if os.path.splitext(os.path.basename(p))[0].isdigit()]
                            if not ocr_hits_reuse:
                                ocr_hits_reuse = all_hits_raw_reuse
                            sel = ocr_hits_reuse[:max_imgs] if max_imgs > 0 else []
                            local_texts, local_pairs = ocr_paths(sel)

            # OCR Excel에서 텍스트를 못 가져온 경우 → 로컬 이미지 OCR 실행

            if cfg.phase != "analysis" and not _used_ocr_excel and use_local and gs_code9:

                all_hits_raw = []



                # OCR용 이미지: 숫자 파일명만 (1.jpg, 2.jpg, 3.jpg...)

                ocr_hits = [

                    p for p in all_hits_raw

                    if os.path.splitext(os.path.basename(p))[0].isdigit()

                ]



                # 숫자파일이 없으면 → O열(상품상세설명)에서 이미지 URL 다운로드 시도

                if not ocr_hits and detail_col and detail_col in df_after.columns:

                    _detail_html_for_dl = str(df_after.at[idx, detail_col]) if pd.notna(df_after.at[idx, detail_col]) else ""

                    if _detail_html_for_dl and "<img" in _detail_html_for_dl.lower():

                        _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — O열에서 상세이미지 다운로드 중...")

                        from app.services.ocr_pipeline import _download_and_save_images

                        _dl_paths = _download_and_save_images(_detail_html_for_dl, f"{gs_code9}A", ocr_temp_root)

                        if _dl_paths:

                            ocr_hits = _dl_paths

                            all_hits_raw = all_hits_raw + _dl_paths

                            _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — 상세이미지 {len(_dl_paths)}개 다운로드 완료")

                        else:

                            _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — O열 이미지 다운로드 실패")



                # 대표이미지용: GS코드가 파일명에 포함된 것만

                _gs_low2 = gs_code9.lower()

                listing_hits = [p for p in all_hits_raw if _gs_low2 in os.path.basename(p).lower()]

                global_listing_sources.extend(listing_hits)



                matched_count = len(all_hits_raw)

                if len(ocr_hits) > 0:

                    sel = ocr_hits[:max_imgs] if max_imgs > 0 else []

                    local_texts, local_pairs = ocr_paths(sel)



                    # OCR 결과 수집 (저장용) - 항상 수집

                    ocr_text_combined = " ".join(local_texts) if local_texts else "(OCR 텍스트 없음)"

                    ocr_results_list.append({

                        "GS코드": gs_code9,

                        "상품명": base_name,

                        "OCR처리이미지수": len(sel),

                        "전체이미지수": matched_count,

                        "대표이미지수": len(listing_hits),

                        "OCR텍스트": ocr_text_combined[:500],  # 500자로 제한

                        "OCR이미지": "; ".join([os.path.basename(p) for p in sel[:5]])  # 처음 5개만

                    })

                elif matched_count > 0:

                    # OCR용 이미지는 없지만 대표이미지는 있는 경우

                    ocr_results_list.append({

                        "GS코드": gs_code9,

                        "상품명": base_name,

                        "OCR처리이미지수": 0,

                        "전체이미지수": matched_count,

                        "대표이미지수": len(listing_hits),

                        "OCR텍스트": "(숫자 파일명 없음 - OCR 불가)",

                        "OCR이미지": "없음"

                    })



            # AU열 (이미지등록(상세)) 순차 다운로드 - 대표이미지용

            # 로컬에 이미지가 없고 AU열에 URL이 있으면 순차 다운로드 (phase=analysis에서는 건너뜀)

            if cfg.phase != "analysis" and listing_img_col and listing_img_col in df_after.columns and gs_code9:

                listing_url = str(df_after.at[idx, listing_img_col]) if pd.notna(df_after.at[idx, listing_img_col]) else ""

                if listing_url and listing_url.startswith("http"):

                    # 로컬에 GS코드 이미지가 없을 때만 다운로드

                    _gs_low_check = gs_code9.lower()

                    _existing_gs_imgs = [p for p in global_listing_sources if _gs_low_check in os.path.basename(p).lower()]

                    if not _existing_gs_imgs:

                        from app.services.ocr_pipeline import _download_sequential_images

                        downloaded = _download_sequential_images(

                            base_url=listing_url,

                            gs_code=f"{gs_code9}A",  # GS코드A 폴더에 저장

                            target_dir=ocr_temp_root,

                            max_fails=3,

                            max_images=100

                        )

                        if downloaded:

                            global_listing_sources.extend(downloaded)

                            _status(status_cb, f"[AU다운] {gs_code9}: {len(downloaded)}개 이미지 다운로드")


            # phase=images → 이미지 수집만 하고 OCR/Vision/키워드 건너뜀
            if cfg.phase == "images":
                continue

            detail_html = str(df_after.at[idx, detail_col]) if (detail_col and detail_col in df_after.columns) else ""

            html_text = re.sub(r"\s+", " ", core.extract_text_from_html(detail_html)) if detail_html else ""

            if korean_only and html_text:

                html_text = re.sub(r"[^가-힣\s]", " ", html_text)



            url_texts = []

            if cfg.merge_ocr_with_name and cfg.use_html_ocr and detail_html:

                srcs = core.extract_img_srcs(detail_html, max_images=max_imgs)

                for url in srcs:

                    ocr_txt = core.ocr_image_url(url, tess_lang, timeout=10, psm=psm, oem=oem, korean_only=korean_only)

                    if ocr_txt:

                        url_texts.append(ocr_txt)



            # OCR Excel 에서 raw 텍스트를 최우선으로 GPT에 전달

            ocr_raw_text = ""

            sum_text = ""

            _ocr_key = None

            if ocr_lookup and gs_code9:

                # GSxxxxxxxA 형태로도 매칭 시도

                for _try_key in [gs_code9, f"{gs_code9}A"]:

                    if _try_key in ocr_lookup:

                        _ocr_key = _try_key

                        break



            # OCR Excel이 로드된 상태에서 해당 상품의 OCR 데이터가 없으면 → 상세 없음 → 스킵

            if ocr_lookup and gs_code9:

                _has_ocr = False

                if _ocr_key:

                    _raw_check = ocr_lookup[_ocr_key].get("raw", "").strip()

                    _has_ocr = bool(_raw_check) and _raw_check != "(옵션상품 — A옵션 결과 참조)"

                if not _has_ocr:

                    no_detail_indices.append(idx)

                    _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — 상세 없음, 스킵")

                    continue



            if _ocr_key and cfg.merge_ocr_with_name:

                ocr_raw_text = ocr_lookup[_ocr_key].get("raw", "")

                # 반복 문구 필터링 후 raw 텍스트를 최대 500자까지 GPT에 직접 전달

                if ocr_raw_text:

                    from app.services.ocr_noise_filter import filter_ocr_text, preprocess_ocr_for_llm

                    _raw0 = ocr_raw_text

                    _f = filter_ocr_text(_raw0)

                    _p = preprocess_ocr_for_llm(_f)

                    ocr_raw_text = _p if _p else (_f if _f else _raw0)

                sum_text = ocr_raw_text[:500] if ocr_raw_text else ""

            elif cfg.merge_ocr_with_name:

                detail_texts = []

                if local_texts:

                    detail_texts.extend(local_texts)

                if html_text:

                    detail_texts.append(html_text)

                if url_texts:

                    detail_texts.extend(url_texts)

                _combined_ocr = " ".join(detail_texts)

                # 반복 문구 필터링 (OCR Excel 경로와 동일하게 적용)

                if _combined_ocr:

                    try:

                        from app.services.ocr_noise_filter import filter_ocr_text, preprocess_ocr_for_llm

                        _raw0 = _combined_ocr

                        _f = filter_ocr_text(_raw0)

                        _p = preprocess_ocr_for_llm(_f)

                        _combined_ocr = _p if _p else (_f if _f else _raw0)

                    except Exception:

                        pass

                # OCR 원문을 500자까지 직접 GPT에 전달 (summarize_features_tokens 미사용)

                _combined_ocr = re.sub(r"\s+", " ", _combined_ocr).strip()

                sum_text = _combined_ocr[:500] if _combined_ocr else ""

            else:

                sum_text = ""





            # 디버그: GPT에 전달되는 OCR 텍스트 확인

            if sum_text:

                _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — GPT에 OCR {len(sum_text)}자 전달")

            else:

                _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — ⚠ GPT에 OCR 텍스트 없음! (local_texts={len(local_texts)}, html_text={len(html_text) if html_text else 0})")



            gpt_model_kw = cfg.model_keyword

            gpt_model_lt = cfg.model_longtail

            stage2_model = cfg.model_keyword_stage2 or gpt_model_kw



            # 대표이미지 Vision 분석 복원: category/구조/설치/재질/색상/유형교정을 추정

            _vision_analysis = {}

            _vision_hint_parts = []

            # OCR 텍스트가 충분하면(300자 이상) Vision 스킵 → 토큰 절약
            _skip_vision = len(sum_text) >= 300 or cfg.phase == "ocr_only"

            if gs_code9 and not _skip_vision:

                _gs_low_v = gs_code9.lower()

                _vision_imgs = [p for p in global_listing_sources if _gs_low_v in os.path.basename(p).lower() and os.path.isfile(p)][:3]

                if _vision_imgs:

                    _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — Vision 분석 중 ({len(_vision_imgs)}장)...")

                    _vision_analysis = analyze_product_images_local(_vision_imgs, prompt_product_name, gpt_model_kw)

                    if _vision_analysis:

                        _vision_hint_parts, _filled_paths = _extract_vision_hints(_vision_analysis)

                        _status(

                            status_cb,

                            f"[{row_i}/{total_rows}] {gs_code9} — Vision 채움 항목: {len(_filled_paths)}개"

                            + (f" ({', '.join(_filled_paths[:8])})" if _filled_paths else "")

                        )



            _sum_text_with_vision = sum_text

            if _vision_hint_parts:

                _vision_hint_text = " ".join(dict.fromkeys(_vision_hint_parts))[:300]

                _sum_text_with_vision = (f"{sum_text} {_vision_hint_text}").strip()



            _vision_payload = _vision_excel_payload(_vision_analysis, _vision_hint_parts)

            for _k, _v in _vision_payload.items():

                df_after.at[idx, _k] = _v

            _stamp_vision_to_ocr_results(gs_code9, _vision_payload)



            # Vision JSON이 실제로 채워졌을 때만 로컬 키워드 빌더 사용.

            # Vision이 비어 있으면 OCR 안내문이 그대로 토큰화될 수 있어 1차 GPT 경로를 우선한다.

            kw_line = ""

            kw_tokens = []

            gpt_err = ""

            # ocr_only 모드: LLM 키워드 생성 전부 스킵
            if cfg.phase == "ocr_only":
                search_keywords = ""
                _status(status_cb, f"[{row_i}/{total_rows}] {gs_code9} — OCR only 모드: 키워드 생성 스킵")
                # 후처리까지 건너뛰기 위해 아래 블록 전체 스킵
                kw_tokens = [t for t in re.split(r"\s+", prompt_product_name) if t][:5]
                kw_line = " ".join(kw_tokens)

                # 이 행의 나머지 키워드/검색어 처리는 불필요하므로 continue 전에 OCR요약만 저장
                if "OCR요약" in df_after.columns:
                    df_after.at[idx, "OCR요약"] = sum_text[:500] if sum_text else ""
                df_after.at[idx, "검색키워드"] = ""
                df_after.at[idx, "검색어설정"] = ""

                debug_rows.append({
                    "GS코드": gs_code9 or "",
                    "상품명": base_name,
                    "OCR길이": len(sum_text),
                    "모드": "ocr_only",
                })
                continue

            builder_source_text = _sum_text_with_vision
            if use_keyword_v2 and str(ocr_raw_text or "").strip():
                builder_source_text = ocr_raw_text

            builder_line = ""
            try:
                builder_line = _build_builder_line(
                    base_name_value=base_name,
                    option_text_value=option_text,
                    ocr_text_value=builder_source_text,
                    vision_analysis_value=_vision_analysis,
                    market_value="A",
                )
            except Exception:
                builder_line = ""

            if use_keyword_v2 and gpt_model_kw == "없음":
                kw_line = builder_line
                kw_tokens = [t for t in re.split(r"\s+", kw_line) if t] if kw_line else []
                gpt_err = ""
            else:
                # GPT 프롬프트(최적화됨)를 우선 사용. Vision 데이터는 context로 전달.
                kw_line, kw_tokens = core.generate_keyword_gpt(
                    prompt_product_name, _sum_text_with_vision, gpt_model_kw, max_words, max_len, min_len,
                    vision_analysis=_vision_analysis,
                )
                gpt_err = getattr(core, "LAST_GPT_ERROR", "")
                if gpt_err:
                    _status(status_cb, f"GPT 오류: {gpt_err}")
                    if _is_fatal_gpt_error(gpt_err):
                        fatal_gpt_error = gpt_err
                        _status(status_cb, "치명적 GPT 오류(404) 감지: 저장 없이 작업 중단")
                        break

            base_anchors, base_baseline = _resolve_topic_refs(base_name)
            kw_quality = core.keyword_local_score(
                kw_line,
                base_name=base_name,
                anchors=base_anchors,
                baseline=base_baseline,
            ) if kw_line else -999

            kw_noise = _keyword_noise_score(kw_line)

            # GPT 실패 시 keyword_builder → heuristic fallback
            if not kw_line or len(kw_tokens) < min_kw_tokens or kw_quality < kw_quality_gate or (use_keyword_v2 and kw_noise >= 6):
                try:
                    builder_tokens = [t for t in re.split(r"\s+", builder_line) if t] if builder_line else []
                    builder_quality = core.keyword_local_score(
                        builder_line,
                        base_name=base_name,
                        anchors=base_anchors,
                        baseline=base_baseline,
                    ) if builder_line else -999
                    _builder_noise = _keyword_noise_score(builder_line)
                    _current_noise = _keyword_noise_score(kw_line)
                    if builder_line and ((builder_quality >= kw_quality or _builder_noise + 2 <= _current_noise or (len(builder_tokens) >= max(3, len(kw_tokens) - 3) and _builder_noise < _current_noise)) if use_keyword_v2 else (builder_quality >= (kw_quality - 4) or len(builder_tokens) >= len(kw_tokens))):
                        kw_line = builder_line
                        kw_tokens = builder_tokens
                        kw_quality = builder_quality
                except Exception:
                    pass

            if not kw_line:
                kw_line = builder_line or core._fallback_heuristic(prompt_product_name, _sum_text_with_vision, max_n=max_words)
                kw_tokens = [t for t in re.split(r"\s+", kw_line) if t]



            search_keywords = ""

            kw_tokens = _prune_v2_tokens(kw_tokens, base_name_value=base_name, option_token_set=option_tokens)

            kw_tokens = [t for t in kw_tokens if t not in core.SIZE_WORDS and t not in core.STOPWORDS]

            if drop_digits:

                kw_tokens = core._filter_tokens_drop_digits(kw_tokens)



            # 네이버 쇼핑 자동완성 키워드 조합

            ac_keywords_debug = []

            if cfg.naver_autocomplete:

                try:

                    ac_raw = core.get_autocomplete_keywords_for_product(base_name, max_queries=2, max_results=10)

                    if ac_raw:

                        anchors_for_ac = core.build_anchors_from_name(base_name)

                        baseline_for_ac = core.build_baseline_tokens_from_name(base_name)

                        ac_cleaned = core.clean_naver_kw_list(ac_raw, anchors=anchors_for_ac, baseline=baseline_for_ac)

                        ac_keywords_debug = list(ac_cleaned)

                        # 자동완성 키워드를 토큰화하여 기존 kw_tokens에 없는 것만 추가

                        existing = set(kw_tokens)

                        for ac_kw in ac_cleaned:

                            ac_toks = core.tokenize_korean_words(ac_kw)

                            for t in ac_toks:

                                if t not in existing and t not in core.STOPWORDS and t not in core.SIZE_WORDS:

                                    if len(t) >= 2:

                                        existing.add(t)

                                        kw_tokens.append(t)

                        _status(status_cb, f"네이버자동완성: {base_name[:20]} → {len(ac_cleaned)}개")

                except Exception as e:

                    _status(status_cb, f"네이버자동완성 오류: {e}")



            # 구글 자동완성 키워드 조합

            google_ac_debug = []

            if cfg.google_autocomplete:

                try:

                    gac_raw = core.get_google_autocomplete_for_product(base_name, max_queries=2, max_results=10)

                    if gac_raw:

                        anchors_for_gac = core.build_anchors_from_name(base_name)

                        baseline_for_gac = core.build_baseline_tokens_from_name(base_name)

                        gac_cleaned = core.clean_naver_kw_list(gac_raw, anchors=anchors_for_gac, baseline=baseline_for_gac)

                        google_ac_debug = list(gac_cleaned)

                        existing = set(kw_tokens)

                        for gac_kw in gac_cleaned:

                            gac_toks = core.tokenize_korean_words(gac_kw)

                            for t in gac_toks:

                                if t not in existing and t not in core.STOPWORDS and t not in core.SIZE_WORDS:

                                    if len(t) >= 2:

                                        existing.add(t)

                                        kw_tokens.append(t)

                        _status(status_cb, f"구글자동완성: {base_name[:20]} → {len(gac_cleaned)}개")

                except Exception as e:

                    _status(status_cb, f"구글자동완성 오류: {e}")



            kw_line = " ".join(kw_tokens)



            _sparse_title_mode = (not _vision_analysis) and (not _has_meaningful_title_evidence(sum_text))
            if use_keyword_v2:
                final_line = _compose_v2_title(
                    base_name_value=base_name,
                    kw_line_value=kw_line,
                    option_text_value=option_text,
                    max_len_value=max_len,
                    market_value="A",
                )
            elif _sparse_title_mode and kw_line:
                final_line = core.normalize_space(str(kw_line or "")).strip()[:max_len]
            else:
                final_line = core.merge_base_name_with_keywords(base_name, kw_line, max_words, max_len, option_tokens=option_tokens, ocr_text=sum_text)



            # min_len 미달 시 OCR 텍스트에서 직접 보충

            if (not use_keyword_v2) and len(final_line) < min_len and sum_text:

                _ocr_sup, _ocr_toks = core.postprocess_keywords_tokens(sum_text, max_words=max_words, max_len=max_len)

                _existing = set(final_line.split())

                _final_tokens = final_line.split()

                for _t in _ocr_toks:

                    _allow_ocr_token = False

                    if (
                        _t not in _existing
                        and len(_t) >= 2
                        and _t not in core.STOPWORDS
                        and _t not in core.SIZE_WORDS
                    ):

                        if use_keyword_v2:

                            _allow_ocr_token = core.is_on_topic(_t, base_anchors, base_baseline)

                        else:

                            _allow_ocr_token = True

                    if _allow_ocr_token:

                        _existing.add(_t)

                        _final_tokens.append(_t)

                    if len(" ".join(_final_tokens)) >= min_len or len(_final_tokens) >= max_words:

                        break

                final_line = " ".join(_final_tokens)[:max_len].rstrip()



            df_after.at[idx, name_col] = final_line

            df_after.at[idx, "1차키워드"] = final_line

            df_after.at[idx, "최종키워드2차"] = final_line

            df_after.at[idx, "OCR요약"] = (sum_text or "")[:500]

            df_after.at[idx, "네이버검색광고데이터"] = ""

            df_after.at[idx, "검색키워드"] = search_keywords  # 검색 키워드 추가

            # B마켓 상품명 생성 (설정된 글자수)
            if cfg.enable_b_market:
                try:
                    b_kw_seed = _build_builder_line(
                        base_name_value=base_name,
                        option_text_value=option_text,
                        ocr_text_value=builder_source_text,
                        vision_analysis_value=_vision_analysis,
                        market_value="B",
                    )
                except Exception:
                    b_kw_seed = ""
                b_final = _compose_b_market_title(
                    base_name=base_name,
                    kw_line=kw_line,
                    option_tokens=option_tokens,
                    option_text=option_text,
                    ocr_text=sum_text,
                    b_kw_line=b_kw_seed,
                    prefer_seed_only=_sparse_title_mode,
                    a_final_name=final_line,
                )
                df_after.at[idx, "B_상품명"] = b_final
                if idx not in b_market_rows:
                    b_market_rows[idx] = {}
                b_market_rows[idx]["상품명"] = b_final



            if bool(cfg.write_to_r):

                TARGET_N = 20



                anchors, baseline = _resolve_topic_refs(base_name, final_line)



                lt10_raw = core.generate_longtail10(base_name, sum_text, client=core.client, model_name=gpt_model_lt)

                lt10 = core.clean_naver_kw_list(lt10_raw, anchors=anchors, baseline=baseline)

                if len(lt10) < 10:

                    name_parts = [p for p in re.sub(r"[^0-9가-힣\sA-Za-z]", " ", base_name).split() if len(p) >= 2]

                    bigrams = [name_parts[i] + name_parts[i + 1] for i in range(len(name_parts) - 1)]

                    lt10_backup = core.clean_naver_kw_list(bigrams, anchors=anchors, baseline=baseline)

                    for k in lt10_backup:

                        if k not in lt10:

                            lt10.append(k)

                        if len(lt10) >= 10:

                            break

                lt10 = lt10[:10]



                naver_pc5, naver_mo5 = [], []

                items, hint_used, naver_err = _query_naver_two_pass(final_line, base_name)



                if items:

                    pc_list = core.rank_and_pick_with_ctr(items, platform="pc", want=5, ctr_threshold=CTR_THR)

                    mo_list = core.rank_and_pick_with_ctr(items, platform="mobile", want=5, ctr_threshold=CTR_THR)

                    pc_list = core.clean_naver_kw_list(pc_list, anchors=anchors, baseline=baseline)

                    mo_list = core.clean_naver_kw_list(mo_list, anchors=anchors, baseline=baseline)



                    inter = set(pc_list) & set(mo_list)

                    if inter:

                        pc_list = [k for k in pc_list if k not in inter]



                    if len(pc_list) < 5:

                        backup_pc = core.rank_and_pick_with_ctr(items, platform="pc", want=15, ctr_threshold=0.0)

                        backup_pc = core.clean_naver_kw_list(backup_pc, anchors=anchors, baseline=baseline)

                        for k in backup_pc:

                            if k not in pc_list and k not in mo_list:

                                pc_list.append(k)

                            if len(pc_list) >= 5:

                                break

                    if len(mo_list) < 5:

                        backup_mo = core.rank_and_pick_with_ctr(items, platform="mobile", want=15, ctr_threshold=0.0)

                        backup_mo = core.clean_naver_kw_list(backup_mo, anchors=anchors, baseline=baseline)

                        for k in backup_mo:

                            if k not in mo_list and k not in pc_list:

                                mo_list.append(k)

                            if len(mo_list) >= 5:

                                break



                    naver_pc5 = pc_list[:5]

                    naver_mo5 = mo_list[:5]



                final_kw, seen = [], set()



                def _clean_search_terms(lst):

                    cleaned = []

                    seen_local = set()

                    for w in (lst or []):

                        k = re.sub(r"\s+", "", str(w or ""))

                        if not k:

                            continue

                        # 조사/어미/문장형 파편 제거

                        if re.search(r"(하다|하는|되어|됨|하고|하기)$", k):

                            continue

                        if re.search(r"(에|에서|으로|로|을|를|이|가|은|는|의|와|과)$", k):

                            continue

                        if re.search(r"[가-힣](에|에서|으로|로|을|를|이|가|은|는|의|와|과)[가-힣]{2,}", k):

                            continue

                        if any(x in k for x in ["설치하고", "차량에", "조명을"]):

                            continue

                        # 판매처/브랜드 워터마크 파편 제거 (예: 홈런마켓브라켓)

                        if any(x in k for x in ["마켓", "스토어", "쇼핑몰", "샵", "몰"]):

                            continue

                        # 비자연 결합어 제거 (의미 약한 합성 파편)

                        if any(x in k for x in ["금속플라스틱", "플라스틱검정", "검정은색", "은색실외", "볼트외부", "외부금속"]):

                            continue

                        # 평가/형용 파편 결합어 제거

                        if re.search(r"(견고한|간편한|강력한|튼튼한|편리한)$", k):

                            continue

                        if re.search(r"(견고한|간편한|강력한|튼튼한|편리한)", k) and len(k) >= 6:

                            if not any(x in k for x in ["브라켓", "마운트", "거치대", "고정", "볼트", "조명"]):

                                continue

                        if "간편설치" in k and len(k) >= 6:

                            continue

                        if "외부간편" in k:

                            continue

                        if re.search(r"설치[가-힣]{2,}", k):

                            continue

                        if any(x in k for x in ["간편설치견고한", "견고한간편한", "간편한금속"]):

                            continue

                        # 재질/색상 단어 2개 이상을 붙여쓴 기계적 결합어 제거

                        mat_color = ["금속", "플라스틱", "스틸", "알루미늄", "철제", "검정", "은색", "실버", "블랙"]

                        hit = sum(1 for t in mat_color if t in k)

                        if hit >= 2 and len(k) >= 6:

                            continue

                        lk = k.lower()

                        if lk in seen_local:

                            continue

                        seen_local.add(lk)

                        cleaned.append(k)

                    return cleaned



                def push(lst):

                    for w in lst:

                        w = core._clean_one_kw(w)

                        if not w:

                            continue

                        if any(b in w for b in core.BAN):

                            continue

                        if not (2 <= len(w) <= 12):

                            continue

                        if not core.is_on_topic(w, anchors, baseline):

                            continue

                        if w in seen:

                            continue

                        seen.add(w)

                        final_kw.append(w)

                        if len(final_kw) >= TARGET_N:

                            break



                push(naver_pc5)

                push(naver_mo5)

                push(lt10)



                if len(final_kw) < TARGET_N and items:

                    backup_mix = core.rank_and_pick_with_ctr(items, platform="pc", want=30, ctr_threshold=0.0)

                    backup_mix += core.rank_and_pick_with_ctr(items, platform="mobile", want=30, ctr_threshold=0.0)

                    backup_mix = core.clean_naver_kw_list(backup_mix, anchors=anchors, baseline=baseline)

                    push(backup_mix)



                if len(final_kw) < TARGET_N:

                    name_parts = [p for p in re.sub(r"[^0-9가-힣\sA-Za-z]", " ", base_name).split() if len(p) >= 2]

                    extra = []

                    for i in range(len(name_parts) - 1):

                        extra.append(name_parts[i] + name_parts[i + 1])

                    for i in range(len(name_parts) - 2):

                        extra.append(name_parts[i] + name_parts[i + 1] + name_parts[i + 2])

                    extra = core.clean_naver_kw_list(extra, anchors=anchors, baseline=baseline)

                    push(extra)



                if len(final_kw) < TARGET_N and core.USE_GPT_BACKFILL:

                    fallback = core._fallback_heuristic(base_name, sum_text, max_n=TARGET_N)

                    push([x for x in fallback.split(",") if x])



                final_kw = _clean_search_terms(final_kw)

                df_after.at[idx, "검색어설정"] = ",".join(final_kw[:TARGET_N])

                naver_info = (

                    f"hint={hint_used} | "

                    f"PC5={','.join(naver_pc5)} | "

                    f"MO5={','.join(naver_mo5)}"

                )

                if naver_err:

                    naver_info += f" | 오류={naver_err[:120]}"

                naver_table = _format_naver_keyword_table(items, limit=15)

                df_after.at[idx, "네이버검색광고데이터"] = naver_table if naver_table else naver_info



                stage2_source = naver_table if naver_table else naver_info

                if use_keyword_v2 and (not stage2_model or stage2_model == "없음"):

                    stage2_kw, _stage2_tokens = "", []

                else:

                    stage2_kw, _stage2_tokens = core.generate_keyword_stage2(

                        seed_keywords=final_line,

                        naver_keyword_table=stage2_source,

                        ocr_text=sum_text,

                        model_name=stage2_model,

                        min_len=50,

                        max_len=max_len if use_keyword_v2 else min(max_len, 90),

                        max_words=max_words,

                    )

                stage2_err = getattr(core, "LAST_GPT_ERROR", "")

                if stage2_err and _is_fatal_gpt_error(stage2_err):

                    fatal_gpt_error = stage2_err

                    _status(status_cb, f"치명적 GPT 오류(2차): {stage2_err}")

                    break

                if stage2_kw:

                    _st2_ok = True

                    _st2_toks = core.tokenize_korean_words(stage2_kw)

                    _base_toks = core.tokenize_korean_words(base_name)

                    _seed_score = core.keyword_local_score(final_line, base_name=base_name, anchors=anchors, baseline=baseline)

                    _stage2_score = core.keyword_local_score(stage2_kw, base_name=base_name, anchors=anchors, baseline=baseline)

                    if len(_st2_toks) < 2:

                        _st2_ok = False

                    if _base_toks and core.semantic_overlap_count(_st2_toks, _base_toks) == 0:

                        _st2_ok = False

                    if not core.is_on_topic(stage2_kw, anchors, baseline):

                        _st2_ok = False

                    if use_keyword_v2 and _stage2_score < _seed_score:

                        _st2_ok = False

                    if (not use_keyword_v2) and _stage2_score < (_seed_score - 4):

                        _st2_ok = False

                    if _st2_ok:

                        df_after.at[idx, "최종키워드2차"] = stage2_kw

                        df_after.at[idx, name_col] = stage2_kw

                    else:

                        stage2_kw = ""

                final_for_search = stage2_kw if stage2_kw else final_line

                market_pkg = _apply_market_keyword_packages(

                    row_idx=idx,

                    product_name=final_for_search,

                    source_text=_sum_text_with_vision,

                    naver_keyword_table=stage2_source,

                    model_name=gpt_model_lt,

                    anchors=anchors,

                    baseline=baseline,

                    market="A",
                    avoid_terms=final_line,

                )

                search_keywords = market_pkg.search_keywords or search_keywords

                # B마켓 전용 상품명 생성 (토큰 순서/글자수 다르게)
                try:
                    b_kw_line = _build_builder_line(
                        base_name_value=base_name,
                        option_text_value=option_text,
                        ocr_text_value=builder_source_text,
                        vision_analysis_value=_vision_analysis,
                        market_value="B",
                    )
                except Exception:
                    b_kw_line = ""

                b_final_line = _compose_b_market_title(
                    base_name=base_name,
                    kw_line=kw_line,
                    option_tokens=option_tokens,
                    option_text=option_text,
                    ocr_text=sum_text,
                    b_kw_line=b_kw_line,
                    prefer_seed_only=_sparse_title_mode,
                    a_final_name=final_line,
                )
                df_after.at[idx, "B_상품명"] = b_final_line
                if cfg.enable_b_market:
                    if idx not in b_market_rows:
                        b_market_rows[idx] = {}
                    b_market_rows[idx]["상품명"] = b_final_line

                # B마켓 키워드도 함께 생성
                _apply_market_keyword_packages(

                    row_idx=idx,

                    product_name=b_final_line,

                    source_text=_sum_text_with_vision,

                    naver_keyword_table=stage2_source,

                    model_name=gpt_model_lt,

                    anchors=anchors,

                    baseline=baseline,

                    market="B",
                    avoid_terms=final_line,

                )



                if debug_on:

                    debug_rows.append({

                        "R열_타깃개수": TARGET_N,

                        "R열_최종개수": len(final_kw),

                        "네이버_DRY": core.DRY_RUN,

                        "네이버_사용": "Y" if cfg.naver_enabled else "N",

                        "네이버_오류": naver_err,

                        "네이버_hintKeywords": hint_used,

                        "네이버_PC5": ",".join(naver_pc5),

                        "네이버_MO5": ",".join(naver_mo5),

                        "롱테일10": ",".join(lt10),

                        "앵커": ",".join(sorted(anchors)),

                        "베이스라인": ",".join(sorted(baseline)),

                        "검색어설정(R)": df_after.at[idx, "검색어설정"],

                    })

            else:

                anchors, baseline = _resolve_topic_refs(base_name, final_line)



                naver_pc5, naver_mo5 = [], []

                items, hint_used, naver_err = _query_naver_two_pass(final_line, base_name)



                if items:

                    pc_list = core.rank_and_pick_with_ctr(items, platform="pc", want=5, ctr_threshold=CTR_THR)

                    mo_list = core.rank_and_pick_with_ctr(items, platform="mobile", want=5, ctr_threshold=CTR_THR)

                    naver_pc5 = core.clean_naver_kw_list(pc_list, anchors=anchors, baseline=baseline)[:5]

                    naver_mo5 = core.clean_naver_kw_list(mo_list, anchors=anchors, baseline=baseline)[:5]



                naver_info = (

                    f"hint={hint_used} | "

                    f"PC5={','.join(naver_pc5)} | "

                    f"MO5={','.join(naver_mo5)}"

                )

                if naver_err:

                    naver_info += f" | 오류={naver_err[:120]}"

                naver_table = _format_naver_keyword_table(items, limit=15)

                df_after.at[idx, "네이버검색광고데이터"] = naver_table if naver_table else naver_info



                stage2_source = naver_table if naver_table else naver_info

                if use_keyword_v2 and (not stage2_model or stage2_model == "없음"):

                    stage2_kw, _stage2_tokens = "", []

                else:

                    stage2_kw, _stage2_tokens = core.generate_keyword_stage2(

                        seed_keywords=final_line,

                        naver_keyword_table=stage2_source,

                        ocr_text=sum_text,

                        model_name=stage2_model,

                        min_len=50,

                        max_len=max_len if use_keyword_v2 else min(max_len, 90),

                        max_words=max_words,

                    )

                stage2_err = getattr(core, "LAST_GPT_ERROR", "")

                if stage2_err and _is_fatal_gpt_error(stage2_err):

                    fatal_gpt_error = stage2_err

                    _status(status_cb, f"치명적 GPT 오류(2차): {stage2_err}")

                    break

                if stage2_kw:

                    _st2_ok = True

                    _st2_toks = core.tokenize_korean_words(stage2_kw)

                    _base_toks = core.tokenize_korean_words(base_name)

                    _seed_score = core.keyword_local_score(final_line, base_name=base_name, anchors=anchors, baseline=baseline)

                    _stage2_score = core.keyword_local_score(stage2_kw, base_name=base_name, anchors=anchors, baseline=baseline)

                    if len(_st2_toks) < 2:

                        _st2_ok = False

                    if _base_toks and core.semantic_overlap_count(_st2_toks, _base_toks) == 0:

                        _st2_ok = False

                    if not core.is_on_topic(stage2_kw, anchors, baseline):

                        _st2_ok = False

                    if use_keyword_v2 and _stage2_score < _seed_score:

                        _st2_ok = False

                    if (not use_keyword_v2) and _stage2_score < (_seed_score - 4):

                        _st2_ok = False

                    if _st2_ok:

                        df_after.at[idx, "최종키워드2차"] = stage2_kw

                        df_after.at[idx, name_col] = stage2_kw

                    else:

                        stage2_kw = ""

                final_for_search = stage2_kw if stage2_kw else final_line

                market_pkg = _apply_market_keyword_packages(

                    row_idx=idx,

                    product_name=final_for_search,

                    source_text=_sum_text_with_vision,

                    naver_keyword_table=stage2_source,

                    model_name=gpt_model_lt,

                    anchors=anchors,

                    baseline=baseline,

                    market="A",
                    avoid_terms=final_line,

                )

                search_keywords = market_pkg.search_keywords or search_keywords

                # B마켓 전용 상품명 생성 (토큰 순서/글자수 다르게)
                try:
                    b_kw_line = _build_builder_line(
                        base_name_value=base_name,
                        option_text_value=option_text,
                        ocr_text_value=builder_source_text,
                        vision_analysis_value=_vision_analysis,
                        market_value="B",
                    )
                except Exception:
                    b_kw_line = ""

                b_final_line = _compose_b_market_title(
                    base_name=base_name,
                    kw_line=kw_line,
                    option_tokens=option_tokens,
                    option_text=option_text,
                    ocr_text=sum_text,
                    b_kw_line=b_kw_line,
                    prefer_seed_only=_sparse_title_mode,
                    a_final_name=final_line,
                )
                df_after.at[idx, "B_상품명"] = b_final_line
                if cfg.enable_b_market:
                    if idx not in b_market_rows:
                        b_market_rows[idx] = {}
                    b_market_rows[idx]["상품명"] = b_final_line

                # B마켓 키워드도 함께 생성
                _apply_market_keyword_packages(

                    row_idx=idx,

                    product_name=b_final_line,

                    source_text=_sum_text_with_vision,

                    naver_keyword_table=stage2_source,

                    model_name=gpt_model_lt,

                    anchors=anchors,

                    baseline=baseline,

                    market="B",
                    avoid_terms=final_line,

                )



            if debug_on:

                ocr_samples = "; ".join([f"{fn}:{snip}" for fn, snip in (local_pairs[:5] if local_pairs else [])])

                debug_rows.append({

                    "상품명(원본)": full_pname,

                    "기본상품명(옵션제외)": base_name,

                    "옵션(본문)": option_text,

                    "옵션_토큰": " ".join(sorted(option_tokens)),

                    "적용 GS코드": gs_code9 or "",

                    "로컬_매칭수": matched_count,

                    "OCR샘플": ocr_samples,

                    "HTML텍스트": html_text[:200] if html_text else "",

                        "상세요약": sum_text[:220],

                        "GPT_ERROR": gpt_err,

                        "로컬빌더": builder_line,

                        "키워드정렬": kw_line,

                        "검색키워드": search_keywords,

                        "네이버자동완성": ",".join(ac_keywords_debug) if ac_keywords_debug else "",

                        "구글자동완성": ",".join(google_ac_debug) if google_ac_debug else "",

                        "최종상품명": final_line,

                    })

        except Exception as e:

            if debug_on:

                debug_rows.append({"오류행": int(idx), "오류": f"{type(e).__name__}: {e}"})

            continue

        if row_i % 5 == 0 or row_i == total_rows:

            pct = 10 + int((row_i / total_rows) * 55)

            _progress(progress_cb, min(65, pct))



    if fatal_gpt_error:

        raise RuntimeError(f"치명적 GPT 오류로 저장 중단: {fatal_gpt_error}")



    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    save_path = os.path.join(export_root, f"상품전처리GPT_{keyword_version_slug}_{timestamp}.xlsx")

    def _seq_path(directory: str, prefix: str, ext: str) -> str:

        """순차 번호 파일명 생성: prefix_01.ext, prefix_02.ext, ..."""

        seq = 1

        while True:

            name = f"{prefix}_{seq:02d}{ext}"

            path = os.path.join(directory, name)

            if not os.path.exists(path):

                return path

            seq += 1



    # 상세 없는 상품은 업로드용에서 제외

    df_upload = df_after.loc[~df_after.index.isin(no_detail_indices)].copy() if no_detail_indices else df_after

    upload_cols = [c for c in input_cols if c in df_upload.columns]

    _market_keyword_output_cols = list(MARKET_KEYWORD_COLUMNS_10) + ["홈런_네이버태그"]

    extra_output_cols = [

        c for c in ["1차키워드", "최종키워드2차", "OCR요약", "검색키워드", "검색어설정", "쿠팡검색태그", "네이버태그", *_market_keyword_output_cols, "네이버검색광고데이터", "옵션추가금"]

        if c in df_upload.columns and c not in upload_cols

    ]

    export_cols = upload_cols + extra_output_cols

    df_upload_export = df_upload.loc[:, export_cols].copy() if export_cols else df_upload.copy()

    # ── B마켓 시트용 DataFrame 생성 ──
    b_extra_cols = [
        c for c in ["B_상품명", "1차키워드", "최종키워드2차", "OCR요약", "B_검색키워드", "B_검색어설정", "B_쿠팡검색태그", "B_네이버태그", "네이버검색광고데이터", "옵션추가금"]
        if c in df_upload.columns and c not in upload_cols
    ]
    b_export_cols = upload_cols + b_extra_cols
    df_upload_export_b = df_upload.loc[:, [c for c in b_export_cols if c in df_upload.columns]].copy()
    _has_b_name_export = "B_상품명" in df_upload_export_b.columns
    # B_ 접두어 제거하여 A시트와 동일한 컬럼명 사용
    df_upload_export_b = df_upload_export_b.rename(columns={
        "B_검색키워드": "검색키워드",
        "B_검색어설정": "검색어설정",
        "B_쿠팡검색태그": "쿠팡검색태그",
        "B_네이버태그": "네이버태그",
    })



    upload_path = _seq_path(export_root, f"업로드용_{date_tag}_{keyword_version_slug}", ".xlsx")

    # B마켓 DataFrame 생성
    df_b_market = None
    if cfg.enable_b_market:
        if not df_upload_export_b.empty:
            df_b_market = df_upload_export_b.copy()
            if "B_상품명" in df_b_market.columns:
                if name_col in df_b_market.columns:
                    _b_name_series = df_b_market["B_상품명"].astype("string").fillna("").str.strip()
                    df_b_market[name_col] = _b_name_series.where(_b_name_series != "", df_b_market[name_col])
                df_b_market.drop(columns=["B_상품명"], inplace=True, errors="ignore")
        if b_market_rows:
            if df_b_market is None:
                df_b_market = df_upload_export.copy()
        for idx, b_data in b_market_rows.items():
            if idx in df_b_market.index:
                if "상품명" in b_data and name_col in df_b_market.columns and not _has_b_name_export:
                    df_b_market.at[idx, name_col] = b_data["상품명"]
                if "검색어설정" in b_data and "검색어설정" in df_b_market.columns:
                    df_b_market.at[idx, "검색어설정"] = b_data["검색어설정"]
                if "검색키워드" in b_data and "검색키워드" in df_b_market.columns:
                    df_b_market.at[idx, "검색키워드"] = b_data["검색키워드"]
                if "네이버태그" in b_data and "네이버태그" in df_b_market.columns:
                    df_b_market.at[idx, "네이버태그"] = b_data["네이버태그"]
        # B마켓 상세 이미지 태그 삽입 (A마켓 태그를 B마켓 태그로 교체)
        if detail_col and detail_col in df_b_market.columns:
            if cfg.img_tag_b:
                # A마켓 태그가 들어있으면 B마켓 태그로 교체, 없으면 B마켓 태그 삽입
                def _replace_detail_tag(html):
                    html = "" if (html is None or (isinstance(html, float))) else str(html)
                    if cfg.img_tag and cfg.img_tag in html:
                        return html.replace(cfg.img_tag, cfg.img_tag_b)
                    return core.insert_img_tag(html, cfg.img_tag_b)
                df_b_market[detail_col] = df_b_market[detail_col].apply(_replace_detail_tag)
            elif cfg.img_tag:
                # B마켓 태그 없으면 A마켓 태그 제거
                df_b_market[detail_col] = df_b_market[detail_col].apply(
                    lambda x: str(x or '').replace(cfg.img_tag, '') if x else x
                )
        # B마켓 옵션값 차별화: 쿠팡 동일상품 매칭 방지
        _opt_col = None
        for _c in df_b_market.columns:
            if str(_c).strip() in ("옵션입력",):
                _opt_col = _c
                break
        if _opt_col:
            _QTY_RE = re.compile(r'\d+\s*[Pp]$|\d+\s*개$|\d+\s*세트$|\d+\s*매$|\d+\s*장$|\d+\s*ea$', re.IGNORECASE)
            def _differentiate_option(val):
                val = str(val or "").strip()
                if not val.startswith("옵션{"):
                    return val
                inner = val[len("옵션{"):-1] if val.endswith("}") else val[len("옵션{"):]
                parts = inner.split("|")
                new_parts = []
                for p in parts:
                    p = p.strip()
                    if not p:
                        continue
                    # "A 6cm" → code="A", opt_val="6cm"
                    opt_val = p[2:] if len(p) > 2 and p[1] == " " else p
                    if _QTY_RE.search(opt_val):
                        new_parts.append(p)  # 이미 수량 있으면 그대로
                    else:
                        new_parts.append(f"{p} 1개")
                return "옵션{" + "|".join(new_parts) + "}"
            df_b_market[_opt_col] = df_b_market[_opt_col].apply(_differentiate_option)

        _status(status_cb, f"B마켓 시트 생성: {len(b_market_rows)}개 상품")

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:

        df.to_excel(writer, sheet_name="분리추출전", index=False)

        df_upload_export.to_excel(writer, sheet_name="분리추출후", index=False)

        if df_b_market is not None:
            df_b_market.to_excel(writer, sheet_name="B마켓", index=False)

        if debug_on and len(debug_rows) > 0:

            pd.DataFrame(debug_rows).to_excel(writer, sheet_name="디버그", index=False)

    try:

        with pd.ExcelWriter(upload_path, engine="openpyxl") as writer:

            df_upload_export.to_excel(writer, sheet_name="분리추출후", index=False)

            if df_b_market is not None:
                df_b_market.to_excel(writer, sheet_name="B마켓", index=False)

            # OCR 결과를 업로드용 엑셀에 함께 저장 (LLM 키워드 생성용)
            if ocr_results_list:
                pd.DataFrame(ocr_results_list).to_excel(writer, sheet_name="OCR결과", index=False)

    except PermissionError:

        # file might be opened by Excel; write a timestamped file instead

        upload_path = _safe_path(upload_path)

        with pd.ExcelWriter(upload_path, engine="openpyxl") as writer:

            df_upload_export.to_excel(writer, sheet_name="분리추출후", index=False)

            if df_b_market is not None:
                df_b_market.to_excel(writer, sheet_name="B마켓", index=False)

            if ocr_results_list:
                pd.DataFrame(ocr_results_list).to_excel(writer, sheet_name="OCR결과", index=False)



    # ── OCR 결과 별도 파일 저장 ──

    if ocr_results_list:

        # legacy-compatible fixed name (overwrite): OCR결과_YYYYMMDD_02.xlsx

        ocr_result_path = os.path.join(export_root, f"OCR결과_{date_tag}_02.xlsx")

        try:

            df_ocr_results = pd.DataFrame(ocr_results_list)

            with pd.ExcelWriter(ocr_result_path, engine="openpyxl") as writer:

                df_ocr_results.to_excel(writer, sheet_name="OCR결과", index=False)

            _status(status_cb, f"OCR 결과 {len(ocr_results_list)}개 → {os.path.basename(ocr_result_path)}")

        except PermissionError:

            # Excel opened/locked: fallback to sequential name

            ocr_result_path = _seq_path(export_root, f"OCR결과_{date_tag}", ".xlsx")

            df_ocr_results = pd.DataFrame(ocr_results_list)

            with pd.ExcelWriter(ocr_result_path, engine="openpyxl") as writer:

                df_ocr_results.to_excel(writer, sheet_name="OCR결과", index=False)

            _status(status_cb, f"OCR 결과(대체 저장) {len(ocr_results_list)}개 → {os.path.basename(ocr_result_path)}")

        except Exception as e:

            _status(status_cb, f"OCR 결과 저장 오류: {e}")



    # ── ocr_only 모드: keyword_skill.md 생성 ──
    if cfg.phase == "ocr_only":
        _generate_keyword_skill_md(export_root, upload_path, date_tag, chunk_size=cfg.chunk_size, status_cb=status_cb,
                                   a_name_min=cfg.a_name_min, a_name_max=cfg.a_name_max,
                                   b_name_min=cfg.b_name_min, b_name_max=cfg.b_name_max,
                                   a_tag_count=cfg.a_tag_count, b_tag_count=cfg.b_tag_count,
                                   keyword_version=keyword_version)

    # ── 상세 없는 상품 별도 파일 저장 (원본 형식 유지) ──

    if no_detail_indices:

        df_no_detail = df.loc[df.index.isin(no_detail_indices)].copy()

        if not df_no_detail.empty:

            no_detail_path = _seq_path(export_root, f"상세없음_{date_tag}", ".xlsx")

            try:

                with pd.ExcelWriter(no_detail_path, engine="openpyxl") as writer:

                    df_no_detail.to_excel(writer, sheet_name="상세없음", index=False)

                _status(status_cb, f"상세 없는 상품 {len(df_no_detail)}개 → {os.path.basename(no_detail_path)}")

            except Exception as e:

                _status(status_cb, f"상세없음 파일 저장 오류: {e}")



    # phase=analysis/ocr_only 이면 리스팅 이미지 처리 건너뜀 (Phase1에서 이미 완료)
    if cfg.phase in ("analysis", "ocr_only"):
        _status(status_cb, "리스팅 이미지 처리 건너뜀 (Phase1에서 완료)")

    elif cfg.make_listing and len(global_listing_sources) > 0:

        _status(status_cb, "처리중... (2/2) 대표이미지 생성")

        _progress(progress_cb, 70)

        listing_out_root = os.path.join(export_root, "listing_images", date_tag)

        os.makedirs(listing_out_root, exist_ok=True)

        total_imgs = len(set(global_listing_sources))

        processed = 0

        def _progress_cb():

            nonlocal processed

            processed += 1

            pct = 70 + int((processed / max(1, total_imgs)) * 25)

            _progress(progress_cb, min(95, pct))

        core.process_listing_images_global(

            src_paths=list(set(global_listing_sources)),

            base_out_root=listing_out_root,

            logo_rgba=logo_rgba,

            size=listing_size,

            pad=listing_pad,

            bg_color=(255, 255, 255),

            pos=logo_pos,

            opacity=logo_opacity,

            logo_ratio=logo_ratio,

            use_auto_contrast=use_auto_contrast,

            use_sharpen=use_sharpen,

            use_small_rotate=use_small_rotate,

            rotate_zoom=rotate_zoom,

            max_images_per_code=listing_max,

            ultra_angle_deg=ultra_angle_deg,

            ultra_translate_px=ultra_translate_px,

            ultra_scale_pct=ultra_scale_pct,

            trim_tol=trim_tol,

            jpeg_q_min=jpeg_q_min,

            jpeg_q_max=jpeg_q_max,

            do_flip_lr=do_flip_lr,

            progress_cb=_progress_cb,

        )

        # B마켓 대표이미지 생성 (별도 로고)
        if cfg.enable_b_market and cfg.logo_path_b and global_listing_sources:
            logo_rgba_b = core._load_logo(cfg.logo_path_b.strip())
            if logo_rgba_b is not None:
                listing_out_root_b = os.path.join(export_root, "listing_images_B", date_tag)
                os.makedirs(listing_out_root_b, exist_ok=True)
                _status(status_cb, "B마켓 대표이미지 생성 중...")
                core.process_listing_images_global(
                    src_paths=list(set(global_listing_sources)),
                    base_out_root=listing_out_root_b,
                    logo_rgba=logo_rgba_b,
                    size=listing_size,
                    pad=listing_pad,
                    bg_color=(255, 255, 255),
                    pos=logo_pos,
                    opacity=logo_opacity,
                    logo_ratio=logo_ratio,
                    use_auto_contrast=use_auto_contrast,
                    use_sharpen=use_sharpen,
                    use_small_rotate=use_small_rotate,
                    rotate_zoom=rotate_zoom,
                    max_images_per_code=listing_max,
                    ultra_angle_deg=ultra_angle_deg,
                    ultra_translate_px=ultra_translate_px,
                    ultra_scale_pct=ultra_scale_pct,
                    trim_tol=trim_tol,
                    jpeg_q_min=jpeg_q_min,
                    jpeg_q_max=jpeg_q_max,
                    do_flip_lr=do_flip_lr,
                )
                _status(status_cb, f"B마켓 대표이미지 완료 → {listing_out_root_b}")



    # phase=images 이면 이미지 처리만 하고 바로 반환 (OCR/키워드 건너뜀)
    if cfg.phase == "images":
        _status(status_cb, "이미지 처리 완료 (Phase 1)")
        _progress(progress_cb, 100)
        return export_root, ""

    _status(status_cb, "처리 완료")

    _progress(progress_cb, 100)

    return export_root, save_path





# ── 대표이미지만 생성하는 독립 함수 ──────────────────────────────────



@dataclass

class ListingOnlyConfig:

    """대표이미지만 생성할 때 필요한 설정."""

    local_img_dir: str = ""

    allow_folder_match: bool = True

    max_depth: int = -1



    listing_size: int = 1000

    listing_pad: int = 20

    listing_max: int = 20



    logo_path: str = ""

    logo_ratio: int = 14

    logo_opacity: int = 65

    logo_pos: str = "tr"



    use_auto_contrast: bool = True

    use_sharpen: bool = True

    use_small_rotate: bool = True

    rotate_zoom: float = 1.04



    ultra_angle_deg: float = 0.35

    ultra_translate_px: float = 0.6

    ultra_scale_pct: float = 0.25



    trim_tol: int = 8

    jpeg_q_min: int = 88

    jpeg_q_max: int = 92

    do_flip_lr: bool = True



    # 소스: CSV 또는 OCR Excel

    file_path: str = ""          # CSV/Excel (GS코드 목록)

    ocr_excel_path: str = ""     # OCR 결과 Excel (이미지 경로 포함)





def run_listing_only(cfg: ListingOnlyConfig, status_cb=None, progress_cb=None) -> str:

    """대표이미지만 생성 — GPT/키워드 없이 이미지 변환만 수행."""



    if not cfg.file_path:

        raise ValueError("CSV/Excel 파일을 선택해 주세요.")

    if not cfg.local_img_dir and not cfg.ocr_excel_path:

        raise ValueError("이미지 폴더 또는 OCR 결과 Excel이 필요합니다.")



    _status(status_cb, "대표이미지 생성 준비...")

    _progress(progress_cb, 5)



    # CSV에서 GS코드 목록 추출

    df = core.safe_read_csv(cfg.file_path)

    if df.empty:

        raise ValueError("CSV/Excel 내용이 비어 있습니다.")



    code_col = None

    for c in df.columns:

        if "코드" in str(c) or "code" in str(c).lower():

            code_col = c

            break



    name_col = None

    for c in df.columns:

        if "상품명" in str(c) or "name" in str(c).lower():

            name_col = c

            break



    # OCR Excel 로드 (이미지 경로)

    ocr_lookup: dict = {}

    if cfg.ocr_excel_path and os.path.isfile(cfg.ocr_excel_path):

        from app.services.ocr_excel import read_ocr_results

        ocr_lookup, _ = read_ocr_results(cfg.ocr_excel_path)

        _status(status_cb, f"OCR 결과 로드: {len(ocr_lookup)}개 ({os.path.basename(cfg.ocr_excel_path)})")



    local_root = cfg.local_img_dir

    allow_folder_match = cfg.allow_folder_match

    max_depth = cfg.max_depth



    # 이미지 소스 수집 — CSV의 모든 행에서 GS코드 추출 → 이미지 경로 수집

    _status(status_cb, f"이미지 소스 수집중... (CSV {len(df)}행, 코드컬럼={code_col}, 이미지폴더={local_root})")

    _progress(progress_cb, 10)

    global_listing_sources = []

    total_rows = len(df)

    found_codes = set()

    skipped_no_code = 0

    skipped_no_img = 0



    for row_i, idx in enumerate(df.index, start=1):

        # GS코드 추출 — 모든 컬럼에서 검색

        gs_code9 = None

        if code_col and code_col in df.columns:

            m = re.search(r"(GS\d{7})", str(df.at[idx, code_col]) or "")

            gs_code9 = m.group(1) if m else None

        if not gs_code9 and name_col and name_col in df.columns:

            m = re.search(r"(GS\d{7})", str(df.at[idx, name_col]) or "")

            gs_code9 = m.group(1) if m else None

        if not gs_code9:

            # 코드 컬럼/상품명 컬럼에 없으면 전체 행에서 검색

            for c in df.columns:

                m = re.search(r"(GS\d{7})", str(df.at[idx, c]) or "")

                if m:

                    gs_code9 = m.group(1)

                    break

        if not gs_code9:

            skipped_no_code += 1

            continue



        # 중복 코드 스킵 (B/C/D 옵션 등)

        if gs_code9 in found_codes:

            continue

        found_codes.add(gs_code9)



        # OCR Excel에서 이미지 경로 가져오기

        found = False

        if ocr_lookup:

            for _mk in [gs_code9, f"{gs_code9}A"]:

                if _mk in ocr_lookup:

                    _imgs = ocr_lookup[_mk].get("images", [])

                    _valid = [p for p in _imgs if p.strip() and p.strip().lower() != "nan" and os.path.isfile(p)]

                    if _valid:

                        global_listing_sources.extend(_valid)

                        found = True

                    break



        # 로컬 폴더에서 직접 검색 — 파일명에 GS코드가 포함된 것만 (대표이미지용)

        # 1.jpg, 2.jpg 등은 OCR 상세페이지용이므로 제외

        if not found and local_root:

            hits = core.find_local_images_for_code(local_root, gs_code9, allow_folder_match=allow_folder_match, max_depth=max_depth)

            gs_lower = gs_code9.lower()

            listing_hits = [p for p in hits if gs_lower in os.path.basename(p).lower()]

            if listing_hits:

                global_listing_sources.extend(listing_hits)

                found = True



        if not found:

            skipped_no_img += 1



        if row_i % 50 == 0:

            _status(status_cb, f"이미지 소스 수집중... {row_i}/{total_rows} (코드 {len(found_codes)}개, 이미지 {len(global_listing_sources)}개)")

            _progress(progress_cb, 10 + int(20 * row_i / max(1, total_rows)))



    _status(status_cb, f"이미지 소스 수집 완료: 코드 {len(found_codes)}개, 이미지 {len(set(global_listing_sources))}개 (코드없음={skipped_no_code}, 이미지없음={skipped_no_img})")

    _progress(progress_cb, 30)



    if not global_listing_sources:

        _status(status_cb, "이미지 소스가 없습니다. 이미지 폴더 경로를 확인해 주세요.")

        _progress(progress_cb, 100)

        return ""



    # 대표이미지 생성

    csv_base = os.path.splitext(os.path.basename(cfg.file_path))[0]

    date_tag = datetime.now().strftime("%Y%m%d")

    export_root = _make_export_root(date_tag, csv_base)

    listing_out_root = os.path.join(export_root, "listing_images", date_tag)

    os.makedirs(listing_out_root, exist_ok=True)



    logo_rgba = core._load_logo(cfg.logo_path.strip())

    listing_size = max(200, int(cfg.listing_size))

    listing_pad = max(0, int(cfg.listing_pad))

    listing_max = max(0, int(cfg.listing_max))

    logo_ratio = max(1, min(60, int(cfg.logo_ratio)))

    logo_opacity = max(0, min(100, int(cfg.logo_opacity)))

    logo_pos = cfg.logo_pos or "tr"

    jpeg_q_min = max(70, min(99, int(cfg.jpeg_q_min)))

    jpeg_q_max = max(jpeg_q_min, min(99, int(cfg.jpeg_q_max)))



    total_imgs = len(set(global_listing_sources))

    processed = 0



    def _pcb():

        nonlocal processed

        processed += 1

        pct = 30 + int((processed / max(1, total_imgs)) * 65)

        _progress(progress_cb, min(95, pct))



    _status(status_cb, f"대표이미지 생성중... ({total_imgs}개)")



    results = core.process_listing_images_global(

        src_paths=list(set(global_listing_sources)),

        base_out_root=listing_out_root,

        logo_rgba=logo_rgba,

        size=listing_size,

        pad=listing_pad,

        bg_color=(255, 255, 255),

        pos=logo_pos,

        opacity=logo_opacity,

        logo_ratio=logo_ratio,

        use_auto_contrast=bool(cfg.use_auto_contrast),

        use_sharpen=bool(cfg.use_sharpen),

        use_small_rotate=bool(cfg.use_small_rotate),

        rotate_zoom=float(cfg.rotate_zoom),

        max_images_per_code=listing_max,

        ultra_angle_deg=float(cfg.ultra_angle_deg),

        ultra_translate_px=float(cfg.ultra_translate_px),

        ultra_scale_pct=float(cfg.ultra_scale_pct),

        trim_tol=int(cfg.trim_tol),

        jpeg_q_min=jpeg_q_min,

        jpeg_q_max=jpeg_q_max,

        do_flip_lr=bool(cfg.do_flip_lr),

        progress_cb=_pcb,

    )



    _status(status_cb, f"대표이미지 생성 완료 - {len(results)}개 -> {listing_out_root}")

    _progress(progress_cb, 100)

    return listing_out_root






