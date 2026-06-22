"""
네이버 커머스 API 클라이언트 + 업로드 파이프라인
- 키: ~/Desktop/key/naver_client_key.txt 에서 로드
- OAuth2 + HMAC-SHA256 인증
- 상품 등록: POST /v2/products
"""
from __future__ import annotations

import base64
import bcrypt
import hashlib
import hmac
import json
import os
import re
import ssl
import time
import urllib.request
import urllib.parse
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Callable

# ─── 키/설정 로드 ───────────────────────────────
_KEY_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "key")

def _load_naver_keys() -> dict[str, str]:
    """naver_client_key.txt 에서 KEY=VALUE 형식으로 로드"""
    path = os.path.join(_KEY_DIR, "naver_client_key.txt")
    keys = {}
    if os.path.isfile(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                keys[k.strip()] = v.strip()
    return keys

_NAVER_KEYS = _load_naver_keys()
CLIENT_ID = _NAVER_KEYS.get("NAVER_COMMERCE_CLIENT_ID", "")
CLIENT_SECRET = _NAVER_KEYS.get("NAVER_COMMERCE_CLIENT_SECRET", "")
ACCOUNT_ID = _NAVER_KEYS.get("NAVER_COMMERCE_ACCOUNT_ID", "")

API_BASE = "https://api.commerce.naver.com/external"


# ─── 인증 (OAuth2 + HMAC-SHA256) ──────────────────

def _generate_sign(client_id: str, client_secret: str, timestamp: int) -> str:
    """bcrypt + base64로 client_secret_sign 생성"""
    password = f"{client_id}_{timestamp}"
    hashed = bcrypt.hashpw(password.encode("utf-8"), client_secret.encode("utf-8"))
    return base64.standard_b64encode(hashed).decode("utf-8")


_token_cache: dict[str, tuple[str, float]] = {}

def get_access_token() -> str:
    """OAuth2 토큰 발급 (캐시)"""
    now = time.time()
    cached = _token_cache.get("token")
    if cached and cached[1] > now:
        return cached[0]

    timestamp = int((now - 3) * 1000)
    sign = _generate_sign(CLIENT_ID, CLIENT_SECRET, timestamp)

    data = urllib.parse.urlencode({
        "client_id": CLIENT_ID,
        "timestamp": timestamp,
        "client_secret_sign": sign,
        "grant_type": "client_credentials",
        "type": "SELF",
    }).encode("utf-8")

    req = urllib.request.Request(
        f"{API_BASE}/v1/oauth2/token",
        data=data,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
    )

    ctx = ssl.create_default_context()
    resp = urllib.request.urlopen(req, context=ctx, timeout=15)
    result = json.loads(resp.read().decode("utf-8"))

    token = result["access_token"]
    expires_in = result.get("expires_in", 600)
    _token_cache["token"] = (token, now + expires_in - 60)
    return token


# ─── API 호출 ──────────────────────────────────

def api_call(method: str, path: str, body: dict | None = None, query: dict | None = None) -> dict:
    token = get_access_token()
    url = f"{API_BASE}{path}"
    if query:
        url += "?" + urllib.parse.urlencode(query)

    json_body = json.dumps(body).encode("utf-8") if body else None
    req = urllib.request.Request(url, data=json_body)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json;charset=UTF-8")
    req.get_method = lambda: method

    ctx = ssl.create_default_context()
    try:
        resp = urllib.request.urlopen(req, context=ctx, timeout=30)
        return json.loads(resp.read().decode("utf-8"))
    except urllib.request.HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        return {"_error": e.code, "_msg": raw[:800]}


# ─── 카테고리 검색 ─────────────────────────────

_all_categories: list[dict] = []

def _load_all_categories() -> list[dict]:
    """전체 카테고리 로드 (캐시)"""
    global _all_categories
    if _all_categories:
        return _all_categories
    result = api_call("GET", "/v1/categories")
    if isinstance(result, list):
        _all_categories = result
    return _all_categories


def predict_category(product_name: str) -> dict:
    """네이버 상품 모델 검색 API로 유사 상품의 카테고리를 가져옴"""
    import re

    # 상품명 정제 (GS코드, 숫자/단위 제거)
    clean = re.sub(r"[A-Z]{1,2}\d{5,}[A-Z]?", "", product_name).strip()
    clean = re.sub(r"\d+(\.\d+)?\s*(cm|mm|m|g|kg|ml|L|개|매|장|ea)", "", clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        clean = product_name

    result = api_call("GET", "/v1/product-models", query={"name": clean})
    if "_error" in result:
        return result

    contents = result.get("contents", [])
    if contents:
        top = contents[0]
        return {
            "id": str(top.get("categoryId", "")),
            "name": top.get("wholeCategoryName", "").split(">")[-1],
            "wholeCategoryName": top.get("wholeCategoryName", ""),
            "matchedProduct": top.get("name", "")[:50],
        }

    return {"_error": 404, "_msg": f"유사 상품 없음: {clean[:30]}"}


# ─── 이미지 업로드 ──────────────────────────────

_SUPPORTED_IMAGE_MIME_TYPES = {
    "image/jpeg": "jpg",
    "image/png": "png",
    "image/gif": "gif",
    "image/bmp": "bmp",
}


def _sniff_image_mime(img_data: bytes) -> str | None:
    if img_data.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if img_data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if img_data.startswith((b"GIF87a", b"GIF89a")):
        return "image/gif"
    if img_data.startswith(b"BM"):
        return "image/bmp"
    if len(img_data) >= 12 and img_data.startswith(b"RIFF") and img_data[8:12] == b"WEBP":
        return "image/webp"
    return None


def _normalize_image_bytes(img_data: bytes, image_url: str, content_type: str = "") -> tuple[bytes, str, str]:
    actual_mime = (content_type or "").split(";", 1)[0].strip().lower()
    sniffed_mime = _sniff_image_mime(img_data)
    if sniffed_mime:
        actual_mime = sniffed_mime

    if actual_mime not in _SUPPORTED_IMAGE_MIME_TYPES:
        try:
            from io import BytesIO
            from PIL import Image

            with Image.open(BytesIO(img_data)) as image:
                if image.mode not in ("RGB", "L"):
                    image = image.convert("RGB")
                output = BytesIO()
                image.save(output, format="JPEG", quality=95)
                img_data = output.getvalue()
            actual_mime = "image/jpeg"
        except Exception as ex:
            raise ValueError(f"이미지 포맷을 확인할 수 없습니다: {image_url[:120]}") from ex

    from pathlib import PurePosixPath
    if os.path.isfile(image_url):
        original_name = os.path.basename(image_url) or "image"
    else:
        original_name = PurePosixPath(urllib.parse.urlparse(image_url).path).name or "image"
    base_name = original_name.rsplit(".", 1)[0] if "." in original_name else original_name
    fname = f"{base_name}.{_SUPPORTED_IMAGE_MIME_TYPES[actual_mime]}"
    return img_data, actual_mime, fname


def upload_image_url(image_url: str) -> str:
    """외부 이미지 URL 또는 로컬 파일을 네이버 이미지 업로드로 전송, 네이버 URL 반환"""
    if not image_url:
        return ""

    # 1) 이미지 다운로드 또는 로컬 파일 로드
    try:
        if os.path.isfile(image_url):
            with open(image_url, "rb") as f:
                img_data = f.read()
            response_content_type = ""
        else:
            ctx = ssl.create_default_context()
            req = urllib.request.Request(
                image_url,
                headers={
                    "User-Agent": "Mozilla/5.0",
                    "Accept": "image/*,*/*;q=0.8",
                },
            )
            resp = urllib.request.urlopen(req, context=ctx, timeout=15)
            img_data = resp.read()
            response_content_type = resp.headers.get("Content-Type", "")
    except Exception as ex:
        raise ValueError(f"이미지 다운로드 실패: {image_url[:120]}") from ex

    img_data, content_type, fname = _normalize_image_bytes(img_data, image_url, response_content_type)

    # 2) multipart/form-data 빌드
    import uuid
    boundary = uuid.uuid4().hex
    body = b""
    body += f"--{boundary}\r\n".encode()
    body += f'Content-Disposition: form-data; name="imageFiles"; filename="{fname}"\r\n'.encode()
    body += f"Content-Type: {content_type}\r\n\r\n".encode()
    body += img_data
    body += f"\r\n--{boundary}--\r\n".encode()

    # 3) 네이버 이미지 업로드 API 호출
    token = get_access_token()
    url = f"{API_BASE}/v1/product-images/upload"
    req2 = urllib.request.Request(url, data=body)
    req2.add_header("Authorization", f"Bearer {token}")
    req2.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    req2.get_method = lambda: "POST"

    try:
        ctx2 = ssl.create_default_context()
        resp2 = urllib.request.urlopen(req2, context=ctx2, timeout=30)
        result = json.loads(resp2.read().decode("utf-8"))
    except urllib.request.HTTPError as ex:
        raw = ex.read().decode("utf-8", errors="replace")
        raise ValueError(f"네이버 이미지 업로드 실패: {raw[:200]}") from ex
    except Exception as ex:
        raise ValueError(f"네이버 이미지 업로드 실패: {image_url[:120]}") from ex

    images = result.get("images", [])
    if images:
        uploaded_url = images[0].get("url", "")
        if uploaded_url:
            return uploaded_url

    raise ValueError(f"네이버 이미지 업로드 응답 이상: {json.dumps(result, ensure_ascii=False)[:200]}")


_IMAGE_SELECTION_CACHE: dict[str, dict[str, dict[str, object]]] = {}


def _extract_gs9(code: str | None) -> str:
    match = re.search(r"GS\d{7,9}", str(code or ""), re.IGNORECASE)
    return match.group(0).upper() if match else ""


def _sanitize_seller_tag(tag: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z가-힣\s]", "", str(tag))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _resolve_export_root(source_file_path: str) -> str:
    from pathlib import Path

    path = Path(source_file_path).resolve()
    parent = path.parent
    parent_name = parent.name.lower()
    grand_name = parent.parent.name.lower() if parent.parent else ""

    if parent_name == "llm_result" and grand_name == "llm_chunks":
        return str(parent.parent.parent)
    if parent_name == "llm_result":
        return str(parent.parent)
    return str(parent)


def _load_image_selections(export_root: str) -> dict[str, dict[str, object]]:
    cached = _IMAGE_SELECTION_CACHE.get(export_root)
    if cached is not None:
        return cached

    path = os.path.join(export_root, "image_selections.json")
    if not os.path.isfile(path):
        _IMAGE_SELECTION_CACHE[export_root] = {}
        return _IMAGE_SELECTION_CACHE[export_root]

    try:
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)
        parsed: dict[str, dict[str, object]] = {}
        for key, value in raw.items():
            if not isinstance(value, dict):
                continue
            parsed[key.upper()] = {
                "main": value.get("main"),
                "mainB": value.get("mainB"),
                "additional": value.get("additional") if isinstance(value.get("additional"), list) else [],
            }
        _IMAGE_SELECTION_CACHE[export_root] = parsed
    except Exception:
        _IMAGE_SELECTION_CACHE[export_root] = {}

    return _IMAGE_SELECTION_CACHE[export_root]


def _find_listing_folder(export_root: str, gs9: str) -> str | None:
    from pathlib import Path

    for root_name in ("listing_images",):
        base = Path(export_root) / root_name
        if not base.is_dir():
            continue
        matches = sorted((p for p in base.rglob(gs9) if p.is_dir()), key=lambda p: str(p))
        if matches:
            return str(matches[0])
    return None


def _pick_local_listing_images(row: dict) -> list[str]:
    source_file_path = str(row.get("_source_file_path") or "")
    gs9 = _extract_gs9(row.get("자체 상품코드"))
    if not source_file_path or not gs9:
        return []

    export_root = str(row.get("_export_root") or _resolve_export_root(source_file_path))
    folder_path = _find_listing_folder(export_root, gs9)
    if not folder_path:
        return []

    files = sorted(
        [
            os.path.join(folder_path, name)
            for name in os.listdir(folder_path)
            if os.path.splitext(name)[1].lower() in (".jpg", ".jpeg", ".png", ".bmp", ".webp")
        ],
        key=lambda p: os.path.basename(p).lower(),
    )
    if not files:
        return []

    selection = _load_image_selections(export_root).get(gs9, {})
    main_index = selection.get("main")
    if not isinstance(main_index, int):
        main_index = 1 if len(files) > 1 else 0
    if main_index < 0 or main_index >= len(files):
        main_index = 1 if len(files) > 1 else 0

    ordered = [files[main_index]]
    seen = {main_index}

    additional = selection.get("additional") if isinstance(selection.get("additional"), list) else []
    for idx in additional:
        if isinstance(idx, int) and 0 <= idx < len(files) and idx not in seen:
            ordered.append(files[idx])
            seen.add(idx)

    for idx, file_path in enumerate(files):
        if idx in seen:
            continue
        ordered.append(file_path)

    return ordered[:9]


# ─── 엑셀 → 네이버 상품 JSON ────────────────────

def read_source_file(file_path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = {c.column: c.value for c in ws[1] if c.value}
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for col, name in headers.items():
            row[name] = ws.cell(row=r, column=col).value
        row["_row_num"] = r
        row["_source_file_path"] = file_path
        row["_export_root"] = _resolve_export_root(file_path)
        rows.append(row)
    return rows


def parse_options(option_str: str | None, extra_price_str: str | None = None) -> list[dict]:
    """옵션 파싱 (쿠팡과 동일 형식)"""
    if not option_str:
        return []
    m = re.findall(r"([A-Z])\s+([^,}]+)", option_str)
    prices = []
    if extra_price_str:
        for p in re.split(r"[,|]", str(extra_price_str)):
            p = p.strip()
            if p:
                try:
                    prices.append(int(float(p)))
                except ValueError:
                    prices.append(0)
    options = []
    for i, (_label, value) in enumerate(m):
        price = prices[i] if i < len(prices) else 0
        options.append({"name": value.strip(), "price": price})
    return options


def build_naver_product(row: dict, category_id: str | None = None) -> dict:
    """가공파일 행 → 네이버 커머스 API 상품 등록 JSON"""
    # 상품명: 키워드 가공된 이름 우선 사용
    product_name = (
        row.get("홈런_스마트스토어상품명")
        or row.get("홈런_네이버상품명")
        or row.get("네이버상품명")
        or row.get("상품명")
        or row.get("최종키워드2차")
        or row.get("1차키워드")
        or ""
    )
    product_name = str(product_name).strip()[:100]
    sale_price = max(int(row.get("판매가", 0) or 0), 100)

    # 상세설명
    detail_html = row.get("상품 상세설명") or row.get("상세설명") or ""

    # 이미지: 일반 가공 대표이미지 우선, 없으면 엑셀 URL과 상세 HTML 사용
    images = {}
    product_img_urls = _pick_local_listing_images(row)

    representative_candidates = []
    for value in (row.get("이미지등록(목록)"), row.get("이미지등록(상세)")):
        if not value:
            continue
        representative_candidates.extend(
            [u.strip() for u in re.split(r"[|\n]", str(value)) if u and str(u).strip()]
        )

    detail_image_candidates = []
    if detail_html:
        html_imgs = re.findall(r"<img[^>]+src=[\"']([^\"']+)", detail_html)
        gs_code = row.get("자체 상품코드") or ""
        for u in html_imgs:
            if gs_code and gs_code in u and re.search(r'/\d+\.(jpg|jpeg|png|bmp|webp)$', u, re.IGNORECASE):
                detail_image_candidates.append(u)

    seen_candidates: set[str] = {str(p).strip() for p in product_img_urls if str(p).strip()}
    for candidate in representative_candidates + detail_image_candidates:
        candidate = str(candidate).strip()
        if not candidate or candidate in seen_candidates:
            continue
        seen_candidates.add(candidate)
        product_img_urls.append(candidate)

    if product_img_urls:
        uploaded_urls: list[str] = []
        image_errors: list[str] = []
        seen_image_urls: set[str] = set()
        for candidate_url in product_img_urls[:9]:
            candidate_url = str(candidate_url).strip()
            if not candidate_url or candidate_url in seen_image_urls:
                continue
            seen_image_urls.add(candidate_url)
            try:
                uploaded_urls.append(upload_image_url(candidate_url))
            except Exception as ex:
                image_errors.append(str(ex))

        if uploaded_urls:
            images["representativeImage"] = {"url": uploaded_urls[0]}
            optional_images = [{"url": u} for u in uploaded_urls[1:9]]
            if optional_images:
                images["optionalImages"] = optional_images
        else:
            raise ValueError(image_errors[0] if image_errors else "대표 이미지 업로드 실패")

    # 검색 태그: 스마트스토어/네이버 전용 태그 우선
    search_tags = str(
        row.get("홈런_스마트스토어태그")
        or row.get("홈런_네이버태그")
        or row.get("네이버태그")
        or row.get("홈런_스마트스토어검색키워드")
        or row.get("검색키워드")
        or ""
    )
    if "|" in search_tags or "," in search_tags or "\n" in search_tags:
        raw_tags = re.split(r"[|,\n]+", search_tags)
    else:
        raw_tags = search_tags.split()

    tag_list: list[str] = []
    seen_tags: set[str] = set()
    for raw_tag in raw_tags:
        tag = _sanitize_seller_tag(raw_tag)
        if not tag or tag in seen_tags:
            continue
        seen_tags.add(tag)
        tag_list.append(tag)
        if len(tag_list) >= 10:
            break

    # 옵션
    options = parse_options(row.get("옵션입력"), row.get("옵션추가금"))
    seller_code = row.get("자체 상품코드") or ""

    # 배송비
    delivery_info = {
        "deliveryType": "DELIVERY",
        "deliveryAttributeType": "NORMAL",
        "deliveryCompany": "CJGLS",
        "deliveryFee": {
            "deliveryFeeType": "CONDITIONAL_FREE",
            "baseFee": 3000,
            "freeConditionalAmount": 50000,
            "deliveryFeePayType": "PREPAID",
        },
        "claimDeliveryInfo": {
            "returnDeliveryFee": 3000,
            "exchangeDeliveryFee": 6000,
        },
    }

    # 상품 정보 제공 고시 (기본 - 기타)
    product_info_provided_notice = {
        "productInfoProvidedNoticeType": "ETC",
        "etc": {
            "returnCostReason": "상세페이지 참조",
            "noRefundReason": "상세페이지 참조",
            "qualityAssuranceStandard": "상세페이지 참조",
            "compensationProcedure": "상세페이지 참조",
            "troubleShootingContents": "상세페이지 참조",
            "itemName": "상세페이지 참조",
            "modelName": "상세페이지 참조",
            "manufacturer": "상세페이지 참조",
            "afterServiceDirector": "상세페이지 참조",
        },
    }

    product = {
        "originProduct": {
            "statusType": "SALE",
            "saleType": "NEW",
            "leafCategoryId": category_id or "",
            "name": product_name,
            "detailContent": detail_html,
            "images": images,
            "salePrice": sale_price,
            "stockQuantity": 999,
            "deliveryInfo": delivery_info,
            "sellerCodeInfo": {
                "sellerManagementCode": seller_code,
            },
            "detailAttribute": {
                "naverShoppingSearchInfo": {
                    "manufacturerName": "상세페이지 참조",
                    "brandName": "",
                },
                "afterServiceInfo": {
                    "afterServiceTelephoneNumber": "010-2324-8352",
                    "afterServiceGuideContent": "전화 문의",
                },
                "originAreaInfo": {
                    "originAreaCode": "0200037",
                    "importer": "상세페이지 참조",
                    "content": "상세설명 참조",
                    "plural": False,
                },
                "productInfoProvidedNotice": product_info_provided_notice,
                "minorPurchasable": True,
                "seoInfo": {
                    "sellerTags": [{"text": t} for t in tag_list],
                },
            },
        },
        "smartstoreChannelProduct": {
            "channelProductDisplayStatusType": "ON",
            "storeKeepExclusiveProduct": False,
            "naverShoppingRegistration": True,
        },
    }

    # 옵션 설정
    if options:
        option_combinations = []
        for i, opt in enumerate(options):
            option_combinations.append({
                "optionName1": opt["name"],
                "stockQuantity": 999,
                "price": sale_price + opt["price"],
                "usable": True,
            })
        product["originProduct"]["optionInfo"] = {
            "optionCombinationSortType": "CREATE",
            "optionCombinationGroupNames": {"optionGroupName1": "옵션"},
            "optionCombinations": option_combinations,
        }

    return product


def register_product(product_json: dict) -> dict:
    """상품 등록 API 호출"""
    return api_call("POST", "/v2/products", body=product_json)


def _extract_restricted_seller_tags(error_text: str) -> list[str]:
    try:
        payload = json.loads(error_text)
    except Exception:
        payload = None

    restricted: list[str] = []
    inputs = payload.get("invalidInputs", []) if isinstance(payload, dict) else []
    for item in inputs:
        if not isinstance(item, dict):
            continue
        item_type = str(item.get("type", ""))
        message = str(item.get("message", ""))
        if item_type != "Restricted.sellerTags" and "등록불가인 단어" not in message:
            continue
        restricted.extend(re.findall(r"등록불가인 단어\(([^)]+)\)", message))

    return [tag.strip() for tag in restricted if tag and tag.strip()]


# ─── 업로드 파이프라인 ─────────────────────────

@dataclass
class NaverUploadConfig:
    file_path: str = ""
    row_start: int = 0
    row_end: int = 0
    dry_run: bool = True

@dataclass
class NaverUploadResult:
    row: int = 0
    name: str = ""
    status: str = ""
    product_id: str = ""
    error: str = ""


def run_naver_upload(
    config: NaverUploadConfig,
    status_cb: Callable[[str], None] | None = None,
    progress_cb: Callable[[int], None] | None = None,
) -> list[NaverUploadResult]:
    """네이버 스마트스토어 업로드 메인 함수"""

    def _status(msg: str) -> None:
        if status_cb:
            status_cb(msg)

    _status("가공파일 읽는 중...")
    rows = read_source_file(config.file_path)
    total = len(rows)
    _status(f"{total}개 상품 로드 완료")

    if config.row_start > 0:
        end = config.row_end if config.row_end > 0 else config.row_start
        target_rows = [r for r in rows if config.row_start <= (r["_row_num"] - 1) <= end]
    else:
        target_rows = rows

    target_count = len(target_rows)
    _status(f"처리 대상: {target_count}개")
    results: list[NaverUploadResult] = []

    # 카테고리 추천 + 등록
    for idx, row in enumerate(target_rows):
        row_num = row.get("_row_num", idx + 2)
        product_name = row.get("상품명", "")
        _status(f"[{idx+1}/{target_count}] {product_name[:30]}...")

        # 엑셀에 네이버카테고리코드가 있으면 그대로 사용
        preset_cat = row.get("네이버카테고리코드") or row.get("네이버카테고리")
        if preset_cat:
            category_id = str(int(float(str(preset_cat))))
            cat_name = f"엑셀지정({category_id})"
            _status(f"  → 엑셀 카테고리 사용: {category_id}")
        else:
            # API 기반 카테고리 추천
            try:
                cat_result = predict_category(product_name)
                if "_error" in cat_result:
                    results.append(NaverUploadResult(
                        row=row_num, name=product_name,
                        status="CATEGORY_FAIL",
                        error=cat_result.get("_msg", "카테고리 추천 실패")[:200],
                    ))
                    continue
                category_id = str(cat_result.get("id", ""))
                cat_name = cat_result.get("wholeCategoryName", "")
            except Exception as ex:
                results.append(NaverUploadResult(
                    row=row_num, name=product_name,
                    status="CATEGORY_FAIL", error=str(ex)[:200],
                ))
                continue

        if config.dry_run:
            _status(f"  → {cat_name}")
            results.append(NaverUploadResult(
                row=row_num, name=product_name,
                status="DRY_RUN_OK",
                product_id=f"{cat_name} ({category_id})",
            ))
            continue

        # 실제 등록
        try:
            product_json = build_naver_product(row, category_id)
            resp = register_product(product_json)

            restricted_tags = _extract_restricted_seller_tags(resp.get("_msg", "")) if "_error" in resp else []
            if restricted_tags:
                seller_tags = product_json.get("originProduct", {}).get("detailAttribute", {}).get("seoInfo", {}).get("sellerTags", [])
                filtered_tags = [tag for tag in seller_tags if str(tag.get("text", "")).strip() not in restricted_tags]
                product_json["originProduct"]["detailAttribute"]["seoInfo"]["sellerTags"] = filtered_tags
                resp = register_product(product_json)

            if "_error" in resp:
                results.append(NaverUploadResult(
                    row=row_num, name=product_name,
                    status="FAIL", error=resp.get("_msg", "등록 실패")[:200],
                ))
            else:
                pid = str(resp.get("smartstoreChannelProductNo",
                          resp.get("originProductNo", "")))
                results.append(NaverUploadResult(
                    row=row_num, name=product_name,
                    status="OK", product_id=pid,
                ))
        except Exception as ex:
            results.append(NaverUploadResult(
                row=row_num, name=product_name,
                status="FAIL", error=str(ex)[:200],
            ))

        # 속도 제한
        if (idx + 1) % 5 == 0:
            time.sleep(1.0)

        if progress_cb:
            progress_cb(int((idx + 1) / target_count * 100))

    return results

