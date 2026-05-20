from __future__ import annotations



import os

import re

from datetime import datetime

import sys

from PySide6 import QtCore, QtGui, QtWidgets



from app.gui.widgets import DropLineEdit

from app.gui.worker import PipelineWorker, ListingWorker, CoupangUploadWorker

from app.gui.image_selector_dialog import ImageSelectorDialog

from app.services.pipeline import PipelineConfig, ListingOnlyConfig
from app.services.coupang import CoupangUploadConfig

from app.services import legacy_core as core





class MainWindow(QtWidgets.QMainWindow):

    def __init__(self) -> None:

        super().__init__()

        self._default_keyword_model = self._normalize_runtime_model_name(

            os.getenv("KEYWORDOCR_DEFAULT_MODEL_KEYWORD", "gpt-4.1"), "gpt-4.1"

        )

        self._default_longtail_model = self._normalize_runtime_model_name(

            os.getenv("KEYWORDOCR_DEFAULT_MODEL_LONGTAIL", "gpt-4.1-mini"), "gpt-4.1-mini"

        )

        settings_org = os.getenv("KEYWORDOCR_SETTINGS_ORG", "KeywordOCR")

        settings_app = os.getenv("KEYWORDOCR_SETTINGS_APP", "PySide6App")

        self.setWindowTitle(os.getenv("KEYWORDOCR_WINDOW_TITLE", "키워드 OCR 도구 (PySide6)"))

        self.resize(1120, 980)

        self.setFixedSize(1120, 980)



        self._thread: QtCore.QThread | None = None

        self._worker: PipelineWorker | None = None

        self._log_file_handle = None

        self._last_output_file = ""

        self._settings = QtCore.QSettings(settings_org, settings_app)

        self._recent_files = []

        self._recent_dirs = []

        self._search_matches = []

        self._search_index = -1

        self._model_desc_map = {

            "gpt-4o": "빠르고 유연한 고성능 범용 모델.",

            "gpt-4o-mini": "가장 비용 효율적인 소형 모델.",

            "gpt-4.1": "코딩/지시이행/긴 컨텍스트 강화 모델.",

            "gpt-4.1-mini": "gpt-4.1 소형 고속 버전.",

            "gpt-4.1-nano": "gpt-4.1 초경량 모델.",

            "o3-mini": "저비용 소형 추론 모델.",

            "o4-mini": "최신 소형 o-series 추론 모델.",

            "claude-opus-4-6": "Claude Opus 4.6 - 최고 성능 추론 모델.",

            "claude-sonnet-4-6": "Claude Sonnet 4.6 - 균형형 고성능 모델.",

            "claude-haiku-4-5": "Claude Haiku 4.5 - 가장 빠른 Claude 모델.",

        }

        self._model_tip_map = {

            "gpt-4o": "고성능/균형형 모델",

            "gpt-4o-mini": "저비용/고속 모델",

            "gpt-4.1": "정확도/지시이행 강화",

            "gpt-4.1-mini": "균형형 소형 모델",

            "gpt-4.1-nano": "초경량/최저지연",

            "o3-mini": "추론 중심 소형 모델",

            "o4-mini": "추론 중심 최신 소형 모델",

            "claude-opus-4-6": "최고 성능 Claude 모델",

            "claude-sonnet-4-6": "균형형 Claude 모델",

            "claude-haiku-4-5": "고속 Claude 모델",

        }



        central = QtWidgets.QWidget()

        self.setCentralWidget(central)



        root_layout = QtWidgets.QHBoxLayout(central)

        root_layout.setContentsMargins(12, 12, 12, 12)

        root_layout.setSpacing(10)



        left = QtWidgets.QWidget()

        left_layout = QtWidgets.QVBoxLayout(left)

        left_layout.setContentsMargins(0, 0, 0, 0)

        left_layout.setSpacing(10)



        left_layout.addWidget(self._build_persistent_bar())

        left_layout.addWidget(self._build_left_tabs(), 0)



        nav_row = QtWidgets.QHBoxLayout()

        self.prev_btn = QtWidgets.QPushButton("이전")

        self.next_btn = QtWidgets.QPushButton("다음")

        self.prev_btn.clicked.connect(self._prev_step)

        self.next_btn.clicked.connect(self._next_step)

        nav_row.addWidget(self.prev_btn)

        nav_row.addWidget(self.next_btn)

        nav_row.addStretch(1)

        left_layout.addLayout(nav_row)



        right = QtWidgets.QWidget()

        right_layout = QtWidgets.QVBoxLayout(right)

        right_layout.setContentsMargins(0, 0, 0, 0)

        right_layout.setSpacing(8)



        self.status_label = QtWidgets.QLabel("대기 중")

        self.status_label.setStyleSheet("color:#1565c0; font-weight:600;")

        self.status_label.setWordWrap(False)

        self.status_label.setMinimumWidth(0)

        self.status_label.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Fixed)

        right_layout.addWidget(self.status_label)



        self.progress = QtWidgets.QProgressBar()

        self.progress.setRange(0, 100)

        self.progress.setValue(0)

        self.progress.setTextVisible(True)

        right_layout.addWidget(self.progress)



        self.stage_label = QtWidgets.QLabel("단계: 대기")

        self.stage_label.setStyleSheet("color:#455a64;")

        self.stage_label.setWordWrap(False)

        self.stage_label.setMinimumWidth(0)

        self.stage_label.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Fixed)

        right_layout.addWidget(self.stage_label)



        search_row = QtWidgets.QHBoxLayout()

        self.search_edit = QtWidgets.QLineEdit()

        self.search_edit.setPlaceholderText("로그 검색")

        self.search_btn = QtWidgets.QPushButton("찾기")

        self.search_btn.clicked.connect(self._find_in_logs)

        self.search_prev = QtWidgets.QPushButton("이전")

        self.search_next = QtWidgets.QPushButton("다음")

        self.search_prev.clicked.connect(self._prev_match)

        self.search_next.clicked.connect(self._next_match)

        self.search_count = QtWidgets.QLabel("0/0")

        self.search_count.setMinimumWidth(60)

        self.search_count.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

        search_row.addWidget(self.search_edit, 1)

        search_row.addWidget(self.search_btn)

        search_row.addWidget(self.search_prev)

        search_row.addWidget(self.search_next)

        search_row.addWidget(self.search_count)

        right_layout.addLayout(search_row)



        self.log_tabs = QtWidgets.QTabWidget()

        self.log_box = QtWidgets.QPlainTextEdit()

        self.log_box.setReadOnly(True)

        self.log_box.setMaximumBlockCount(2000)

        self.log_box.setPlaceholderText("로그가 여기에 표시됩니다.")



        self.error_box = QtWidgets.QPlainTextEdit()

        self.error_box.setReadOnly(True)

        self.error_box.setMaximumBlockCount(2000)

        self.error_box.setPlaceholderText("오류 로그가 여기에 표시됩니다.")



        self.debug_box = QtWidgets.QPlainTextEdit()

        self.debug_box.setReadOnly(True)

        self.debug_box.setMaximumBlockCount(2000)

        self.debug_box.setPlaceholderText("디버그 로그가 여기에 표시됩니다.")



        self.log_tabs.addTab(self.log_box, "요약")

        self.log_tabs.addTab(self.error_box, "오류")

        self.log_tabs.addTab(self.debug_box, "디버그")



        right_layout.addWidget(self.log_tabs, 1)



        splitter = QtWidgets.QSplitter()

        splitter.setOrientation(QtCore.Qt.Horizontal)

        splitter.addWidget(left)

        splitter.addWidget(right)

        splitter.setStretchFactor(0, 3)

        splitter.setStretchFactor(1, 2)

        splitter.setSizes([760, 420])

        root_layout.addWidget(splitter, 1)



        self._update_nav_buttons()

        self._load_settings()

        self._update_model_desc(self.model_keyword_combo.currentText())

        self._apply_model_tooltips()

        self._connect_setting_signals()  # 설정값 변경 시 자동 저장 연결



    def _build_persistent_bar(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("필수")

        grid = QtWidgets.QGridLayout(box)



        self.file_edit = DropLineEdit(mode="file")

        self.file_edit.setPlaceholderText("CSV/Excel 파일을 드래그앤드롭 또는 선택")

        self.file_btn = QtWidgets.QPushButton("파일 선택")

        self.file_btn.clicked.connect(self._browse_file)



        grid.addWidget(QtWidgets.QLabel("CSV/Excel"), 0, 0)

        grid.addWidget(self.file_edit, 0, 1)

        grid.addWidget(self.file_btn, 0, 2)



        self.run_btn = QtWidgets.QPushButton("전처리 + OCR + GPT + (대표이미지) 실행")

        self.run_btn.setStyleSheet("QPushButton { background:#2e7d32; color:#fff; padding:8px; font-weight:600; }")

        self.run_btn.clicked.connect(self._on_run)

        grid.addWidget(self.run_btn, 1, 0, 1, 3)



        self.run_excel_only_btn = QtWidgets.QPushButton("전처리 + OCR + GPT (엑셀만 생성)")

        self.run_excel_only_btn.setStyleSheet("QPushButton { background:#f57c00; color:#fff; padding:8px; font-weight:600; }")

        self.run_excel_only_btn.clicked.connect(self._on_run_excel_only)

        grid.addWidget(self.run_excel_only_btn, 2, 0, 1, 3)



        self.run_keyword_from_ocr_btn = QtWidgets.QPushButton("OCR결과 사용 키워드만 생성")

        self.run_keyword_from_ocr_btn.setStyleSheet("QPushButton { background:#1565c0; color:#fff; padding:8px; font-weight:600; }")

        self.run_keyword_from_ocr_btn.clicked.connect(self._on_run_keyword_from_ocr_only)

        self.run_keyword_from_ocr_btn.setToolTip("구글 OCR 재실행 없이 OCR 결과 Excel을 사용하여 키워드만 생성합니다.")

        grid.addWidget(self.run_keyword_from_ocr_btn, 3, 0, 1, 3)



        self.naver_enable = QtWidgets.QCheckBox("네이버 상위 키워드 사용")

        self.naver_enable.setChecked(False)

        self.naver_enable.setEnabled(False)

        self.naver_enable.setToolTip("현재 비활성화됨")

        self.naver_dry = QtWidgets.QCheckBox("DRY-RUN(모의응답)")

        self.naver_dry.setChecked(False)

        self.naver_dry.setEnabled(False)

        self.naver_dry.setToolTip("현재 비활성화됨")

        self.debug = QtWidgets.QCheckBox("디버그 시트 저장")

        self.debug.setChecked(True)



        grid.addWidget(self.naver_enable, 4, 0)

        grid.addWidget(self.naver_dry, 4, 1)

        grid.addWidget(self.debug, 4, 2)



        return box



    def _build_wizard(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("단계별 설정")

        v = QtWidgets.QVBoxLayout(box)



        header = QtWidgets.QHBoxLayout()

        self.step_label = QtWidgets.QLabel("Step 1/5: 입력")

        self.step_label.setStyleSheet("font-weight:600;")

        header.addWidget(self.step_label)

        header.addStretch(1)

        v.addLayout(header)



        self.steps = QtWidgets.QStackedWidget()

        self.steps.addWidget(self._step_input())

        self.steps.addWidget(self._step_keyword())

        self.steps.addWidget(self._step_ocr())

        self.steps.addWidget(self._step_naver())

        self.steps.addWidget(self._step_listing())

        v.addWidget(self.steps)



        return box



    def _build_left_tabs(self) -> QtWidgets.QTabWidget:

        tabs = QtWidgets.QTabWidget()

        tabs.addTab(self._build_integrated_tab(), "통합 실행")

        tabs.addTab(self._build_wizard(), "전처리")

        tabs.addTab(self._build_cafe24_tab(), "Cafe24 업로드")

        tabs.addTab(self._build_coupang_tab(), "쿠팡 업로드")

        tabs.addTab(self._build_price_review_tab(), "옵션 가격 확인")

        tabs.addTab(self._build_excluded_words_tab(), "제외단어")

        return tabs



    def _build_input_group(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("입력")

        grid = QtWidgets.QGridLayout(box)

        grid.setHorizontalSpacing(8)

        grid.setVerticalSpacing(8)



        self.local_edit = DropLineEdit(mode="dir")

        self.local_edit.setPlaceholderText("로컬 이미지 루트 폴더")

        self.local_btn = QtWidgets.QPushButton("폴더 선택")

        self.local_btn.clicked.connect(self._browse_dir)



        default_local = r"D:\pp"

        if os.path.isdir(default_local):

            self.local_edit.setText(default_local)



        grid.addWidget(QtWidgets.QLabel("로컬 이미지"), 1, 0)

        grid.addWidget(self.local_edit, 1, 1)

        grid.addWidget(self.local_btn, 1, 2)



        self.img_tag_edit = QtWidgets.QLineEdit()

        self.img_tag_edit.setPlaceholderText("상세 이미지 태그(없으면 비워도 됨)")

        self.img_tag_edit.setText("<img src='https://gi.esmplus.com/rkghrud/1.jpg' />")

        grid.addWidget(QtWidgets.QLabel("상세 이미지 태그"), 2, 0)

        grid.addWidget(self.img_tag_edit, 2, 1, 1, 2)



        self.tess_edit = QtWidgets.QLineEdit()

        self.tess_edit.setPlaceholderText("Tesseract 경로(옵션)")

        self.tess_btn = QtWidgets.QPushButton("자동탐색")

        self.tess_btn.clicked.connect(self._auto_tesseract)



        grid.addWidget(QtWidgets.QLabel("Tesseract"), 3, 0)

        grid.addWidget(self.tess_edit, 3, 1)

        grid.addWidget(self.tess_btn, 3, 2)



        return box



    def _build_keyword_group(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("키워드 설정")

        v = QtWidgets.QVBoxLayout(box)



        row = QtWidgets.QHBoxLayout()



        self.max_words = QtWidgets.QSpinBox()

        self.max_words.setRange(5, 60)

        self.max_words.setValue(24)



        self.max_len = QtWidgets.QSpinBox()

        self.max_len.setRange(30, 200)

        self.max_len.setValue(140)



        self.min_len = QtWidgets.QSpinBox()

        self.min_len.setRange(0, 200)

        self.min_len.setValue(90)



        _MODEL_LIST = ["gpt-4.1", "gpt-4.1-mini", "gpt-4.1-nano", "claude-haiku-4-5", "claude-sonnet-4-6", "claude-opus-4-6"]



        self.model_keyword_combo = QtWidgets.QComboBox()

        self.model_keyword_combo.addItems(_MODEL_LIST)

        self.model_keyword_combo.setCurrentText(self._default_keyword_model)

        self.model_keyword_combo.setToolTip("상품 키워드 생성에 사용할 모델")

        self.model_keyword_combo.currentTextChanged.connect(self._update_model_desc)



        self.model_longtail_combo = QtWidgets.QComboBox()

        self.model_longtail_combo.addItems(_MODEL_LIST)

        self.model_longtail_combo.setCurrentText(self._default_longtail_model)

        self.model_longtail_combo.setToolTip("롱테일/R열 키워드 생성에 사용할 모델")

        self.model_longtail_combo.currentTextChanged.connect(self._update_model_desc)



        self.keyword_version_combo = QtWidgets.QComboBox()

        self.keyword_version_combo.addItems(["3.0", "2.0", "1.0"])

        self.keyword_version_combo.setCurrentText("2.0")

        self.keyword_version_combo.setToolTip("3.0=타겟형(소스엄격/A·B 독립), 2.0=근거 우선 정제형, 1.0=확장형")



        self.model_desc = QtWidgets.QLabel("")

        self.model_desc.setStyleSheet("color:#546e7a;")

        self.model_desc.setWordWrap(True)



        row.addWidget(QtWidgets.QLabel("최대 단어수"))

        row.addWidget(self.max_words)

        row.addSpacing(12)

        row.addWidget(QtWidgets.QLabel("최대 글자수"))

        row.addWidget(self.max_len)

        row.addSpacing(12)

        row.addWidget(QtWidgets.QLabel("최소 글자수"))

        row.addWidget(self.min_len)

        row.addStretch(1)

        v.addLayout(row)



        row_model = QtWidgets.QHBoxLayout()

        row_model.addWidget(QtWidgets.QLabel("키워드 모델"))

        row_model.addWidget(self.model_keyword_combo)

        row_model.addSpacing(16)

        row_model.addWidget(QtWidgets.QLabel("롱테일 모델"))

        row_model.addWidget(self.model_longtail_combo)

        row_model.addSpacing(16)

        row_model.addWidget(QtWidgets.QLabel("키워드 버전"))

        row_model.addWidget(self.keyword_version_combo)

        row_model.addSpacing(12)

        row_model.addWidget(self.model_desc, 1)

        row_model.addStretch(1)

        v.addLayout(row_model)



        row2 = QtWidgets.QHBoxLayout()

        self.naver_autocomplete = QtWidgets.QCheckBox("네이버 쇼핑 자동완성 키워드 사용")

        self.naver_autocomplete.setChecked(False)

        self.naver_autocomplete.setEnabled(False)

        self.naver_autocomplete.setToolTip("현재 비활성화됨")

        row2.addWidget(self.naver_autocomplete)



        self.google_autocomplete = QtWidgets.QCheckBox("구글 자동완성 키워드 사용")

        self.google_autocomplete.setChecked(True)

        self.google_autocomplete.setToolTip("상품명 핵심어로 구글 자동완성을 조회하여 네이버에 없는 키워드를 추가합니다.")

        row2.addWidget(self.google_autocomplete)



        row2.addStretch(1)

        v.addLayout(row2)



        return box



    def _build_ocr_group(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("OCR 결과 불러오기")

        grid = QtWidgets.QGridLayout(box)



        grid.addWidget(QtWidgets.QLabel("OCR 결과 Excel"), 0, 0)

        self.ocr_excel_edit = DropLineEdit(mode="file")

        self.ocr_excel_edit.setPlaceholderText("OCR 결과 .xlsx 파일 (자동 탐색 또는 직접 선택)")

        grid.addWidget(self.ocr_excel_edit, 0, 1)



        ocr_btn_lay = QtWidgets.QHBoxLayout()

        ocr_browse_btn = QtWidgets.QPushButton("파일 선택")

        ocr_browse_btn.clicked.connect(self._browse_ocr_excel)

        ocr_btn_lay.addWidget(ocr_browse_btn)

        ocr_auto_btn = QtWidgets.QPushButton("자동 탐색")

        ocr_auto_btn.clicked.connect(self._auto_find_ocr_excel)

        ocr_btn_lay.addWidget(ocr_auto_btn)

        grid.addLayout(ocr_btn_lay, 0, 2)



        self.ocr_status_label = QtWidgets.QLabel("OCR 파일 미선택")

        self.ocr_status_label.setStyleSheet("color:#888;")

        self.ocr_status_label.setWordWrap(False)

        self.ocr_status_label.setMinimumWidth(0)

        self.ocr_status_label.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Fixed)

        grid.addWidget(self.ocr_status_label, 1, 0, 1, 3)



        self.merge_ocr = QtWidgets.QCheckBox("OCR 텍스트 + 상품명 결합하여 GPT 키워드 생성")

        self.merge_ocr.setChecked(True)

        grid.addWidget(self.merge_ocr, 2, 0, 1, 3)



        self.drop_digits = QtWidgets.QCheckBox("키워드 생성 시 숫자 토큰 제외")

        self.drop_digits.setChecked(True)

        grid.addWidget(self.drop_digits, 3, 0, 1, 3)



        note = QtWidgets.QLabel("※ OCR 전용 프로그램(main_ocr.py)에서 먼저 OCR을 실행한 후 결과 파일을 불러오세요.")

        note.setStyleSheet("color:#999; font-size:11px;")

        note.setWordWrap(True)

        grid.addWidget(note, 4, 0, 1, 3)



        return box



    def _browse_ocr_excel(self) -> None:

        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "OCR 결과 Excel 선택", "", "Excel (*.xlsx)")

        if path:

            self.ocr_excel_edit.setText(path)

            self.ocr_status_label.setText(f"선택됨: {os.path.basename(path)}")

            self.ocr_status_label.setStyleSheet("color:#1565c0; font-weight:bold;")



    def _auto_find_ocr_excel(self) -> None:

        csv_path = self.file_edit.text().strip()

        if not csv_path:

            QtWidgets.QMessageBox.information(self, "안내", "먼저 CSV 파일을 선택해 주세요.")

            return

        from app.services.ocr_excel import find_matching_ocr_file

        found = find_matching_ocr_file(csv_path)

        if found:

            self.ocr_excel_edit.setText(found)

            self.ocr_status_label.setText(f"자동 매칭: {os.path.basename(found)}")

            self.ocr_status_label.setStyleSheet("color:#2e7d32; font-weight:bold;")

        else:

            self.ocr_status_label.setText("매칭되는 OCR 결과 파일을 찾지 못했습니다.")

            self.ocr_status_label.setStyleSheet("color:#c62828;")



    def _build_naver_group(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("네이버 검색광고 API")

        v = QtWidgets.QVBoxLayout(box)

        row = QtWidgets.QHBoxLayout()



        row.addWidget(self.naver_enable)

        row.addWidget(self.naver_dry)

        self.naver_retry = QtWidgets.QCheckBox("네이버 재시도")

        self.naver_retry.setChecked(False)

        self.naver_retry.setEnabled(False)

        self.naver_retry_count = QtWidgets.QSpinBox()

        self.naver_retry_count.setRange(0, 5)

        self.naver_retry_count.setValue(2)

        self.naver_retry_count.setEnabled(False)

        self.naver_retry_delay = QtWidgets.QDoubleSpinBox()

        self.naver_retry_delay.setRange(0.2, 5.0)

        self.naver_retry_delay.setSingleStep(0.2)

        self.naver_retry_delay.setValue(0.8)

        self.naver_retry_delay.setEnabled(False)

        row.addWidget(self.naver_retry)

        row.addWidget(QtWidgets.QLabel("횟수"))

        row.addWidget(self.naver_retry_count)

        row.addWidget(QtWidgets.QLabel("지연(s)"))

        row.addWidget(self.naver_retry_delay)

        row.addStretch(1)



        v.addLayout(row)



        return box



    def _build_listing_group(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("대표이미지 생성")

        grid = QtWidgets.QGridLayout(box)



        self.make_listing = QtWidgets.QCheckBox("대표이미지 생성 실행")

        self.make_listing.setChecked(True)

        grid.addWidget(self.make_listing, 0, 0, 1, 2)



        self.listing_size = QtWidgets.QSpinBox()

        self.listing_size.setRange(200, 4000)

        self.listing_size.setValue(1000)



        self.listing_pad = QtWidgets.QSpinBox()

        self.listing_pad.setRange(0, 200)

        self.listing_pad.setValue(20)



        self.listing_max = QtWidgets.QSpinBox()

        self.listing_max.setRange(0, 100)

        self.listing_max.setValue(20)



        grid.addWidget(QtWidgets.QLabel("사이즈(px)"), 1, 0)

        grid.addWidget(self.listing_size, 1, 1)

        grid.addWidget(QtWidgets.QLabel("여백(px)"), 1, 2)

        grid.addWidget(self.listing_pad, 1, 3)

        grid.addWidget(QtWidgets.QLabel("최대 개수/코드"), 1, 4)

        grid.addWidget(self.listing_max, 1, 5)



        self.logo_edit = QtWidgets.QLineEdit()

        self.logo_btn = QtWidgets.QPushButton("찾기")

        self.logo_btn.clicked.connect(self._browse_logo)



        grid.addWidget(QtWidgets.QLabel("로고 이미지"), 2, 0)

        grid.addWidget(self.logo_edit, 2, 1, 1, 3)

        grid.addWidget(self.logo_btn, 2, 4)



        self.logo_ratio = QtWidgets.QSpinBox()

        self.logo_ratio.setRange(1, 60)

        self.logo_ratio.setValue(14)



        self.logo_opacity = QtWidgets.QSpinBox()

        self.logo_opacity.setRange(0, 100)

        self.logo_opacity.setValue(65)



        self.logo_pos = QtWidgets.QComboBox()

        self.logo_pos.addItems(["br", "tr", "bl", "tl", "c"])

        self.logo_pos.setCurrentText("tr")



        grid.addWidget(QtWidgets.QLabel("로고 비율(%)"), 3, 0)

        grid.addWidget(self.logo_ratio, 3, 1)

        grid.addWidget(QtWidgets.QLabel("투명도(%)"), 3, 2)

        grid.addWidget(self.logo_opacity, 3, 3)

        grid.addWidget(QtWidgets.QLabel("위치"), 3, 4)

        grid.addWidget(self.logo_pos, 3, 5)



        self.auto_contrast = QtWidgets.QCheckBox("오토컨트라스트")

        self.auto_contrast.setChecked(True)

        self.sharpen = QtWidgets.QCheckBox("샤픈")

        self.sharpen.setChecked(True)

        self.rotate = QtWidgets.QCheckBox("초미세 회전")

        self.rotate.setChecked(True)



        grid.addWidget(self.auto_contrast, 4, 0)

        grid.addWidget(self.sharpen, 4, 1)

        grid.addWidget(self.rotate, 4, 2)



        self.ultra_angle = QtWidgets.QDoubleSpinBox()

        self.ultra_angle.setRange(0.0, 5.0)

        self.ultra_angle.setSingleStep(0.05)

        self.ultra_angle.setValue(0.35)



        self.ultra_translate = QtWidgets.QDoubleSpinBox()

        self.ultra_translate.setRange(0.0, 10.0)

        self.ultra_translate.setSingleStep(0.05)

        self.ultra_translate.setValue(0.6)



        self.ultra_scale = QtWidgets.QDoubleSpinBox()

        self.ultra_scale.setRange(0.0, 5.0)

        self.ultra_scale.setSingleStep(0.05)

        self.ultra_scale.setValue(0.25)



        self.rotate_zoom = QtWidgets.QDoubleSpinBox()

        self.rotate_zoom.setRange(1.0, 2.0)

        self.rotate_zoom.setSingleStep(0.01)

        self.rotate_zoom.setValue(1.04)



        grid.addWidget(QtWidgets.QLabel("초미세 회전(±도)"), 5, 0)

        grid.addWidget(self.ultra_angle, 5, 1)

        grid.addWidget(QtWidgets.QLabel("초미세 이동(±px)"), 5, 2)

        grid.addWidget(self.ultra_translate, 5, 3)

        grid.addWidget(QtWidgets.QLabel("초미세 스케일(±%)"), 5, 4)

        grid.addWidget(self.ultra_scale, 5, 5)



        self.trim_tol = QtWidgets.QSpinBox()

        self.trim_tol.setRange(0, 50)

        self.trim_tol.setValue(8)



        self.jpeg_q_min = QtWidgets.QSpinBox()

        self.jpeg_q_min.setRange(70, 99)

        self.jpeg_q_min.setValue(88)



        self.jpeg_q_max = QtWidgets.QSpinBox()

        self.jpeg_q_max.setRange(70, 99)

        self.jpeg_q_max.setValue(92)



        self.flip_lr = QtWidgets.QCheckBox("좌우반전")

        self.flip_lr.setChecked(True)



        grid.addWidget(QtWidgets.QLabel("회전-확대 배율"), 6, 0)

        grid.addWidget(self.rotate_zoom, 6, 1)

        grid.addWidget(QtWidgets.QLabel("트림 허용오차"), 6, 2)

        grid.addWidget(self.trim_tol, 6, 3)

        grid.addWidget(QtWidgets.QLabel("JPEG 품질(최소/최대)"), 6, 4)

        hq = QtWidgets.QHBoxLayout()

        hq.addWidget(self.jpeg_q_min)

        hq.addWidget(QtWidgets.QLabel("~"))

        hq.addWidget(self.jpeg_q_max)

        grid.addLayout(hq, 6, 5)



        grid.addWidget(self.flip_lr, 7, 0)



        self.listing_only_btn = QtWidgets.QPushButton("대표이미지만 생성")

        self.listing_only_btn.setStyleSheet(

            "QPushButton { background:#1565c0; color:#fff; padding:8px; font-weight:600; }"

        )

        self.listing_only_btn.clicked.connect(self._on_run_listing_only)

        grid.addWidget(self.listing_only_btn, 8, 0, 1, 6)



        return box



    def _build_integrated_tab(self) -> QtWidgets.QWidget:

        """통합 실행 탭 - 모듈화된 4개 버튼"""

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)

        v.setContentsMargins(12, 12, 12, 12)

        v.setSpacing(16)



        # 안내 문구

        info_label = QtWidgets.QLabel(

            "✨ CSV 파일만 넣으면 끝! ESM 폴더에서 이미지 자동 다운로드\n"

            "📋 O열(상품 상세설명) → OCR용 이미지\n"

            "🖼️ AU열(이미지등록 상세) → 대표이미지용"

        )

        info_label.setStyleSheet("color:#1565c0; font-size:13px; padding:10px; background:#e3f2fd; border-radius:4px;")

        v.addWidget(info_label)



        # CSV 파일 입력

        file_group = QtWidgets.QGroupBox("CSV 파일")

        file_layout = QtWidgets.QHBoxLayout(file_group)

        self.integrated_file_edit = DropLineEdit(mode="file")

        self.integrated_file_edit.setPlaceholderText("CSV/Excel 파일을 선택하거나 드래그하세요")

        integrated_file_btn = QtWidgets.QPushButton("파일 선택")

        integrated_file_btn.clicked.connect(self._browse_integrated_file)

        file_layout.addWidget(self.integrated_file_edit, 1)

        file_layout.addWidget(integrated_file_btn)

        v.addWidget(file_group)



        # 로컬 이미지 폴더 (옵션)

        local_group = QtWidgets.QGroupBox("로컬 이미지 폴더 (선택사항)")

        local_layout = QtWidgets.QHBoxLayout(local_group)

        self.integrated_local_edit = DropLineEdit(mode="dir")

        self.integrated_local_edit.setPlaceholderText("D:\\Pp (비어있어도 됨, URL에서 자동 다운로드)")

        integrated_local_btn = QtWidgets.QPushButton("폴더 선택")

        integrated_local_btn.clicked.connect(self._browse_integrated_dir)



        default_local = r"D:\pp"

        if os.path.isdir(default_local):

            self.integrated_local_edit.setText(default_local)



        local_layout.addWidget(self.integrated_local_edit, 1)

        local_layout.addWidget(integrated_local_btn)

        v.addWidget(local_group)



        # 4개 모듈 버튼

        buttons_group = QtWidgets.QGroupBox("실행 모드 선택")

        buttons_layout = QtWidgets.QVBoxLayout(buttons_group)

        buttons_layout.setSpacing(10)



        # ① OCR만 실행

        self.integrated_ocr_btn = QtWidgets.QPushButton("① OCR만 실행")

        self.integrated_ocr_btn.setStyleSheet(

            "QPushButton { background:#f57c00; color:#fff; padding:12px; font-size:14px; font-weight:600; }"

        )

        self.integrated_ocr_btn.setToolTip("O열(상품 상세설명)에서 이미지 다운로드 → Google Vision OCR → OCR 결과 Excel 저장")

        self.integrated_ocr_btn.clicked.connect(self._on_integrated_ocr_only)

        buttons_layout.addWidget(self.integrated_ocr_btn)



        # ② 키워드만 생성

        self.integrated_keyword_btn = QtWidgets.QPushButton("② 키워드만 생성")

        self.integrated_keyword_btn.setStyleSheet(

            "QPushButton { background:#1976d2; color:#fff; padding:12px; font-size:14px; font-weight:600; }"

        )

        self.integrated_keyword_btn.setToolTip("OCR 결과 Excel 읽기 → GPT로 키워드 생성 → 업로드용 Excel 저장")

        self.integrated_keyword_btn.clicked.connect(self._on_integrated_keyword_only)

        buttons_layout.addWidget(self.integrated_keyword_btn)



        # ③ 대표이미지만 생성

        self.integrated_listing_btn = QtWidgets.QPushButton("③ 대표이미지만 생성")

        self.integrated_listing_btn.setStyleSheet(

            "QPushButton { background:#7b1fa2; color:#fff; padding:12px; font-size:14px; font-weight:600; }"

        )

        self.integrated_listing_btn.setToolTip("AU열(이미지등록 상세)에서 순차 다운로드 → 로고 삽입 + 미세 변형 → listing_images 폴더 저장")

        self.integrated_listing_btn.clicked.connect(self._on_integrated_listing_only)

        buttons_layout.addWidget(self.integrated_listing_btn)



        # ④ 통합 실행 (OCR → 키워드 → 대표이미지)

        separator = QtWidgets.QFrame()

        separator.setFrameShape(QtWidgets.QFrame.HLine)

        separator.setFrameShadow(QtWidgets.QFrame.Sunken)

        buttons_layout.addWidget(separator)



        self.integrated_full_btn = QtWidgets.QPushButton("④ 통합 실행 (OCR → 키워드 → 대표이미지)")

        self.integrated_full_btn.setStyleSheet(

            "QPushButton { background:#2e7d32; color:#fff; padding:16px; font-size:15px; font-weight:600; }"

        )

        self.integrated_full_btn.setToolTip("CSV만 넣으면 모든 과정 자동 실행: OCR → 키워드 → 대표이미지 완성!")

        self.integrated_full_btn.clicked.connect(self._on_integrated_full)

        buttons_layout.addWidget(self.integrated_full_btn)



        v.addWidget(buttons_group)

        v.addStretch(1)



        return w



    def _build_cafe24_tab(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)



        form = QtWidgets.QFormLayout()



        self.cafe24_token_path = QtWidgets.QLineEdit(self._desktop_key_path("cafe24_token.txt"))

        token_btn = QtWidgets.QPushButton("찾기")

        token_btn.clicked.connect(self._browse_token)

        token_row = QtWidgets.QHBoxLayout()

        token_row.addWidget(self.cafe24_token_path, 1)

        token_row.addWidget(token_btn)

        form.addRow("토큰 파일", token_row)



        self.cafe24_date_tag = QtWidgets.QLineEdit("")

        self.cafe24_date_tag.setPlaceholderText("YYYYMMDD (비우면 오늘 날짜)")

        form.addRow("날짜 폴더", self.cafe24_date_tag)



        self.cafe24_main_idx = QtWidgets.QSpinBox()

        self.cafe24_main_idx.setRange(1, 20)

        self.cafe24_main_idx.setValue(2)

        form.addRow("대표이미지 인덱스", self.cafe24_main_idx)



        self.cafe24_add_start = QtWidgets.QSpinBox()

        self.cafe24_add_start.setRange(1, 20)

        self.cafe24_add_start.setValue(3)

        form.addRow("추가이미지 시작", self.cafe24_add_start)



        self.cafe24_add_max = QtWidgets.QSpinBox()

        self.cafe24_add_max.setRange(1, 20)

        self.cafe24_add_max.setValue(10)

        form.addRow("추가이미지 최대", self.cafe24_add_max)



        self.cafe24_export_dir = QtWidgets.QLineEdit("")

        self.cafe24_export_dir.setPlaceholderText("비우면 C:\\code\\exports")

        export_btn = QtWidgets.QPushButton("찾기")

        export_btn.clicked.connect(self._browse_export_dir)

        export_row = QtWidgets.QHBoxLayout()

        export_row.addWidget(self.cafe24_export_dir, 1)

        export_row.addWidget(export_btn)

        form.addRow("exports 경로", export_row)



        self.cafe24_image_root = QtWidgets.QLineEdit("")

        self.cafe24_image_root.setPlaceholderText("비우면 자동 탐색")

        img_btn = QtWidgets.QPushButton("찾기")

        img_btn.clicked.connect(self._browse_image_root)

        img_row = QtWidgets.QHBoxLayout()

        img_row.addWidget(self.cafe24_image_root, 1)

        img_row.addWidget(img_btn)

        form.addRow("이미지 루트", img_row)



        self.cafe24_retry = QtWidgets.QSpinBox()

        self.cafe24_retry.setRange(0, 5)

        self.cafe24_retry.setValue(1)

        self.cafe24_retry_delay = QtWidgets.QDoubleSpinBox()

        self.cafe24_retry_delay.setRange(0.2, 5.0)

        self.cafe24_retry_delay.setSingleStep(0.2)

        self.cafe24_retry_delay.setValue(1.0)

        retry_row = QtWidgets.QHBoxLayout()

        retry_row.addWidget(QtWidgets.QLabel("재시도 횟수"))

        retry_row.addWidget(self.cafe24_retry)

        retry_row.addWidget(QtWidgets.QLabel("지연(s)"))

        retry_row.addWidget(self.cafe24_retry_delay)

        form.addRow("재시도", retry_row)



        self.cafe24_match_mode = QtWidgets.QComboBox()

        self.cafe24_match_mode.addItems(["PREFIX", "CONTAINS", "EXACT"])

        self.cafe24_match_mode.setCurrentText("PREFIX")

        self.cafe24_match_prefix = QtWidgets.QSpinBox()

        self.cafe24_match_prefix.setRange(5, 50)

        self.cafe24_match_prefix.setValue(20)

        match_row = QtWidgets.QHBoxLayout()

        match_row.addWidget(QtWidgets.QLabel("매칭 방식"))

        match_row.addWidget(self.cafe24_match_mode)

        match_row.addWidget(QtWidgets.QLabel("프리픽스 길이"))

        match_row.addWidget(self.cafe24_match_prefix)

        form.addRow("상품 매칭", match_row)



        v.addLayout(form)



        self.cafe24_upload_btn = QtWidgets.QPushButton("대표/추가 이미지 업로드 실행")

        self.cafe24_upload_btn.clicked.connect(self._run_cafe24_upload)

        v.addWidget(self.cafe24_upload_btn)



        self.cafe24_status = QtWidgets.QLabel("대기 중")

        v.addWidget(self.cafe24_status)



        self.cafe24_progress = QtWidgets.QProgressBar()

        self.cafe24_progress.setRange(0, 100)

        self.cafe24_progress.setValue(0)

        v.addWidget(self.cafe24_progress)



        self.cafe24_log_path = QtWidgets.QLineEdit("")

        self.cafe24_log_path.setPlaceholderText("업로드 로그 경로(비우면 자동)")

        log_btn = QtWidgets.QPushButton("로그 저장 위치")

        log_btn.clicked.connect(self._browse_log_output)

        log_row = QtWidgets.QHBoxLayout()

        log_row.addWidget(self.cafe24_log_path, 1)

        log_row.addWidget(log_btn)

        v.addLayout(log_row)



        self.cafe24_log_open = QtWidgets.QPushButton("로그 파일 열기")

        self.cafe24_log_open.clicked.connect(self._open_upload_log)

        v.addWidget(self.cafe24_log_open)



        self.cafe24_table = QtWidgets.QTableWidget(0, 5)

        self.cafe24_table.setHorizontalHeaderLabels(["GS", "PRODUCT_NO", "STATUS", "MAIN", "ERROR"])

        self.cafe24_table.horizontalHeader().setStretchLastSection(True)

        self.cafe24_table.itemDoubleClicked.connect(self._open_gs_folder)

        v.addWidget(self.cafe24_table)



        table_btn_row = QtWidgets.QHBoxLayout()

        self.cafe24_filter_miss = QtWidgets.QPushButton("미스만 보기")

        self.cafe24_filter_miss.clicked.connect(self._filter_misses)

        self.cafe24_retry_miss = QtWidgets.QPushButton("미스 재시도")

        self.cafe24_retry_miss.clicked.connect(self._retry_misses)

        table_btn_row.addWidget(self.cafe24_filter_miss)

        table_btn_row.addWidget(self.cafe24_retry_miss)

        table_btn_row.addStretch(1)

        v.addLayout(table_btn_row)



        self.cafe24_summary = QtWidgets.QLabel("미스 사유 요약: -")

        v.addWidget(self.cafe24_summary)



        v.addStretch(1)

        return w



    # ═══════════════════════════════════════════════════════════════

    # ── 제외 단어 관리 탭 ─────────────────────────────────────────

    # ═══════════════════════════════════════════════════════════════



    def _build_excluded_words_tab(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        main_layout = QtWidgets.QVBoxLayout(w)



        # ── 상단: 파일 불러오기 ──

        load_row = QtWidgets.QHBoxLayout()

        self.excl_file_path = QtWidgets.QLineEdit("")

        self.excl_file_path.setPlaceholderText("결과 Excel 경로 (비우면 최근 결과 자동)")

        excl_browse = QtWidgets.QPushButton("찾기")

        excl_browse.clicked.connect(self._browse_excl_file)

        excl_load = QtWidgets.QPushButton("불러오기")

        excl_load.setStyleSheet("QPushButton{background:#1565c0;color:#fff;padding:5px 14px;font-weight:600;}")

        excl_load.clicked.connect(self._load_keywords_for_exclusion)

        load_row.addWidget(self.excl_file_path, 1)

        load_row.addWidget(excl_browse)

        load_row.addWidget(excl_load)

        main_layout.addLayout(load_row)



        self.excl_status = QtWidgets.QLabel("결과 파일을 불러와 주세요.")

        self.excl_status.setStyleSheet("color:#1565c0; padding:2px;")

        main_layout.addWidget(self.excl_status)



        # ── 중앙: 좌우 패널 ──

        panels = QtWidgets.QHBoxLayout()



        # 왼쪽: 키워드 목록

        left_panel = QtWidgets.QVBoxLayout()

        left_label = QtWidgets.QLabel("키워드 목록")

        left_label.setStyleSheet("font-weight:600; font-size:13px;")

        left_panel.addWidget(left_label)



        left_btn_row = QtWidgets.QHBoxLayout()

        excl_sel_all = QtWidgets.QPushButton("전체 선택")

        excl_sel_all.clicked.connect(self._excl_select_all)

        excl_desel_all = QtWidgets.QPushButton("전체 해제")

        excl_desel_all.clicked.connect(self._excl_deselect_all)

        left_btn_row.addWidget(excl_sel_all)

        left_btn_row.addWidget(excl_desel_all)

        left_btn_row.addStretch(1)

        left_panel.addLayout(left_btn_row)



        self.excl_left_search = QtWidgets.QLineEdit()

        self.excl_left_search.setPlaceholderText("검색...")

        self.excl_left_search.textChanged.connect(self._filter_excl_left_list)

        left_panel.addWidget(self.excl_left_search)



        self.excl_left_list = QtWidgets.QListWidget()

        self.excl_left_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)

        left_panel.addWidget(self.excl_left_list, 1)



        self.excl_left_count = QtWidgets.QLabel("0개")

        left_panel.addWidget(self.excl_left_count)

        panels.addLayout(left_panel, 1)



        # 가운데: 이동 버튼

        center_btns = QtWidgets.QVBoxLayout()

        center_btns.addStretch(1)

        excl_add_btn = QtWidgets.QPushButton("  →  ")

        excl_add_btn.setToolTip("선택 단어를 제외 목록에 추가")

        excl_add_btn.setStyleSheet("QPushButton{font-size:16px;font-weight:700;padding:10px;}")

        excl_add_btn.clicked.connect(self._excl_move_to_right)

        excl_remove_btn = QtWidgets.QPushButton("  ←  ")

        excl_remove_btn.setToolTip("제외 목록에서 제거")

        excl_remove_btn.setStyleSheet("QPushButton{font-size:16px;font-weight:700;padding:10px;}")

        excl_remove_btn.clicked.connect(self._excl_move_to_left)

        center_btns.addWidget(excl_add_btn)

        center_btns.addSpacing(8)

        center_btns.addWidget(excl_remove_btn)

        center_btns.addStretch(1)

        panels.addLayout(center_btns)



        # 오른쪽: 제외 목록

        right_panel = QtWidgets.QVBoxLayout()

        right_label = QtWidgets.QLabel("제외 목록 (다음 실행 시 자동 적용)")

        right_label.setStyleSheet("font-weight:600; font-size:13px; color:#c62828;")

        right_panel.addWidget(right_label)



        self.excl_right_list = QtWidgets.QListWidget()

        self.excl_right_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)

        right_panel.addWidget(self.excl_right_list, 1)



        self.excl_right_count = QtWidgets.QLabel("0개")

        self.excl_right_count.setStyleSheet("color:#c62828;")

        right_panel.addWidget(self.excl_right_count)

        panels.addLayout(right_panel, 1)



        main_layout.addLayout(panels, 1)



        # ── 하단: 저장 버튼 ──

        bottom_row = QtWidgets.QHBoxLayout()

        bottom_row.addStretch(1)

        excl_save_btn = QtWidgets.QPushButton("업로드 (저장)")

        excl_save_btn.setStyleSheet(

            "QPushButton{background:#2e7d32;color:#fff;padding:8px 28px;font-weight:600;font-size:14px;}"

            "QPushButton:hover{background:#388e3c;}"

        )

        excl_save_btn.clicked.connect(self._excl_save)

        bottom_row.addWidget(excl_save_btn)

        main_layout.addLayout(bottom_row)



        # 기존 사용자 제외 단어 로드

        self._excl_user_words = set()

        self._excl_all_keywords = set()

        self._excl_load_existing()



        return w



    # ── 제외 단어 핸들러 ──



    def _excl_load_existing(self) -> None:

        """앱 시작 시 기존 user_stopwords.json 로드하여 오른쪽 패널에 표시"""

        from app.services import legacy_core as core

        self._excl_user_words = core.load_user_stopwords()

        self.excl_right_list.clear()

        for word in sorted(self._excl_user_words):

            self.excl_right_list.addItem(word)

        self.excl_right_count.setText(f"{len(self._excl_user_words)}개")



    def _browse_excl_file(self) -> None:

        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "결과 Excel 선택", "", "Excel (*.xlsx)")

        if path:

            self.excl_file_path.setText(path)



    def _load_keywords_for_exclusion(self) -> None:

        """결과 Excel에서 키워드 불러오기 → 왼쪽 패널에 표시"""

        import glob as _glob

        path = self.excl_file_path.text().strip()



        # 경로 비었으면 자동 탐색

        if not path:

            if self._last_output_file and os.path.isfile(self._last_output_file):

                path = self._last_output_file

            else:

                base = r"C:\code\exports"

                if os.path.isdir(base):

                    dirs = sorted(

                        [d for d in os.listdir(base) if os.path.isdir(os.path.join(base, d))],

                        reverse=True)

                    for d in dirs:

                        gpt_files = sorted(_glob.glob(os.path.join(base, d, "상품전처리GPT_*.xlsx")))

                        if gpt_files:

                            path = gpt_files[-1]

                            break



        if not path or not os.path.isfile(path):

            self.excl_status.setText("⚠ 결과 파일을 찾을 수 없습니다.")

            return



        self.excl_file_path.setText(path)

        try:

            import pandas as pd

            df = pd.read_excel(path, sheet_name="분리추출후")

            if "상품명" not in df.columns:

                self.excl_status.setText("⚠ '상품명' 컬럼이 없습니다.")

                return



            all_tokens = set()

            for val in df["상품명"].dropna():

                for t in str(val).split():

                    t = t.strip()

                    if len(t) >= 2:

                        all_tokens.add(t)



            self._excl_all_keywords = all_tokens

            available = sorted(all_tokens - self._excl_user_words)



            self.excl_left_list.clear()

            for word in available:

                item = QtWidgets.QListWidgetItem(word)

                item.setFlags(item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)

                item.setCheckState(QtCore.Qt.CheckState.Unchecked)

                self.excl_left_list.addItem(item)



            self.excl_left_count.setText(f"{len(available)}개")

            self.excl_status.setText(

                f"불러오기 완료: {len(all_tokens)}개 고유 토큰 (제외목록 {len(self._excl_user_words)}개 제외)")

        except Exception as e:

            self.excl_status.setText(f"⚠ 로드 오류: {e}")



    def _filter_excl_left_list(self, text: str) -> None:

        search = text.strip().lower()

        for i in range(self.excl_left_list.count()):

            item = self.excl_left_list.item(i)

            item.setHidden(bool(search) and search not in item.text().lower())



    def _excl_select_all(self) -> None:

        for i in range(self.excl_left_list.count()):

            item = self.excl_left_list.item(i)

            if not item.isHidden():

                item.setCheckState(QtCore.Qt.CheckState.Checked)



    def _excl_deselect_all(self) -> None:

        for i in range(self.excl_left_list.count()):

            item = self.excl_left_list.item(i)

            item.setCheckState(QtCore.Qt.CheckState.Unchecked)



    def _excl_move_to_right(self) -> None:

        """체크된 단어를 제외 목록으로 이동"""

        to_remove = []

        for i in range(self.excl_left_list.count()):

            item = self.excl_left_list.item(i)

            if item.checkState() == QtCore.Qt.CheckState.Checked:

                self._excl_user_words.add(item.text())

                to_remove.append(i)



        for i in reversed(to_remove):

            self.excl_left_list.takeItem(i)



        self.excl_right_list.clear()

        for word in sorted(self._excl_user_words):

            self.excl_right_list.addItem(word)



        self.excl_left_count.setText(f"{self.excl_left_list.count()}개")

        self.excl_right_count.setText(f"{len(self._excl_user_words)}개")



    def _excl_move_to_left(self) -> None:

        """제외 목록에서 선택된 단어를 복원"""

        selected = self.excl_right_list.selectedItems()

        if not selected:

            return



        for item in selected:

            word = item.text()

            self._excl_user_words.discard(word)

            if word in self._excl_all_keywords:

                new_item = QtWidgets.QListWidgetItem(word)

                new_item.setFlags(new_item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)

                new_item.setCheckState(QtCore.Qt.CheckState.Unchecked)

                self.excl_left_list.addItem(new_item)



        self.excl_right_list.clear()

        for word in sorted(self._excl_user_words):

            self.excl_right_list.addItem(word)



        self.excl_left_count.setText(f"{self.excl_left_list.count()}개")

        self.excl_right_count.setText(f"{len(self._excl_user_words)}개")



    def _excl_save(self) -> None:

        """제외 단어 JSON 저장 + STOPWORDS 합치기"""

        from app.services import legacy_core as core

        core.save_user_stopwords(self._excl_user_words)

        core.merge_user_stopwords()

        self.excl_status.setText(

            f"✅ 저장 완료: {len(self._excl_user_words)}개 제외 단어가 적용되었습니다.")

        QtWidgets.QMessageBox.information(

            self, "저장 완료",

            f"{len(self._excl_user_words)}개 제외 단어가 저장되었습니다.\n"

            "다음 전처리 실행 시 자동 적용됩니다.")



    # ─── 쿠팡 업로드 탭 ──────────────────────────────────

    def _build_coupang_tab(self) -> QtWidgets.QWidget:
        w = QtWidgets.QWidget()
        v = QtWidgets.QVBoxLayout(w)

        # 파일 선택
        file_group = QtWidgets.QGroupBox("가공파일 (업로드용 엑셀)")
        fg = QtWidgets.QHBoxLayout(file_group)
        self.coupang_file_edit = DropLineEdit()
        self.coupang_file_edit.setPlaceholderText("업로드용_YYYYMMDD_XX_llm.xlsx 파일 경로")
        fg.addWidget(self.coupang_file_edit, 1)
        btn_browse = QtWidgets.QPushButton("찾아보기")
        btn_browse.clicked.connect(self._browse_coupang_file)
        fg.addWidget(btn_browse)
        v.addWidget(file_group)

        # 행 범위 & 모드
        opt_group = QtWidgets.QGroupBox("업로드 설정")
        og = QtWidgets.QGridLayout(opt_group)

        og.addWidget(QtWidgets.QLabel("시작 행:"), 0, 0)
        self.coupang_row_start = QtWidgets.QSpinBox()
        self.coupang_row_start.setRange(0, 9999)
        self.coupang_row_start.setSpecialValueText("전체")
        self.coupang_row_start.setToolTip("0이면 전체 행 처리")
        og.addWidget(self.coupang_row_start, 0, 1)

        og.addWidget(QtWidgets.QLabel("끝 행:"), 0, 2)
        self.coupang_row_end = QtWidgets.QSpinBox()
        self.coupang_row_end.setRange(0, 9999)
        self.coupang_row_end.setSpecialValueText("시작행과 동일")
        og.addWidget(self.coupang_row_end, 0, 3)

        self.coupang_dry_run = QtWidgets.QCheckBox("DRY RUN (등록 안함, 미리보기만)")
        self.coupang_dry_run.setChecked(True)
        og.addWidget(self.coupang_dry_run, 1, 0, 1, 4)

        v.addWidget(opt_group)

        # 실행 버튼
        btn_layout = QtWidgets.QHBoxLayout()
        self.coupang_run_btn = QtWidgets.QPushButton("쿠팡 업로드 실행")
        self.coupang_run_btn.setFixedHeight(38)
        self.coupang_run_btn.setStyleSheet("font-weight:bold; font-size:13px;")
        self.coupang_run_btn.clicked.connect(self._run_coupang_upload)
        btn_layout.addWidget(self.coupang_run_btn)
        v.addLayout(btn_layout)

        # 진행 상태
        self.coupang_status = QtWidgets.QLabel("대기 중")
        v.addWidget(self.coupang_status)
        self.coupang_progress = QtWidgets.QProgressBar()
        self.coupang_progress.setValue(0)
        v.addWidget(self.coupang_progress)

        # 결과 테이블
        self.coupang_table = QtWidgets.QTableWidget()
        self.coupang_table.setColumnCount(5)
        self.coupang_table.setHorizontalHeaderLabels(["행", "상품명", "상태", "카테고리/ID", "오류"])
        self.coupang_table.horizontalHeader().setStretchLastSection(True)
        self.coupang_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v.addWidget(self.coupang_table, 1)

        # 요약
        self.coupang_summary = QtWidgets.QLabel("")
        v.addWidget(self.coupang_summary)

        return w

    def _browse_coupang_file(self) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "가공파일 선택", r"C:\code\exports", "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self.coupang_file_edit.setText(path)
            self._save_settings()

    def _run_coupang_upload(self) -> None:
        file_path = self.coupang_file_edit.text().strip()
        if not file_path or not os.path.isfile(file_path):
            QtWidgets.QMessageBox.warning(self, "오류", "가공파일 경로를 확인해주세요.")
            return

        dry_run = self.coupang_dry_run.isChecked()
        if not dry_run:
            reply = QtWidgets.QMessageBox.question(
                self, "쿠팡 실제 등록",
                "실제로 쿠팡에 상품을 등록합니다.\n계속하시겠습니까?",
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
            )
            if reply != QtWidgets.QMessageBox.StandardButton.Yes:
                return

        config = CoupangUploadConfig(
            file_path=file_path,
            row_start=self.coupang_row_start.value(),
            row_end=self.coupang_row_end.value(),
            dry_run=dry_run,
        )

        self.coupang_run_btn.setEnabled(False)
        self.coupang_status.setText("업로드 준비 중...")
        self.coupang_progress.setValue(0)
        self.coupang_table.setRowCount(0)
        self.coupang_summary.setText("")

        self._coupang_thread = QtCore.QThread(self)
        self._coupang_worker = CoupangUploadWorker(config)
        self._coupang_worker.moveToThread(self._coupang_thread)

        self._coupang_worker.status.connect(self._on_coupang_status)
        self._coupang_worker.progress.connect(self.coupang_progress.setValue)
        self._coupang_worker.finished.connect(self._on_coupang_finished)
        self._coupang_worker.error.connect(self._on_coupang_error)

        self._coupang_thread.started.connect(self._coupang_worker.run)
        self._coupang_thread.start()

    def _on_coupang_status(self, msg: str) -> None:
        self.coupang_status.setText(msg)
        self.log_box.appendPlainText(f"[쿠팡] {msg}")

    def _on_coupang_finished(self, results: list) -> None:
        self._coupang_thread.quit()
        self._coupang_thread.wait()
        self.coupang_run_btn.setEnabled(True)
        self.coupang_progress.setValue(100)

        self.coupang_table.setRowCount(0)
        ok_count = 0
        fail_count = 0
        for r in results:
            row_idx = self.coupang_table.rowCount()
            self.coupang_table.insertRow(row_idx)
            self.coupang_table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(str(r.row)))
            self.coupang_table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(r.name))

            status_item = QtWidgets.QTableWidgetItem(r.status)
            if r.status == "SUCCESS":
                status_item.setBackground(QtGui.QColor("#e8f5e9"))
                ok_count += 1
            elif r.status == "DRY_RUN":
                status_item.setBackground(QtGui.QColor("#e3f2fd"))
                ok_count += 1
            else:
                status_item.setBackground(QtGui.QColor("#ffebee"))
                fail_count += 1
            self.coupang_table.setItem(row_idx, 2, status_item)

            info = r.seller_product_id or r.category
            self.coupang_table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(info))
            self.coupang_table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(r.error))

        total = ok_count + fail_count
        self.coupang_summary.setText(f"결과: 성공 {ok_count} / 실패 {fail_count} / 전체 {total}")
        self.coupang_status.setText("완료")

    def _on_coupang_error(self, tb: str) -> None:
        self._coupang_thread.quit()
        self._coupang_thread.wait()
        self.coupang_run_btn.setEnabled(True)
        self.coupang_status.setText("오류 발생")
        self.error_box.appendPlainText(f"[쿠팡 오류]\n{tb}")
        QtWidgets.QMessageBox.critical(self, "쿠팡 업로드 오류", tb[:500])



    def _build_price_review_tab(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)



        # 데이터 로드 영역

        load_row = QtWidgets.QHBoxLayout()

        self.price_data_path = QtWidgets.QLineEdit("")

        self.price_data_path.setPlaceholderText("상품전처리GPT 파일 경로 (비우면 최신 자동 탐색)")

        price_browse = QtWidgets.QPushButton("찾기")

        price_browse.clicked.connect(self._browse_price_data)

        price_load = QtWidgets.QPushButton("데이터 로드")

        price_load.clicked.connect(self._load_price_data)

        load_row.addWidget(self.price_data_path, 1)

        load_row.addWidget(price_browse)

        load_row.addWidget(price_load)

        v.addLayout(load_row)



        self.price_status_label = QtWidgets.QLabel("데이터를 로드해 주세요.")

        self.price_status_label.setStyleSheet("color:#1565c0;")

        v.addWidget(self.price_status_label)



        # 전체 선택/해제 + 업로드 버튼

        action_row = QtWidgets.QHBoxLayout()

        self.price_check_all = QtWidgets.QPushButton("전체 선택")

        self.price_check_all.clicked.connect(self._price_check_all)

        self.price_uncheck_all = QtWidgets.QPushButton("전체 해제")

        self.price_uncheck_all.clicked.connect(self._price_uncheck_all)

        self.price_recalc_btn = QtWidgets.QPushButton("가격 재계산")

        self.price_recalc_btn.clicked.connect(self._price_recalc)

        self.price_upload_btn = QtWidgets.QPushButton("선택 상품 업로드")

        self.price_upload_btn.setStyleSheet("QPushButton { background:#2e7d32; color:#fff; padding:6px; font-weight:600; }")

        self.price_upload_btn.clicked.connect(self._price_upload)

        action_row.addWidget(self.price_check_all)

        action_row.addWidget(self.price_uncheck_all)

        action_row.addWidget(self.price_recalc_btn)

        action_row.addStretch(1)

        action_row.addWidget(self.price_upload_btn)

        v.addLayout(action_row)



        # 이미지 설정 영역

        img_settings_row = QtWidgets.QHBoxLayout()

        img_settings_row.addWidget(QtWidgets.QLabel("대표 이미지 인덱스"))

        self.price_main_idx = QtWidgets.QSpinBox()

        self.price_main_idx.setRange(1, 20)

        self.price_main_idx.setValue(2)

        self.price_main_idx.setToolTip("listing_images 폴더 내 대표 이미지 순서 (기본: 2번째)")

        img_settings_row.addWidget(self.price_main_idx)



        img_settings_row.addWidget(QtWidgets.QLabel("추가 이미지 시작"))

        self.price_add_start = QtWidgets.QSpinBox()

        self.price_add_start.setRange(1, 20)

        self.price_add_start.setValue(3)

        self.price_add_start.setToolTip("추가 이미지 시작 인덱스")

        img_settings_row.addWidget(self.price_add_start)



        img_settings_row.addWidget(QtWidgets.QLabel("추가 이미지 최대"))

        self.price_add_max = QtWidgets.QSpinBox()

        self.price_add_max.setRange(1, 20)

        self.price_add_max.setValue(10)

        self.price_add_max.setToolTip("추가 이미지 최대 개수")

        img_settings_row.addWidget(self.price_add_max)



        img_settings_row.addStretch(1)

        v.addLayout(img_settings_row)



        # 분할 설정 영역

        split_row = QtWidgets.QHBoxLayout()

        split_row.addWidget(QtWidgets.QLabel("추가금액 한도"))

        self.split_threshold_spin = QtWidgets.QSpinBox()

        self.split_threshold_spin.setRange(50, 500)

        self.split_threshold_spin.setValue(100)

        self.split_threshold_spin.setSuffix("%")

        self.split_threshold_spin.setToolTip("기본 판매가 대비 추가금액 한도 비율 (100% = 추가금액 ≤ 판매가)")

        split_row.addWidget(self.split_threshold_spin)



        self.split_warn_btn = QtWidgets.QPushButton("초과 경고 확인")

        self.split_warn_btn.clicked.connect(self._check_price_warnings)

        split_row.addWidget(self.split_warn_btn)



        self.split_auto_btn = QtWidgets.QPushButton("분할")

        self.split_auto_btn.setStyleSheet(

            "QPushButton { background:#e65100; color:#fff; padding:6px; font-weight:600; }"

        )

        self.split_auto_btn.clicked.connect(self._auto_split_all)

        self.split_auto_btn.setEnabled(False)

        split_row.addWidget(self.split_auto_btn)



        self.split_undo_btn = QtWidgets.QPushButton("분할 취소")

        self.split_undo_btn.clicked.connect(self._undo_split)

        self.split_undo_btn.setEnabled(False)

        split_row.addWidget(self.split_undo_btn)



        split_row.addStretch(1)

        v.addLayout(split_row)



        # 가격 테이블

        self.price_table = QtWidgets.QTableWidget(0, 7)

        self.price_table.setHorizontalHeaderLabels([

            "선택", "GS코드", "옵션명", "공급가", "판매가", "추가금액", "소비자가"

        ])

        self.price_table.horizontalHeader().setStretchLastSection(True)

        self.price_table.setColumnWidth(0, 40)

        self.price_table.setColumnWidth(1, 90)

        self.price_table.setColumnWidth(2, 160)

        self.price_table.setColumnWidth(3, 80)

        self.price_table.setColumnWidth(4, 80)

        self.price_table.setColumnWidth(5, 80)

        self.price_table.setColumnWidth(6, 80)

        self.price_table.cellChanged.connect(self._on_price_cell_changed)

        v.addWidget(self.price_table, 1)



        # 업로드 진행

        self.price_progress = QtWidgets.QProgressBar()

        self.price_progress.setRange(0, 100)

        self.price_progress.setValue(0)

        v.addWidget(self.price_progress)



        # 내부 데이터

        self._price_review_data = []  # [(gs9, option_name, supply_price), ...]

        self._price_review_gs_map = {}  # {gs9: [(suffix, supply_price), ...]}

        self._price_split_groups = {}   # {gs9: [[idx, ...], [idx, ...], ...]}

        self._price_is_split = {}       # {gs9: True/False}



        return w



    def _image_selection_history_path(self) -> str:

        return self._app_root_path("cafe24_image_selection_history.json")



    def _load_image_selection_history(self) -> dict:

        import json

        p = self._image_selection_history_path()

        if not os.path.isfile(p):

            return {}

        try:

            with open(p, "r", encoding="utf-8") as f:

                data = json.load(f)

            return data if isinstance(data, dict) else {}

        except Exception:

            return {}



    def _save_image_selection_history(self, products_for_selector: list, image_selections: dict) -> None:

        import json

        hist = self._load_image_selection_history()

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")



        for prod in products_for_selector:

            gs = str(prod.get("gs_code") or "").strip()

            images = list(prod.get("images") or [])

            sel = image_selections.get(gs)

            if not gs or not isinstance(sel, dict):

                continue



            main_idx = sel.get("main")

            add_idx = list(sel.get("additional") or [])

            main_name = images[main_idx] if isinstance(main_idx, int) and 0 <= main_idx < len(images) else ""

            add_names = [images[i] for i in add_idx if isinstance(i, int) and 0 <= i < len(images)]



            hist[gs] = {

                "main": main_idx,

                "additional": add_idx,

                "main_name": main_name,

                "additional_names": add_names,

                "updated_at": now,

            }



        try:

            with open(self._image_selection_history_path(), "w", encoding="utf-8") as f:

                json.dump(hist, f, ensure_ascii=False, indent=2)

        except Exception:

            pass



    def _build_preselected_image_map(self, products_for_selector: list, history: dict) -> dict:

        out = {}

        for prod in products_for_selector:

            gs = str(prod.get("gs_code") or "").strip()

            images = list(prod.get("images") or [])

            item = history.get(gs)

            if not gs or not isinstance(item, dict):

                continue



            main_idx = None

            add_indices = []



            main_name = str(item.get("main_name") or "").strip()

            if main_name and main_name in images:

                main_idx = images.index(main_name)

            elif isinstance(item.get("main"), int) and 0 <= int(item.get("main")) < len(images):

                main_idx = int(item.get("main"))



            for nm in list(item.get("additional_names") or []):

                if nm in images:

                    add_indices.append(images.index(nm))



            if not add_indices:

                for i in list(item.get("additional") or []):

                    if isinstance(i, int) and 0 <= i < len(images):

                        add_indices.append(i)



            add_indices = sorted(set(add_indices))

            out[gs] = {"main": main_idx, "additional": add_indices}



        return out



    def _browse_price_data(self) -> None:

        base = self.cafe24_export_dir.text().strip() or "C:\\code\\exports"

        date_keys = self._list_price_date_keys(base)

        date_options = ["전체(최신순)"] + date_keys

        date_selected, ok = QtWidgets.QInputDialog.getItem(

            self,

            "날짜 선택",

            "조회 날짜(YYYYMMDD)",

            date_options,

            0,

            False,

        )

        if not ok:

            return



        selected_date = "" if date_selected == "전체(최신순)" else date_selected

        candidates = self._list_price_data_candidates(base, limit=40, date_key=selected_date)



        if candidates:

            items = [label for label, _, _, _ in candidates]

            items.append("직접 파일 선택...")

            selected, ok = QtWidgets.QInputDialog.getItem(

                self,

                "상품전처리GPT 파일 선택",

                "파일명 / 행 수",

                items,

                0,

                False,

            )

            if not ok:

                return

            if selected != "직접 파일 선택...":

                for label, path, row_count, folder_name in candidates:

                    if label == selected:

                        self.price_data_path.setText(path)

                        preview_names = self._preview_product_names(path, limit=5)

                        preview_text = " | ".join(preview_names) if preview_names else "미리보기 없음"

                        self.price_status_label.setText(

                            f"선택됨: {os.path.basename(path)} ({row_count}행, {folder_name})"

                        )

                        self.price_status_label.setStyleSheet("color:#1565c0;")

                        QtWidgets.QMessageBox.information(

                            self,

                            "상품명 미리보기",

                            f"{os.path.basename(path)}\\n\\n상품명 샘플:\\n{preview_text}",

                        )

                        return



        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "상품전처리GPT 파일 선택", "", "Excel (*.xlsx)"

        )

        if path:

            self.price_data_path.setText(path)



    def _count_excel_rows_quick(self, path: str) -> int:

        try:

            from openpyxl import load_workbook



            wb = load_workbook(path, read_only=True, data_only=True)

            ws = wb.active

            rows = max(0, int(ws.max_row) - 1)

            wb.close()

            return rows

        except Exception:

            try:

                import pandas as pd



                return int(len(pd.read_excel(path)))

            except Exception:

                return -1



    def _list_price_date_keys(self, base_root: str) -> list[str]:

        if not os.path.isdir(base_root):

            return []



        keys = set()

        for d in os.listdir(base_root):

            full = os.path.join(base_root, d)

            if not os.path.isdir(full):

                continue

            m = re.match(r"^(\\d{8})", str(d))

            if m:

                keys.add(m.group(1))

        return sorted(keys, reverse=True)



    def _preview_product_names(self, path: str, limit: int = 5) -> list[str]:

        try:

            from openpyxl import load_workbook



            wb = load_workbook(path, read_only=True, data_only=True)

            ws = wb.active

            header = [str(c.value or "").strip() for c in ws[1]]

            idx = -1

            for i, name in enumerate(header):

                if name == "상품명":

                    idx = i + 1

                    break

            if idx < 0:

                idx = 8



            out = []

            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):

                v = str(row[0] or "").strip()

                if not v:

                    continue

                out.append(v)

                if len(out) >= limit:

                    break

            wb.close()

            return out

        except Exception:

            return []



    def _list_price_data_candidates(self, base_root: str, limit: int = 30, date_key: str = "") -> list[tuple[str, str, int, str]]:

        import glob as _glob



        if not os.path.isdir(base_root):

            return []



        out: list[tuple[str, str, int, str]] = []

        export_dirs = [d for d in os.listdir(base_root) if os.path.isdir(os.path.join(base_root, d))]

        export_dirs.sort(reverse=True)



        for d in export_dirs:

            if date_key and not str(d).startswith(date_key):

                continue

            folder = os.path.join(base_root, d)

            files = sorted(_glob.glob(os.path.join(folder, "상품전처리GPT_*.xlsx")), reverse=True)

            for p in files:

                row_count = self._count_excel_rows_quick(p)

                row_text = f"{row_count}행" if row_count >= 0 else "행수확인실패"

                label = f"{os.path.basename(p)} / {row_text} / {d}"

                out.append((label, p, row_count, str(d)))

                if len(out) >= limit:

                    return out



        return out



    def _load_price_data(self) -> None:

        import glob as _glob

        path = self.price_data_path.text().strip()

        if not path:

            # 자동 탐색: exports 최신 폴더에서 상품전처리GPT_*.xlsx

            base = self.cafe24_export_dir.text().strip() or "C:\\code\\exports"

            if not os.path.isdir(base):

                self.price_status_label.setText("exports 폴더를 찾을 수 없습니다.")

                return

            dirs = [d for d in os.listdir(base) if os.path.isdir(os.path.join(base, d))]

            dirs.sort(reverse=True)

            if not dirs:

                self.price_status_label.setText("exports 하위 폴더가 없습니다.")

                return

            latest = os.path.join(base, dirs[0])

            gpt_files = sorted(_glob.glob(os.path.join(latest, "상품전처리GPT_*.xlsx")))

            if not gpt_files:

                self.price_status_label.setText("상품전처리GPT 파일을 찾을 수 없습니다.")

                return

            path = gpt_files[-1]

            self.price_data_path.setText(path)



        if not os.path.isfile(path):

            self.price_status_label.setText(f"파일 없음: {path}")

            return



        try:

            from cafe24_upload import load_option_supply_prices, calc_option_prices

            gs_map = load_option_supply_prices(path)

        except Exception as e:

            self.price_status_label.setText(f"로드 오류: {e}")

            return



        # 전체 상품 로드 (단일 옵션 포함)

        multi_gs = gs_map

        if not multi_gs:

            self.price_status_label.setText("로드된 상품이 없습니다.")

            self.price_table.setRowCount(0)

            return



        self._price_review_gs_map = multi_gs

        self._price_review_data = []



        self.price_table.setRowCount(0)

        self.price_table.blockSignals(True)



        row = 0

        for gs9 in sorted(multi_gs.keys()):

            opts = multi_gs[gs9]

            supply_prices = [sp for _, sp in opts]

            try:

                from cafe24_upload import calc_option_prices

                prices = calc_option_prices(supply_prices)

            except Exception:

                prices = {"base_selling": 0, "base_consumer": 0, "additional_amounts": [0] * len(opts)}



            base_selling = prices["base_selling"]

            consumer = prices["base_consumer"]

            additionals = prices["additional_amounts"]



            for i, (suffix, sp) in enumerate(opts):

                self.price_table.insertRow(row)

                sell_price = base_selling + additionals[i]



                # 체크박스 (첫 번째 옵션에만 표시)

                chk_item = QtWidgets.QTableWidgetItem()

                if i == 0:

                    chk_item.setFlags(chk_item.flags() | QtCore.Qt.ItemIsUserCheckable)

                    chk_item.setCheckState(QtCore.Qt.Checked)

                else:

                    chk_item.setFlags(chk_item.flags() & ~QtCore.Qt.ItemIsUserCheckable)

                    chk_item.setCheckState(QtCore.Qt.Unchecked)

                self.price_table.setItem(row, 0, chk_item)



                # GS코드

                gs_item = QtWidgets.QTableWidgetItem(gs9 if i == 0 else "")

                gs_item.setFlags(gs_item.flags() & ~QtCore.Qt.ItemIsEditable)

                self.price_table.setItem(row, 1, gs_item)



                # 옵션명

                opt_item = QtWidgets.QTableWidgetItem(suffix or f"옵션{chr(65 + i)}")

                opt_item.setFlags(opt_item.flags() & ~QtCore.Qt.ItemIsEditable)

                self.price_table.setItem(row, 2, opt_item)



                # 공급가 (읽기 전용)

                sp_item = QtWidgets.QTableWidgetItem(f"{int(sp):,}")

                sp_item.setFlags(sp_item.flags() & ~QtCore.Qt.ItemIsEditable)

                sp_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                self.price_table.setItem(row, 3, sp_item)



                # 판매가 (편집 가능)

                sell_item = QtWidgets.QTableWidgetItem(str(int(sell_price)))

                sell_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                self.price_table.setItem(row, 4, sell_item)



                # 추가금액 (편집 가능)

                add_item = QtWidgets.QTableWidgetItem(str(int(additionals[i])))

                add_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                self.price_table.setItem(row, 5, add_item)



                # 소비자가 (첫 번째 옵션에만 표시)

                con_item = QtWidgets.QTableWidgetItem(f"{int(consumer):,}" if i == 0 else "")

                con_item.setFlags(con_item.flags() & ~QtCore.Qt.ItemIsEditable)

                con_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                self.price_table.setItem(row, 6, con_item)



                # 그룹 구분: GS코드별 배경색 교대

                if list(sorted(multi_gs.keys())).index(gs9) % 2 == 1:

                    for c in range(7):

                        it = self.price_table.item(row, c)

                        if it:

                            it.setBackground(QtGui.QColor("#f5f5f5"))



                self._price_review_data.append((gs9, suffix, sp))

                row += 1



        self.price_table.blockSignals(False)

        _multi_cnt = sum(1 for v in multi_gs.values() if len(v) > 1)

        _single_cnt = len(multi_gs) - _multi_cnt

        self.price_status_label.setText(

            f"전체 {len(multi_gs)}개 상품 ({_multi_cnt}개 다중옵션, {_single_cnt}개 단일), {row}개 옵션 로드 완료. "

            f"판매가/추가금액을 수정 후 업로드하세요."

        )

        # 분할 상태 초기화 & 초과 경고 자동 확인

        self._price_split_groups.clear()

        self._price_is_split.clear()

        self.split_undo_btn.setEnabled(False)

        self._check_price_warnings()



    def _price_check_all(self) -> None:

        for r in range(self.price_table.rowCount()):

            item = self.price_table.item(r, 0)

            if item and (item.flags() & QtCore.Qt.ItemIsUserCheckable):

                item.setCheckState(QtCore.Qt.Checked)



    def _price_uncheck_all(self) -> None:

        for r in range(self.price_table.rowCount()):

            item = self.price_table.item(r, 0)

            if item and (item.flags() & QtCore.Qt.ItemIsUserCheckable):

                item.setCheckState(QtCore.Qt.Unchecked)



    # ── 가격 초과 경고 / 분할 ──────────────────────────────────



    def _check_price_warnings(self) -> None:

        """추가금액이 기본판매가 × threshold%를 초과하는 옵션을 빨간 배경으로 표시."""

        if not self._price_review_gs_map:

            return

        threshold_pct = self.split_threshold_spin.value()

        threshold_ratio = threshold_pct / 100.0

        has_warnings = False

        gs_keys = sorted(self._price_review_gs_map.keys())



        row = 0

        for gi, gs9 in enumerate(gs_keys):

            opts = self._price_review_gs_map[gs9]



            # 분할 헤더행 건너뛰기 (분할 상태일 때)

            if self._price_is_split.get(gs9):

                # 분할 상태에서는 그룹별로 검사

                groups = self._price_split_groups.get(gs9, [])

                for g_idx, group_indices in enumerate(groups):

                    # 분할 헤더행 건너뛰기

                    header_item = self.price_table.item(row, 1)

                    if header_item and header_item.data(QtCore.Qt.UserRole) == "split_header":

                        row += 1

                    # 그룹 내 첫 번째 옵션이 base

                    try:

                        group_base = int(self.price_table.item(row, 4).text())

                    except (ValueError, AttributeError):

                        row += len(group_indices)

                        continue

                    limit = group_base * threshold_ratio

                    for j, opt_idx in enumerate(group_indices):

                        if row >= self.price_table.rowCount():

                            break

                        try:

                            additional = int(self.price_table.item(row, 5).text())

                        except (ValueError, AttributeError):

                            additional = 0

                        if additional > limit:

                            for c in range(7):

                                it = self.price_table.item(row, c)

                                if it:

                                    it.setBackground(QtGui.QColor("#ffcdd2"))

                            has_warnings = True

                        else:

                            bg = QtGui.QColor("#f5f5f5") if gi % 2 == 1 else QtGui.QColor("#ffffff")

                            for c in range(7):

                                it = self.price_table.item(row, c)

                                if it:

                                    it.setBackground(bg)

                        row += 1

            else:

                # 일반 상태: 첫 번째 옵션의 판매가가 base

                try:

                    base_sell = int(self.price_table.item(row, 4).text())

                except (ValueError, AttributeError):

                    row += len(opts)

                    continue

                limit = base_sell * threshold_ratio

                for i in range(len(opts)):

                    r = row + i

                    if r >= self.price_table.rowCount():

                        break

                    try:

                        additional = int(self.price_table.item(r, 5).text())

                    except (ValueError, AttributeError):

                        additional = 0

                    if additional > limit:

                        for c in range(7):

                            it = self.price_table.item(r, c)

                            if it:

                                it.setBackground(QtGui.QColor("#ffcdd2"))

                        has_warnings = True

                    else:

                        bg = QtGui.QColor("#f5f5f5") if gi % 2 == 1 else QtGui.QColor("#ffffff")

                        for c in range(7):

                            it = self.price_table.item(r, c)

                            if it:

                                it.setBackground(bg)

                row += len(opts)



        self.split_auto_btn.setEnabled(has_warnings)

        if has_warnings:

            self.price_status_label.setText(

                f"⚠ 추가금액이 판매가의 {threshold_pct}%를 초과하는 옵션이 있습니다. "

                f"'분할' 버튼으로 자동 분할할 수 있습니다."

            )

            self.price_status_label.setStyleSheet("color:#c62828; font-weight:600;")

        else:

            self.price_status_label.setText("모든 옵션의 추가금액이 한도 이내입니다.")

            self.price_status_label.setStyleSheet("color:#2e7d32;")



    def _auto_split_all(self) -> None:

        """초과하는 GS그룹을 자동 분할하여 테이블을 재렌더링."""

        if not self._price_review_gs_map:

            return

        from cafe24_upload import calc_option_prices, compute_split_groups



        threshold_pct = self.split_threshold_spin.value()

        threshold_ratio = threshold_pct / 100.0

        split_count = 0



        for gs9 in sorted(self._price_review_gs_map.keys()):

            opts = self._price_review_gs_map[gs9]

            supply_prices = [sp for _, sp in opts]

            prices = calc_option_prices(supply_prices)

            base_selling = prices["base_selling"]

            additionals = prices["additional_amounts"]



            # 초과 여부 확인

            limit = base_selling * threshold_ratio

            if not any(a > limit for a in additionals):

                continue



            # 판매가 리스트 구성

            sell_prices = [base_selling + a for a in additionals]

            groups = compute_split_groups(sell_prices, threshold_pct=threshold_pct)



            if len(groups) > 1:

                self._price_split_groups[gs9] = groups

                self._price_is_split[gs9] = True

                split_count += 1



        if split_count > 0:

            self._render_split_table()

            self.split_undo_btn.setEnabled(True)

            self.split_auto_btn.setEnabled(False)

            self.price_status_label.setText(f"{split_count}개 상품이 분할되었습니다.")

            self.price_status_label.setStyleSheet("color:#e65100; font-weight:600;")

        else:

            self.price_status_label.setText("분할이 필요한 상품이 없습니다.")

            self.price_status_label.setStyleSheet("color:#2e7d32;")



    def _render_split_table(self) -> None:

        """분할 정보를 반영하여 가격 테이블을 재렌더링."""

        if not self._price_review_gs_map:

            return

        from cafe24_upload import calc_option_prices, _ceil100



        self.price_table.setRowCount(0)

        self.price_table.blockSignals(True)

        self._price_review_data = []



        row = 0

        gs_keys = sorted(self._price_review_gs_map.keys())

        for gi, gs9 in enumerate(gs_keys):

            opts = self._price_review_gs_map[gs9]

            supply_prices = [sp for _, sp in opts]

            prices = calc_option_prices(supply_prices)

            base_selling_orig = prices["base_selling"]

            additionals_orig = prices["additional_amounts"]

            sell_all = [base_selling_orig + a for a in additionals_orig]



            is_split = self._price_is_split.get(gs9, False)



            if is_split and gs9 in self._price_split_groups:

                groups = self._price_split_groups[gs9]

                for g_idx, group_indices in enumerate(groups):

                    # ── 분할 헤더행 삽입 ──

                    self.price_table.insertRow(row)

                    # 옵션 이름 범위 계산

                    first_opt = opts[group_indices[0]][0] or chr(65 + group_indices[0])

                    last_opt = opts[group_indices[-1]][0] or chr(65 + group_indices[-1])

                    first_label = first_opt.split()[0] if first_opt.split() else first_opt[:1]

                    last_label = last_opt.split()[0] if last_opt.split() else last_opt[:1]

                    range_text = f"({first_label}~{last_label})" if len(group_indices) > 1 else f"({first_label})"



                    group_sell = [sell_all[idx] for idx in group_indices]

                    group_base = group_sell[0]

                    group_consumer = _ceil100(group_base * 1.2)



                    # 체크박스 (분할 그룹의 첫 번째 헤더에만)

                    chk_item = QtWidgets.QTableWidgetItem()

                    if g_idx == 0:

                        chk_item.setFlags(chk_item.flags() | QtCore.Qt.ItemIsUserCheckable)

                        chk_item.setCheckState(QtCore.Qt.Checked)

                    else:

                        chk_item.setFlags(chk_item.flags() & ~QtCore.Qt.ItemIsUserCheckable)

                    self.price_table.setItem(row, 0, chk_item)



                    # GS코드 (헤더행 마킹용)

                    gs_item = QtWidgets.QTableWidgetItem(gs9 if g_idx == 0 else "")

                    gs_item.setFlags(gs_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    gs_item.setData(QtCore.Qt.UserRole, "split_header")

                    self.price_table.setItem(row, 1, gs_item)



                    # 헤더 텍스트

                    hdr_item = QtWidgets.QTableWidgetItem(

                        f"━ 분할{g_idx + 1}/{len(groups)} {range_text} ━"

                    )

                    hdr_item.setFlags(hdr_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    self.price_table.setItem(row, 2, hdr_item)



                    # 공급가 칸 비움

                    sp_hdr = QtWidgets.QTableWidgetItem("")

                    sp_hdr.setFlags(sp_hdr.flags() & ~QtCore.Qt.ItemIsEditable)

                    self.price_table.setItem(row, 3, sp_hdr)



                    # base 판매가

                    base_item = QtWidgets.QTableWidgetItem(f"base:{group_base}")

                    base_item.setFlags(base_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    base_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 4, base_item)



                    # 한도

                    limit_val = int(group_base * (self.split_threshold_spin.value() / 100.0))

                    limit_item = QtWidgets.QTableWidgetItem(f"한도:{limit_val}")

                    limit_item.setFlags(limit_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    limit_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 5, limit_item)



                    # 소비자가

                    con_hdr = QtWidgets.QTableWidgetItem(f"{group_consumer:,}")

                    con_hdr.setFlags(con_hdr.flags() & ~QtCore.Qt.ItemIsEditable)

                    con_hdr.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 6, con_hdr)



                    # 헤더행 배경색 (주황)

                    for c in range(7):

                        it = self.price_table.item(row, c)

                        if it:

                            it.setBackground(QtGui.QColor("#fff3e0"))

                            it.setForeground(QtGui.QColor("#e65100"))

                    row += 1



                    # ── 그룹 내 옵션 행 ──

                    for j, opt_idx in enumerate(group_indices):

                        self.price_table.insertRow(row)

                        suffix, sp = opts[opt_idx]

                        sell_price = sell_all[opt_idx]

                        additional = sell_price - group_base



                        # 체크박스 (옵션행에는 비활성)

                        opt_chk = QtWidgets.QTableWidgetItem()

                        opt_chk.setFlags(opt_chk.flags() & ~QtCore.Qt.ItemIsUserCheckable)

                        self.price_table.setItem(row, 0, opt_chk)



                        # GS코드 (비움)

                        opt_gs = QtWidgets.QTableWidgetItem("")

                        opt_gs.setFlags(opt_gs.flags() & ~QtCore.Qt.ItemIsEditable)

                        self.price_table.setItem(row, 1, opt_gs)



                        # 옵션명

                        opt_name_item = QtWidgets.QTableWidgetItem(

                            suffix or f"옵션{chr(65 + opt_idx)}"

                        )

                        opt_name_item.setFlags(opt_name_item.flags() & ~QtCore.Qt.ItemIsEditable)

                        self.price_table.setItem(row, 2, opt_name_item)



                        # 공급가 (읽기 전용)

                        sp_item = QtWidgets.QTableWidgetItem(f"{int(sp):,}")

                        sp_item.setFlags(sp_item.flags() & ~QtCore.Qt.ItemIsEditable)

                        sp_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                        self.price_table.setItem(row, 3, sp_item)



                        # 판매가 (편집 가능)

                        sell_item = QtWidgets.QTableWidgetItem(str(int(sell_price)))

                        sell_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                        self.price_table.setItem(row, 4, sell_item)



                        # 추가금액 (편집 가능)

                        add_item = QtWidgets.QTableWidgetItem(str(int(additional)))

                        add_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                        self.price_table.setItem(row, 5, add_item)



                        # 소비자가 (비움)

                        con_item = QtWidgets.QTableWidgetItem("")

                        con_item.setFlags(con_item.flags() & ~QtCore.Qt.ItemIsEditable)

                        self.price_table.setItem(row, 6, con_item)



                        # 배경색

                        if gi % 2 == 1:

                            for c in range(7):

                                it = self.price_table.item(row, c)

                                if it:

                                    it.setBackground(QtGui.QColor("#f5f5f5"))



                        self._price_review_data.append((gs9, suffix, sp))

                        row += 1

            else:

                # ── 분할 안 된 상품: 기존 방식대로 표시 ──

                consumer = prices["base_consumer"]

                for i, (suffix, sp) in enumerate(opts):

                    self.price_table.insertRow(row)

                    sell_price = base_selling_orig + additionals_orig[i]



                    # 체크박스

                    chk_item = QtWidgets.QTableWidgetItem()

                    if i == 0:

                        chk_item.setFlags(chk_item.flags() | QtCore.Qt.ItemIsUserCheckable)

                        chk_item.setCheckState(QtCore.Qt.Checked)

                    else:

                        chk_item.setFlags(chk_item.flags() & ~QtCore.Qt.ItemIsUserCheckable)

                        chk_item.setCheckState(QtCore.Qt.Unchecked)

                    self.price_table.setItem(row, 0, chk_item)



                    gs_item = QtWidgets.QTableWidgetItem(gs9 if i == 0 else "")

                    gs_item.setFlags(gs_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    self.price_table.setItem(row, 1, gs_item)



                    opt_item = QtWidgets.QTableWidgetItem(suffix or f"옵션{chr(65 + i)}")

                    opt_item.setFlags(opt_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    self.price_table.setItem(row, 2, opt_item)



                    sp_item = QtWidgets.QTableWidgetItem(f"{int(sp):,}")

                    sp_item.setFlags(sp_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    sp_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 3, sp_item)



                    sell_item = QtWidgets.QTableWidgetItem(str(int(sell_price)))

                    sell_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 4, sell_item)



                    add_item = QtWidgets.QTableWidgetItem(str(int(additionals_orig[i])))

                    add_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 5, add_item)



                    con_item = QtWidgets.QTableWidgetItem(f"{int(consumer):,}" if i == 0 else "")

                    con_item.setFlags(con_item.flags() & ~QtCore.Qt.ItemIsEditable)

                    con_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

                    self.price_table.setItem(row, 6, con_item)



                    if gi % 2 == 1:

                        for c in range(7):

                            it = self.price_table.item(row, c)

                            if it:

                                it.setBackground(QtGui.QColor("#f5f5f5"))



                    self._price_review_data.append((gs9, suffix, sp))

                    row += 1



        self.price_table.blockSignals(False)



    def _undo_split(self) -> None:

        """분할을 취소하고 원래 테이블로 복원."""

        self._price_split_groups.clear()

        self._price_is_split.clear()

        self.split_undo_btn.setEnabled(False)

        self.split_auto_btn.setEnabled(False)

        self._load_price_data()

        self.price_status_label.setText("분할이 취소되었습니다.")

        self.price_status_label.setStyleSheet("color:#1565c0;")



    def _on_price_cell_changed(self, changed_row: int, changed_col: int) -> None:

        """판매가(col 4) 수정 시 같은 그룹의 추가금액(col 5)을 자동 재계산.



        분할 상태에서는 분할 그룹 단위로, 일반 상태에서는 GS코드 단위로 동작.

        """

        if changed_col != 4:

            return

        if not self._price_review_gs_map:

            return



        # 분할 헤더행이면 무시 (편집 불가이므로 보통 호출 안 되지만 안전장치)

        gs_item = self.price_table.item(changed_row, 1)

        if gs_item and gs_item.data(QtCore.Qt.UserRole) == "split_header":

            return



        # 분할 상태인 상품이 있는 경우: 테이블 행 구조가 달라지므로 별도 처리

        any_split = any(self._price_is_split.values())



        if any_split:

            # 분할 테이블에서 changed_row가 속한 그룹(분할 서브그룹 또는 일반 GS그룹) 찾기

            row = 0

            for gs9 in sorted(self._price_review_gs_map.keys()):

                opts = self._price_review_gs_map[gs9]

                is_split = self._price_is_split.get(gs9, False)



                if is_split and gs9 in self._price_split_groups:

                    groups = self._price_split_groups[gs9]

                    for group_indices in groups:

                        header_row = row

                        row += 1  # 헤더행 건너뛰기

                        opt_start = row

                        opt_count = len(group_indices)

                        if opt_start <= changed_row < opt_start + opt_count:

                            # 이 서브그룹 내에서 재계산

                            try:

                                base_sell = int(self.price_table.item(opt_start, 4).text())

                            except (ValueError, AttributeError):

                                return

                            self.price_table.blockSignals(True)

                            for k in range(opt_count):

                                r = opt_start + k

                                try:

                                    sell = int(self.price_table.item(r, 4).text())

                                except (ValueError, AttributeError):

                                    sell = base_sell

                                self.price_table.item(r, 5).setText(str(sell - base_sell))

                            self.price_table.blockSignals(False)

                            return

                        row += opt_count

                else:

                    group_start = row

                    group_len = len(opts)

                    if group_start <= changed_row < group_start + group_len:

                        try:

                            base_sell = int(self.price_table.item(group_start, 4).text())

                        except (ValueError, AttributeError):

                            return

                        self.price_table.blockSignals(True)

                        for k in range(group_len):

                            r = group_start + k

                            try:

                                sell = int(self.price_table.item(r, 4).text())

                            except (ValueError, AttributeError):

                                sell = base_sell

                            self.price_table.item(r, 5).setText(str(sell - base_sell))

                        self.price_table.blockSignals(False)

                        return

                    row += group_len

        else:

            # 일반 상태: 기존 로직

            row = 0

            for gs9 in sorted(self._price_review_gs_map.keys()):

                opts = self._price_review_gs_map[gs9]

                if row <= changed_row < row + len(opts):

                    try:

                        base_sell = int(self.price_table.item(row, 4).text())

                    except (ValueError, AttributeError):

                        return

                    self.price_table.blockSignals(True)

                    for i in range(len(opts)):

                        r = row + i

                        try:

                            sell = int(self.price_table.item(r, 4).text())

                        except (ValueError, AttributeError):

                            sell = base_sell

                        self.price_table.item(r, 5).setText(str(sell - base_sell))

                    self.price_table.blockSignals(False)

                    return

                row += len(opts)



    def _price_recalc(self) -> None:

        """현재 테이블의 공급가를 기반으로 가격을 재계산."""

        if not self._price_review_gs_map:

            return

        from cafe24_upload import calc_option_prices



        row = 0

        for gs9 in sorted(self._price_review_gs_map.keys()):

            opts = self._price_review_gs_map[gs9]

            supply_prices = [sp for _, sp in opts]

            prices = calc_option_prices(supply_prices)

            base_selling = prices["base_selling"]

            consumer = prices["base_consumer"]

            additionals = prices["additional_amounts"]



            for i in range(len(opts)):

                if row >= self.price_table.rowCount():

                    break

                sell_price = base_selling + additionals[i]

                self.price_table.item(row, 4).setText(str(int(sell_price)))

                self.price_table.item(row, 5).setText(str(int(additionals[i])))

                if i == 0:

                    self.price_table.item(row, 6).setText(f"{int(consumer):,}")

                row += 1



        self.price_status_label.setText("가격 재계산 완료.")



    def _price_upload(self) -> None:

        """선택된 상품의 이미지 + 가격을 Cafe24에 업로드."""

        if not self._price_review_gs_map:

            QtWidgets.QMessageBox.warning(self, "데이터 없음", "먼저 데이터를 로드해 주세요.")

            return



        any_split = any(self._price_is_split.values())



        # 테이블에서 체크된 GS코드와 수정된 추가금액 수집

        checked_gs = set()

        edited_amounts = {}   # {gs9: [additional_amount, ...]}  원본 인덱스 순서

        split_data = {}       # {gs9: {groups: [...]}}



        row = 0

        for gs9 in sorted(self._price_review_gs_map.keys()):

            opts = self._price_review_gs_map[gs9]

            is_split = self._price_is_split.get(gs9, False) and gs9 in self._price_split_groups



            if is_split:

                groups = self._price_split_groups[gs9]

                # 첫 번째 분할 헤더행의 체크박스 확인

                chk_item = self.price_table.item(row, 0)

                if chk_item and chk_item.checkState() == QtCore.Qt.Checked:

                    checked_gs.add(gs9)



                # 분할 그룹별로 추가금액 수집

                amounts_by_orig_idx = {}  # {원본인덱스: 추가금액}

                group_data_list = []

                for g_idx, group_indices in enumerate(groups):

                    row += 1  # 헤더행 건너뛰기

                    group_amounts = []

                    group_sell_prices = []

                    for j, opt_idx in enumerate(group_indices):

                        try:

                            amt = int(self.price_table.item(row, 5).text())

                        except (ValueError, AttributeError):

                            amt = 0

                        try:

                            sell = int(self.price_table.item(row, 4).text())

                        except (ValueError, AttributeError):

                            sell = 0

                        amounts_by_orig_idx[opt_idx] = amt

                        group_amounts.append(amt)

                        group_sell_prices.append(sell)

                        row += 1



                    # 그룹 정보 저장

                    first_opt = opts[group_indices[0]][0] or chr(65 + group_indices[0])

                    last_opt = opts[group_indices[-1]][0] or chr(65 + group_indices[-1])

                    first_label = first_opt.split()[0] if first_opt.split() else first_opt[:1]

                    last_label = last_opt.split()[0] if last_opt.split() else last_opt[:1]

                    name_suffix = f"({first_label}~{last_label})" if len(group_indices) > 1 else f"({first_label})"



                    group_data_list.append({

                        "option_indices": group_indices,

                        "base_selling": group_sell_prices[0] if group_sell_prices else 0,

                        "additional_amounts": group_amounts,

                        "name_suffix": name_suffix,

                    })



                # edited_amounts는 원본 옵션 순서로 복원

                if gs9 in checked_gs:

                    edited_amounts[gs9] = [

                        amounts_by_orig_idx.get(i, 0) for i in range(len(opts))

                    ]

                    split_data[gs9] = {"groups": group_data_list}

            else:

                # 일반 상태

                chk_item = self.price_table.item(row, 0)

                if chk_item and chk_item.checkState() == QtCore.Qt.Checked:

                    checked_gs.add(gs9)

                if gs9 in checked_gs:

                    amounts = []

                    for i in range(len(opts)):

                        try:

                            amt = int(self.price_table.item(row + i, 5).text())

                        except (ValueError, AttributeError):

                            amt = 0

                        amounts.append(amt)

                    edited_amounts[gs9] = amounts

                row += len(opts)



        if not checked_gs:

            QtWidgets.QMessageBox.warning(self, "선택 없음", "업로드할 상품을 선택해 주세요.")

            return



        # 이미지 폴더 경로 확인

        export_dir_val = self.cafe24_export_dir.text().strip() or ""

        image_root_val = self.cafe24_image_root.text().strip() or ""

        date_tag = self.cafe24_date_tag.text().strip() or QtCore.QDate.currentDate().toString("yyyyMMdd")



        if image_root_val:

            img_root = image_root_val

        else:

            # 1순위: 현재 로드한 상품전처리 파일 폴더 기준

            loaded_path = self.price_data_path.text().strip()

            loaded_dir = os.path.dirname(loaded_path) if loaded_path else ""

            if loaded_dir and os.path.isdir(loaded_dir):

                img_root = os.path.join(loaded_dir, "listing_images", date_tag)

                if not os.path.isdir(img_root):

                    img_root = os.path.join(loaded_dir, "listing_images")

            else:

                # 2순위: export 루트의 최신 폴더

                base_root = export_dir_val or os.path.join("C:\\code", "exports")

                if os.path.isdir(base_root):

                    export_dirs = [d for d in os.listdir(base_root) if os.path.isdir(os.path.join(base_root, d))]

                    export_dirs.sort(reverse=True)

                    if export_dirs:

                        latest = os.path.join(base_root, export_dirs[0])

                        img_root = os.path.join(latest, "listing_images", date_tag)

                        if not os.path.isdir(img_root):

                            img_root = os.path.join(latest, "listing_images")

                    else:

                        QtWidgets.QMessageBox.warning(self, "폴더 없음", "이미지 폴더를 찾을 수 없습니다.")

                        return

                else:

                    QtWidgets.QMessageBox.warning(self, "폴더 없음", "export 폴더를 찾을 수 없습니다.")

                    return



        candidate_roots = []

        if os.path.isdir(img_root):

            candidate_roots.append(img_root)



        dated_root = os.path.join(img_root, date_tag)

        if os.path.isdir(dated_root):

            candidate_roots.append(dated_root)



        if os.path.isdir(img_root):

            subdirs = [

                d for d in os.listdir(img_root)

                if os.path.isdir(os.path.join(img_root, d))

            ]

            date_subdirs = [d for d in subdirs if re.match(r"^\\d{8}$", str(d))]

            for d in sorted(date_subdirs, reverse=True):

                candidate_roots.append(os.path.join(img_root, d))



        # 중복 제거 (순서 유지)

        _seen_roots = set()

        roots = []

        for r in candidate_roots:

            rr = os.path.normpath(r)

            if rr in _seen_roots:

                continue

            _seen_roots.add(rr)

            roots.append(r)



        products_for_selector = []

        missed_gs = []

        for gs9 in sorted(checked_gs):

            gs_folder = None



            for root in roots:

                if not os.path.isdir(root):

                    continue

                for folder_name in os.listdir(root):

                    folder_path = os.path.join(root, folder_name)

                    if os.path.isdir(folder_path) and folder_name.upper().startswith(gs9.upper()):

                        gs_folder = folder_path

                        break

                if gs_folder:

                    break



            if not gs_folder:

                missed_gs.append(gs9)

                continue



            image_files = [

                f for f in os.listdir(gs_folder)

                if f.lower().endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp"))

            ]

            image_files.sort()



            if image_files:

                products_for_selector.append({

                    "gs_code": gs9,

                    "image_folder": gs_folder,

                    "images": image_files

                })

            else:

                missed_gs.append(gs9)



        if not products_for_selector:

            roots_msg = "\\n".join([f"- {r}" for r in roots[:5]])

            QtWidgets.QMessageBox.warning(

                self,

                "이미지 없음",

                "선택한 상품의 이미지를 찾을 수 없습니다.\\n"

                f"검색 루트:\\n{roots_msg}"

            )

            return



        # 이미지 선택 대화상자 표시

        history_map = self._load_image_selection_history()

        preselected = self._build_preselected_image_map(products_for_selector, history_map)

        dialog = ImageSelectorDialog(products_for_selector, self, initial_selections=preselected)

        if dialog.exec() != QtWidgets.QDialog.Accepted:

            return  # 취소



        image_selections = dialog.get_selections()

        self._save_image_selection_history(products_for_selector, image_selections)



        # 설정값 수집 (옵션 가격 확인 탭의 이미지 설정 사용)

        cfg_path = self._desktop_key_path("cafe24_token.txt")

        date_tag = self.cafe24_date_tag.text().strip() or QtCore.QDate.currentDate().toString("yyyyMMdd")

        main_index = self.price_main_idx.value()  # 옵션 가격 확인 탭에서 가져옴

        add_start = self.price_add_start.value()  # 옵션 가격 확인 탭에서 가져옴

        add_max = self.price_add_max.value()      # 옵션 가격 확인 탭에서 가져옴

        match_mode = self.cafe24_match_mode.currentText().strip()

        match_prefix = self.cafe24_match_prefix.value()



        # GS 리스트와 수정된 금액을 임시 파일로 저장

        import json

        price_data_path = self._app_root_path("cafe24_price_upload_data.json")

        export_dir_val = self.cafe24_export_dir.text().strip() or ""

        image_root_val = self.cafe24_image_root.text().strip() or ""

        loaded_path = self.price_data_path.text().strip()

        loaded_dir = os.path.dirname(loaded_path) if loaded_path else ""

        effective_export_dir = export_dir_val or loaded_dir

        data = {

            "checked_gs": list(checked_gs),

            "edited_amounts": edited_amounts,

            "date_tag": date_tag,

            "main_index": main_index,

            "add_start": add_start,

            "add_max": add_max,

            "match_mode": match_mode,

            "match_prefix": match_prefix,

            "export_dir": effective_export_dir,

            "image_root": image_root_val,

            "image_selections": image_selections,  # 이미지 선택 정보 추가

        }

        # 분할 정보가 있으면 추가 (하위 호환 유지)

        if split_data:

            data["split_groups"] = split_data

        with open(price_data_path, "w", encoding="utf-8") as f:

            json.dump(data, f, ensure_ascii=False, indent=2)



        # cafe24_upload_config.txt 작성

        up_cfg_path = self._app_root_path("cafe24_upload_config.txt")

        with open(up_cfg_path, "w", encoding="utf-8") as f:

            if date_tag:

                f.write(f"DATE_TAG={date_tag}\n")

            f.write(f"MAIN_INDEX={main_index}\n")

            f.write(f"ADD_START={add_start}\n")

            f.write(f"ADD_MAX={add_max}\n")

            if effective_export_dir:

                f.write(f"EXPORT_DIR={effective_export_dir}\n")

            if image_root_val:

                f.write(f"IMAGE_ROOT={image_root_val}\n")

            f.write(f"RETRY_COUNT={self.cafe24_retry.value()}\n")

            f.write(f"RETRY_DELAY={self.cafe24_retry_delay.value()}\n")

            f.write(f"MATCH_MODE={match_mode}\n")

            f.write(f"MATCH_PREFIX={match_prefix}\n")

            f.write(f"PRICE_DATA={price_data_path}\n")



        # 토큰 파일 복사

        token_src = self.cafe24_token_path.text().strip()

        token_dst = self._desktop_key_path("cafe24_token.txt")

        if token_src and os.path.isfile(token_src) and os.path.abspath(token_src) != os.path.abspath(token_dst):

            try:

                os.makedirs(os.path.dirname(token_dst), exist_ok=True)

                with open(token_src, "r", encoding="utf-8") as rf:

                    td = rf.read()

                with open(token_dst, "w", encoding="utf-8") as wf:

                    wf.write(td)

            except Exception:

                pass



        # 실제 업로드될 상품 수 표시 (옵션 상품 + 단일 상품)

        upload_count_msg = f"업로드 시작: 옵션상품 {len(checked_gs)}개"

        if checked_gs:

            upload_count_msg += " + 단일상품 (이미지 폴더 기준)"

        else:

            upload_count_msg = "업로드 시작: 이미지 폴더 전체 상품..."

        self.price_status_label.setText(upload_count_msg)

        self.price_progress.setValue(0)



        self._price_upload_proc = QtCore.QProcess(self)

        self._price_upload_proc.setProgram(sys.executable)

        upload_script = self._app_root_path("cafe24_upload.py")

        self._price_upload_proc.setArguments([upload_script])

        self._price_upload_proc.setWorkingDirectory(self._app_root_path(""))

        self._price_upload_proc.readyReadStandardOutput.connect(self._on_price_upload_stdout)

        self._price_upload_proc.readyReadStandardError.connect(self._on_price_upload_stderr)

        self._price_upload_proc.finished.connect(self._on_price_upload_finished)



        # 디버그 로그

        self.log_box.appendPlainText("=" * 60)

        self.log_box.appendPlainText("🚀 Cafe24 업로드 프로세스 시작")

        self.log_box.appendPlainText(f"Python: {sys.executable}")

        self.log_box.appendPlainText(f"Script: {upload_script}")

        self.log_box.appendPlainText(f"설정 파일: {up_cfg_path}")

        self.log_box.appendPlainText(f"가격 데이터: {price_data_path}")

        self.log_box.appendPlainText(f"선택된 상품: {len(checked_gs)}개")

        self.log_box.appendPlainText(f"이미지 선택: {len(image_selections)}개")

        self.log_box.appendPlainText("=" * 60)



        self._price_upload_proc.start()



        # 프로세스 시작 확인

        if not self._price_upload_proc.waitForStarted(3000):

            error_msg = self._price_upload_proc.errorString()

            self.log_box.appendPlainText(f"❌ 프로세스 시작 실패: {error_msg}")

            QtWidgets.QMessageBox.critical(self, "업로드 실패", f"업로드 프로세스를 시작할 수 없습니다:\n{error_msg}")

        else:

            self.log_box.appendPlainText("✅ 업로드 프로세스 시작됨")



    def _on_price_upload_stdout(self) -> None:

        data = bytes(self._price_upload_proc.readAllStandardOutput()).decode("utf-8", errors="ignore")

        if data.strip():

            for line in data.splitlines():

                if line.startswith("PROGRESS "):

                    try:

                        cur, total = line.replace("PROGRESS ", "").split("/")

                        cur = int(cur.strip()); total = int(total.strip())

                        if total > 0:

                            self.price_progress.setValue(int(cur * 100 / total))

                    except Exception:

                        pass

                elif line.startswith("STATUS "):

                    self.price_status_label.setText(line[7:])

            self.log_box.appendPlainText(data.strip())



    def _on_price_upload_stderr(self) -> None:

        data = bytes(self._price_upload_proc.readAllStandardError()).decode("utf-8", errors="ignore")

        if data.strip():

            self.error_box.appendPlainText(data.strip())



    def _on_price_upload_finished(self) -> None:

        self.price_status_label.setText("업로드 완료")

        self.price_progress.setValue(100)



    def _build_misc_group(self) -> QtWidgets.QGroupBox:

        box = QtWidgets.QGroupBox("기타")

        v = QtWidgets.QVBoxLayout(box)

        row = QtWidgets.QHBoxLayout()



        self.write_r = QtWidgets.QCheckBox("R열(검색어설정) 작성")

        self.write_r.setChecked(True)



        self.threads = QtWidgets.QSpinBox()

        self.threads.setRange(1, 16)

        self.threads.setValue(6)



        self.max_depth = QtWidgets.QSpinBox()

        self.max_depth.setRange(-1, 50)

        self.max_depth.setValue(-1)



        self.allow_folder_match = QtWidgets.QCheckBox("폴더 경로에 GS코드가 있어도 매칭 허용")

        self.allow_folder_match.setChecked(True)



        row.addWidget(self.write_r)

        row.addWidget(self.debug)

        row.addStretch(1)

        row.addWidget(QtWidgets.QLabel("OCR 스레드"))

        row.addWidget(self.threads)

        row.addSpacing(8)

        row.addWidget(QtWidgets.QLabel("재귀 깊이"))

        row.addWidget(self.max_depth)

        row.addSpacing(8)

        row.addWidget(self.allow_folder_match)



        v.addLayout(row)



        recent_tabs = QtWidgets.QTabWidget()

        self.recent_file_list = QtWidgets.QListWidget()

        self.recent_file_list.setMaximumHeight(90)

        self.recent_file_list.itemDoubleClicked.connect(self._apply_recent_file_item)

        self.recent_dir_list = QtWidgets.QListWidget()

        self.recent_dir_list.setMaximumHeight(90)

        self.recent_dir_list.itemDoubleClicked.connect(self._apply_recent_dir_item)

        recent_tabs.addTab(self.recent_file_list, "최근 파일")

        recent_tabs.addTab(self.recent_dir_list, "최근 폴더")

        v.addWidget(recent_tabs)



        log_row = QtWidgets.QHBoxLayout()

        self.auto_log = QtWidgets.QCheckBox("로그 자동 저장")

        self.auto_log.setChecked(False)

        self.log_path_edit = QtWidgets.QLineEdit()

        self.log_path_edit.setPlaceholderText("로그 파일 경로 (예: C:\\code\\logs\\run.log)")

        self.log_path_btn = QtWidgets.QPushButton("로그 경로")

        self.log_path_btn.clicked.connect(self._browse_log_path)

        log_row.addWidget(self.auto_log)

        log_row.addWidget(self.log_path_edit, 1)

        log_row.addWidget(self.log_path_btn)

        v.addLayout(log_row)



        return box



    def _step_input(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)

        v.addWidget(self._build_input_group())

        v.addWidget(self._build_misc_group())

        v.addStretch(1)

        return w



    def _step_keyword(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)

        v.addWidget(self._build_keyword_group())

        v.addStretch(1)

        return w



    def _step_ocr(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)

        v.addWidget(self._build_ocr_group())

        v.addStretch(1)

        return w



    def _step_naver(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)

        v.addWidget(self._build_naver_group())

        v.addStretch(1)

        return w



    def _step_listing(self) -> QtWidgets.QWidget:

        w = QtWidgets.QWidget()

        v = QtWidgets.QVBoxLayout(w)

        v.addWidget(self._build_listing_group())

        v.addStretch(1)

        return w



    def _update_nav_buttons(self) -> None:

        idx = self.steps.currentIndex()

        total = self.steps.count()

        self.prev_btn.setEnabled(idx > 0)

        self.next_btn.setEnabled(idx < total - 1)

        titles = ["입력", "키워드", "OCR", "네이버", "대표이미지"]

        name = titles[idx] if idx < len(titles) else ""

        self.step_label.setText(f"Step {idx + 1}/{total}: {name}")



    def _prev_step(self) -> None:

        idx = self.steps.currentIndex()

        if idx > 0:

            self.steps.setCurrentIndex(idx - 1)

            self._update_nav_buttons()



    def _next_step(self) -> None:

        idx = self.steps.currentIndex()

        if idx < self.steps.count() - 1:

            self.steps.setCurrentIndex(idx + 1)

            self._update_nav_buttons()



    def _browse_file(self) -> None:

        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "CSV/Excel 선택", "", "CSV/Excel (*.csv *.xlsx *.xls)"

        )

        if path:

            self.file_edit.setText(path)

            self._save_settings()

            self._try_auto_match_ocr(path)



    def _try_auto_match_ocr(self, csv_path: str) -> None:

        """CSV 파일 선택 시 OCR 결과 Excel 자동 탐색."""

        try:

            from app.services.ocr_excel import find_matching_ocr_file

            found = find_matching_ocr_file(csv_path)

            if found:

                self.ocr_excel_edit.setText(found)

                self.ocr_status_label.setText(f"자동 매칭: {os.path.basename(found)}")

                self.ocr_status_label.setStyleSheet("color:#2e7d32; font-weight:bold;")

        except Exception:

            pass



    def _browse_dir(self) -> None:

        path = QtWidgets.QFileDialog.getExistingDirectory(self, "로컬 이미지 폴더 선택")

        if path:

            self.local_edit.setText(path)

            self._push_recent_dir(path)

            self._save_settings()



    def _browse_integrated_file(self) -> None:

        """통합 탭 CSV 파일 선택"""

        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "CSV/Excel 선택", "", "CSV/Excel (*.csv *.xlsx *.xls)"

        )

        if path:

            self.integrated_file_edit.setText(path)

            self._push_recent_file(path)

            self._save_settings()



    def _browse_integrated_dir(self) -> None:

        """통합 탭 로컬 이미지 폴더 선택"""

        path = QtWidgets.QFileDialog.getExistingDirectory(self, "로컬 이미지 폴더 선택 (선택사항)")

        if path:

            self.integrated_local_edit.setText(path)

            self._push_recent_dir(path)

            self._save_settings()



    def _browse_logo(self) -> None:

        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "로고 이미지 선택", "", "Images (*.png *.jpg *.jpeg *.webp *.bmp)"

        )

        if path:

            self.logo_edit.setText(path)

            self._save_settings()  # 로고 경로 저장



    def _browse_token(self) -> None:

        key_dir = self._desktop_key_path("")

        path, _ = QtWidgets.QFileDialog.getOpenFileName(

            self, "Cafe24 토큰 파일 선택", key_dir, "Token Files (*.json *.txt)"

        )

        if path:

            self.cafe24_token_path.setText(path)



    def _browse_export_dir(self) -> None:

        path = QtWidgets.QFileDialog.getExistingDirectory(self, "exports 경로 선택")

        if path:

            self.cafe24_export_dir.setText(path)



    def _browse_image_root(self) -> None:

        path = QtWidgets.QFileDialog.getExistingDirectory(self, "이미지 루트 선택")

        if path:

            self.cafe24_image_root.setText(path)



    def _browse_log_output(self) -> None:

        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "로그 저장 위치", "", "Excel (*.xlsx)")

        if path:

            self.cafe24_log_path.setText(path)



    def _app_root_path(self, name: str) -> str:

        here = os.path.abspath(os.path.dirname(__file__))

        root = os.path.abspath(os.path.join(here, "..", ".."))

        return os.path.join(root, name)


    def _desktop_key_path(self, name: str) -> str:

        return os.path.join(os.path.expanduser("~"), "Desktop", "key", name)



    def _auto_tesseract(self) -> None:

        path = core.setup_tesseract()

        if path:

            self.tess_edit.setText(path)

            self._log(f"Tesseract 자동탐색 성공: {path}")

        else:

            QtWidgets.QMessageBox.warning(self, "Tesseract", "자동탐색 실패. 경로를 직접 입력해 주세요.")



    def _log(self, msg: str) -> None:

        self.log_box.appendPlainText(msg)

        self.status_label.setText(msg)

        self._update_progress(msg)

        self._append_log_file(msg)



    def _log_debug(self, msg: str) -> None:

        self.debug_box.appendPlainText(msg)

        self._append_log_file(f"[DEBUG] {msg}")



    def _log_error(self, msg: str) -> None:

        self.error_box.appendPlainText(msg)

        self._append_log_file(f"[ERROR] {msg}")



    def _update_model_desc(self, model_name: str) -> None:

        kw_model = self.model_keyword_combo.currentText()

        lt_model = self.model_longtail_combo.currentText()

        kw_desc = self._model_desc_map.get(kw_model, "")

        lt_desc = self._model_desc_map.get(lt_model, "")

        self.model_desc.setText(f"키워드: {kw_desc}  |  롱테일: {lt_desc}")



    def _apply_model_tooltips(self) -> None:

        for combo in (self.model_keyword_combo, self.model_longtail_combo):

            for i in range(combo.count()):

                name = combo.itemText(i)

                tip = self._model_tip_map.get(name, "")

                if tip:

                    combo.setItemData(i, tip, QtCore.Qt.ToolTipRole)



    def _update_progress(self, msg: str) -> None:

        if "처리중... (1/2)" in msg:

            self.progress.setValue(35)

            self.stage_label.setText("단계: OCR/키워드")

            return

        if "네이버 키워드 조회" in msg:

            self.stage_label.setText("단계: 네이버 조회")

            return

        if "처리중... (2/2)" in msg:

            self.progress.setValue(75)

            self.stage_label.setText("단계: 대표이미지")

            return

        if "처리 완료" in msg or "완료" in msg:

            self.progress.setValue(100)

            self.stage_label.setText("단계: 완료")

            return

        if "오류" in msg:

            self.progress.setValue(0)

            self.stage_label.setText("단계: 오류")

            return

        if "작업 시작" in msg:

            self.stage_label.setText("단계: 시작")



    def _append_log_file(self, msg: str) -> None:

        if not self._log_file_handle:

            return

        try:

            self._log_file_handle.write(msg + "\n")

            self._log_file_handle.flush()

        except Exception:

            pass



    def _open_log_file(self) -> None:

        if not self.auto_log.isChecked():

            return

        path = self.log_path_edit.text().strip()

        if not path:

            base = os.path.join("C:\\code", "logs")

            ts = QtCore.QDateTime.currentDateTime().toString("yyyyMMdd_HHmmss")

            path = os.path.join(base, f"run_{ts}.log")

            self.log_path_edit.setText(path)

        try:

            os.makedirs(os.path.dirname(path), exist_ok=True)

            self._log_file_handle = open(path, "a", encoding="utf-8")

        except Exception:

            self._log_file_handle = None



    def _close_log_file(self) -> None:

        try:

            if self._log_file_handle:

                self._log_file_handle.close()

        finally:

            self._log_file_handle = None



    def _select_a_products_for_run(self, source_path: str, run_label: str) -> str | None:

        """실행 대상 상품을 선택하고, GS+7자리+A 행만 임시 파일로 반환."""

        try:

            df = core.safe_read_csv(source_path)

        except Exception as e:

            QtWidgets.QMessageBox.warning(self, "파일 오류", f"입력 파일을 읽을 수 없습니다.\n{e}")

            return None



        if df is None or df.empty:

            QtWidgets.QMessageBox.warning(self, "데이터 없음", "입력 파일에 데이터가 없습니다.")

            return None



        name_col = "상품명" if "상품명" in df.columns else None

        if not name_col:

            for c in df.columns:

                cs = str(c).strip()

                if "상품명" in cs or "name" in cs.lower():

                    name_col = c

                    break



        code_col = None

        for c in df.columns:

            if str(c).strip() in ["자체상품코드", "자체 상품코드", "상품코드B", "코드", "코드B"]:

                code_col = c

                break



        if not name_col and not code_col:

            QtWidgets.QMessageBox.warning(self, "컬럼 없음", "상품명/코드 컬럼을 찾을 수 없습니다.")

            return None



        name_s = df[name_col].astype(str) if name_col else None

        code_s = df[code_col].astype(str) if code_col else None



        mask_name = name_s.str.contains(r"GS\d{7}A\b", na=False, regex=True) if name_s is not None else False

        mask_code = code_s.str.contains(r"GS\d{7}A$", na=False, regex=True) if code_s is not None else False

        mask = (mask_name | mask_code)



        candidates = df.loc[mask].copy()

        if candidates.empty:

            QtWidgets.QMessageBox.information(

                self,

                "대상 없음",

                "GS + 7자리 + A로 끝나는 상품이 없습니다."

            )

            return None



        dialog = QtWidgets.QDialog(self)

        dialog.setWindowTitle(f"{run_label} - 상품 선택 (A코드만)")

        dialog.resize(860, 620)



        layout = QtWidgets.QVBoxLayout(dialog)

        info = QtWidgets.QLabel(

            f"조회 {len(candidates):,}개 (GSxxxxxxxA). 실행할 상품만 체크하세요."

        )

        layout.addWidget(info)



        action_row = QtWidgets.QHBoxLayout()

        sel_all_btn = QtWidgets.QPushButton("전체 선택")

        unsel_all_btn = QtWidgets.QPushButton("전체 해제")

        action_row.addWidget(sel_all_btn)

        action_row.addWidget(unsel_all_btn)

        action_row.addStretch(1)

        layout.addLayout(action_row)



        table = QtWidgets.QTableWidget(len(candidates), 3)

        table.setHorizontalHeaderLabels(["선택", "GS코드", "상품명"])

        table.setColumnWidth(0, 60)

        table.setColumnWidth(1, 120)

        table.horizontalHeader().setStretchLastSection(True)

        layout.addWidget(table, 1)



        candidate_indices = list(candidates.index)

        for row_i, src_idx in enumerate(candidate_indices):

            row = candidates.loc[src_idx]

            raw_name = str(row.get(name_col, "")) if name_col else ""

            raw_code = str(row.get(code_col, "")) if code_col else ""



            m = re.search(r"(GS\d{7}A)\b", raw_name)

            if not m:

                m = re.search(r"(GS\d{7}A)$", raw_code)

            gs_a = m.group(1) if m else ""



            chk = QtWidgets.QTableWidgetItem()

            chk.setFlags(chk.flags() | QtCore.Qt.ItemIsUserCheckable)

            chk.setCheckState(QtCore.Qt.Checked)

            chk.setData(QtCore.Qt.UserRole, int(src_idx))

            table.setItem(row_i, 0, chk)



            gs_item = QtWidgets.QTableWidgetItem(gs_a)

            gs_item.setFlags(gs_item.flags() & ~QtCore.Qt.ItemIsEditable)

            table.setItem(row_i, 1, gs_item)



            name_item = QtWidgets.QTableWidgetItem(raw_name)

            name_item.setFlags(name_item.flags() & ~QtCore.Qt.ItemIsEditable)

            table.setItem(row_i, 2, name_item)



        def _set_all(state: QtCore.Qt.CheckState) -> None:

            for r in range(table.rowCount()):

                item = table.item(r, 0)

                if item:

                    item.setCheckState(state)



        sel_all_btn.clicked.connect(lambda: _set_all(QtCore.Qt.Checked))

        unsel_all_btn.clicked.connect(lambda: _set_all(QtCore.Qt.Unchecked))



        buttons = QtWidgets.QDialogButtonBox(

            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel

        )

        ok_btn = buttons.button(QtWidgets.QDialogButtonBox.Ok)

        if ok_btn:

            ok_btn.setText("실행")

        buttons.accepted.connect(dialog.accept)

        buttons.rejected.connect(dialog.reject)

        layout.addWidget(buttons)



        if dialog.exec() != QtWidgets.QDialog.Accepted:

            return None



        selected_src_indices = []

        for r in range(table.rowCount()):

            item = table.item(r, 0)

            if item and item.checkState() == QtCore.Qt.Checked:

                src_idx = item.data(QtCore.Qt.UserRole)

                selected_src_indices.append(int(src_idx))



        if not selected_src_indices:

            QtWidgets.QMessageBox.warning(self, "선택 필요", "최소 1개 상품을 선택해 주세요.")

            return None



        selected_df = df.loc[selected_src_indices].copy()



        ts = QtCore.QDateTime.currentDateTime().toString("yyyyMMdd_HHmmss")

        selected_path = self._app_root_path(f"_selected_input_A_{ts}.xlsx")

        selected_df.to_excel(selected_path, index=False)



        self._log(

            f"선택 실행: {len(selected_df):,}개 / A코드 후보 {len(candidates):,}개"

        )

        return selected_path



    def _on_run(self) -> None:

        if not self.file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV/Excel 파일을 선택해 주세요.")

            return



        selected_input_path = self._select_a_products_for_run(

            self.file_edit.text().strip(),

            "전처리 + OCR + GPT + (대표이미지)",

        )

        if not selected_input_path:

            return



        cfg = PipelineConfig(

            file_path=selected_input_path,

            img_tag=self.img_tag_edit.text().strip(),

            tesseract_path=self.tess_edit.text().strip(),

            model_keyword=self._normalize_runtime_model_name(self.model_keyword_combo.currentText(), self._default_keyword_model),

            model_longtail=self._normalize_runtime_model_name(self.model_longtail_combo.currentText(), self._default_longtail_model),

            keyword_version=self.keyword_version_combo.currentText().strip(),

            max_words=self.max_words.value(),

            max_len=self.max_len.value(),

            min_len=self.min_len.value(),

            use_html_ocr=False,

            use_local_ocr=True,

            merge_ocr_with_name=self.merge_ocr.isChecked(),

            max_imgs=999,

            threads=self.threads.value(),

            max_depth=self.max_depth.value(),

            local_img_dir=self.local_edit.text().strip(),

            allow_folder_match=self.allow_folder_match.isChecked(),

            korean_only=False,

            drop_digits=self.drop_digits.isChecked(),

            psm=11,

            oem=3,

            ocr_excel_path="",

            write_to_r=self.write_r.isChecked(),

            debug=self.debug.isChecked(),

            naver_enabled=self.naver_enable.isChecked(),

            naver_dry_run=self.naver_dry.isChecked(),

            naver_retry=self.naver_retry.isChecked(),

            naver_retry_count=self.naver_retry_count.value(),

            naver_retry_delay=self.naver_retry_delay.value(),

            naver_autocomplete=self.naver_autocomplete.isChecked(),

            google_autocomplete=self.google_autocomplete.isChecked(),

            make_listing=self.make_listing.isChecked(),

            listing_size=self.listing_size.value(),

            listing_pad=self.listing_pad.value(),

            listing_max=self.listing_max.value(),

            logo_path=self.logo_edit.text().strip(),

            logo_ratio=self.logo_ratio.value(),

            logo_opacity=self.logo_opacity.value(),

            logo_pos=self.logo_pos.currentText(),

            use_auto_contrast=self.auto_contrast.isChecked(),

            use_sharpen=self.sharpen.isChecked(),

            use_small_rotate=self.rotate.isChecked(),

            rotate_zoom=self.rotate_zoom.value(),

            ultra_angle_deg=self.ultra_angle.value(),

            ultra_translate_px=self.ultra_translate.value(),

            ultra_scale_pct=self.ultra_scale.value(),

            trim_tol=self.trim_tol.value(),

            jpeg_q_min=self.jpeg_q_min.value(),

            jpeg_q_max=self.jpeg_q_max.value(),

            do_flip_lr=self.flip_lr.isChecked(),

        )



        self.run_btn.setEnabled(False)

        self.listing_only_btn.setEnabled(False)

        self._open_log_file()

        self.progress.setValue(5)

        self._log("작업 시작")

        self._push_recent(self.file_edit.text().strip())

        if self.local_edit.text().strip():

            self._push_recent_dir(self.local_edit.text().strip())



        self._thread = QtCore.QThread(self)

        self._worker = PipelineWorker(cfg)

        self._worker.moveToThread(self._thread)



        self._thread.started.connect(self._worker.run)

        self._worker.status.connect(self._log)

        self._worker.progress.connect(self.progress.setValue)

        self._worker.finished.connect(self._on_finished)

        self._worker.error.connect(self._on_error)



        self._worker.finished.connect(self._thread.quit)

        self._worker.finished.connect(self._worker.deleteLater)

        self._thread.finished.connect(self._thread.deleteLater)

        self._worker.error.connect(self._thread.quit)

        self._worker.error.connect(self._worker.deleteLater)



        self._thread.start()



    def _on_run_excel_only(self) -> None:

        """엑셀만 생성 (대표이미지 생성 스킵)"""

        if not self.file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV/Excel 파일을 선택해 주세요.")

            return



        selected_input_path = self._select_a_products_for_run(

            self.file_edit.text().strip(),

            "전처리 + OCR + GPT (엑셀만 생성)",

        )

        if not selected_input_path:

            return



        cfg = PipelineConfig(

            file_path=selected_input_path,

            img_tag=self.img_tag_edit.text().strip(),

            tesseract_path=self.tess_edit.text().strip(),

            model_keyword=self._normalize_runtime_model_name(self.model_keyword_combo.currentText(), self._default_keyword_model),

            model_longtail=self._normalize_runtime_model_name(self.model_longtail_combo.currentText(), self._default_longtail_model),

            keyword_version=self.keyword_version_combo.currentText().strip(),

            max_words=self.max_words.value(),

            max_len=self.max_len.value(),

            min_len=self.min_len.value(),

            use_html_ocr=False,

            use_local_ocr=True,

            merge_ocr_with_name=self.merge_ocr.isChecked(),

            max_imgs=999,

            threads=self.threads.value(),

            max_depth=self.max_depth.value(),

            local_img_dir=self.local_edit.text().strip(),

            allow_folder_match=self.allow_folder_match.isChecked(),

            korean_only=False,

            drop_digits=self.drop_digits.isChecked(),

            psm=11,

            oem=3,

            ocr_excel_path="",

            write_to_r=self.write_r.isChecked(),

            debug=self.debug.isChecked(),

            naver_enabled=self.naver_enable.isChecked(),

            naver_dry_run=self.naver_dry.isChecked(),

            naver_retry=self.naver_retry.isChecked(),

            naver_retry_count=self.naver_retry_count.value(),

            naver_retry_delay=self.naver_retry_delay.value(),

            naver_autocomplete=self.naver_autocomplete.isChecked(),

            google_autocomplete=self.google_autocomplete.isChecked(),

            make_listing=False,  # ← 이미지 생성 비활성화

            listing_size=self.listing_size.value(),

            listing_pad=self.listing_pad.value(),

            listing_max=self.listing_max.value(),

            logo_path=self.logo_edit.text().strip(),

            logo_ratio=self.logo_ratio.value(),

            logo_opacity=self.logo_opacity.value(),

            logo_pos=self.logo_pos.currentText(),

            use_auto_contrast=self.auto_contrast.isChecked(),

            use_sharpen=self.sharpen.isChecked(),

            use_small_rotate=self.rotate.isChecked(),

            rotate_zoom=self.rotate_zoom.value(),

            ultra_angle_deg=self.ultra_angle.value(),

            ultra_translate_px=self.ultra_translate.value(),

            ultra_scale_pct=self.ultra_scale.value(),

            trim_tol=self.trim_tol.value(),

            jpeg_q_min=self.jpeg_q_min.value(),

            jpeg_q_max=self.jpeg_q_max.value(),

            do_flip_lr=self.flip_lr.isChecked(),

        )



        self.run_btn.setEnabled(False)

        self.run_excel_only_btn.setEnabled(False)

        self.listing_only_btn.setEnabled(False)

        self._open_log_file()

        self.progress.setValue(5)

        self._log("엑셀만 생성 시작")

        self._push_recent(self.file_edit.text().strip())

        if self.local_edit.text().strip():

            self._push_recent_dir(self.local_edit.text().strip())



        self._thread = QtCore.QThread(self)

        self._worker = PipelineWorker(cfg)

        self._worker.moveToThread(self._thread)



        self._thread.started.connect(self._worker.run)

        self._worker.status.connect(self._log)

        self._worker.progress.connect(self.progress.setValue)

        self._worker.finished.connect(self._on_finished)

        self._worker.error.connect(self._on_error)



        self._worker.finished.connect(self._thread.quit)

        self._worker.finished.connect(self._worker.deleteLater)

        self._thread.finished.connect(self._thread.deleteLater)

        self._worker.error.connect(self._thread.quit)

        self._worker.error.connect(self._worker.deleteLater)



        self._thread.start()



    def _on_run_keyword_from_ocr_only(self) -> None:

        """OCR 결과 Excel만 사용하여 키워드/엑셀만 생성 (OCR 재실행 없음)."""

        if not self.file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV/Excel 파일을 선택해 주세요.")

            return

        if not self.ocr_excel_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "OCR 결과 Excel 파일을 선택해 주세요.")

            return



        cfg = PipelineConfig(

            file_path=self.file_edit.text().strip(),

            img_tag=self.img_tag_edit.text().strip(),

            tesseract_path=self.tess_edit.text().strip(),

            model_keyword=self._normalize_runtime_model_name(self.model_keyword_combo.currentText(), self._default_keyword_model),

            model_longtail=self._normalize_runtime_model_name(self.model_longtail_combo.currentText(), self._default_longtail_model),

            keyword_version=self.keyword_version_combo.currentText().strip(),

            max_words=self.max_words.value(),

            max_len=self.max_len.value(),

            min_len=self.min_len.value(),

            use_html_ocr=False,

            use_local_ocr=False,  # OCR 재실행 금지

            merge_ocr_with_name=True,

            max_imgs=0,

            threads=self.threads.value(),

            max_depth=self.max_depth.value(),

            local_img_dir="",

            allow_folder_match=self.allow_folder_match.isChecked(),

            korean_only=True,

            drop_digits=self.drop_digits.isChecked(),

            psm=11,

            oem=3,

            ocr_excel_path=self.ocr_excel_edit.text().strip(),

            write_to_r=self.write_r.isChecked(),

            debug=self.debug.isChecked(),

            naver_enabled=False,  # OCR 기반 추론 전용: 네이버 검색광고 미사용

            naver_dry_run=self.naver_dry.isChecked(),

            naver_retry=self.naver_retry.isChecked(),

            naver_retry_count=self.naver_retry_count.value(),

            naver_retry_delay=self.naver_retry_delay.value(),

            naver_autocomplete=self.naver_autocomplete.isChecked(),

            google_autocomplete=self.google_autocomplete.isChecked(),

            make_listing=False,

            listing_size=self.listing_size.value(),

            listing_pad=self.listing_pad.value(),

            listing_max=self.listing_max.value(),

            logo_path=self.logo_edit.text().strip(),

            logo_ratio=self.logo_ratio.value(),

            logo_opacity=self.logo_opacity.value(),

            logo_pos=self.logo_pos.currentText(),

            use_auto_contrast=self.auto_contrast.isChecked(),

            use_sharpen=self.sharpen.isChecked(),

            use_small_rotate=self.rotate.isChecked(),

            rotate_zoom=self.rotate_zoom.value(),

            ultra_angle_deg=self.ultra_angle.value(),

            ultra_translate_px=self.ultra_translate.value(),

            ultra_scale_pct=self.ultra_scale.value(),

            trim_tol=self.trim_tol.value(),

            jpeg_q_min=self.jpeg_q_min.value(),

            jpeg_q_max=self.jpeg_q_max.value(),

            do_flip_lr=self.flip_lr.isChecked(),

        )



        self.run_btn.setEnabled(False)

        self.run_excel_only_btn.setEnabled(False)

        self.run_keyword_from_ocr_btn.setEnabled(False)

        self.listing_only_btn.setEnabled(False)

        self._open_log_file()

        self.progress.setValue(5)

        self._log("OCR결과 사용 키워드 생성 시작")

        self._push_recent(self.file_edit.text().strip())



        self._thread = QtCore.QThread(self)

        self._worker = PipelineWorker(cfg)

        self._worker.moveToThread(self._thread)



        self._thread.started.connect(self._worker.run)

        self._worker.status.connect(self._log)

        self._worker.progress.connect(self.progress.setValue)

        self._worker.finished.connect(self._on_finished)

        self._worker.error.connect(self._on_error)



        self._worker.finished.connect(self._thread.quit)

        self._worker.finished.connect(self._worker.deleteLater)

        self._thread.finished.connect(self._thread.deleteLater)

        self._worker.error.connect(self._thread.quit)

        self._worker.error.connect(self._worker.deleteLater)



        self._thread.start()



    def _on_finished(self, out_root: str, out_file: str) -> None:

        self._last_output_file = out_file

        self.run_btn.setEnabled(True)

        self.run_excel_only_btn.setEnabled(True)

        self.run_keyword_from_ocr_btn.setEnabled(True)

        self.listing_only_btn.setEnabled(True)

        self._log(f"완료: {out_file}")

        self._log_debug(f"결과 폴더: {out_root}")

        self._close_log_file()

        self._save_settings()

        QtWidgets.QMessageBox.information(self, "완료", f"결과 폴더:\n{out_root}")

        try:

            os.startfile(out_root)

        except Exception:

            pass



    def _on_error(self, err: str) -> None:

        self.run_btn.setEnabled(True)

        self.run_excel_only_btn.setEnabled(True)

        self.run_keyword_from_ocr_btn.setEnabled(True)

        self.listing_only_btn.setEnabled(True)

        self._log("오류 발생")

        self._log_error(err)

        self._close_log_file()

        self._save_settings()

        QtWidgets.QMessageBox.critical(self, "오류", "처리 중 오류가 발생했습니다.\n로그를 확인하세요.")



    # ── 대표이미지만 생성 ──────────────────────────────────────────────



    def _on_run_listing_only(self) -> None:

        if not self.file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV/Excel 파일을 선택해 주세요.")

            return

        if not self.local_edit.text().strip() and not self.ocr_excel_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "이미지 폴더 또는 OCR 결과 Excel이 필요합니다.")

            return



        cfg = ListingOnlyConfig(

            file_path=self.file_edit.text().strip(),

            local_img_dir=self.local_edit.text().strip(),

            allow_folder_match=self.allow_folder_match.isChecked(),

            max_depth=self.max_depth.value(),

            ocr_excel_path=self.ocr_excel_edit.text().strip(),

            listing_size=self.listing_size.value(),

            listing_pad=self.listing_pad.value(),

            listing_max=self.listing_max.value(),

            logo_path=self.logo_edit.text().strip(),

            logo_ratio=self.logo_ratio.value(),

            logo_opacity=self.logo_opacity.value(),

            logo_pos=self.logo_pos.currentText(),

            use_auto_contrast=self.auto_contrast.isChecked(),

            use_sharpen=self.sharpen.isChecked(),

            use_small_rotate=self.rotate.isChecked(),

            rotate_zoom=self.rotate_zoom.value(),

            ultra_angle_deg=self.ultra_angle.value(),

            ultra_translate_px=self.ultra_translate.value(),

            ultra_scale_pct=self.ultra_scale.value(),

            trim_tol=self.trim_tol.value(),

            jpeg_q_min=self.jpeg_q_min.value(),

            jpeg_q_max=self.jpeg_q_max.value(),

            do_flip_lr=self.flip_lr.isChecked(),

        )



        self.run_btn.setEnabled(False)

        self.listing_only_btn.setEnabled(False)

        self._open_log_file()

        self.progress.setValue(5)

        self._log("대표이미지만 생성 시작")



        self._thread = QtCore.QThread(self)

        self._worker = ListingWorker(cfg)

        self._worker.moveToThread(self._thread)



        self._thread.started.connect(self._worker.run)

        self._worker.status.connect(self._log)

        self._worker.progress.connect(self.progress.setValue)

        self._worker.finished_listing.connect(self._on_listing_finished)

        self._worker.error.connect(self._on_listing_error)



        self._worker.finished_listing.connect(self._thread.quit)

        self._worker.finished_listing.connect(self._worker.deleteLater)

        self._thread.finished.connect(self._thread.deleteLater)

        self._worker.error.connect(self._thread.quit)

        self._worker.error.connect(self._worker.deleteLater)



        self._thread.start()



    def _on_listing_finished(self, out_root: str) -> None:

        self.run_btn.setEnabled(True)

        self.listing_only_btn.setEnabled(True)

        self._log(f"대표이미지 생성 완료: {out_root}")

        self._close_log_file()

        self._save_settings()

        if out_root:

            QtWidgets.QMessageBox.information(self, "완료", f"대표이미지 생성 완료:\n{out_root}")

            try:

                os.startfile(out_root)

            except Exception:

                pass

        else:

            QtWidgets.QMessageBox.warning(self, "완료", "이미지 소스가 없어 생성된 이미지가 없습니다.")



    def _on_listing_error(self, err: str) -> None:

        self.run_btn.setEnabled(True)

        self.listing_only_btn.setEnabled(True)

        self._log("대표이미지 생성 오류")

        self._log_error(err)

        self._close_log_file()

        self._save_settings()

        QtWidgets.QMessageBox.critical(self, "오류", "대표이미지 생성 중 오류가 발생했습니다.\n로그를 확인하세요.")



    def _run_cafe24_upload(self) -> None:

        # write config file for cafe24_upload.py

        cfg_path = self._app_root_path("cafe24_upload_config.txt")

        date_tag = self.cafe24_date_tag.text().strip()

        main_idx = self.cafe24_main_idx.value()

        add_start = self.cafe24_add_start.value()

        add_max = self.cafe24_add_max.value()

        export_dir = self.cafe24_export_dir.text().strip()

        image_root = self.cafe24_image_root.text().strip()

        retry_count = self.cafe24_retry.value()

        retry_delay = self.cafe24_retry_delay.value()

        log_path = self.cafe24_log_path.text().strip()

        match_mode = self.cafe24_match_mode.currentText().strip()

        match_prefix = self.cafe24_match_prefix.value()



        with open(cfg_path, "w", encoding="utf-8") as f:

            if date_tag:

                f.write(f"DATE_TAG={date_tag}\n")

            f.write(f"MAIN_INDEX={main_idx}\n")

            f.write(f"ADD_START={add_start}\n")

            f.write(f"ADD_MAX={add_max}\n")

            if export_dir:

                f.write(f"EXPORT_DIR={export_dir}\n")

            if image_root:

                f.write(f"IMAGE_ROOT={image_root}\n")

            f.write(f"RETRY_COUNT={retry_count}\n")

            f.write(f"RETRY_DELAY={retry_delay}\n")

            if log_path:

                f.write(f"LOG_PATH={log_path}\n")

            f.write(f"MATCH_MODE={match_mode}\n")

            f.write(f"MATCH_PREFIX={match_prefix}\n")



        # ensure token file is in Desktop/key (copy if needed)

        token_src = self.cafe24_token_path.text().strip()

        token_dst = self._desktop_key_path("cafe24_token.txt")

        if token_src and os.path.isfile(token_src) and os.path.abspath(token_src) != os.path.abspath(token_dst):

            try:

                os.makedirs(os.path.dirname(token_dst), exist_ok=True)

                with open(token_src, "r", encoding="utf-8") as rf:

                    data = rf.read()

                with open(token_dst, "w", encoding="utf-8") as wf:

                    wf.write(data)

            except Exception:

                pass



        self.cafe24_status.setText("업로드 실행 중...")

        self.cafe24_progress.setValue(0)



        self._upload_proc = QtCore.QProcess(self)

        self._upload_proc.setProgram(sys.executable)

        self._upload_proc.setArguments([self._app_root_path("cafe24_upload.py")])

        self._upload_proc.setWorkingDirectory(self._app_root_path(""))

        self._upload_proc.readyReadStandardOutput.connect(self._on_upload_stdout)

        self._upload_proc.readyReadStandardError.connect(self._on_upload_stderr)

        self._upload_proc.finished.connect(self._on_upload_finished)

        self._upload_proc.start()



    def _on_upload_stdout(self) -> None:

        data = bytes(self._upload_proc.readAllStandardOutput()).decode("utf-8", errors="ignore")

        if data.strip():

            for line in data.splitlines():

                if line.startswith("PROGRESS "):

                    try:

                        cur, total = line.replace("PROGRESS ", "").split("/")

                        cur = int(cur.strip()); total = int(total.strip())

                        if total > 0:

                            self.cafe24_progress.setValue(int(cur * 100 / total))

                    except Exception:

                        pass

                elif line.startswith("STATUS "):

                    self.cafe24_status.setText(line[7:])

            self.log_box.appendPlainText(data.strip())

            for line in data.splitlines():

                if line.startswith("[LOG]"):

                    parts = line.split(" ", 1)

                    if len(parts) == 2:

                        self._last_upload_log = parts[1].strip()



    def _on_upload_stderr(self) -> None:

        data = bytes(self._upload_proc.readAllStandardError()).decode("utf-8", errors="ignore")

        if data.strip():

            self.error_box.appendPlainText(data.strip())



    def _on_upload_finished(self) -> None:

        self.cafe24_status.setText("업로드 완료")

        self.cafe24_progress.setValue(100)

        self._load_upload_log()



    def _open_upload_log(self) -> None:

        # open latest log if path not provided

        if self.cafe24_log_path.text().strip():

            path = self.cafe24_log_path.text().strip()

        else:

            base = self.cafe24_export_dir.text().strip() or "C:\\code\\exports"

            if not os.path.isdir(base):

                return

            dirs = [d for d in os.listdir(base) if os.path.isdir(os.path.join(base, d))]

            dirs.sort(reverse=True)

            if not dirs:

                return

            latest = os.path.join(base, dirs[0])

            logs = [f for f in os.listdir(latest) if f.startswith("cafe24_upload_log_") and f.endswith(".xlsx")]

            logs.sort(reverse=True)

            if not logs:

                return

            path = os.path.join(latest, logs[0])

        try:

            os.startfile(path)

        except Exception:

            pass



    def _load_upload_log(self) -> None:

        path = getattr(self, "_last_upload_log", "")

        if not path:

            return

        if not os.path.isfile(path):

            return

        try:

            import pandas as pd

            df = pd.read_excel(path)

        except Exception:

            return

        self._last_log_df = df

        self._render_log_table(df)



    def _render_log_table(self, df) -> None:

        self.cafe24_table.setRowCount(0)

        cols = ["GS", "PRODUCT_NO", "STATUS", "MAIN", "ERROR"]

        for _, row in df.iterrows():

            r = self.cafe24_table.rowCount()

            self.cafe24_table.insertRow(r)

            for c, name in enumerate(cols):

                val = row.get(name, "")

                item = QtWidgets.QTableWidgetItem(str(val))

                if name == "STATUS":

                    if str(val) == "OK":

                        item.setBackground(QtGui.QColor("#e8f5e9"))

                    else:

                        item.setBackground(QtGui.QColor("#ffebee"))

                self.cafe24_table.setItem(r, c, item)

        # summary

        if "STATUS" in df.columns:

            ok = int((df["STATUS"] == "OK").sum())

            miss = int((df["STATUS"] != "OK").sum())

            miss_types = df[df["STATUS"] != "OK"]["STATUS"].value_counts().to_dict()

            self.cafe24_summary.setText(f"미스 사유 요약: OK {ok} / MISS {miss} / {miss_types}")



    def _filter_misses(self) -> None:

        df = getattr(self, "_last_log_df", None)

        if df is None:

            return

        miss = df[df["STATUS"] != "OK"]

        self._render_log_table(miss)



    def _retry_misses(self) -> None:

        df = getattr(self, "_last_log_df", None)

        if df is None:

            return

        miss = df[df["STATUS"] != "OK"]

        if miss.empty:

            return

        gs_list_path = self._app_root_path("cafe24_retry_gs.txt")

        with open(gs_list_path, "w", encoding="utf-8") as f:

            for gs in miss["GS"].dropna().unique().tolist():

                f.write(f"{gs}\n")

        cfg_path = self._app_root_path("cafe24_upload_config.txt")

        with open(cfg_path, "a", encoding="utf-8") as f:

            f.write(f"GS_LIST={gs_list_path}\n")

        self._run_cafe24_upload()



    def _open_gs_folder(self) -> None:

        item = self.cafe24_table.currentItem()

        if not item:

            return

        row = item.row()

        gs_item = self.cafe24_table.item(row, 0)

        if not gs_item:

            return

        gs = gs_item.text().strip()

        # try to open image folder

        base = self.cafe24_image_root.text().strip()

        if base and os.path.isdir(base):

            folder = os.path.join(base, gs)

        else:

            export_base = self.cafe24_export_dir.text().strip() or "C:\\code\\exports"

            if not os.path.isdir(export_base):

                return

            dirs = [d for d in os.listdir(export_base) if os.path.isdir(os.path.join(export_base, d))]

            dirs.sort(reverse=True)

            if not dirs:

                return

            latest = os.path.join(export_base, dirs[0], "listing_images")

            date_tag = self.cafe24_date_tag.text().strip()

            if date_tag and os.path.isdir(os.path.join(latest, date_tag)):

                folder = os.path.join(latest, date_tag, gs)

            else:

                folder = os.path.join(latest, gs)

        if os.path.isdir(folder):

            try:

                os.startfile(folder)

            except Exception:

                pass



    def _browse_log_path(self) -> None:

        path, _ = QtWidgets.QFileDialog.getSaveFileName(

            self, "로그 파일 저장 위치", "", "Text Files (*.log *.txt)"

        )

        if path:

            self.log_path_edit.setText(path)

            self._save_settings()



    def _normalize_runtime_model_name(self, model_name: str, fallback: str) -> str:

        model = str(model_name or "").strip()

        aliases = {

            "gpt-4o": "gpt-4.1",

            "gpt-4o-mini": "gpt-4.1-mini",

            "claude-sonnet-4": "claude-sonnet-4-6",

            "claude-sonnet-4.6": "claude-sonnet-4-6",

            "claude-opus-4.6": "claude-opus-4-6",

            "claude-haiku-4.5": "claude-haiku-4-5",

            "claude-haiku-4-5-20251001": "claude-haiku-4-5",

            "claude-haiku-4.5-20251001": "claude-haiku-4-5",

        }

        model = aliases.get(model, model)

        return model or fallback



    def _save_settings(self) -> None:

        self._settings.setValue("recent_file", self.file_edit.text().strip())

        self._settings.setValue("recent_local_dir", self.local_edit.text().strip())

        self._settings.setValue("log_path", self.log_path_edit.text().strip())

        self._settings.setValue("auto_log", self.auto_log.isChecked())

        self._settings.setValue("recent_files", self._recent_files)

        self._settings.setValue("recent_dirs", self._recent_dirs)

        self._settings.setValue("model_keyword", self._normalize_runtime_model_name(self.model_keyword_combo.currentText(), self._default_keyword_model))

        self._settings.setValue("model_longtail", self._normalize_runtime_model_name(self.model_longtail_combo.currentText(), self._default_longtail_model))

        self._settings.setValue("keyword_version", self.keyword_version_combo.currentText().strip())

        self._settings.setValue("cafe24_match_mode", self.cafe24_match_mode.currentText().strip())

        self._settings.setValue("cafe24_match_prefix", self.cafe24_match_prefix.value())



        # 로고 및 이미지 설정

        self._settings.setValue("logo_path", self.logo_edit.text().strip())

        self._settings.setValue("logo_position", self.logo_pos.currentText().strip())

        self._settings.setValue("logo_opacity", self.logo_opacity.value())

        self._settings.setValue("logo_ratio", self.logo_ratio.value())



        # 키워드 생성 옵션

        self._settings.setValue("merge_ocr", self.merge_ocr.isChecked())

        self._settings.setValue("drop_digits", self.drop_digits.isChecked())

        self._settings.setValue("max_words", self.max_words.value())

        self._settings.setValue("max_len", self.max_len.value())

        self._settings.setValue("min_len", self.min_len.value())



        # 네이버 API 설정

        self._settings.setValue("naver_retry", self.naver_retry.isChecked())

        self._settings.setValue("naver_retry_count", self.naver_retry_count.value())

        self._settings.setValue("naver_retry_delay", self.naver_retry_delay.value())



        # 대표이미지 생성 설정

        self._settings.setValue("make_listing", self.make_listing.isChecked())

        # 쿠팡 업로드 설정
        self._settings.setValue("coupang_file", self.coupang_file_edit.text().strip())
        self._settings.setValue("coupang_dry_run", self.coupang_dry_run.isChecked())



    def _load_settings(self) -> None:

        rf = self._settings.value("recent_file", "")

        rd = self._settings.value("recent_local_dir", "")

        lp = self._settings.value("log_path", "")

        al = self._settings.value("auto_log", False, type=bool)

        recent = self._settings.value("recent_files", [])

        recent_dirs = self._settings.value("recent_dirs", [])

        model_keyword = self._normalize_runtime_model_name(self._settings.value("model_keyword", ""), self._default_keyword_model)

        model_longtail = self._normalize_runtime_model_name(self._settings.value("model_longtail", ""), self._default_longtail_model)

        keyword_version = str(self._settings.value("keyword_version", "2.0") or "2.0").strip()

        cafe24_match_mode = self._settings.value("cafe24_match_mode", "PREFIX")

        cafe24_match_prefix = self._settings.value("cafe24_match_prefix", 20, type=int)



        # 로고 및 이미지 설정 불러오기

        logo_path = self._settings.value("logo_path", "")

        logo_position = self._settings.value("logo_position", "br")

        logo_opacity = self._settings.value("logo_opacity", 65, type=int)

        logo_ratio = self._settings.value("logo_ratio", 14, type=int)



        # 키워드 생성 옵션 불러오기

        merge_ocr = self._settings.value("merge_ocr", True, type=bool)

        drop_digits = self._settings.value("drop_digits", False, type=bool)

        max_words = self._settings.value("max_words", 22, type=int)

        max_len = self._settings.value("max_len", 120, type=int)

        min_len = self._settings.value("min_len", 40, type=int)



        # 네이버 API 설정 불러오기

        naver_retry = self._settings.value("naver_retry", False, type=bool)

        naver_retry_count = self._settings.value("naver_retry_count", 2, type=int)

        naver_retry_delay = self._settings.value("naver_retry_delay", 3.0, type=float)



        # 대표이미지 생성 설정 불러오기

        make_listing = self._settings.value("make_listing", True, type=bool)



        if rf and os.path.isfile(rf):

            self.file_edit.setText(rf)

            self._try_auto_match_ocr(rf)

        if rd and os.path.isdir(rd):

            self.local_edit.setText(rd)

        if lp:

            self.log_path_edit.setText(lp)

        self.auto_log.setChecked(bool(al))

        if model_keyword and self.model_keyword_combo.findText(str(model_keyword)) >= 0:

            self.model_keyword_combo.setCurrentText(str(model_keyword))

        else:

            self.model_keyword_combo.setCurrentText(self._default_keyword_model)

        if model_longtail and self.model_longtail_combo.findText(str(model_longtail)) >= 0:

            self.model_longtail_combo.setCurrentText(str(model_longtail))

        else:

            self.model_longtail_combo.setCurrentText(self._default_longtail_model)

        if keyword_version and self.keyword_version_combo.findText(str(keyword_version)) >= 0:

            self.keyword_version_combo.setCurrentText(str(keyword_version))

        else:

            self.keyword_version_combo.setCurrentText("2.0")

        self.cafe24_match_mode.setCurrentText(str(cafe24_match_mode))

        self.cafe24_match_prefix.setValue(int(cafe24_match_prefix))



        # 불러온 설정값 적용

        # 저장된 로고가 없으면 기본 로고 경로 사용

        default_logo = r"C:\Users\가호경\OneDrive\Desktop\홈런마켓\image.png"

        if logo_path and os.path.isfile(logo_path):

            self.logo_edit.setText(logo_path)

        elif os.path.isfile(default_logo):

            self.logo_edit.setText(default_logo)

            self._save_settings()  # 기본 로고 경로 저장



        if self.logo_pos.findText(str(logo_position)) >= 0:

            self.logo_pos.setCurrentText(str(logo_position))

        self.logo_opacity.setValue(int(logo_opacity))

        self.logo_ratio.setValue(int(logo_ratio))



        self.merge_ocr.setChecked(bool(merge_ocr))

        self.drop_digits.setChecked(bool(drop_digits))

        self.max_words.setValue(int(max_words))

        self.max_len.setValue(int(max_len))

        self.min_len.setValue(int(min_len))



        self.naver_retry.setChecked(bool(naver_retry))

        self.naver_retry_count.setValue(int(naver_retry_count))

        self.naver_retry_delay.setValue(float(naver_retry_delay))



        self.make_listing.setChecked(bool(make_listing))

        # 쿠팡 업로드 설정 복원
        coupang_file = self._settings.value("coupang_file", "")
        if coupang_file and os.path.isfile(coupang_file):
            self.coupang_file_edit.setText(coupang_file)
        coupang_dry = self._settings.value("coupang_dry_run", True, type=bool)
        self.coupang_dry_run.setChecked(coupang_dry)

        if isinstance(recent, list):

            self._recent_files = recent

            self._refresh_recent_combo()

        if isinstance(recent_dirs, list):

            self._recent_dirs = recent_dirs

            self._refresh_recent_combo()



    def _connect_setting_signals(self) -> None:

        """설정값 변경 시 자동 저장 연결"""

        # 콤보박스

        self.model_keyword_combo.currentTextChanged.connect(self._save_settings)

        self.model_longtail_combo.currentTextChanged.connect(self._save_settings)

        self.keyword_version_combo.currentTextChanged.connect(self._save_settings)

        self.logo_pos.currentTextChanged.connect(self._save_settings)

        self.cafe24_match_mode.currentTextChanged.connect(self._save_settings)



        # 스핀박스

        self.logo_opacity.valueChanged.connect(self._save_settings)

        self.logo_ratio.valueChanged.connect(self._save_settings)

        self.max_words.valueChanged.connect(self._save_settings)

        self.max_len.valueChanged.connect(self._save_settings)

        self.min_len.valueChanged.connect(self._save_settings)

        self.naver_retry_count.valueChanged.connect(self._save_settings)

        self.naver_retry_delay.valueChanged.connect(self._save_settings)

        self.cafe24_match_prefix.valueChanged.connect(self._save_settings)



        # 체크박스

        self.merge_ocr.stateChanged.connect(self._save_settings)

        self.drop_digits.stateChanged.connect(self._save_settings)

        self.naver_retry.stateChanged.connect(self._save_settings)

        self.make_listing.stateChanged.connect(self._save_settings)

        self.auto_log.stateChanged.connect(self._save_settings)



    def _refresh_recent_combo(self) -> None:

        self.recent_file_list.clear()

        for p in self._recent_files[:5]:

            self.recent_file_list.addItem(p)

        self.recent_dir_list.clear()

        for p in self._recent_dirs[:5]:

            self.recent_dir_list.addItem(p)



    def _push_recent(self, path: str) -> None:

        if not path:

            return

        if path in self._recent_files:

            self._recent_files.remove(path)

        self._recent_files.insert(0, path)

        self._recent_files = self._recent_files[:5]

        self._refresh_recent_combo()

        self._save_settings()



    def _push_recent_dir(self, path: str) -> None:

        if not path:

            return

        if path in self._recent_dirs:

            self._recent_dirs.remove(path)

        self._recent_dirs.insert(0, path)

        self._recent_dirs = self._recent_dirs[:5]

        self._refresh_recent_combo()

        self._save_settings()



    def _apply_recent_file_item(self, item: QtWidgets.QListWidgetItem) -> None:

        path = item.text()

        if path and os.path.isfile(path):

            self.file_edit.setText(path)

            self._save_settings()



    def _apply_recent_dir_item(self, item: QtWidgets.QListWidgetItem) -> None:

        path = item.text()

        if path and os.path.isdir(path):

            self.local_edit.setText(path)

            self._save_settings()



    def _find_in_logs(self) -> None:

        text = self.search_edit.text().strip()

        if not text:

            self.search_count.setText("0/0")

            return

        current = self.log_tabs.currentWidget()

        if not isinstance(current, QtWidgets.QPlainTextEdit):

            self.search_count.setText("0/0")

            return

        self._search_matches = []

        self._search_index = -1

        doc = current.document()

        cursor = QtGui.QTextCursor(doc)

        extra = []

        fmt = QtGui.QTextCharFormat()

        fmt.setBackground(QtGui.QColor("#fff59d"))

        while True:

            cursor = doc.find(text, cursor)

            if cursor.isNull():

                break

            sel = QtWidgets.QTextEdit.ExtraSelection()

            sel.cursor = cursor

            sel.format = fmt

            extra.append(sel)

            self._search_matches.append(cursor)

        current.setExtraSelections(extra)

        if self._search_matches:

            self._search_index = 0

            current.setTextCursor(self._search_matches[0])

            self.search_count.setText(f"{self._search_index + 1}/{len(self._search_matches)}")

        else:

            self.search_count.setText("0/0")



    def _next_match(self) -> None:

        current = self.log_tabs.currentWidget()

        if not isinstance(current, QtWidgets.QPlainTextEdit):

            return

        if not self._search_matches:

            self._find_in_logs()

            return

        self._search_index = (self._search_index + 1) % len(self._search_matches)

        current.setTextCursor(self._search_matches[self._search_index])

        self.search_count.setText(f"{self._search_index + 1}/{len(self._search_matches)}")



    def _prev_match(self) -> None:

        current = self.log_tabs.currentWidget()

        if not isinstance(current, QtWidgets.QPlainTextEdit):

            return

        if not self._search_matches:

            self._find_in_logs()

            return

        self._search_index = (self._search_index - 1) % len(self._search_matches)

        current.setTextCursor(self._search_matches[self._search_index])

        self.search_count.setText(f"{self._search_index + 1}/{len(self._search_matches)}")



    # ═══════════════════════════════════════════════════════════════════

    # 통합 실행 탭 핸들러

    # ═══════════════════════════════════════════════════════════════════



    def _on_integrated_ocr_only(self) -> None:

        """① OCR만 실행 - O열 이미지 다운로드 → OCR → Excel 저장"""

        if not self.integrated_file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV 파일을 선택해 주세요.")

            return



        QtWidgets.QMessageBox.information(

            self, "OCR 실행",

            "OCR 전용 기능은 현재 개발 중입니다.\n"

            "잠시만 기다려주세요!"

        )



    def _on_integrated_keyword_only(self) -> None:

        """② 키워드만 생성 - OCR Excel → GPT 키워드 → 업로드용 Excel"""

        if not self.integrated_file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV 파일을 선택해 주세요.")

            return



        csv_path = self.integrated_file_edit.text().strip()

        ocr_excel_path = self.ocr_excel_edit.text().strip()



        if not ocr_excel_path:

            try:

                from app.services.ocr_excel import find_matching_ocr_file

                found = find_matching_ocr_file(csv_path)

                if found:

                    ocr_excel_path = found

                    self.ocr_excel_edit.setText(found)

            except Exception:

                pass



        if not ocr_excel_path:

            QtWidgets.QMessageBox.warning(

                self,

                "입력 필요",

                "OCR 결과 Excel 파일이 필요합니다.\n"

                "상단의 OCR 결과 파일을 선택하거나 자동 찾기를 먼저 실행해 주세요."

            )

            return



        cfg = PipelineConfig(

            file_path=csv_path,

            img_tag="",

            tesseract_path="",

            model_keyword=self._normalize_runtime_model_name(self.model_keyword_combo.currentText(), self._default_keyword_model),

            model_longtail=self._normalize_runtime_model_name(self.model_longtail_combo.currentText(), self._default_longtail_model),

            keyword_version=self.keyword_version_combo.currentText().strip(),

            max_words=self.max_words.value(),

            max_len=self.max_len.value(),

            min_len=self.min_len.value(),

            use_html_ocr=False,

            use_local_ocr=False,  # OCR 재실행 없이 Excel 결과만 사용

            merge_ocr_with_name=True,

            max_imgs=0,

            threads=self.threads.value(),

            max_depth=-1,

            local_img_dir="",

            allow_folder_match=True,

            korean_only=True,

            drop_digits=self.drop_digits.isChecked(),

            psm=11,

            oem=3,

            ocr_excel_path=ocr_excel_path,

            write_to_r=True,

            debug=True,

            naver_enabled=False,  # OCR 기반 추론 전용: 네이버 검색광고 미사용

            naver_dry_run=self.naver_dry.isChecked(),

            naver_retry=self.naver_retry.isChecked(),

            naver_retry_count=self.naver_retry_count.value(),

            naver_retry_delay=self.naver_retry_delay.value(),

            naver_autocomplete=self.naver_autocomplete.isChecked(),

            google_autocomplete=self.google_autocomplete.isChecked(),

            make_listing=False,  # 키워드 전용

            listing_size=self.listing_size.value(),

            listing_pad=self.listing_pad.value(),

            listing_max=self.listing_max.value(),

            logo_path=self.logo_edit.text().strip(),

            logo_ratio=self.logo_ratio.value(),

            logo_opacity=self.logo_opacity.value(),

            logo_pos=self.logo_pos.currentText(),

            use_auto_contrast=self.auto_contrast.isChecked(),

            use_sharpen=self.sharpen.isChecked(),

            use_small_rotate=self.rotate.isChecked(),

            rotate_zoom=self.rotate_zoom.value(),

            ultra_angle_deg=self.ultra_angle.value(),

            ultra_translate_px=self.ultra_translate.value(),

            ultra_scale_pct=self.ultra_scale.value(),

            trim_tol=self.trim_tol.value(),

            jpeg_q_min=self.jpeg_q_min.value(),

            jpeg_q_max=self.jpeg_q_max.value(),

            do_flip_lr=self.flip_lr.isChecked(),

        )



        self.integrated_ocr_btn.setEnabled(False)

        self.integrated_keyword_btn.setEnabled(False)

        self.integrated_listing_btn.setEnabled(False)

        self.integrated_full_btn.setEnabled(False)



        self._open_log_file()

        self.progress.setValue(5)

        self._log("═" * 50)

        self._log("키워드 전용 실행 시작 (OCR 결과 Excel 사용)")

        self._log(f"CSV: {os.path.basename(csv_path)}")

        self._log(f"OCR Excel: {os.path.basename(ocr_excel_path)}")

        self._log("═" * 50)



        self._thread = QtCore.QThread(self)

        self._worker = PipelineWorker(cfg)

        self._worker.moveToThread(self._thread)



        self._thread.started.connect(self._worker.run)

        self._worker.status.connect(self._log)

        self._worker.progress.connect(self.progress.setValue)

        self._worker.finished.connect(self._on_integrated_finished)

        self._worker.error.connect(self._on_integrated_error)



        self._worker.finished.connect(self._thread.quit)

        self._worker.finished.connect(self._worker.deleteLater)

        self._thread.finished.connect(self._thread.deleteLater)

        self._worker.error.connect(self._thread.quit)

        self._worker.error.connect(self._worker.deleteLater)



        self._thread.start()



    def _on_integrated_listing_only(self) -> None:

        """③ 대표이미지만 생성 - AU열 다운로드 → 이미지 변형 → listing_images 저장"""

        if not self.integrated_file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV 파일을 선택해 주세요.")

            return



        QtWidgets.QMessageBox.information(

            self, "대표이미지 생성",

            "대표이미지 전용 기능은 현재 개발 중입니다.\n"

            "잠시만 기다려주세요!"

        )



    def _on_integrated_full(self) -> None:

        """④ 통합 실행 - OCR → 키워드 → 대표이미지 순차 실행"""

        if not self.integrated_file_edit.text().strip():

            QtWidgets.QMessageBox.warning(self, "입력 필요", "CSV 파일을 선택해 주세요.")

            return



        # 로컬 이미지 폴더 (비어있으면 임시 폴더 자동 생성)

        local_dir = self.integrated_local_edit.text().strip()

        if not local_dir:

            # 임시 폴더 생성: C:\code\temp_images\YYYYMMDD

            date_tag = datetime.now().strftime("%Y%m%d")

            local_dir = os.path.join("C:\\code", "temp_images", date_tag)

            os.makedirs(local_dir, exist_ok=True)



        # PipelineConfig 생성 (전처리 탭 설정값 사용)

        cfg = PipelineConfig(

            file_path=self.integrated_file_edit.text().strip(),

            img_tag="",

            tesseract_path="",

            model_keyword=self._normalize_runtime_model_name(self.model_keyword_combo.currentText(), self._default_keyword_model),

            model_longtail=self._normalize_runtime_model_name(self.model_longtail_combo.currentText(), self._default_longtail_model),

            keyword_version=self.keyword_version_combo.currentText().strip(),

            max_words=self.max_words.value(),

            max_len=self.max_len.value(),

            min_len=self.min_len.value(),

            use_html_ocr=False,

            use_local_ocr=True,

            merge_ocr_with_name=self.merge_ocr.isChecked(),

            max_imgs=999,

            threads=self.threads.value(),

            max_depth=-1,

            local_img_dir=local_dir,

            allow_folder_match=True,

            korean_only=True,

            drop_digits=self.drop_digits.isChecked(),

            psm=11,

            oem=3,

            ocr_excel_path="",

            write_to_r=True,

            debug=True,

            naver_enabled=self.naver_enable.isChecked(),

            naver_dry_run=False,

            naver_retry=self.naver_retry.isChecked(),

            naver_retry_count=self.naver_retry_count.value(),

            naver_retry_delay=self.naver_retry_delay.value(),

            naver_autocomplete=True,

            google_autocomplete=True,

            make_listing=self.make_listing.isChecked(),

            listing_size=self.listing_size.value(),

            listing_pad=self.listing_pad.value(),

            listing_max=self.listing_max.value(),

            logo_path=self.logo_edit.text().strip(),

            logo_ratio=self.logo_ratio.value(),

            logo_opacity=self.logo_opacity.value(),

            logo_pos=self.logo_pos.currentText(),

            use_auto_contrast=self.auto_contrast.isChecked(),

            use_sharpen=self.sharpen.isChecked(),

            use_small_rotate=self.rotate.isChecked(),

            rotate_zoom=self.rotate_zoom.value(),

            ultra_angle_deg=self.ultra_angle.value(),

            ultra_translate_px=self.ultra_translate.value(),

            ultra_scale_pct=self.ultra_scale.value(),

            trim_tol=self.trim_tol.value(),

            jpeg_q_min=self.jpeg_q_min.value(),

            jpeg_q_max=self.jpeg_q_max.value(),

            do_flip_lr=self.flip_lr.isChecked(),

        )



        # 실행 전 버튼 비활성화

        self.integrated_ocr_btn.setEnabled(False)

        self.integrated_keyword_btn.setEnabled(False)

        self.integrated_listing_btn.setEnabled(False)

        self.integrated_full_btn.setEnabled(False)



        self._open_log_file()

        self.progress.setValue(0)

        self._log("═" * 50)

        self._log("통합 실행 시작: CSV → 완성!")

        self._log("═" * 50)

        self._log(f"CSV: {os.path.basename(cfg.file_path)}")

        self._log(f"이미지 폴더: {local_dir}")

        self._log("O열 → OCR, AU열 → 대표이미지 자동 다운로드")



        # Worker 스레드 생성

        self._thread = QtCore.QThread(self)

        self._worker = PipelineWorker(cfg)

        self._worker.moveToThread(self._thread)



        # 시그널 연결

        self._thread.started.connect(self._worker.run)

        self._worker.status.connect(self._log)

        self._worker.progress.connect(self.progress.setValue)

        self._worker.finished.connect(self._on_integrated_finished)

        self._worker.error.connect(self._on_integrated_error)



        self._worker.finished.connect(self._thread.quit)

        self._worker.finished.connect(self._worker.deleteLater)

        self._thread.finished.connect(self._thread.deleteLater)

        self._worker.error.connect(self._thread.quit)

        self._worker.error.connect(self._worker.deleteLater)



        self._log("🔄 워커 스레드 시작...")

        self._thread.start()

        self._log("✅ 스레드 start() 호출 완료")



    def _on_integrated_finished(self, out_root: str, out_file: str) -> None:

        """통합 실행 완료"""

        self._last_output_file = out_file

        self.integrated_ocr_btn.setEnabled(True)

        self.integrated_keyword_btn.setEnabled(True)

        self.integrated_listing_btn.setEnabled(True)

        self.integrated_full_btn.setEnabled(True)



        self._log("═" * 50)

        self._log("✅ 통합 실행 완료!")

        self._log(f"결과: {out_root}")

        self._log("═" * 50)

        self._close_log_file()

        self._save_settings()



        QtWidgets.QMessageBox.information(

            self, "완료",

            f"✅ 모든 과정 완료!\n\n결과 폴더:\n{out_root}"

        )



        # 결과 폴더 자동 열기

        try:

            os.startfile(out_root)

        except Exception:

            pass



    def _on_integrated_error(self, error_msg: str) -> None:

        """통합 실행 오류"""

        self.integrated_ocr_btn.setEnabled(True)

        self.integrated_keyword_btn.setEnabled(True)

        self.integrated_listing_btn.setEnabled(True)

        self.integrated_full_btn.setEnabled(True)



        self._log(f"❌ 오류: {error_msg}")

        self._close_log_file()



        QtWidgets.QMessageBox.critical(self, "오류", f"오류:\n\n{error_msg}")

















