from __future__ import annotations

import argparse
import base64
from contextlib import contextmanager
import csv
import hashlib
import hmac
import json
import math
import mimetypes
import os
import queue
import re
import sqlite3
import ssl
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
import urllib.parse
import urllib.request
import urllib.error


ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT.parent
BACKEND_ROOT = PROJECT_ROOT / "backend"
BRIDGE_SCRIPT = PROJECT_ROOT / "KeywordOcr.App" / "Bridge" / "run_pipeline_bridge.py"
DATA_ROOT = ROOT / "data"
UPLOAD_ROOT = DATA_ROOT / "uploads"
LOGO_ROOT = DATA_ROOT / "logos"
JOBS_ROOT = DATA_ROOT / "jobs"
SEED_ROOT = DATA_ROOT / "seeds"
EXPORT_ROOT = DATA_ROOT / "exports"
EMERGENCY_ROOT = DATA_ROOT / "emergency"
MARKET_KEY_ROOT = DATA_ROOT / "market_keys"
MARKET_KEY_SETTINGS = MARKET_KEY_ROOT / "settings.json"
CATEGORY_REFERENCE_ROOT = PROJECT_ROOT / "data" / "category_reference"
DOTNET_UPLOAD_PROJECT = PROJECT_ROOT / "KeywordOcr.App.Tests" / "KeywordOcr.App.Tests.csproj"
DOTNET_UPLOAD_EXE = PROJECT_ROOT / "KeywordOcr.App.Tests" / "bin" / "Debug" / "net8.0-windows" / "win-x64" / "KeywordOcr.App.Tests.exe"
DESKTOP_KEY_ROOT = Path(os.environ.get("WEBOCR_KEY_ROOT") or os.environ.get("KEYWORDOCR_KEY_DIR") or (Path.home() / "Desktop" / "key"))
PRODUCT_MANAGER_ROOT = Path(os.environ.get("WEBOCR_PRODUCT_MANAGER_ROOT") or (Path.home() / "Desktop" / "ProductManager"))
PRODUCT_MANAGER_DB = PRODUCT_MANAGER_ROOT / "data" / "products.db"

for directory in (UPLOAD_ROOT, LOGO_ROOT, JOBS_ROOT, SEED_ROOT, EXPORT_ROOT, EMERGENCY_ROOT, MARKET_KEY_ROOT):
    directory.mkdir(parents=True, exist_ok=True)

ACTIVE_PROCESSES: dict[str, list[subprocess.Popen]] = {}
ACTIVE_PROCESS_LOCK = threading.Lock()
CANCELLED_JOB_IDS: set[str] = set()
CATEGORY_CACHE: dict[str, list[dict[str, object]]] = {}


def category_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def category_compact(value: object) -> str:
    return re.sub(r"\s+", "", category_text(value)).lower()


def normalise_category_market(value: object) -> str:
    text = category_text(value).lower()
    aliases = {
        "naver": "naver",
        "네이버": "naver",
        "smartstore": "naver",
        "스마트스토어": "naver",
        "coupang": "coupang",
        "쿠팡": "coupang",
        "lotteon": "lotteon",
        "lotte": "lotteon",
        "롯데": "lotteon",
        "롯데on": "lotteon",
        "11st": "11st",
        "11번가": "11st",
        "eleven": "11st",
        "esm": "esm",
        "gmarket": "esm",
        "g마켓": "esm",
        "auction": "esm",
        "옥션": "esm",
    }
    return aliases.get(text, text)


def parse_category_depth(value: object, fallback: int = 0) -> int:
    try:
        return int(float(category_text(value)))
    except Exception:
        return fallback


def parse_category_leaf(value: object) -> bool:
    text = category_text(value).lower()
    return text in {"y", "yes", "true", "1", "leaf"}


def category_path_parts(path: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*>\s*|/", path or "") if part.strip()]


def category_query_tokens(value: object) -> list[str]:
    text = category_text(value)
    text = re.sub(r"[A-Z]{1,3}\d{4,}[A-Z]?", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\d+(\.\d+)?\s*(cm|mm|m|g|kg|ml|l|개|매|장|ea|set|세트)", " ", text, flags=re.IGNORECASE)
    raw_tokens = re.split(r"[\s>/,|;:()\[\]{}+·._-]+", text)
    stopwords = {
        "상품", "제품", "선택", "옵션", "세트", "블랙", "화이트", "그레이", "실버", "골드",
        "대형", "소형", "중형", "미니", "고급", "국내", "수입", "정리", "방지", "보호",
        "홈런마켓", "급배송", "국내배송", "배송", "무료배송", "택배사", "주문", "발송",
        "구매대행", "창고", "결제", "문의", "상담", "주의사항", "교환", "반품", "수령",
        "검색", "리뷰", "후기", "판매자", "고객님", "사용", "흔적", "단순변심",
    }
    tokens: list[str] = []
    for token in raw_tokens:
        clean = category_compact(re.sub(r"^[\"']+|[\"']+$", "", token))
        if len(clean) < 2 or clean in stopwords:
            continue
        tokens.append(clean)
    return list(dict.fromkeys(tokens))[:32]


def category_has_any(text: str, terms: tuple[str, ...]) -> bool:
    return any(term in text for term in terms)


DOG_QUERY_TOKENS = ("강아지", "반려견", "애견", "댕댕이", "소형견", "중형견", "대형견")
CAT_QUERY_TOKENS = ("고양이", "반려묘", "애묘", "냥이")
OTHER_ANIMAL_QUERY_TOKENS = ("햄스터", "토끼", "기니피그", "고슴도치", "파충류", "장수풍뎅이", "곤충", "소동물")
GENERIC_PET_QUERY_TOKENS = ("반려동물", "애완", "펫")
PET_QUERY_TOKENS = (*GENERIC_PET_QUERY_TOKENS, *OTHER_ANIMAL_QUERY_TOKENS)
DOG_PATH_TOKENS = ("강아지", "강아지용품", "반려견", "애견", "고양이겸용")
CAT_PATH_TOKENS = ("고양이", "고양이용품", "반려묘", "애묘", "고양이겸용")
OTHER_ANIMAL_PATH_TOKENS = ("조류", "파충류", "햄스터", "토끼", "기니피그", "고슴도치", "장수풍뎅이", "곤충", "어류")
PET_PATH_TOKENS = ("반려", "애완용품", "펫", "강아지", "고양이", *OTHER_ANIMAL_PATH_TOKENS)
NON_PET_PATH_TOKENS = ("자동차용품", "출산", "육아", "화장지", "세차", "도서", "문구", "사무", "가전")
PET_PRODUCT_TOKENS = (
    "배변", "발캡", "발커버", "발바닥", "발도장", "발자국", "산책", "목줄", "하네스", "리드줄",
    "휘슬", "훈련", "장난감", "노즈워크", "간식", "사료", "물티슈", "샴푸", "미용", "브러쉬",
    "유산균", "건강", "위생", "의류", "옷", "신발", "방석", "매트", "케이지", "이동장",
    "급수", "급식", "식기", "목욕", "커버",
)
DESIGN_OBJECT_TOKENS = ("후크", "걸이", "선반", "책상", "소품", "수납", "키링", "고리", "참", "장식", "스티커", "인형", "로봇", "완구")
PET_MATCH_TERMS = ("배변", "발캡", "발커버", "발바닥", "발도장", "산책", "휘슬", "훈련", "장난감", "물티슈", "미용", "위생", "건강", "목욕", "커버")
PET_SPECIFIC_CATEGORY_TERMS = (
    "사료", "간식", "유산균", "영양제", "개껌", "동결건조", "비스킷", "스낵", "육포", "음료",
    "치약", "칫솔", "구강", "눈", "귀", "배변패드", "배변판", "기저귀", "탈취제", "소독제",
    "배변", "배변봉투", "배변유도제", "노즈워크", "장난감", "훈련", "샴푸", "브러시", "발톱", "목욕", "미용",
)
CATEGORY_STRONG_TERMS = (
    "물총", "변기솔", "스티커", "꽃다발", "플라워", "후크", "걸이", "양말", "파일", "모자",
    "확대경", "휘슬", "옷핀", "가방", "프라이팬", "필터", "손잡이", "호스", "북마크",
    "발톱칼", "니플", "발캡", "발도장", "조명", "부쉬", "호루라기",
)
BOOK_PATH_TOKENS = ("도서", "잡지", "ebook", "국내도서", "외국도서")
BOOK_QUERY_TOKENS = ("도서", "책", "잡지", "교재", "문제집", "소설", "에세이", "사진집")
FOOD_PATH_TOKENS = ("식품", "음료", "우유", "커피", "과자", "간식")
FOOD_QUERY_TOKENS = ("식품", "음료", "우유", "커피", "과자", "간식", "사료")
SERVICE_PATH_TOKENS = ("여가", "생활편의", "홈케어서비스", "이사", "청소대행", "서비스")
SERVICE_QUERY_TOKENS = ("서비스", "대행", "업체", "방문청소", "입주청소", "이사청소", "홈케어", "청소대행", "견적", "예약")
BATHROOM_QUERY_TOKENS = ("화장실", "욕실", "변기", "세면대", "샤워", "목욕")
BATHROOM_PATH_TOKENS = (
    "욕실용품", "욕실청소용품", "욕실잡화", "욕실세트", "욕실수납",
    "변기솔", "변기커버", "변기세정제", "양변기", "뚫어뻥", "세면대", "샤워",
)
BATHROOM_FOOTWEAR_PATH_TOKENS = ("욕실화", "욕실발판", "욕실매트", "실내화", "슬리퍼")
BATHROOM_FOOTWEAR_QUERY_TOKENS = ("욕실화", "욕실발판", "욕실매트", "실내화", "슬리퍼", "발판", "매트")
OUTDOOR_BATHROOM_PATH_TOKENS = (
    "캠핑", "스포츠", "레저", "캠핑용품", "캠핑용변기", "휴대용변기", "이동식변기",
    "간이화장실", "이동식화장실", "샤워텐트", "캠핑샤워기", "야외샤워",
)
OUTDOOR_BATHROOM_QUERY_TOKENS = ("캠핑", "야외", "아웃도어", "차박", "등산", "낚시", "휴대용", "이동식", "간이")
BATHROOM_BEAUTY_PATH_TOKENS = ("뷰티", "바디", "바디케어", "샤워코롱", "바디워시", "바디스크럽", "입욕용품", "향수")
BATHROOM_BEAUTY_QUERY_TOKENS = ("뷰티", "바디", "향수", "코롱", "워시", "스크럽", "입욕", "목욕용품")
COLLECTIBLE_PATH_TOKENS = ("수집품", "아이돌굿즈", "포토카드", "브로마이드", "응원봉", "굿즈")
COLLECTIBLE_QUERY_TOKENS = ("수집품", "아이돌", "굿즈", "포토카드", "브로마이드", "응원봉", "팬덤", "케이팝", "kpop", "앨범", "연예인")
GENERIC_STICKER_BAD_PATH_TOKENS = (
    "자동차용품", "세차용품", "타르제거제", "디지털", "가전", "pc액세서리", "키보드키스킨",
    "카메라", "캠코더", "보호필름", "휴대폰액세서리", "홈버튼스티커", "diy자재", "인테리어",
    "출산", "육아", "돌기념품",
)
GENERIC_STICKER_GOOD_PATH_TOKENS = ("문구", "사무", "팬시", "스티커", "네임스티커", "다꾸스티커", "주문제작스티커", "카드스티커")


CATEGORY_INTENT_PROFILES: tuple[dict[str, object], ...] = (
    {
        "id": "bathroom_toilet_brush",
        "label": "욕실청소/변기솔",
        "triggers": ("변기솔",),
        "primary": ("변기솔", "청소솔", "욕실청소도구", "욕실청소용품"),
        "positive": ("욕실", "화장실", "변기", "청소솔", "청소용품", "욕실청소", "욕실잡화", "뚫어뻥"),
        "negative": ("스티커", "문구", "자동차", "오토바이", "세차", "타르", "골프", "프라모델", "디지털", "가전", "컴퓨터", "완구", "스포츠"),
        "require": ("욕실", "화장실", "변기", "청소솔", "청소용품", "욕실청소"),
    },
    {
        "id": "fashion_hat",
        "label": "패션잡화/모자",
        "triggers": ("농사모자",),
        "primary": ("모자", "사파리모자", "챙모자", "선캡", "썬캡", "왕골", "밀짚", "벙거지", "버킷햇"),
        "positive": ("패션잡화", "잡화", "모자", "사파리", "챙모자", "선캡", "썬캡", "왕골", "밀짚", "벙거지", "버킷"),
        "negative": ("군모", "야구모자", "스냅백", "귀달이", "귀마개", "털모자", "방울", "비니", "베레모", "두건", "반다나", "페도라", "헌팅캡", "골프", "낚시", "등산", "유아동", "신생아", "도서", "원데이", "클래스", "플라워", "꽃/가드닝", "원피스"),
        "require": ("패션잡화", "패션의류잡화", "잡화", "여성패션", "남성패션", "유니섹스"),
    },
    {
        "id": "car_wash_hose",
        "label": "자동차/세차 호스",
        "triggers": ("호스관", "호스받침"),
        "primary": ("호스", "물호스"),
        "positive": ("자동차", "자동차용품", "차량", "차량용", "세차", "세차용품", "호스", "물호스"),
        "negative": ("스킨스쿠버", "관상어", "반려", "애완", "샤워", "욕실", "청소기", "생활가전", "가전", "원예", "정원", "가드닝", "튜닝", "흡기", "배기"),
        "require": ("자동차", "차량", "세차", "호스"),
    },
    {
        "id": "garden_hose",
        "label": "원예/정원 호스",
        "triggers": ("호스가이드", "정원호스"),
        "primary": ("호스", "호스카트", "물호스", "릴호스", "정원호스"),
        "positive": ("원예", "정원", "가드닝", "화단", "잔디", "공구", "호스", "호스카트", "물호스", "릴호스", "분사기"),
        "negative": ("스킨스쿠버", "관상어", "반려", "애완", "샤워", "욕실", "청소기", "생활가전", "가전", "자동차", "차량", "튜닝", "흡기", "배기"),
        "require": ("원예", "정원", "가드닝", "공구", "호스"),
    },
    {
        "id": "packaging_sticker",
        "label": "문구/포장 스티커",
        "triggers": ("포장봉인스티커", "고급포장스티커"),
        "primary": ("스티커", "포장데코스티커", "데코스티커", "네임스티커"),
        "positive": ("문구", "사무", "팬시", "포장", "스티커", "데코", "포장용품", "학용품"),
        "negative": ("자동차", "오토바이", "세차", "타르", "골프", "프라모델", "디지털", "가전", "아이돌", "수집품", "도서", "유아", "가구", "인테리어", "DIY자재", "시트지", "안내"),
        "require": ("문구", "팬시", "포장", "스티커", "데코"),
    },
    {
        "id": "automotive_sticker",
        "label": "자동차/바이크 스티커",
        "triggers": ("차량용스티커", "자동차스티커", "배트입체스티커"),
        "primary": ("차량용스티커", "자동차스티커", "스티커", "엠블럼"),
        "positive": ("자동차", "차량", "차량용", "오토바이", "바이크", "익스테리어", "스티커", "엠블럼", "튜닝"),
        "negative": ("문구", "팬시", "포장", "도서", "골프", "프라모델", "아이돌", "수집품"),
        "require": ("자동차", "차량", "오토바이", "바이크", "익스테리어", "스티커"),
    },
    {
        "id": "toy_water_gun",
        "label": "완구/물총",
        "triggers": ("물총",),
        "primary": ("물총", "물놀이용품"),
        "positive": ("완구", "장난감", "물놀이", "물총", "유아", "육아"),
        "negative": ("도서", "사진", "자동차", "세차", "스킨스쿠버", "욕실", "문구"),
        "require": ("완구", "장난감", "물놀이", "물총"),
    },
    {
        "id": "gift_flower",
        "label": "조화/선물꽃",
        "triggers": ("뜨개꽃다발", "꽃다발"),
        "primary": ("꽃다발", "조화", "선물용꽃", "생화다발"),
        "positive": ("꽃", "조화", "꽃다발", "선물용꽃", "이벤트꽃", "가드닝", "생화", "플라워"),
        "negative": ("원피스", "패션", "의류", "도서", "수공예클래스"),
        "require": ("꽃", "조화", "꽃다발", "선물용꽃", "생화"),
    },
    {
        "id": "party_balloon",
        "label": "파티/풍선용품",
        "triggers": ("러브플라워", "알루미늄풍선", "풍선"),
        "primary": ("풍선", "풍선용품", "파티용품", "이벤트", "데코용품"),
        "positive": ("문구", "사무", "이벤트", "파티", "풍선", "데코용품", "알루미늄풍선", "가랜드"),
        "negative": ("플라워", "꽃", "원예", "가드닝", "원피스", "의류", "원데이", "클래스", "드라이플라워"),
        "require": ("이벤트", "파티", "풍선", "문구", "사무"),
    },
    {
        "id": "yoga_socks",
        "label": "요가/필라테스 양말",
        "triggers": ("요가양말", "필라테스양말", "토삭스", "논슬립요가양말"),
        "primary": ("요가양말", "필라테스양말", "요가용품", "기타요가용품", "발가락양말", "스포츠양말"),
        "positive": ("요가", "필라테스", "양말", "발가락양말", "스포츠양말", "토삭스", "논슬립"),
        "negative": ("골프", "야구", "축구", "스키", "보드", "등산", "유아동", "신생아", "도서"),
        "require": ("요가", "필라테스", "양말"),
    },
    {
        "id": "storage_hook",
        "label": "수납/소품걸이",
        "triggers": ("선반후크", "책상후크", "소품걸이"),
        "primary": ("소품걸이", "후크", "기타걸이", "잡화걸이", "다용도걸이", "벽걸이선반", "선반"),
        "positive": ("생활용품", "생활", "수납", "정리", "소품걸이", "후크", "걸이", "선반", "옷걸이", "다용도걸이"),
        "negative": ("자동차", "차량", "오토바이", "유모차", "반려", "애완", "욕실", "샤워", "골프", "카메라"),
        "require": ("생활", "수납", "정리", "후크", "걸이", "선반"),
    },
    {
        "id": "pan_cover",
        "label": "주방/프라이팬 덮개",
        "triggers": ("프라이팬커버", "프라이팬덮개"),
        "primary": ("프라이팬덮개", "프라이팬", "냄비뚜껑", "뚜껑", "커버"),
        "positive": ("주방", "주방용품", "프라이팬", "팬", "덮개", "뚜껑", "커버"),
        "negative": ("스포츠", "캠핑", "침구", "소파", "자동차", "유모차", "반려", "욕실"),
        "require": ("주방", "프라이팬", "팬", "덮개", "뚜껑"),
    },
    {
        "id": "eyeglass_light",
        "label": "작업 조명/안경 클립 조명",
        "triggers": ("안경클립조명", "클립조명", "안경조명"),
        "primary": ("헤드랜턴", "작업등", "랜턴", "손전등", "북라이트", "안경소품"),
        "positive": ("조명", "랜턴", "헤드랜턴", "작업등", "캠핑", "공구", "안경", "안경소품"),
        "negative": ("자동차", "차량", "후크", "옷걸이", "행거", "수납", "카메라", "패션안경테", "관상어", "반려", "애완", "RC완구", "완구", "수조등", "거실조명", "욕실조명", "크리스마스"),
        "require": ("조명", "랜턴", "작업등", "안경", "공구", "캠핑"),
    },
    {
        "id": "magnifier",
        "label": "돋보기/확대경",
        "triggers": ("확대경", "돋보기", "안경부착확대경"),
        "primary": ("돋보기", "확대경", "안경소품"),
        "positive": ("돋보기", "확대경", "안경", "실버용품", "의료", "공구", "측정"),
        "negative": ("자동차", "차량", "후크", "옷걸이", "행거", "수납", "카메라"),
        "require": ("돋보기", "확대경", "안경", "실버", "의료", "공구"),
    },
    {
        "id": "washer_filter",
        "label": "세탁기/급수 필터",
        "triggers": ("세탁기필터",),
        "primary": ("세탁기", "필터", "녹물제거필터", "리필필터", "샤워기필터"),
        "positive": ("생활가전", "세탁", "건조기", "세탁기", "필터", "욕실", "수전", "녹물", "급수"),
        "negative": ("반려", "애완", "정수기", "가습기", "자동차", "카메라", "PC", "공기정화기"),
        "require": ("세탁", "세탁기", "필터", "욕실", "수전", "녹물"),
    },
    {
        "id": "brake_nipple",
        "label": "자동차/브레이크 니플",
        "triggers": ("브레이크니플",),
        "primary": ("브레이크", "니플", "브레이크용품", "튜닝용품"),
        "positive": ("자동차", "차량", "브레이크", "니플", "튜닝", "공구", "배관"),
        "negative": ("뷰티", "바디", "몸매", "패치", "니플밴드", "속옷", "언더웨어", "자전거"),
        "require": ("자동차", "차량", "브레이크", "니플", "공구", "배관"),
    },
    {
        "id": "safety_pin",
        "label": "수선/옷핀",
        "triggers": ("옷핀", "안전핀"),
        "primary": ("옷핀", "브로치핀", "시침핀", "클립", "핀", "브로치"),
        "positive": ("수선", "수예", "재단", "문구", "클립", "핀", "옷핀", "브로치", "패션소품"),
        "negative": ("카메라", "게임기", "가방", "케이스", "스포츠", "낚시", "자동차", "도서"),
        "require": ("수선", "수예", "재단", "문구", "클립", "핀", "브로치", "패션"),
    },
    {
        "id": "phone_wrist_bag",
        "label": "휴대폰/스마트폰 파우치",
        "triggers": ("손목폰가방", "폰가방", "휴대폰가방", "스마트폰가방"),
        "primary": ("스마트폰카드지갑", "휴대폰", "스마트폰", "파우치", "가방", "지갑", "힙색", "슬링백"),
        "positive": ("패션", "잡화", "여성가방", "남성가방", "휴대폰", "스마트폰", "폰", "파우치", "가방", "지갑", "힙색", "슬링백", "손목스트랩"),
        "negative": ("카메라", "캠코더", "노트북", "게임기", "여행", "스포츠", "골프", "등산", "반려", "출산", "육아", "유아동", "미아방지"),
        "require": ("패션", "잡화", "휴대폰", "스마트폰", "폰", "파우치", "가방", "지갑", "힙색", "슬링백"),
    },
    {
        "id": "document_file",
        "label": "문구/파일",
        "triggers": ("A4파일", "a4파일"),
        "primary": ("파일", "바인더", "클리어파일", "화일"),
        "positive": ("문구", "사무", "파일", "바인더", "클립보드", "오피스"),
        "negative": ("디지털", "카메라", "가방", "케이스", "도서", "음반"),
        "require": ("문구", "사무", "파일", "바인더", "오피스"),
    },
    {
        "id": "dog_training_whistle",
        "label": "반려견 훈련/고주파 휘슬",
        "triggers": ("고주파휘슬", "강아지휘슬", "애견휘슬"),
        "primary": ("훈련용품", "휘슬", "호루라기", "강아지장난감", "반려동물용품"),
        "positive": ("반려", "애완", "강아지", "강아지용품", "훈련", "훈련용품", "휘슬", "호루라기"),
        "negative": ("악기", "관악기", "스포츠", "심판", "호신", "안전보호구", "소리", "요가", "침구"),
        "require": ("반려", "애완", "강아지", "훈련", "휘슬", "호루라기"),
    },
    {
        "id": "safety_whistle",
        "label": "안전/비상 호루라기",
        "triggers": ("비상휘슬", "안전휘슬", "호신휘슬"),
        "primary": ("호루라기", "휘슬", "호신용호루라기", "스포츠액세서리"),
        "positive": ("안전", "호신", "호루라기", "휘슬", "스포츠액세서리", "심판용품"),
        "negative": ("악기", "관악기", "음반", "국악", "소리", "디자인", "팬시", "자동차", "경광봉"),
        "require": ("안전", "호신", "호루라기", "휘슬", "스포츠"),
    },
    {
        "id": "pull_push_handle",
        "label": "DIY/문 손잡이",
        "triggers": ("풀푸쉬손잡이", "pullpush손잡이", "pushpull손잡이"),
        "primary": ("손잡이", "문고리", "방문손잡이", "가구손잡이", "DIY자재"),
        "positive": ("가구", "인테리어", "DIY", "DIY자재", "건축자재", "손잡이", "문고리", "방문손잡이"),
        "negative": ("욕실", "변기커버", "변기시트", "주방", "뚜껑", "프라이팬", "손발건강", "출산", "육아"),
        "require": ("가구", "인테리어", "DIY", "건축자재", "손잡이", "문고리"),
    },
    {
        "id": "dog_paw_cover",
        "label": "반려견 발바닥 보호/발캡",
        "triggers": ("강아지일회용발캡", "강아지발캡", "발캡", "발바닥커버"),
        "primary": ("발바닥보호제", "피부관리", "기타반려동물용품", "강아지위생용품"),
        "positive": ("반려", "애완", "강아지", "강아지용품", "위생", "건강", "피부관리", "발바닥보호제", "발바닥", "발캡"),
        "negative": ("간식", "사료", "개껌", "동결건조", "가발", "헤어", "출산", "육아", "자동차"),
        "require": ("반려", "애완", "강아지", "위생", "건강", "피부관리", "발바닥", "발캡"),
    },
    {
        "id": "dog_paw_stamp",
        "label": "반려견 발도장/기념 키트",
        "triggers": ("강아지발도장", "발도장"),
        "primary": ("기타반려동물용품", "반려동물용품", "강아지용품", "기념", "DIY"),
        "positive": ("반려", "애완", "강아지", "발도장", "발자국", "기타반려동물용품", "강아지용품", "DIY"),
        "negative": ("간식", "사료", "식기", "물통", "목줄", "자동줄", "철장", "울타리", "하우스", "수납용품"),
        "require": ("반려", "애완", "강아지", "발도장", "발자국", "기타반려동물"),
    },
    {
        "id": "wire_bushing",
        "label": "전기/전선 보호 부싱",
        "triggers": ("코드부쉬", "코드부싱", "전선부쉬", "전선부싱", "케이블부쉬", "케이블부싱"),
        "primary": ("전선", "케이블", "전선정리용품", "전기설비자재", "부싱", "부쉬"),
        "positive": ("공구", "전기", "전기용품", "전선", "케이블", "전선정리", "전기설비", "산업자재", "전기생활용품"),
        "negative": ("조화", "꽃", "원예", "가드닝", "바코드", "게임", "디지털코드", "수영", "수상스포츠", "서핑", "등산", "스노클링", "다이빙"),
        "require": ("전기", "전선", "케이블", "공구", "산업", "전기용품"),
    },
    {
        "id": "mini_keychain_knife",
        "label": "휴대용 미니 칼/커터",
        "triggers": ("발톱칼", "황동발톱칼", "키링칼", "미니칼"),
        "primary": ("커터칼", "칼", "커팅기구", "수공구", "키링"),
        "positive": ("문구", "커터", "칼", "커팅", "공구", "수공구", "캠핑", "키링", "키홀더"),
        "negative": ("출산", "육아", "유아동", "반려", "애완", "발톱", "발관리", "주얼리", "목걸이", "펜던트", "명품", "럭셔리"),
        "require": ("문구", "커터", "칼", "공구", "수공구", "캠핑", "키링"),
    },
)


def category_query_context(compact_query: str) -> dict[str, bool]:
    dog_intent = category_has_any(compact_query, DOG_QUERY_TOKENS)
    cat_intent = category_has_any(compact_query, CAT_QUERY_TOKENS)
    other_animal_intent = category_has_any(compact_query, OTHER_ANIMAL_QUERY_TOKENS)
    generic_pet = category_has_any(compact_query, GENERIC_PET_QUERY_TOKENS)
    product_intent = category_has_any(compact_query, PET_PRODUCT_TOKENS)
    design_object = category_has_any(compact_query, DESIGN_OBJECT_TOKENS)
    if design_object and not product_intent and not generic_pet:
        dog_intent = False
        cat_intent = False
    return {
        "pet": dog_intent or cat_intent or other_animal_intent or generic_pet,
        "dog": dog_intent,
        "cat": cat_intent,
        "other_animal": other_animal_intent,
        "generic_pet": generic_pet,
    }


def category_intent_profile(compact_query: str) -> dict[str, object] | None:
    if not compact_query:
        return None

    if "발톱칼" in compact_query and not category_has_any(compact_query, ("강아지", "고양이", "반려", "애완", "발톱깎", "발톱깍")):
        return next(profile for profile in CATEGORY_INTENT_PROFILES if profile["id"] == "mini_keychain_knife")

    if "스티커" in compact_query:
        if category_has_any(compact_query, ("자동차", "차량", "바이크", "오토바이", "엠블럼", "익스테리어")):
            return next(profile for profile in CATEGORY_INTENT_PROFILES if profile["id"] == "automotive_sticker")
        if category_has_any(compact_query, ("포장", "봉인", "thank", "order", "쇼핑백", "봉투", "쿠키")):
            return next(profile for profile in CATEGORY_INTENT_PROFILES if profile["id"] == "packaging_sticker")

    if "호스" in compact_query:
        if category_has_any(compact_query, ("호스관", "호스받침", "세차", "자동차", "차량")):
            return next(profile for profile in CATEGORY_INTENT_PROFILES if profile["id"] == "car_wash_hose")
        if category_has_any(compact_query, ("호스가이드", "정원", "화단", "잔디", "원예", "가드닝", "꼬임")):
            return next(profile for profile in CATEGORY_INTENT_PROFILES if profile["id"] == "garden_hose")

    if "모자" in compact_query and category_has_any(compact_query, ("농사", "챙", "자외선", "야외", "얼굴", "목둘레", "하얀꽃", "패턴", "패션잡화")):
        return next(profile for profile in CATEGORY_INTENT_PROFILES if profile["id"] == "fashion_hat")

    for profile in CATEGORY_INTENT_PROFILES:
        if category_has_any(compact_query, tuple(profile.get("triggers") or ())):
            return profile
    return None


def category_profile_adjustment(profile: dict[str, object] | None, compact_path: str, compact_name: str) -> int:
    if not profile:
        return 0

    text = f"{compact_path}{compact_name}"
    primary = tuple(profile.get("primary") or ())
    positive = tuple(profile.get("positive") or ())
    negative = tuple(profile.get("negative") or ())
    required = tuple(profile.get("require") or ())
    adjustment = 0

    if category_has_any(text, primary):
        adjustment += 150
    if category_has_any(text, positive):
        adjustment += 95
    if required and not category_has_any(text, required):
        adjustment -= 115
    if category_has_any(text, negative):
        adjustment -= 165

    profile_id = category_text(profile.get("id"))
    if profile_id == "fashion_hat" and ("군모" in text or "야구모자" in text) and not category_has_any(text, ("사파리", "챙모자", "선캡", "썬캡", "왕골", "밀짚")):
        adjustment -= 140
    if profile_id == "fashion_hat" and category_has_any(text, ("양봉", "가축사료", "가축용품", "농업용")):
        adjustment -= 220
    if profile_id == "fashion_hat" and category_has_any(text, ("패션의류잡화", "여성패션", "남성패션", "패션잡화", "모자", "사파리모자", "챙모자")):
        adjustment += 130
    if profile_id == "car_wash_hose" and "호스" in text and category_has_any(text, ("자동차", "차량", "세차")):
        adjustment += 120
    if profile_id == "car_wash_hose" and "호스" in text and not category_has_any(text, ("자동차", "차량", "세차")):
        adjustment -= 120
    if profile_id == "car_wash_hose" and category_has_any(text, ("자동차", "차량", "세차")):
        adjustment += 120
    if profile_id == "garden_hose" and "호스" in text and category_has_any(text, ("원예", "정원", "가드닝", "화단", "잔디", "공구")):
        adjustment += 120
    if profile_id == "garden_hose" and category_has_any(text, ("배관", "건축자재")) and not category_has_any(text, ("물호스", "릴호스", "분사기", "원예", "정원", "가드닝")):
        adjustment -= 120
    if profile_id == "garden_hose" and category_has_any(text, ("물호스", "릴호스", "분사기", "정원호스")):
        adjustment += 150
    if profile_id == "bathroom_toilet_brush" and "스티커" in text and "변기솔" not in text and "청소솔" not in text:
        adjustment -= 160
    if profile_id == "bathroom_toilet_brush" and category_has_any(text, ("주변기기", "모니터", "디지털", "가전")):
        adjustment -= 220
    if profile_id == "storage_hook" and category_has_any(text, ("자동차", "차량", "오토바이")):
        adjustment -= 180
    if profile_id == "storage_hook" and category_has_any(text, ("캠핑", "스포츠", "레저")):
        adjustment -= 150
    if profile_id == "storage_hook" and category_has_any(text, ("생활용품", "수납정리", "수납", "정리용품", "소품걸이", "후크", "걸이")):
        adjustment += 140
    if profile_id == "party_balloon" and category_has_any(text, ("풍선", "파티", "이벤트")):
        adjustment += 130
    if profile_id == "party_balloon" and "풍선" in text:
        adjustment += 200
    if profile_id == "party_balloon" and category_has_any(text, ("보드게임", "룰렛", "다트룰렛", "원돌림판", "복불복")):
        adjustment -= 240
    if profile_id == "packaging_sticker" and category_has_any(text, ("포장", "문구", "팬시")):
        adjustment += 130
    if profile_id == "phone_wrist_bag" and category_has_any(text, ("카메라", "캠코더", "노트북", "게임기")):
        adjustment -= 180
    if profile_id == "phone_wrist_bag" and category_has_any(text, ("유아동", "미아방지", "출산", "육아", "수예", "뜨개질", "diy패키지", "반려", "애완")):
        adjustment -= 220
    if profile_id == "phone_wrist_bag" and category_has_any(text, ("여성가방", "남성가방", "파우치", "힙색", "슬링백", "스마트폰카드지갑")):
        adjustment += 140
    if profile_id == "washer_filter" and "필터" not in text and not category_has_any(text, ("세탁용품", "세탁잡화")):
        adjustment -= 180
    if profile_id == "washer_filter" and "필터" in text:
        adjustment += 130
    if profile_id == "washer_filter" and category_has_any(text, ("드럼세탁기", "일반세탁기", "세탁기+건조기")) and "필터" not in text:
        adjustment -= 220
    if profile_id == "washer_filter" and category_has_any(text, ("비데", "공기청정기", "에어컨", "가습기", "정수기", "카메라", "팬필터", "pc부품")):
        adjustment -= 220
    if profile_id == "washer_filter" and category_has_any(text, ("세탁기필터", "세탁용품", "녹물제거필터", "욕실보수용품")):
        adjustment += 180
    if profile_id == "washer_filter" and category_has_any(text, ("세탁용품", "세탁잡화")):
        adjustment += 180
    if profile_id == "eyeglass_light" and category_has_any(text, ("가구", "인테리어", "거실조명", "욕실조명", "야외조명", "벽조명", "전구", "크리스마스", "트리", "관상어", "수조등", "rc완구", "완구")):
        adjustment -= 240
    if profile_id == "eyeglass_light" and category_has_any(text, ("헤드랜턴", "손전등", "랜턴", "작업등", "북라이트", "안경소품")):
        adjustment += 170
    if profile_id == "pan_cover" and category_has_any(text, ("덮개", "뚜껑")):
        adjustment += 170
    if profile_id == "pan_cover" and "프라이팬" in text:
        adjustment += 130
    if profile_id == "pan_cover" and category_has_any(text, ("김치냉장고", "냉장고", "밥솥", "가전", "침구", "소파")):
        adjustment -= 240
    if profile_id == "yoga_socks" and category_has_any(text, ("요가", "필라테스", "요가양말", "필라테스양말")):
        adjustment += 140
    if profile_id == "yoga_socks" and category_has_any(text, ("스포츠>레저>요가", "요가>필라테스", "기타요가용품", "요가용품")):
        adjustment += 230
    if profile_id == "yoga_socks" and category_has_any(text, ("기타요가", "필라테스용품")):
        adjustment += 90
    if profile_id == "yoga_socks" and category_has_any(text, ("요가매트", "짐볼", "폼롤러", "필라테스링", "요가블럭", "볼스터")):
        adjustment -= 190
    if profile_id == "yoga_socks" and category_has_any(text, ("남성양말", "여성양말", "스포츠양말", "발가락양말")) and not category_has_any(text, ("요가", "필라테스")):
        adjustment -= 140
    if profile_id == "yoga_socks" and category_has_any(text, ("골프", "야구", "축구", "스키", "보드")):
        adjustment -= 180
    if profile_id == "wire_bushing" and category_has_any(text, ("전선", "케이블", "전기용품", "전기설비", "전기생활용품", "전선정리")):
        adjustment += 170
    if profile_id == "wire_bushing" and category_has_any(text, ("조화", "꽃", "가드닝", "바코드", "게임", "스포츠", "수영", "등산", "다이빙", "자동차", "배터리", "usb", "저장장치")):
        adjustment -= 220
    if profile_id == "mini_keychain_knife" and category_has_any(text, ("주얼리", "목걸이", "펜던트", "출산", "육아", "유아동", "반려", "애완", "발관리")):
        adjustment -= 220
    if profile_id == "mini_keychain_knife" and category_has_any(text, ("커터칼", "칼", "커팅기구", "수공구", "캠핑용품", "키링", "키홀더")):
        adjustment += 150
    if profile_id == "brake_nipple" and category_has_any(text, ("rc", "무선", "취미")):
        adjustment -= 220
    if profile_id == "brake_nipple" and category_has_any(text, ("자동차", "차량", "브레이크", "튜닝용품")):
        adjustment += 150
    if profile_id == "safety_pin" and category_has_any(text, ("파라핀", "손발건강", "볼링핀", "헤어핀")):
        adjustment -= 220
    if profile_id == "safety_pin" and category_has_any(text, ("옷핀", "문구", "핀", "브로치", "수선", "수예")):
        adjustment += 140
    if profile_id == "dog_training_whistle" and category_has_any(text, ("강아지", "반려", "애완", "훈련용품")):
        adjustment += 190
    if profile_id == "dog_training_whistle" and category_has_any(text, ("악기", "관악기", "요가", "침구", "스포츠", "심판", "호신")):
        adjustment -= 220
    if profile_id == "safety_whistle" and category_has_any(text, ("호루라기", "휘슬", "안전", "호신", "스포츠액세서리", "심판용품")):
        adjustment += 170
    if profile_id == "safety_whistle" and category_has_any(text, ("악기", "관악기", "음반", "국악", "디자인", "팬시", "자동차", "경광봉")):
        adjustment -= 220
    if profile_id == "pull_push_handle" and category_has_any(text, ("손잡이", "문고리", "방문손잡이", "가구손잡이", "diy자재", "건축자재")):
        adjustment += 170
    if profile_id == "pull_push_handle" and category_has_any(text, ("욕실", "변기커버", "변기시트", "프라이팬", "뚜껑", "손발건강", "출산", "육아", "가구다리", "발통", "시트지", "스티커", "필름지", "냄비", "주방")):
        adjustment -= 240
    if profile_id == "pull_push_handle" and category_has_any(text, ("도어용품", "철물", "인테리어자재", "diy자재>용품>자재")):
        adjustment += 120
    if profile_id == "dog_paw_cover" and category_has_any(text, ("강아지", "반려", "애완", "피부관리", "발바닥보호제", "위생", "건강")):
        adjustment += 180
    if profile_id == "dog_paw_cover" and category_has_any(text, ("발바닥보호제", "위생용품")):
        adjustment += 170
    if profile_id == "dog_paw_cover" and category_has_any(text, ("소동물", "가축", "기타반려동물용품")):
        adjustment -= 170
    if profile_id == "dog_paw_cover" and category_has_any(text, ("간식", "사료", "개껌", "동결건조", "가발", "출산", "육아", "뷰티", "피부관리기")):
        adjustment -= 220
    if profile_id == "dog_paw_cover" and not category_has_any(text, ("반려", "애완", "강아지", "고양이겸용")):
        adjustment -= 180
    if profile_id == "dog_paw_cover" and "해외직구" in text:
        adjustment -= 70
    if profile_id == "dog_paw_stamp" and category_has_any(text, ("강아지", "반려", "애완", "기타반려동물용품", "강아지용품")):
        adjustment += 160
    if profile_id == "dog_paw_stamp" and category_has_any(text, ("기타리빙용품", "기타반려동물용품")):
        adjustment += 140
    if profile_id == "dog_paw_stamp" and category_has_any(text, ("소동물", "가축", "계단", "슬라이드", "매트", "침대", "방석", "쿠션", "목걸이", "인식표", "캐리어", "이동장", "패션", "티셔츠", "우비")):
        adjustment -= 120
    if profile_id == "dog_paw_stamp" and category_has_any(text, ("간식", "사료", "식기", "목줄", "자동줄", "가슴줄", "하네스", "철장", "울타리", "하우스", "수납용품")):
        adjustment -= 220
    if profile_id == "dog_paw_stamp" and not category_has_any(text, ("반려", "애완", "강아지", "기타반려동물")):
        adjustment -= 150
    if profile_id == "dog_paw_stamp" and "해외직구" in text:
        adjustment -= 70
    if profile_id == "phone_wrist_bag" and category_has_any(text, ("디지털", "태블릿", "태블릿pc", "게임기", "카메라", "등산", "자전거", "수영", "문구", "필통")):
        adjustment -= 210
    return adjustment


def category_profile_summary(profile: dict[str, object] | None) -> dict[str, object] | None:
    if not profile:
        return None
    return {
        "id": profile.get("id"),
        "label": profile.get("label"),
        "primary": list(profile.get("primary") or ()),
        "positive": list(profile.get("positive") or ()),
        "negative": list(profile.get("negative") or ()),
    }


CATEGORY_BASIS_SKIP_LINE = re.compile(
    r"^(?:[•ㆍ\-]\s*)?(SPECIFICATION|제품\s*상세\s*정보|제품상세정보|제품\s*상세|제품상세|상세정보|"
    r"소재|구성|수량|판매수량|제조국|수입사|색상|사이즈|SIZE|Size Info|사이즈\s*안내|"
    r"참고사항|주의\s*사항|국내배송|빠르고\s*안전하게|홈런마켓|급배송|REAL\s*REVIEW|실제고객님들의|"
    r"판매자|고객님|평일|택배사|퀵서비스|상품\/대량구매|모든\s*제품|무료|국내\s*초대형|"
    r"사진은|모니터의|해상도에|상기\s*정보|옵션)$",
    re.IGNORECASE,
)
CATEGORY_BASIS_SKIP_NEXT_LABEL = re.compile(
    r"^(?:[•ㆍ\-]\s*)?(소재|구성|수량|판매수량|제조국|수입사|색상|사이즈|옵션)$",
    re.IGNORECASE,
)
CATEGORY_BASIS_NOISE = re.compile(
    r"(홈런마켓|급배송|택배사|구매대행|대량구매|무료|퀵서비스|REAL\s*REVIEW|실제고객|"
    r"주의\s*사항|참고사항|교환|반품|색상\s*관련|모니터|해상도|고객님|판매자)",
    re.IGNORECASE,
)


def clean_category_basis_text(value: object) -> str:
    text = category_text(value)
    if text in {'""', "''", "{}", "[]"}:
        return ""
    text = re.sub(r"[{}\"`]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def category_ocr_signal_lines(raw_text: object, limit: int = 24) -> tuple[list[str], list[str]]:
    lines: list[str] = []
    excluded: list[str] = []
    skip_next = 0
    for raw_line in category_text(raw_text).splitlines():
        line = clean_category_basis_text(raw_line.replace("•", " ").replace("ㆍ", " "))
        line = line.strip(" -:;|")
        if not line:
            continue
        compact = category_compact(line)
        if re.fullmatch(r"[a-z]{0,3}\d+[a-z]?", compact, re.IGNORECASE):
            continue
        if re.search(r"[★☆✰⭑]{2,}", line) or re.fullmatch(r"[a-z0-9*]{2,}", compact, re.IGNORECASE):
            continue
        if re.fullmatch(r"[\d\s./x×~\-()]+", line):
            continue
        if skip_next:
            excluded.append(line)
            skip_next -= 1
            continue
        if CATEGORY_BASIS_SKIP_NEXT_LABEL.match(line):
            skip_next = 1
            continue
        if CATEGORY_BASIS_SKIP_LINE.match(line):
            continue
        if CATEGORY_BASIS_NOISE.search(line):
            continue
        if category_has_any(compact, ("평일", "문자상담", "친절하게상담", "초대형창고", "묶음배송", "소중한리뷰", "항상건강", "에서만가능", "많은상품")):
            continue
        if len(line) > 90:
            continue
        lines.append(line)
        if len(lines) >= limit:
            break
    return lines, excluded[:20]


def category_basis_from_product(product: dict[str, object]) -> dict[str, object]:
    if not isinstance(product, dict):
        return {}
    ocr = product.get("ocrAnalysis") if isinstance(product.get("ocrAnalysis"), dict) else {}
    fields = ocr.get("fields", {}) if isinstance(ocr.get("fields"), dict) else {}
    source_name = clean_product_title(product.get("sourceName") or product.get("name") or product.get("gs") or "")
    identity = clean_product_title(fields.get("상품명")) or source_name
    raw_text = category_text(ocr.get("rawText") or fields.get("OCR텍스트") or "")
    signals, excluded = category_ocr_signal_lines(raw_text)
    vision_signals = [
        clean_category_basis_text(fields.get(key))
        for key in (
            "Vision_core_identity",
            "Vision_usage_context",
            "Vision_functional_inference",
            "Vision_search_boost_elements",
            "Vision_market_expansion",
        )
    ]
    vision_signals = [value for value in vision_signals if value]
    basis_text = " ".join([identity, source_name, *signals[:14], *vision_signals[:5]])
    profile = category_intent_profile(category_compact(basis_text))
    profile_terms: list[str] = []
    if profile:
        profile_terms.extend(category_text(term) for term in profile.get("primary") or ())
        profile_terms.extend(category_text(term) for term in list(profile.get("positive") or ())[:10])
    query_parts = [identity, source_name, *signals[:12], *profile_terms]
    query_tokens = category_query_tokens(" ".join(part for part in query_parts if part))
    query = " ".join(query_tokens[:90])
    return {
        "source": "ocr_analysis",
        "identity": identity,
        "profile": category_profile_summary(profile),
        "signals": signals[:18],
        "excludedAccessoryLines": excluded,
        "query": query,
    }


def category_context_adjustment(context: dict[str, bool], compact_query: str, compact_path: str) -> int:
    if not context.get("pet"):
        return 0

    adjustment = 0
    pet_path = category_has_any(compact_path, PET_PATH_TOKENS)
    other_animal_path = category_has_any(compact_path, OTHER_ANIMAL_PATH_TOKENS)
    non_pet_path = category_has_any(compact_path, NON_PET_PATH_TOKENS)

    if pet_path:
        adjustment += 35
    if non_pet_path:
        adjustment -= 140

    if context.get("dog"):
        dog_path = category_has_any(compact_path, DOG_PATH_TOKENS)
        cat_only_path = category_has_any(compact_path, ("고양이용품", "고양이간식", "고양이장난감")) and not dog_path
        if dog_path:
            adjustment += 130
        elif pet_path:
            adjustment += 20
        if other_animal_path:
            adjustment -= 160
        if cat_only_path:
            adjustment -= 60

    if context.get("cat"):
        cat_path = category_has_any(compact_path, CAT_PATH_TOKENS)
        dog_only_path = category_has_any(compact_path, ("강아지용품", "강아지간식", "강아지장난감")) and not cat_path
        if cat_path:
            adjustment += 130
        elif pet_path:
            adjustment += 20
        if other_animal_path:
            adjustment -= 160
        if dog_only_path:
            adjustment -= 50

    if context.get("other_animal"):
        dog_or_cat_path = category_has_any(compact_path, DOG_PATH_TOKENS + CAT_PATH_TOKENS)
        if other_animal_path:
            adjustment += 170
        elif pet_path:
            adjustment += 25
        if dog_or_cat_path and not other_animal_path:
            adjustment -= 90

    if context.get("generic_pet") and other_animal_path and not (context.get("dog") or context.get("cat") or context.get("other_animal")):
        adjustment -= 45

    matched_specific = any(term in compact_query and term in compact_path for term in PET_SPECIFIC_CATEGORY_TERMS)
    unmatched_specific = any(term not in compact_query and term in compact_path for term in PET_SPECIFIC_CATEGORY_TERMS)
    if unmatched_specific and not matched_specific:
        adjustment -= 75
    if not matched_specific and ("기타반려동물용품" in compact_path or "기타강아지용품" in compact_path or "기타애완용품" in compact_path):
        adjustment += 60

    for term in PET_MATCH_TERMS:
        if term in compact_query and term in compact_path:
            adjustment += 24
    return adjustment


def category_domain_adjustment(compact_query: str, compact_path: str) -> int:
    adjustment = 0
    profile = category_intent_profile(compact_query)
    strong_terms = [term for term in CATEGORY_STRONG_TERMS if term in compact_query]
    if profile:
        primary_terms = tuple(profile.get("primary") or ())
        if primary_terms and category_has_any(compact_path, primary_terms):
            adjustment += 70
        elif strong_terms and any(term in compact_path for term in strong_terms):
            adjustment -= 95
    elif strong_terms:
        if any(term in compact_path for term in strong_terms):
            adjustment += 115
        else:
            adjustment -= 75

    if category_has_any(compact_path, BOOK_PATH_TOKENS) and not category_has_any(compact_query, BOOK_QUERY_TOKENS):
        adjustment -= 140
    if category_has_any(compact_path, FOOD_PATH_TOKENS) and not category_has_any(compact_query, FOOD_QUERY_TOKENS):
        adjustment -= 120
    if category_has_any(compact_path, SERVICE_PATH_TOKENS) and not category_has_any(compact_query, SERVICE_QUERY_TOKENS):
        adjustment -= 260
    if category_has_any(compact_path, COLLECTIBLE_PATH_TOKENS) and not category_has_any(compact_query, COLLECTIBLE_QUERY_TOKENS):
        adjustment -= 260
    if category_has_any(compact_query, OTHER_ANIMAL_QUERY_TOKENS):
        if category_has_any(compact_path, OTHER_ANIMAL_PATH_TOKENS):
            adjustment += 170
        if "화장실" in compact_query and "화장실" in compact_path and category_has_any(compact_path, OTHER_ANIMAL_PATH_TOKENS):
            adjustment += 120
    if category_has_any(compact_query, BATHROOM_QUERY_TOKENS) and not category_has_any(compact_query, PET_QUERY_TOKENS + DOG_QUERY_TOKENS + CAT_QUERY_TOKENS):
        if category_has_any(compact_path, BATHROOM_PATH_TOKENS):
            adjustment += 160
        if category_has_any(compact_path, PET_PATH_TOKENS):
            adjustment -= 320
        if "화장실" in compact_path and category_has_any(compact_path, PET_PATH_TOKENS):
            adjustment -= 360
        if category_has_any(compact_path, OTHER_ANIMAL_PATH_TOKENS):
            adjustment -= 420
        if category_has_any(compact_path, OUTDOOR_BATHROOM_PATH_TOKENS) and not category_has_any(compact_query, OUTDOOR_BATHROOM_QUERY_TOKENS):
            adjustment -= 260
        if category_has_any(compact_path, BATHROOM_BEAUTY_PATH_TOKENS) and not category_has_any(compact_query, BATHROOM_BEAUTY_QUERY_TOKENS):
            adjustment -= 360
        if category_has_any(compact_path, BATHROOM_FOOTWEAR_PATH_TOKENS) and not category_has_any(compact_query, BATHROOM_FOOTWEAR_QUERY_TOKENS):
            adjustment -= 180
        if "화장실" in compact_query and "목욕" in compact_path and "목욕" not in compact_query:
            adjustment -= 220
        if category_has_any(compact_path, ("출산", "육아", "유아", "신생아")) and not category_has_any(compact_query, ("유아", "아기", "신생아", "출산", "육아")):
            adjustment -= 190
        if category_has_any(compact_path, ("실버용품", "목욕보조용품")) and not category_has_any(compact_query, ("실버", "노인", "어르신", "목욕보조", "장애인")):
            adjustment -= 190
        if "청소" in compact_query and category_has_any(compact_path, ("청소기", "로봇청소기", "물걸레청소기", "스팀청소기", "업소용청소기", "창문청소기", "침구청소기")):
            adjustment -= 220
        if "청소" in compact_query and "청소용품" in compact_path and not category_has_any(compact_path, BATHROOM_PATH_TOKENS + ("청소솔",)):
            adjustment -= 160
        if "청소" in compact_query and category_has_any(compact_path, ("욕실청소용품", "욕실청소도구", "변기솔", "뚫어뻥", "청소솔")):
            adjustment += 150
    if "비즈11번가" in compact_path:
        adjustment -= 240

    if "스티커" in compact_query and not profile:
        if category_has_any(compact_path, GENERIC_STICKER_GOOD_PATH_TOKENS):
            adjustment += 140
        if category_has_any(compact_path, GENERIC_STICKER_BAD_PATH_TOKENS):
            adjustment -= 170

    if ("후크" in compact_query or "걸이" in compact_query or "선반" in compact_query) and category_has_any(compact_path, ("후크", "걸이", "선반", "수납", "정리용품", "소품걸이")):
        adjustment += 90
    if "꽃다발" in compact_query and category_has_any(compact_path, ("꽃다발", "꽃", "조화", "이벤트꽃", "파티")):
        adjustment += 110
    if "물총" in compact_query and "물총" in compact_path:
        adjustment += 130
    if "변기솔" in compact_query and "변기솔" in compact_path:
        adjustment += 130
    if "스티커" in compact_query and "스티커" in compact_path:
        adjustment += 95
    return adjustment


CATEGORY_BUNDLE_MARKETS = (
    ("naver", "네이버"),
    ("coupang", "쿠팡"),
    ("lotteon", "롯데ON"),
    ("11st", "11번가"),
    ("esm", "ESM"),
)


def category_selection_key(market_key: str, item: dict[str, object] | None = None) -> str:
    market_key = normalise_category_market(market_key)
    item = item or {}
    if market_key == "naver":
        return "naver"
    if market_key == "coupang":
        return "coupang"
    if market_key == "11st":
        return "elevenst"
    if market_key == "esm":
        return "esm"
    if market_key == "lotteon":
        item_type = category_text(item.get("type")).lower()
        code = category_text(item.get("code"))
        if "display" in item_type or code.startswith("FC"):
            return "lotte_display"
        return "lotte_standard"
    return market_key or "category"


def category_bundle_search_query(query: object, naver_anchor: dict[str, object] | None) -> str:
    query_text = category_text(query)
    if not naver_anchor:
        return query_text
    broad_parts = {"생활", "건강", "패션", "잡화", "가구", "인테리어", "디지털", "가전", "출산", "육아"}
    anchor_parts = [
        part for part in category_path_parts(category_text(naver_anchor.get("path")))
        if part and part not in broad_parts
    ]
    anchor_name = category_text(naver_anchor.get("name"))
    return " ".join(part for part in [query_text, *anchor_parts, anchor_name] if part).strip()[:1200]


def category_bundle_reference(query: object, limit: int = 8) -> dict[str, object]:
    query_text = category_text(query)
    profile = category_intent_profile(category_compact(query_text))
    naver_items = search_category_reference("naver", query_text, limit)
    naver_anchor = naver_items[0] if naver_items else None
    bundle_query = category_bundle_search_query(query_text, naver_anchor)
    items_by_market: dict[str, list[dict[str, object]]] = {}
    selections: dict[str, dict[str, object]] = {}

    for market_key, _label in CATEGORY_BUNDLE_MARKETS:
        search_limit = max(limit, 40) if market_key == "lotteon" else limit
        items = naver_items if market_key == "naver" else search_category_reference(market_key, bundle_query, search_limit)
        items_by_market[market_key] = items
        if items:
            if market_key == "lotteon":
                standard = next((item for item in items if category_text(item.get("type")).lower() == "standard" or category_text(item.get("code")).startswith("BC")), None)
                display = next((item for item in items if category_text(item.get("type")).lower() == "display" or category_text(item.get("code")).startswith(("FC", "EC"))), None)
                selected = dict(display or standard or items[0])
                if standard:
                    selected["standardCode"] = category_text(standard.get("code"))
                    selected["standardPath"] = category_text(standard.get("path"))
                    selected["standardName"] = category_text(standard.get("name"))
                if display:
                    selected["displayCode"] = category_text(display.get("code"))
                    selected["displayPath"] = category_text(display.get("path"))
                    selected["displayName"] = category_text(display.get("name"))
                if not standard:
                    compact_bundle = category_compact(bundle_query)
                    if category_has_any(compact_bundle, ("요가양말", "필라테스양말", "토삭스", "논슬립양말")):
                        selected["standardCode"] = "BC80070100"
                        selected["standardPath"] = "헬스 > 수영용품 > 요가 > 필라테스용품 > 기타요가 > 필라테스용품"
                    elif category_has_any(compact_bundle, ("강아지발도장", "반려견발도장", "발도장키트", "발자국키트", "펫발도장")):
                        selected["standardCode"] = "BC26140000"
                        selected["standardPath"] = "반려동물 > 기타반려동물용품"
                selected["itemCode"] = "04" if category_has_any(category_compact(bundle_query), ("장갑", "글러브", "반장갑", "작업장갑", "코팅장갑")) else "38"
            else:
                selected = dict(items[0])
            selected["selectionKey"] = category_selection_key(market_key, selected)
            selected["bundleMarket"] = market_key
            selections[market_key] = selected

    return {
        "query": query_text,
        "bundleQuery": bundle_query,
        "profile": category_profile_summary(profile),
        "anchor": naver_anchor,
        "items": items_by_market,
        "selections": selections,
    }


def load_category_reference(market_key: str) -> list[dict[str, object]]:
    market_key = normalise_category_market(market_key)
    if market_key in CATEGORY_CACHE:
        return CATEGORY_CACHE[market_key]

    files: list[tuple[str, Path]] = []
    if market_key == "naver":
        files = [("naver", CATEGORY_REFERENCE_ROOT / "naver_categories.csv")]
    elif market_key == "coupang":
        files = [("coupang", CATEGORY_REFERENCE_ROOT / "coupang_categories.csv")]
    elif market_key == "lotteon":
        files = [("lotteon", CATEGORY_REFERENCE_ROOT / "lotteon_categories.csv")]
    elif market_key == "11st":
        files = [("11st", CATEGORY_REFERENCE_ROOT / "11st_categories.csv")]
    elif market_key == "esm":
        files = [("esm", CATEGORY_REFERENCE_ROOT / "auction_categories.csv")]

    items: list[dict[str, object]] = []
    for source, path in files:
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for raw in reader:
                if not raw:
                    continue
                code = category_text(
                    raw.get("category_code")
                    or raw.get("category_id")
                    or raw.get("code")
                    or raw.get("카테고리코드")
                )
                name = category_text(raw.get("category_name") or raw.get("name") or raw.get("카테고리명"))
                path_text = category_text(raw.get("full_path") or raw.get("category_path") or raw.get("path") or name)
                if not code or not path_text:
                    continue
                parts = category_path_parts(path_text)
                if not name and parts:
                    name = parts[-1]
                depth = parse_category_depth(raw.get("depth") or raw.get("level"), len(parts))
                is_leaf = parse_category_leaf(raw.get("is_leaf") or raw.get("leaf_yn") or raw.get("is_display"))
                category_type = category_text(raw.get("category_type") or source)
                items.append({
                    "market": source,
                    "type": category_type,
                    "code": code,
                    "name": name or code,
                    "path": " > ".join(parts) if parts else name or code,
                    "depth": depth or len(parts),
                    "parentCode": category_text(raw.get("parent_code") or raw.get("parent_id")),
                    "isLeaf": is_leaf,
                    "parts": parts,
                })

    CATEGORY_CACHE[market_key] = items
    return items


def search_category_reference(market: object, query: object, limit: int = 80) -> list[dict[str, object]]:
    market_key = normalise_category_market(market)
    items = load_category_reference(market_key)
    query_text = category_text(query)
    compact_query = category_compact(query_text)
    tokens = [category_compact(token) for token in re.split(r"[\s>/,]+", query_text) if category_compact(token)]
    smart_tokens = category_query_tokens(query_text)
    query_context = category_query_context(compact_query)
    query_profile = category_intent_profile(compact_query)
    scored: list[tuple[int, dict[str, object]]] = []

    for item in items:
        path_text = category_text(item.get("path"))
        name_text = category_text(item.get("name"))
        compact_path = category_compact(path_text)
        compact_name = category_compact(name_text)
        depth = int(item.get("depth") or 0)
        leaf_bonus = 3 if item.get("isLeaf") else 0
        score = 0
        if not compact_query:
            score = 40 - min(depth, 8) + leaf_bonus
        elif compact_name == compact_query:
            score = 120 + leaf_bonus
        elif compact_name.startswith(compact_query):
            score = 105 + leaf_bonus
        elif compact_query in compact_name:
            score = 92 + leaf_bonus
        elif compact_query in compact_path:
            score = 78 + leaf_bonus
        elif tokens and all(token in compact_path for token in tokens):
            score = 70 + leaf_bonus
        if compact_query and smart_tokens:
            token_score = 0
            for token in smart_tokens:
                if token == compact_name:
                    token_score += 70
                elif token in compact_name:
                    token_score += 46
                elif token in compact_path:
                    token_score += 22
            for part in category_path_parts(path_text):
                compact_part = category_compact(part)
                if compact_part and compact_part in compact_query:
                    token_score += 36 if part == name_text else 18
            if "양말" in compact_query and "양말" in compact_name:
                token_score += 42
            if ("요가" in compact_query or "필라테스" in compact_query) and ("요가" in compact_path or "필라테스" in compact_path):
                token_score += 28
            if "발가락" in compact_query and "발가락" in compact_path:
                token_score += 32
            if token_score:
                score = max(score, token_score + leaf_bonus)
        if score and "양말" in compact_query and "양말" not in compact_path:
            score -= 90
        if score and "양말" in compact_query and "발가락" in compact_query and "양말" in compact_path and "발가락" in compact_path:
            score += 46
        domain_adjustment = category_domain_adjustment(compact_query, compact_path)
        if domain_adjustment and (score or domain_adjustment > 0):
            score = max(score, leaf_bonus) + domain_adjustment
        profile_adjustment = category_profile_adjustment(query_profile, compact_path, compact_name)
        if profile_adjustment and (score or profile_adjustment > 0):
            score = max(score, leaf_bonus) + profile_adjustment
        context_adjustment = category_context_adjustment(query_context, compact_query, compact_path)
        if context_adjustment and (score or context_adjustment > 0):
            score = max(score, leaf_bonus) + context_adjustment
        if market_key == "lotteon" and score:
            code_text = category_text(item.get("code")).upper()
            if "(구)" in path_text:
                score -= 70
            if compact_path.startswith("해외직구"):
                score -= 90
            if code_text.startswith("FC"):
                score += 35
            elif code_text.startswith("EC"):
                score -= 25
        if score:
            scored.append((score - min(depth, 10), item))

    scored.sort(key=lambda pair: (-pair[0], category_text(pair[1].get("path"))))
    return [item for _, item in scored[:max(1, min(limit, 200))]]


def browse_category_reference(market: object, parent_path: object, limit: int = 120) -> list[dict[str, object]]:
    parent = " > ".join(category_path_parts(category_text(parent_path)))
    if not parent:
        return search_category_reference(market, "", limit)
    parent_compact = category_compact(parent)
    items = []
    for item in load_category_reference(normalise_category_market(market)):
        path_text = category_text(item.get("path"))
        path_compact = category_compact(path_text)
        if path_compact == parent_compact or path_compact.startswith(f"{parent_compact}>"):
            items.append(item)
    items.sort(key=lambda item: (int(item.get("depth") or 0), category_text(item.get("path"))))
    return items[:max(1, min(limit, 300))]


MARKET_KEY_OVERLAY_LOCK = threading.Lock()
ACTIVE_MARKET_UPLOAD_LOCK = threading.Lock()
MARKET_UPLOAD_TIMEOUT_SECONDS = 900


KEYWORD_POOL_CATEGORIES = [
    {
        "id": "identity",
        "label": "상품 정체성",
        "description": "상품명, 표준명, 카테고리성 단어",
    },
    {
        "id": "function",
        "label": "기능",
        "description": "고정, 보강, 방지, 연결, 수선 등",
    },
    {
        "id": "usePlace",
        "label": "사용처",
        "description": "신발, 콘센트함, 스위치 박스, 가구 등",
    },
    {
        "id": "problemSolving",
        "label": "문제 해결",
        "description": "미끄럼방지, 누수방지, 흔들림방지 등",
    },
    {
        "id": "materialSpec",
        "label": "재질/식별 규격",
        "description": "EVA, ABS, PA66, M8, 86형 등",
    },
    {
        "id": "userSituation",
        "label": "사용자/상황",
        "description": "DIY, 수리, 시공 등",
    },
    {
        "id": "synonyms",
        "label": "동의어/다른 명칭",
        "description": "현장명, 별칭, 다른 표기",
    },
]

SALES_MARKETS = ["네이버", "쿠팡", "롯데ON", "11번가", "ESM"]
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
KEYWORD_PRODUCT_CHUNK_SIZE = 5
KEYWORD_CODEX_STALL_SECONDS = 600
KEYWORD_CODEX_MAX_WORKERS = 2
MARKET_TITLE_TARGET_MIN = 30
MARKET_TITLE_TARGET_MAX = 40
NAVER_TITLE_TARGET_MAX = 40

COMPOUND_SPACING_RULES = [
    ("카라비너릴고리", "카라비너 릴고리"),
    ("와이어릴고리", "와이어 릴고리"),
    ("릴고리카라비너", "릴고리 카라비너"),
    ("카라비너릴홀더", "카라비너 릴홀더"),
    ("릴홀더와이어", "릴홀더 와이어"),
    ("와이어릴홀더", "와이어 릴홀더"),
    ("카라비너클립", "카라비너 클립"),
    ("백팩카라비너", "백팩 카라비너"),
    ("ABS카라비너", "ABS 카라비너"),
    ("카라비너고리", "카라비너 고리"),
    ("카라비너열쇠고리", "카라비너 열쇠고리"),
    ("카라비너가방고리", "카라비너 가방고리"),
    ("카라비너와이어", "카라비너 와이어"),
    ("키고리릴", "키고리 릴"),
    ("쿠션깔창", "쿠션 깔창"),
    ("신발깔창", "신발 깔창"),
    ("소프트깔창", "소프트 깔창"),
    ("쿠션인솔", "쿠션 인솔"),
    ("보강패드", "보강 패드"),
    ("충격완화", "충격 완화"),
    ("미끄럼방지", "미끄럼 방지"),
    ("먼지방지", "먼지 방지"),
    ("방수방진", "방수 방진"),
    ("열기차단", "열기 차단"),
    ("햇빛차단", "햇빛 차단"),
    ("사계절보호", "사계절 보호"),
]

BANNED_MARKETING_TERMS = [
    "발편한", "발 편한", "편한발", "편한 발",
    "무료배송", "할인", "세일", "특가", "추천", "인기", "베스트", "핫템",
    "가성비", "저렴한", "예쁜", "프리미엄", "고급", "고급형", "최고급",
    "판매", "가격", "문의", "상담",
    "입니다", "되어드립니다", "보호하세요", "있을",
]

LOW_QUALITY_KEYWORD_TITLE_PATTERNS = [
    r"\*\*",
    r"#",
    r"가이드가이드",
    r"연결찰탁",
    r"물막이소",
    r"스티커입니다",
    r"브라켓입니다",
    r"되어드립니다",
    r"사용자\s*현장용",
    r"생활\s*현장용",
    r"교체\s*보수\s*DIY$",
    r"있을$",
    r"^[A-Z]\s*\d",
]

LISTING_IMAGE_COLUMNS = [
    "이미지등록(목록)",
    "목록이미지",
    "대표이미지",
    "대표 이미지",
    "상품이미지",
    "상품 이미지",
    "이미지",
    "썸네일",
]

ADDITIONAL_IMAGE_COLUMNS = [
    "이미지등록(추가)",
    "추가이미지",
    "추가 이미지",
    "권장이미지",
    "권장 이미지",
    "부가이미지",
]

DETAIL_IMAGE_COLUMNS = [
    "이미지등록(상세)",
    "상세이미지",
    "상품 상세설명",
    "모바일 상품 상세설명",
    "상세설명",
    "상품상세설명",
]

DETAIL_HTML_COLUMNS = [
    "상품 상세설명",
    "모바일 상품 상세설명",
    "상세설명",
    "상품상세설명",
]

SEED_ANALYSIS_POLICY = {
    "purpose": "1차 시드는 원본 상품을 마켓별 작업 전에 정리하는 기준 데이터셋이다.",
    "keywordCategories": KEYWORD_POOL_CATEGORIES,
    "candidateOrder": [category["id"] for category in KEYWORD_POOL_CATEGORIES],
    "inputPriority": [
        "원본 상품명과 GS코드",
        "원본 옵션 컬럼과 optionItems",
        "OCR/사진 분석에서 상품과 직접 연결되는 사실",
        "이전 V파이프라인 키워드 후보",
    ],
    "searchTermRule": {
        "productName": "상품명은 표준명과 대표 검색어 중심으로 정확하게 만든다.",
        "searchTerms": "검색어설정은 표준어, 현장명, 별칭을 넓게 섞되 무관 인기어와 일부러 만든 오타는 넣지 않는다.",
        "marketSplit": "Cafe24는 실제 판매 채널이 아니라 네이버/쿠팡/ESM/11번가/롯데ON으로 나누기 전의 공통 키워드 풀로 본다.",
        "synonymExamples": {
            "나사": ["피스", "볼트", "체결나사"],
            "앵커": ["앙카", "칼블럭", "벽고정"],
            "육각렌치": ["알렌키", "L렌치", "렌치"],
            "소켓": ["복스알", "복스", "소켓렌치"],
            "라쳇핸들": ["깔깔이", "라쳇렌치"],
            "몽키스패너": ["몽키", "조절렌치"],
            "타정기": ["타카", "에어타카"],
            "그라인더": ["핸드그라인더", "절단기"],
            "드라이버": ["도라이바", "십자드라이버", "일자드라이버"],
            "콘센트함 고정나사": ["항공나사", "고정핀", "스위치박스 나사"],
        },
    },
    "sizeQuantityRule": {
        "quantity": "수량/구성은 검색어에 유지 가능하다. 예: 100p, 300p, 10개, 3세트, 5매, 2입",
        "singleSpec": "단일 규격 사이즈와 상품 식별 규격은 유지 가능하다. 예: 1M, 2mm, 35mm, M8, 86형, PA66, 304, ABS, EVA",
        "optionSize": "옵션형 상품의 숫자/색상은 원본 옵션 컬럼을 기준으로 판단한다. OCR에서만 나온 숫자 규격은 상품명/검색어/태그에서 제외한다.",
        "optionRange": "옵션 숫자가 230/240/250/260/270처럼 명확하면 필요 시 230-270 사이즈 선택형처럼 압축하고, 전체 옵션값을 나열하지 않는다.",
        "colorOption": "색상 옵션은 상품명에 넣지 않고 옵션 컬럼에서 처리한다.",
    },
    "rejectRule": {
        "ocrNoise": "OCR 눈금, 이미지 배경 숫자, 배송/상담/주소/전화/판매자 안내 문구는 후보에서 제외한다.",
        "salesNoise": "무료배송, 급배송, 할인, 추천, 가격, 인기, 베스트, 문의, 상담 등 판매 문구는 제외한다.",
        "duplication": "같은 뜻의 단어를 반복하거나 같은 구를 두 번 붙인 후보는 제외한다.",
    },
    "imageRule": {
        "listingSize": "대표/추가 이미지는 1000x1000 기준으로 가공한다.",
        "representativeDefault": "별도 선택 전에는 1번 이미지를 대표이미지로 둔다.",
    },
}


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_name(name: str, fallback: str = "upload.bin") -> str:
    base = Path(name or fallback).name
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", base).strip(" .")
    return base or fallback


def read_json(path: Path, default: dict | None = None) -> dict:
    if not path.exists():
        return default or {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default or {}


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{os.getpid()}.{threading.get_ident()}.tmp")
    try:
        tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        last_error: Exception | None = None
        for attempt in range(8):
            try:
                os.replace(tmp, path)
                last_error = None
                break
            except PermissionError as exc:
                last_error = exc
                time.sleep(0.05 * (attempt + 1))
        if last_error is not None:
            raise last_error
    finally:
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception:
            pass


def collect_upload_history(limit_jobs: int = 80) -> list[dict]:
    entries: dict[str, dict] = {}
    job_paths = sorted(JOBS_ROOT.glob("*.json"), key=lambda path: path.stat().st_mtime, reverse=True)[:limit_jobs]
    for job_path in reversed(job_paths):
        job = read_json(job_path, {})
        if not isinstance(job, dict):
            continue
        if job.get("action") != "marketUpload":
            continue
        results = []
        if isinstance(job.get("results"), list):
            results.extend(job.get("results") or [])
        result_payload = job.get("result") if isinstance(job.get("result"), dict) else {}
        if isinstance(result_payload.get("results"), list):
            results.extend(result_payload.get("results") or [])
        updated_at = job.get("finishedAt") or job.get("updatedAt") or datetime.fromtimestamp(job_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        for item in results:
            if not isinstance(item, dict):
                continue
            channel = text_value(item.get("channel") or item.get("channelKey"))
            gs = text_value(item.get("gs")).upper()
            if not channel or not gs:
                continue
            key = f"{channel}:{gs}"
            normalised = dict(item)
            normalised["channelKey"] = channel
            normalised["historyKey"] = key
            normalised["gs"] = gs
            normalised["productId"] = text_value(item.get("productId") or item.get("sellerProductId") or item.get("spdNo"))
            status = text_value(item.get("status") or "failed") or "failed"
            if status == "skipped" and normalised["productId"]:
                status = "uploaded"
            normalised["status"] = status
            normalised["updatedAt"] = text_value(item.get("updatedAt") or updated_at)
            normalised["jobId"] = job_path.stem
            entries[key] = normalised
    return list(entries.values())


def is_within(parent: Path, child: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def runtime_web_url(path: Path) -> str:
    resolved = path.resolve()
    if not is_within(ROOT, resolved):
        return str(resolved)
    relative = resolved.relative_to(ROOT).as_posix()
    return "/" + urllib.parse.quote(relative, safe="/:._-")


def image_file_sort_key(path: Path) -> tuple[int, str]:
    match = re.search(r"_(\d+)(?:\.[^.]+)$", path.name)
    if match:
        return int(match.group(1)), path.name.lower()
    numbers = re.findall(r"\d+", path.stem)
    return int(numbers[-1]) if numbers else 9999, path.name.lower()


def resolve_seed_path(value: str) -> Path:
    raw = text_value(value)
    if not raw:
        raise ValueError("seed path missing")
    path = Path(raw)
    if not path.is_absolute():
        path = SEED_ROOT / safe_name(raw)
    path = path.resolve()
    if not is_within(SEED_ROOT, path):
        raise ValueError("seed path outside data/seeds")
    if path.suffix.lower() != ".json" or ".webseed" not in path.name.lower():
        raise ValueError("not a webseed json file")
    return path


def seed_progress_summary(payload: dict, upload_index: dict[str, list[dict]] | None = None) -> dict:
    products = payload.get("products") if isinstance(payload.get("products"), list) else []
    product_count = len(products)
    gs_codes = [text_value(item.get("gs")).upper() for item in products if isinstance(item, dict) and text_value(item.get("gs"))]
    keyword_products = 0
    keyword_channels = 0
    for product in products:
        if not isinstance(product, dict):
            continue
        market_keywords = product.get("marketKeywords") if isinstance(product.get("marketKeywords"), dict) else {}
        filled = [
            key for key, value in market_keywords.items()
            if isinstance(value, dict) and (text_value(value.get("title")) or text_value(value.get("searchTerms")) or value.get("tags"))
        ]
        if filled:
            keyword_products += 1
            keyword_channels += len(filled)

    upload_items: list[dict] = []
    if upload_index:
        for gs in gs_codes:
            upload_items.extend(upload_index.get(gs, []))
    upload_total = len(upload_items)
    upload_done = sum(1 for item in upload_items if text_value(item.get("status")) in {"uploaded", "skipped", "exported"})
    upload_failed = sum(1 for item in upload_items if text_value(item.get("status")) == "failed")
    by_channel: dict[str, dict] = {}
    for item in upload_items:
        channel = text_value(item.get("channelKey") or item.get("channel"))
        if not channel:
            continue
        bucket = by_channel.setdefault(channel, {"total": 0, "done": 0, "failed": 0, "status": "대기"})
        bucket["total"] += 1
        status = text_value(item.get("status"))
        if status in {"uploaded", "skipped", "exported"}:
            bucket["done"] += 1
        if status == "failed":
            bucket["failed"] += 1
    for bucket in by_channel.values():
        if bucket["failed"]:
            bucket["status"] = "실패 있음"
        elif bucket["done"] and bucket["done"] >= bucket["total"]:
            bucket["status"] = "완료"
        elif bucket["done"]:
            bucket["status"] = "일부 완료"

    if upload_failed:
        badge = f"실패 {upload_failed} · 업로드 {upload_done}/{upload_total}"
        state = "failed"
    elif upload_total and upload_done >= upload_total:
        badge = f"업로드 완료 {upload_done}/{upload_total}"
        state = "uploaded"
    elif upload_total:
        badge = f"업로드 {upload_done}/{upload_total}"
        state = "partial-upload"
    elif product_count and keyword_products >= product_count:
        badge = "키워드 완료"
        state = "keyworded"
    else:
        badge = "전처리 완료"
        state = "seeded"

    return {
        "state": state,
        "badge": badge,
        "productCount": product_count,
        "keywordProducts": keyword_products,
        "keywordChannels": keyword_channels,
        "uploadTotal": upload_total,
        "uploadDone": upload_done,
        "uploadFailed": upload_failed,
        "channels": by_channel,
        "sourceFile": text_value((payload.get("sourceFilter") or {}).get("sourceFile")) if isinstance(payload.get("sourceFilter"), dict) else "",
        "pipelineOutput": text_value((payload.get("pipelineResult") or {}).get("output_file")) if isinstance(payload.get("pipelineResult"), dict) else "",
    }


def seed_summary(path: Path, upload_index: dict[str, list[dict]] | None = None) -> dict:
    payload = hydrate_seed_payload(read_json(path, {}))
    products = payload.get("products") if isinstance(payload.get("products"), list) else []
    first = products[0] if products else {}
    images = first.get("images", {}) if isinstance(first, dict) else {}
    return {
        "id": path.stem,
        "name": path.name,
        "createdAt": payload.get("createdAt") or datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "rows": payload.get("sourceFilter", {}).get("filteredRows", len(products)) if isinstance(payload.get("sourceFilter"), dict) else len(products),
        "gsCodes": len(products),
        "size": path.stat().st_size,
        "thumbnail": images.get("representative") or images.get("sourceThumb") or "",
        "path": str(path),
        "progress": seed_progress_summary(payload, upload_index),
    }


def list_seed_summaries() -> list[dict]:
    items = []
    upload_index: dict[str, list[dict]] = {}
    for item in collect_upload_history():
        gs = text_value(item.get("gs")).upper()
        if gs:
            upload_index.setdefault(gs, []).append(item)
    for path in sorted(SEED_ROOT.glob("*.webseed.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            items.append(seed_summary(path, upload_index))
        except Exception:
            continue
    return items


def market_key_id(account: str, market: str) -> str:
    return f"{text_value(account).upper()}:{text_value(market)}"


ACCOUNT_KEY_DIR_ALIASES = {
    "A": "홈런",
    "B": "준비",
}


def key_account_dir_name(account: str) -> str:
    clean = text_value(account).upper()
    return ACCOUNT_KEY_DIR_ALIASES.get(clean, safe_name(clean or account or "A"))


def account_key_dir_candidates(account: str) -> list[str]:
    clean = text_value(account)
    names = [key_account_dir_name(clean), safe_name(clean), safe_name(clean.lower())]
    deduped: list[str] = []
    for name in names:
        if name and name not in deduped:
            deduped.append(name)
    return deduped


def market_dir_candidates(market: str) -> list[str]:
    clean = text_value(market)
    names = [safe_name(clean)]
    if clean == "롯데ON":
        names.append("롯데온")
    deduped: list[str] = []
    for name in names:
        if name and name not in deduped:
            deduped.append(name)
    return deduped


def organized_market_key_dir(account: str, market: str) -> Path:
    return DESKTOP_KEY_ROOT / key_account_dir_name(account) / safe_name(market)


def read_market_key_settings() -> dict:
    payload = read_json(MARKET_KEY_SETTINGS, {"items": {}})
    if not isinstance(payload.get("items"), dict):
        payload["items"] = {}
    return payload


def write_market_key_settings(payload: dict) -> None:
    write_json(MARKET_KEY_SETTINGS, payload)


def is_allowed_key_path(path: Path) -> bool:
    return is_within(MARKET_KEY_ROOT, path) or is_within(DESKTOP_KEY_ROOT, path)


def market_key_candidate_paths(item: dict) -> list[Path]:
    account = text_value(item.get("account"))
    market = text_value(item.get("market"))
    file_name = safe_name(item.get("fileName", ""))
    candidates: list[Path] = []

    raw_path = text_value(item.get("path"))
    if raw_path:
        candidates.append(Path(raw_path).resolve())

    if file_name:
        for account_dir in account_key_dir_candidates(account):
            for market_dir in market_dir_candidates(market):
                candidates.append((DESKTOP_KEY_ROOT / account_dir / market_dir / file_name).resolve())
                candidates.append((MARKET_KEY_ROOT / account_dir / market_dir / file_name).resolve())
        candidates.append((DESKTOP_KEY_ROOT / file_name).resolve())

    deduped: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        key = str(path).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(path)
    return deduped


def resolve_market_key_item_path(item: dict) -> Path | None:
    for path in market_key_candidate_paths(item):
        if path.is_file() and is_allowed_key_path(path):
            return path
    return None


def market_key_summaries() -> list[dict]:
    payload = read_market_key_settings()
    items: list[dict] = []
    for key, item in payload.get("items", {}).items():
        path = resolve_market_key_item_path(item)
        exists = path is not None
        items.append({
            "key": key,
            "account": item.get("account", ""),
            "market": item.get("market", ""),
            "mode": item.get("mode", "key"),
            "fileName": item.get("fileName", ""),
            "size": path.stat().st_size if exists else 0,
            "updatedAt": item.get("updatedAt", ""),
            "exists": exists,
            "path": str(path) if path is not None else text_value(item.get("path")),
        })
    return sorted(items, key=lambda item: (item.get("market", ""), item.get("account", "")))


def read_secret_payload(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8-sig", errors="replace").strip()
    out: dict[str, str] = {}
    try:
        raw = json.loads(text)
        if isinstance(raw, dict):
            for key, value in raw.items():
                if isinstance(value, (str, int, float, bool)):
                    out[str(key)] = str(value)
                elif isinstance(value, dict):
                    for child_key, child_value in value.items():
                        if isinstance(child_value, (str, int, float, bool)):
                            out[str(child_key)] = str(child_value)
                            out[f"{key}.{child_key}"] = str(child_value)
    except Exception:
        pass
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        out[key.strip()] = value.strip().strip('"').strip("'")
    if not out and text and len(text) < 400:
        out["value"] = text
    return out


def pick_secret(payload: dict[str, str], *names: str) -> str:
    lower = {str(key).lower(): value for key, value in payload.items()}
    for name in names:
        if payload.get(name):
            return payload[name]
        value = lower.get(name.lower())
        if value:
            return value
    return ""


def scrub_secret(value: object, limit: int = 180) -> str:
    text = str(value or "")
    text = re.sub(r"(?i)(access[_-]?token|refresh[_-]?token|client[_-]?secret|secret[_-]?key|api[_-]?key|signature)[^,}\n]*", r"\1=***", text)
    text = re.sub(r"Bearer\s+[A-Za-z0-9._~+/=-]+", "Bearer ***", text)
    return text[:limit]


def request_text(method: str, url: str, headers: dict | None = None, body: bytes | None = None, timeout: int = 20) -> tuple[bool, int | str, str]:
    req = urllib.request.Request(url, data=body, headers=headers or {})
    req.get_method = lambda: method
    ctx = ssl.create_default_context()
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=timeout) as resp:
            return True, resp.status, resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        return False, exc.code, exc.read().decode("utf-8", errors="replace")
    except Exception as exc:
        return False, "ERR", str(exc)


def set_secret_alias(raw: dict, aliases: list[str], value: str) -> None:
    for alias in aliases:
        if alias in raw:
            raw[alias] = value
            return
    raw[aliases[0]] = value


def refresh_cafe24_access_token(path: Path, cfg: dict[str, str], mall_id: str) -> tuple[bool, str]:
    refresh_token = pick_secret(cfg, "REFRESH_TOKEN", "refresh_token", "refreshToken")
    client_id = pick_secret(cfg, "CLIENT_ID", "client_id", "clientId")
    client_secret = pick_secret(cfg, "CLIENT_SECRET", "client_secret", "clientSecret")
    if not refresh_token or not client_id or not client_secret:
        return False, "refresh_token/client 정보 없음"
    basic = base64.b64encode(f"{client_id}:{client_secret}".encode("utf-8")).decode("ascii")
    body = urllib.parse.urlencode({
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
    }).encode("utf-8")
    ok, status, response = request_text("POST", f"https://{mall_id}.cafe24api.com/api/v2/oauth/token", {
        "Authorization": f"Basic {basic}",
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
    }, body)
    if not ok:
        return False, f"토큰 갱신 실패 status={status}: {scrub_secret(response)}"
    try:
        token_payload = json.loads(response)
    except Exception:
        return False, "토큰 갱신 응답 파싱 실패"
    access_token = token_payload.get("access_token")
    if not access_token:
        return False, "토큰 갱신 응답에 access_token 없음"
    try:
        raw = json.loads(path.read_text(encoding="utf-8-sig", errors="replace"))
        if isinstance(raw, dict):
            set_secret_alias(raw, ["AccessToken", "ACCESS_TOKEN", "access_token"], access_token)
            if token_payload.get("refresh_token"):
                set_secret_alias(raw, ["RefreshToken", "REFRESH_TOKEN", "refresh_token"], token_payload["refresh_token"])
            set_secret_alias(raw, ["UpdatedAt", "UPDATED_AT", "updated_at"], now_text())
            path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        return True, access_token
    return True, access_token


def refresh_cafe24_upload_token(path: Path) -> tuple[bool, str]:
    cfg = read_secret_payload(path)
    mall_id = pick_secret(cfg, "MALL_ID", "mall_id", "mallId")
    if not mall_id:
        return False, "MALL_ID 없음"
    refreshed, message = refresh_cafe24_access_token(path, cfg, mall_id)
    if refreshed:
        return True, "토큰 갱신 완료"
    return False, message


def test_cafe24_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    mall_id = pick_secret(cfg, "MALL_ID", "mall_id", "mallId")
    token = pick_secret(cfg, "ACCESS_TOKEN", "access_token", "accessToken")
    api_version = pick_secret(cfg, "API_VERSION", "api_version", "apiVersion") or "2025-12-01"
    shop_no = pick_secret(cfg, "SHOP_NO", "shop_no", "shopNo") or "1"
    if not mall_id or not token:
        return False, "CONFIG", "MALL_ID/ACCESS_TOKEN 없음"
    query = urllib.parse.urlencode({"shop_no": shop_no, "limit": 1})
    url = f"https://{mall_id}.cafe24api.com/api/v2/admin/products?{query}"
    ok, status, body = request_text("GET", url, {
        "Authorization": f"Bearer {token}",
        "X-Cafe24-Api-Version": api_version,
        "Accept": "application/json",
    })
    if not ok and status == 401:
        refreshed, new_token_or_message = refresh_cafe24_access_token(path, cfg, mall_id)
        if refreshed:
            ok, status, body = request_text("GET", url, {
                "Authorization": f"Bearer {new_token_or_message}",
                "X-Cafe24-Api-Version": api_version,
                "Accept": "application/json",
            })
            return ok, status, "token 갱신 후 products 조회 OK" if ok else scrub_secret(body)
        return False, status, new_token_or_message
    return ok, status, "products 조회 OK" if ok else scrub_secret(body)


def test_naver_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    client_id = pick_secret(cfg, "NAVER_COMMERCE_CLIENT_ID", "CLIENT_ID", "client_id")
    client_secret = pick_secret(cfg, "NAVER_COMMERCE_CLIENT_SECRET", "CLIENT_SECRET", "client_secret")
    if not client_id or not client_secret:
        return False, "CONFIG", "CLIENT_ID/CLIENT_SECRET 없음"
    try:
        import bcrypt
    except Exception as exc:
        return False, "LOCAL", f"bcrypt 모듈 없음: {exc}"
    timestamp = int((time.time() - 3) * 1000)
    sign = base64.standard_b64encode(
        bcrypt.hashpw(f"{client_id}_{timestamp}".encode("utf-8"), client_secret.encode("utf-8"))
    ).decode("utf-8")
    body = urllib.parse.urlencode({
        "client_id": client_id,
        "timestamp": timestamp,
        "client_secret_sign": sign,
        "grant_type": "client_credentials",
        "type": "SELF",
    }).encode("utf-8")
    ok, status, response = request_text("POST", "https://api.commerce.naver.com/external/v1/oauth2/token", {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
    }, body)
    return ok, status, "token 발급 OK" if ok else scrub_secret(response)


def coupang_auth(access_key: str, secret_key: str, method: str, path: str, query: str = "") -> str:
    stamp = datetime.utcnow().strftime("%y%m%dT%H%M%SZ")
    message = stamp + method + path + query
    signature = hmac.new(secret_key.encode("utf-8"), message.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"CEA algorithm=HmacSHA256, access-key={access_key}, signed-date={stamp}, signature={signature}"


def test_coupang_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    access_key = pick_secret(cfg, "access_key", "ACCESS_KEY", "COUPANG_ACCESS_KEY")
    secret_key = pick_secret(cfg, "secret_key", "SECRET_KEY", "COUPANG_SECRET_KEY")
    if not access_key or not secret_key:
        return False, "CONFIG", "access_key/secret_key 없음"
    api_path = "/v2/providers/seller_api/apis/api/v1/marketplace/meta/category-related-metas/display-category-codes/56137"
    ok, status, body = request_text("GET", "https://api-gateway.coupang.com" + api_path, {
        "Authorization": coupang_auth(access_key, secret_key, "GET", api_path),
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json",
    })
    return ok, status, "category meta 조회 OK" if ok else scrub_secret(body)


def test_lotteon_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    api_key = pick_secret(cfg, "api_key", "LOTTEON_API_KEY", "ApiKey", "apiKey", "value")
    if not api_key:
        return False, "CONFIG", "LOTTEON_API_KEY/api_key 없음"
    ok, status, body = request_text("GET", "https://openapi.lotteon.com/v1/openapi/common/v1/identity", {
        "Authorization": f"Bearer {api_key}",
        "Accept": "application/json",
    })
    return ok, status, "identity 조회 OK" if ok else scrub_secret(body)


def test_market_key(item: dict) -> dict:
    path = resolve_market_key_item_path(item)
    market = item.get("market", "")
    result = {
        "key": market_key_id(item.get("account", ""), market),
        "account": item.get("account", ""),
        "market": market,
        "fileName": item.get("fileName", ""),
        "ok": False,
        "status": "SKIP",
        "message": "",
        "testedAt": now_text(),
    }
    if path is None:
        result.update({"status": "MISSING", "message": "저장된 파일 없음"})
        return result
    testers = {
        "Cafe24": test_cafe24_key,
        "네이버": test_naver_key,
        "쿠팡": test_coupang_key,
        "롯데ON": test_lotteon_key,
    }
    tester = testers.get(market)
    if not tester:
        result.update({"status": "SKIP", "message": "엑셀 서식 마켓"})
        return result
    ok, status, message = tester(path)
    result.update({"ok": ok, "status": status, "message": message})
    return result


def infer_seed_progress(line: str, fallback: int = 12) -> tuple[int, str, str]:
    clean = text_value(line)
    progress = fallback
    stage = "파이프라인 실행 중"
    current_gs = ""
    gs_match = re.search(r"(GS\d{7}[A-Z0-9]*)", clean, re.IGNORECASE)
    if gs_match:
        current_gs = gs_match.group(1).upper()
    count_match = re.search(r"\[(\d+)\s*/\s*(\d+)\]", clean)
    if count_match:
        done = int(count_match.group(1))
        total = max(1, int(count_match.group(2)))
        progress = max(progress, min(86, 18 + int((done / total) * 60)))
    lowered = clean.lower()
    if "다운로드" in clean or "download" in lowered:
        stage = "이미지 다운로드"
        progress = max(progress, 24)
    if "ocr" in lowered:
        stage = "OCR 분석"
        progress = max(progress, 45)
    if "gpt" in lowered or "사진" in clean or "분석" in clean:
        stage = "사진/상품 분석"
        progress = max(progress, 58)
    if "엑셀" in clean or "excel" in lowered:
        stage = "기본 데이터 정리"
        progress = max(progress, 72)
    if clean.startswith("__RESULT__"):
        stage = "결과 파일 수집"
        progress = 88
    return progress, stage, current_gs


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def text_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def price_value(value: object) -> int:
    text = text_value(value)
    text = re.sub(r"[^\d.-]", "", text)
    if not text:
        return 0
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def v3_price_multiplier(value: object) -> float:
    try:
        amount = float(value)
    except Exception:
        return 2.0
    if amount >= 20000:
        return 1.6
    if amount >= 10000:
        return 1.8
    return 2.0


def round_100(value: object) -> int:
    try:
        return int(round(float(value), -2))
    except Exception:
        return 0


def source_price_info(row: dict, headers: dict[str, str]) -> dict[str, int]:
    supply = first_price(row, headers, ["공급가", "공급 가격", "공급가격", "매입가"])
    raw_sale = first_price(row, headers, ["판매가", "상품가", "가격"])
    raw_consumer = first_price(row, headers, ["소비자가", "정가"])
    if supply:
        sale = round_100(supply * v3_price_multiplier(supply))
        consumer = round_100(sale * 1.2)
    else:
        sale = raw_sale or raw_consumer
        consumer = raw_consumer or (round_100(sale * 1.2) if sale else 0)
    return {
        "supplyPrice": supply,
        "salePrice": sale,
        "consumerPrice": consumer,
    }


def first_value(row: dict, headers: dict[str, str], names: list[str]) -> str:
    for name in names:
        key = headers.get(normalize_header(name))
        if key is not None:
            value = text_value(row.get(key, ""))
            if value:
                return value
    return ""


def first_price(row: dict, headers: dict[str, str], names: list[str]) -> int:
    for name in names:
        key = headers.get(normalize_header(name))
        if key is not None:
            value = price_value(row.get(key, ""))
            if value:
                return value
    return 0


GS_CODE_RE = re.compile(r"(GS\d{7})([A-Z0-9]*)", re.IGNORECASE)


def extract_gs(*values: object) -> str:
    for value in values:
        match = GS_CODE_RE.search(text_value(value))
        if match:
            return match.group(0).upper()
    return ""


def split_gs(gs: str) -> tuple[str, str]:
    match = GS_CODE_RE.search(gs or "")
    if not match:
        return gs.upper(), ""
    return match.group(1).upper(), (match.group(2) or "").upper()


def short_name(name: str, gs: str) -> str:
    text = GS_CODE_RE.sub("", name or "")
    text = re.sub(r"\s+", " ", text).strip(" -_/|")
    return text or name or gs


def common_word_prefix(values: list[str]) -> str:
    token_sets = [value.split() for value in values if value]
    if len(token_sets) < 2:
        return ""
    prefix: list[str] = []
    for tokens in zip(*token_sets):
        if len(set(tokens)) != 1:
            break
        prefix.append(tokens[0])
    return " ".join(prefix).strip()


def strip_option_label(name: str, base_name: str) -> str:
    if base_name and name.startswith(base_name):
        return name[len(base_name):].strip(" -_/|")
    return ""


def compact_option_summary(labels: list[str], total: int) -> str:
    labels = [label for label in labels if label and label != "단일"]
    if total <= 1:
        return labels[0] if labels else "단일"
    if not labels:
        return f"{total} 옵션"
    shown = " / ".join(labels[:8])
    more = f" 외 {len(labels) - 8}" if len(labels) > 8 else ""
    return f"{total} 옵션 · {shown}{more}"


def option_label_with_index(label: object, index: int) -> str:
    cleaned = re.sub(r"\s+", " ", text_value(label)).strip(" ,/|")
    cleaned = re.sub(r"^옵션\s*\{|\}$", "", cleaned).strip(" ,/|")
    if not cleaned or cleaned == "단일":
        return ""
    if re.match(r"^[A-Z]\s+", cleaned, re.IGNORECASE):
        return cleaned[0].upper() + cleaned[1:]
    prefix = chr(65 + index) if index < 26 else f"OPT{index + 1}"
    return f"{prefix} {cleaned}"


def option_labels_from_input(value: object) -> list[str]:
    text = text_value(value)
    if not text or "단일" in text:
        return []
    if "·" in text:
        text = text.split("·", 1)[1]
    match = re.search(r"옵션\s*\{(.+?)\}", text, re.IGNORECASE)
    body = match.group(1) if match else text
    body = re.sub(r"^\s*\d+\s*옵션\s*[:：-]?\s*", "", body).strip()
    parts = [
        re.sub(r"\s+", " ", part).strip(" ,/|")
        for part in re.split(r"\s*\|\s*|\s*/\s*|,\s*", body)
    ]
    return [
        option_label_with_index(part, index)
        for index, part in enumerate(parts)
        if option_label_with_index(part, index)
    ]


def option_input_from_labels(labels: list[str]) -> str:
    clean_labels = [option_label_with_index(label, index) for index, label in enumerate(labels)]
    clean_labels = [label for label in clean_labels if label]
    return "옵션{" + "|".join(clean_labels[:26]) + "}" if clean_labels else ""


def option_price_list(value: object) -> list[int]:
    if isinstance(value, list):
        return [price_value(item) for item in value]
    text = text_value(value)
    if not text:
        return []
    return [price_value(part) for part in re.split(r"[,|\n]+", text) if text_value(part)]


def pm_option_items_for_base(base_gs: str) -> list[dict]:
    base_value, _suffix = split_gs(base_gs)
    if not base_value:
        return []
    conn = pm_db()
    if not conn:
        return []
    try:
        rows = conn.execute("""
            SELECT product_code, sku_group, supplier_code, option_code,
                   product_name, price, image_url, raw_data
            FROM products
            WHERE sku_group = ?
            ORDER BY option_code, product_code
        """, (base_value,)).fetchall()
    except Exception:
        return []
    finally:
        conn.close()

    products = [dict(row) for row in rows]
    products = sorted(products, key=pm_option_sort_key)
    option_items: list[dict] = []
    for product in products:
        raw: dict = {}
        try:
            raw = json.loads(product.get("raw_data") or "{}")
        except Exception:
            raw = {}
        headers = {normalize_header(col): col for col in raw.keys()}
        price_info = source_price_info(raw, headers) if raw else {
            "supplyPrice": 0,
            "salePrice": price_value(product.get("price")),
            "consumerPrice": 0,
        }
        label = pm_option_label(product)
        option_items.append({
            "gs": product.get("product_code", ""),
            "suffix": product.get("option_code", ""),
            "name": product.get("product_name", ""),
            "option": label,
            "price": price_info.get("salePrice") or price_value(product.get("price")),
            "supplyPrice": price_info.get("supplyPrice", 0),
            "salePrice": price_info.get("salePrice") or price_value(product.get("price")),
            "consumerPrice": price_info.get("consumerPrice", 0),
            "thumb": product.get("image_url", ""),
        })
    return option_items


def option_meta(option_summary: str, option_items: list[dict]) -> dict:
    summary = text_value(option_summary) or "단일"
    match = re.search(r"(\d+)\s*옵션", summary)
    summary_count = int(match.group(1)) if match else 0
    count = max(len(option_items or []), summary_count)
    has_options = count > 1 or (summary != "단일" and bool(re.search(r"옵션|/|,", summary)))
    return {
        "optionType": "option" if has_options else "single",
        "optionCount": max(count, 1) if has_options else 1,
    }


def empty_keyword_pool() -> dict[str, list[str]]:
    return {category["id"]: [] for category in KEYWORD_POOL_CATEGORIES}


TERM_STOPWORDS = {
    "상품", "제품", "옵션", "선택", "수량", "소재", "사이즈", "참고사항", "제조국", "중국",
    "수입사", "굿셀러스", "Product", "Profile", "SIZE", "Advantage", "000", "http", "https",
    "홈런마켓", "급배송", "무료", "묶음배송", "대량구매", "문의", "상담", "문자상담",
    "택배사", "방문수령", "경기도", "광주시", "친절하게", "환영합니다", "주의", "사항",
    "앞면", "뒷면", "참고용", "측정방법", "오차", "실제", "선택하세요",
}

CATEGORY_HINTS = {
    "identity": [
        "카라비너", "릴홀더", "릴고리", "고리", "라벨", "태그", "이름표", "커버", "브러시",
        "장갑", "스티커", "파우치", "클립", "후크", "나사", "브라켓", "정리함",
    ],
    "function": [
        "고정", "보강", "방지", "연결", "수선", "탈부착", "체결", "걸이", "걸어",
        "늘어나는", "간편", "빠르고", "장식", "보관", "수납", "보호", "충전", "거치",
    ],
    "usePlace": [
        "가방", "백팩", "텐트", "스트랩", "벨트", "하네스", "손전등", "열쇠", "신발",
        "콘센트", "스위치", "가구", "차량", "현장", "아웃도어", "일상", "캠핑",
    ],
    "problemSolving": [
        "미끄럼방지", "누수방지", "흔들림방지", "분실방지", "낙하방지", "방지",
        "스크래치", "보호", "오차", "편의", "간편",
    ],
    "userSituation": [
        "DIY", "수리", "시공", "작업", "현장", "아웃도어", "캠핑", "등산", "일상",
        "휴대", "여행", "장갑", "활동",
    ],
}

SYNONYM_SEEDS = {
    "카라비너": ["카라비너고리", "카라비너클립", "카라비너열쇠고리", "카라비너가방고리", "카라비너와이어", "백팩카라비너", "아웃도어카라비너"],
    "릴홀더": ["릴고리", "릴홀더와이어", "와이어릴홀더", "키릴", "키고리릴"],
    "릴고리": ["릴홀더", "와이어고리", "키고리릴"],
    "클립": ["집게", "고정클립", "후크"],
    "나사": ["피스", "볼트", "체결나사"],
    "앵커": ["앙카", "칼블럭", "벽고정"],
    "장갑": ["작업장갑", "현장장갑", "보호장갑"],
}


def unique_terms(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = re.sub(r"\s+", " ", text_value(value)).strip(" ,/|")
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def normalize_term(term: object) -> str:
    text = text_value(term)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[{}\"'`]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,./|·:;_-")
    if not text or len(text) < 2:
        return ""
    if text in TERM_STOPWORDS or text.upper() in TERM_STOPWORDS:
        return ""
    if GS_CODE_RE.fullmatch(text):
        return ""
    if re.fullmatch(r"\d+", text):
        return ""
    if len(text) > 36:
        return ""
    return text


def apply_compound_spacing(value: object) -> str:
    text = text_value(value)
    for source, target in COMPOUND_SPACING_RULES:
        text = text.replace(source, target)
    return re.sub(r"\s+", " ", text).strip()


def remove_marketing_terms(value: object) -> str:
    text = apply_compound_spacing(value)
    for word in BANNED_MARKETING_TERMS:
        text = re.sub(re.escape(word), " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip(" ,/|")


def clean_product_title(value: object) -> str:
    text = remove_marketing_terms(value)
    text = re.sub(r"\b1\s*(?:개|입|매|p|P|pcs|PCS|pc|PC)\b", " ", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip(" ,/|")


def compact_search_keyword(value: object) -> str:
    text = remove_marketing_terms(value)
    text = re.sub(r"\s+", "", text)
    return text.strip(" ,/|")


def split_candidate_terms(*values: object) -> list[str]:
    terms: list[str] = []
    for value in values:
        raw = text_value(value)
        if not raw:
            continue
        raw = re.sub(r"<[^>]+>", " ", raw)
        for chunk in re.split(r"[,/|;·\n\r\t]+", raw):
            chunk = normalize_term(chunk)
            if not chunk:
                continue
            terms.append(chunk)
            words = [normalize_term(word) for word in re.split(r"\s+", chunk)]
            words = [word for word in words if word]
            terms.extend(words)
            for i in range(len(words) - 1):
                joined = normalize_term(f"{words[i]} {words[i + 1]}")
                if joined:
                    terms.append(joined)
    return unique_terms(terms)


NOTICE_COMMON_DEFAULTS = {
    "returnCostReason": "0",
    "noRefundReason": "0",
    "qualityAssuranceStandard": "0",
    "compensationProcedure": "0",
    "troubleShootingContents": "0",
}

NOTICE_REVIEW_DEFAULTS = {
    "warrantyPolicy": "0",
    "afterServiceDirector": "0",
    "caution": "0",
}

NOTICE_TEXT_DEFAULT = "상세페이지 참조"
NOTICE_CERT_DEFAULT = "해당 없음"
NOTICE_AS_DEFAULT = "010-2324-8352"

NOTICE_FIELD_LABELS = [
    "소재", "재질", "수량", "색상", "사이즈", "수입사", "수입원", "제조사",
    "제조자", "제조원", "제조국", "원산지", "A/S", "AS", "고객센터",
]

NOTICE_NOISE_PATTERNS = [
    re.compile(r"^(?:high|bullet|advantage|product\s*profile|size)$", re.IGNORECASE),
    re.compile(r"^(?:in/mm|on|off|zero)$", re.IGNORECASE),
    re.compile(r"^[{}\"']+$"),
    re.compile(r"^\d{1,4}$"),
    re.compile(r"^\d{1,4}\s*mm$", re.IGNORECASE),
]


def clean_notice_value(value: object, max_length: int = 160) -> str:
    text = text_value(value)
    text = re.sub(r"[{}\"`]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,./|·:-")
    text = re.split(
        r"\s+(?:홈런마켓|급배송|평일|택배사|구매대행|국내|모든\s*제품|상품/대량구매|대량구매|퀵서비스|방문수령)\b",
        text,
        maxsplit=1,
    )[0].strip(" ,./|·:-")
    return text[:max_length].strip()


def notice_is_noise_line(value: object) -> bool:
    text = clean_notice_value(value, max_length=80)
    if not text:
        return True
    return any(pattern.search(text) for pattern in NOTICE_NOISE_PATTERNS)


def ocr_notice_lines(raw_text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in text_value(raw_text).splitlines():
        line = clean_notice_value(raw_line, max_length=220)
        if line:
            lines.append(line)
    return lines


def extract_labeled_notice_value(lines: list[str], labels: list[str], stop_labels: list[str] | None = None) -> str:
    stop_labels = stop_labels or NOTICE_FIELD_LABELS
    label_pattern = "|".join(re.escape(label) for label in labels)
    stop_pattern = "|".join(re.escape(label) for label in stop_labels if label not in labels)
    joined = " ".join(lines)
    match = re.search(
        rf"(?:^|\s)(?:{label_pattern})\s*[:：]?\s*(.+?)(?=\s+(?:{stop_pattern})\s*[:：]?|\s+(?:SIZE|Product Profile|홈런마켓|급배송)\b|$)",
        joined,
        re.IGNORECASE,
    )
    if match:
        value = clean_notice_value(match.group(1))
        if value and not notice_is_noise_line(value):
            return value

    exact_pattern = re.compile(rf"^(?:{label_pattern})\s*[:：]?$", re.IGNORECASE)
    inline_pattern = re.compile(rf"^(?:{label_pattern})\s*[:：]?\s*(.+)$", re.IGNORECASE)
    stop_line_pattern = re.compile(rf"^(?:{'|'.join(re.escape(label) for label in stop_labels)})\b", re.IGNORECASE)
    for index, line in enumerate(lines):
        inline = inline_pattern.search(line)
        if inline:
            value = clean_notice_value(inline.group(1))
            if value and not notice_is_noise_line(value):
                return value
        if exact_pattern.search(line):
            for next_line in lines[index + 1:index + 4]:
                if stop_line_pattern.search(next_line):
                    break
                if not notice_is_noise_line(next_line):
                    return clean_notice_value(next_line)
    return ""


def normalize_notice_size(raw_size: str, option_items: list[dict]) -> str:
    option_labels = [
        clean_notice_value(item.get("option"))
        for item in option_items or []
        if clean_notice_value(item.get("option")) and clean_notice_value(item.get("option")) != "단일"
    ]
    if option_labels:
        return " / ".join(unique_terms(option_labels[:20]))

    text = clean_notice_value(raw_size, max_length=240)
    text = re.split(r"\s+-?\s*위\s*\d*종|\s+수입사|\s+제조국|\s+SIZE\b", text, maxsplit=1, flags=re.IGNORECASE)[0]
    numbers = re.findall(r"(?<!\d)(?:[A-Z]\s*)?(\d{2,4})(?!\d)", text, flags=re.IGNORECASE)
    shoe_sizes = [number for number in numbers if 150 <= int(number) <= 350]
    if shoe_sizes:
        return " / ".join(unique_terms(shoe_sizes))
    return clean_notice_value(text)


def extract_notice_dimensions(raw_text: str) -> list[str]:
    values = re.findall(r"\b[A-Z]?\s*\d{2,4}\s*-\s*\d{1,4}\s*mm\s*[xX]\s*\d{1,4}\s*mm\b", raw_text, flags=re.IGNORECASE)
    return unique_terms([clean_notice_value(value).replace(" x ", "X").replace("x", "X") for value in values])[:12]


def infer_notice_type(row: dict, extracted: dict[str, str], raw_text: str) -> tuple[str, str]:
    joined = " ".join([
        text_value(row.get("name")),
        text_value(row.get("sourceName")),
        text_value(row.get("opt")),
        text_value(raw_text),
    ])
    if re.search(r"깔창|인솔|신발|운동화|구두|슬리퍼|부츠", joined):
        return "SHOES", "shoes"
    if re.search(r"가방|백팩|파우치|숄더백|토트백", joined):
        return "BAG", "bag"
    if re.search(r"모자|벨트|액세서리|악세사리|키링|브로치|헤어|머리|집게|핀|고리", joined):
        return "FASHION_ITEMS", "fashionItems"
    if re.search(r"스포츠|운동|헬스|요가|필라테스|테이핑|보호대|밴드|스트랩|고정밴드|발목밴드|공|라켓|골프|등산|자전거", joined):
        return "SPORTS_EQUIPMENT", "sportsEquipment"
    if re.search(r"의류|티셔츠|셔츠|바지|자켓|재킷|점퍼|원피스|스커트", joined):
        return "WEAR", "wear"
    if re.search(r"주방|키친|냄비|프라이팬|후라이팬|식기|컵|접시|조리|칼|도마|수저|주걱", joined):
        return "KITCHEN_UTENSILS", "kitchenUtensils"
    if re.search(r"자동차|차량|차종|세차|와이퍼|타이어|핸들|대시보드|카매트|오토바이", joined):
        return "CAR_ARTICLES", "carArticles"
    if re.search(r"가구|의자|책상|테이블|선반|수납장|침대|소파|브라켓|행거", joined):
        return "FURNITURE", "furniture"
    return "ETC", "etc"


def build_naver_provided_notice(row: dict, ocr_record: dict) -> dict:
    fields = ocr_record.get("fields", {}) if isinstance(ocr_record, dict) else {}
    raw_text = text_value(ocr_record.get("rawText")) or text_value(fields.get("OCR텍스트"))
    lines = ocr_notice_lines(raw_text)
    source_name = clean_product_title(row.get("name")) or clean_notice_value(fields.get("상품명")) or text_value(row.get("gs"))
    option_items = row.get("optionItems", []) if isinstance(row.get("optionItems"), list) else []

    extracted = {
        "itemName": source_name,
        "modelName": text_value(row.get("gs")) or text_value(row.get("baseGs")),
        "material": extract_labeled_notice_value(lines, ["소재", "재질"]),
        "quantity": extract_labeled_notice_value(lines, ["수량"]),
        "color": extract_labeled_notice_value(lines, ["색상"]),
        "size": normalize_notice_size(extract_labeled_notice_value(lines, ["사이즈"]), option_items),
        "importer": extract_labeled_notice_value(lines, ["수입사", "수입원"]),
        "manufacturer": extract_labeled_notice_value(lines, ["제조사", "제조자", "제조원"]),
        "origin": extract_labeled_notice_value(lines, ["제조국", "원산지"]),
        "customerServicePhoneNumber": extract_labeled_notice_value(lines, ["A/S", "AS", "고객센터"]),
    }
    dimensions = extract_notice_dimensions(raw_text)
    if dimensions:
        extracted["sizeDetail"] = " / ".join(dimensions)
    extracted = {key: value for key, value in extracted.items() if value}

    notice_type, object_key = infer_notice_type(row, extracted, raw_text)
    manufacturer = extracted.get("manufacturer") or extracted.get("importer") or NOTICE_TEXT_DEFAULT
    material = extracted.get("material") or NOTICE_TEXT_DEFAULT
    color = extracted.get("color") or NOTICE_TEXT_DEFAULT
    size = extracted.get("size") or extracted.get("sizeDetail") or NOTICE_TEXT_DEFAULT
    model_name = extracted.get("modelName") or text_value(row.get("baseGs")) or text_value(row.get("gs")) or source_name
    as_director = extracted.get("customerServicePhoneNumber") or NOTICE_AS_DEFAULT

    if notice_type == "SHOES":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "material": material,
            "color": color,
            "size": size,
            "height": "해당사항 없음",
            "manufacturer": manufacturer,
        }
    elif notice_type == "WEAR":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "material": material,
            "color": color,
            "size": size,
            "manufacturer": manufacturer,
        }
    elif notice_type == "BAG":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "type": source_name,
            "material": material,
            "color": color,
            "size": size,
            "manufacturer": manufacturer,
        }
    elif notice_type == "FASHION_ITEMS":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "type": source_name,
            "material": material,
            "size": size,
            "manufacturer": manufacturer,
        }
    elif notice_type == "SPORTS_EQUIPMENT":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            "itemName": source_name,
            "modelName": model_name,
            "certificationType": NOTICE_CERT_DEFAULT,
            "size": size,
            "weight": NOTICE_TEXT_DEFAULT,
            "color": color,
            "material": material,
            "components": "본품",
            "releaseDateText": NOTICE_TEXT_DEFAULT,
            "manufacturer": manufacturer,
            "detailContent": NOTICE_TEXT_DEFAULT,
            "warrantyPolicy": "관련 법 및 소비자분쟁해결기준에 따름",
            "afterServiceDirector": as_director,
        }
    elif notice_type == "KITCHEN_UTENSILS":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            "itemName": source_name,
            "modelName": model_name,
            "material": material,
            "component": "본품",
            "size": size,
            "releaseDateText": NOTICE_TEXT_DEFAULT,
            "manufacturer": manufacturer,
            "producer": extracted.get("origin") or NOTICE_TEXT_DEFAULT,
            "importDeclaration": NOTICE_CERT_DEFAULT,
            "warrantyPolicy": "관련 법 및 소비자분쟁해결기준에 따름",
            "afterServiceDirector": as_director,
        }
    elif notice_type == "CAR_ARTICLES":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            "itemName": source_name,
            "modelName": model_name,
            "releaseDateText": NOTICE_TEXT_DEFAULT,
            "certificationType": NOTICE_CERT_DEFAULT,
            "caution": NOTICE_TEXT_DEFAULT,
            "manufacturer": manufacturer,
            "size": size,
            "applyModel": NOTICE_TEXT_DEFAULT,
            "warrantyPolicy": "관련 법 및 소비자분쟁해결기준에 따름",
            "roadWorthyCertification": NOTICE_CERT_DEFAULT,
            "afterServiceDirector": as_director,
        }
    elif notice_type == "FURNITURE":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            "itemName": source_name,
            "certificationType": NOTICE_CERT_DEFAULT,
            "color": color,
            "components": "본품",
            "material": material,
            "manufacturer": manufacturer,
            "importer": extracted.get("importer") or NOTICE_TEXT_DEFAULT,
            "producer": extracted.get("origin") or NOTICE_TEXT_DEFAULT,
            "size": size,
            "installedCharge": NOTICE_TEXT_DEFAULT,
            "warrantyPolicy": "관련 법 및 소비자분쟁해결기준에 따름",
            "refurb": NOTICE_CERT_DEFAULT,
            "afterServiceDirector": as_director,
        }
    else:
        detail = {
            "itemName": source_name,
            "modelName": model_name,
            "certificateDetails": NOTICE_CERT_DEFAULT,
            "manufacturer": manufacturer,
            "customerServicePhoneNumber": as_director,
        }

    needs_review: list[str] = []
    if extracted.get("importer") and not extracted.get("manufacturer"):
        needs_review.append("수입사를 제조자/수입자로 임시 사용했습니다.")
    if not extracted.get("customerServicePhoneNumber"):
        needs_review.append("A/S 전화번호는 OCR에서 명확히 확인되지 않았습니다.")
    if notice_type != "ETC":
        needs_review.append("카테고리 API 연동 전까지 OCR 문구로 고시 상품군을 추정했습니다.")
    if extracted.get("sizeDetail") and extracted.get("size"):
        needs_review.append("상세 치수는 보조값으로만 보관하고 옵션 사이즈를 우선했습니다.")

    matched_fields = {
        key: extracted[key]
        for key in ("material", "quantity", "color", "size", "sizeDetail", "importer", "manufacturer", "origin")
        if extracted.get(key)
    }
    status = "matched" if matched_fields else "empty"
    if needs_review:
        status = "partial"

    return {
        "status": status,
        "source": "ocr_label_match",
        "productInfoProvidedNoticeType": notice_type,
        "objectKey": object_key,
        "productInfoProvidedNotice": {
            "productInfoProvidedNoticeType": notice_type,
            object_key: detail,
        },
        "extractedFields": extracted,
        "matchedFields": matched_fields,
        "needsReview": needs_review,
    }


def find_material_spec_terms(*values: object, include_free_numeric: bool = True) -> list[str]:
    text = " ".join(text_value(value) for value in values)
    patterns = [
        r"\b(?:ABS|EVA|PA66|PVC|PP|PE|PU|SUS|STS)\b",
        r"(?:에어메쉬|폴리에스테르|나일론|스판|면|패브릭|부직포|실리콘|고무)",
        r"\b(?:304|316)\b",
        r"\bM\d+(?:\.\d+)?\b",
        r"\b\d+\s?형\b",
    ]
    if include_free_numeric:
        patterns.append(r"\b\d+(?:\.\d+)?\s?(?:mm|cm|m|M)\b")
    terms: list[str] = []
    for pattern in patterns:
        terms.extend(match.group(0).replace(" ", "") for match in re.finditer(pattern, text, re.IGNORECASE))
    return unique_terms(terms)


def terms_by_hints(terms: list[str], hints: list[str]) -> list[str]:
    return unique_terms([term for term in terms if any(hint.lower() in term.lower() for hint in hints)])


def synonym_terms(terms: list[str]) -> list[str]:
    joined = " ".join(terms)
    out: list[str] = []
    for trigger, synonyms in SYNONYM_SEEDS.items():
        if trigger in joined:
            out.extend(synonyms)
    return unique_terms(out)


def seed_keyword_pool_from_product(row: dict, ocr_text: str = "", keyword_record: dict | None = None) -> dict[str, list[str]]:
    pool = empty_keyword_pool()
    option_terms = [item.get("option", "") for item in row.get("optionItems", []) if item.get("option") != "단일"]
    keyword_record = keyword_record or {}
    generated_product_names = keyword_record.get("productNames", [])
    generated_search_terms = keyword_record.get("searchTerms", [])
    generated_longtails = keyword_record.get("longtails", [])
    generated_debug_terms = keyword_record.get("debugTerms", [])
    all_terms = split_candidate_terms(
        row.get("name", ""),
        row.get("opt", ""),
        " ".join(option_terms),
        ocr_text,
        " ".join(generated_product_names),
        " ".join(generated_search_terms),
        " ".join(generated_longtails),
        " ".join(generated_debug_terms),
    )
    pool["identity"] = unique_terms(
        [row.get("name", "")]
        + generated_product_names[:4]
        + terms_by_hints(all_terms, CATEGORY_HINTS["identity"])
    )[:18]
    pool["function"] = terms_by_hints(all_terms, CATEGORY_HINTS["function"])[:20]
    pool["usePlace"] = terms_by_hints(all_terms, CATEGORY_HINTS["usePlace"])[:20]
    pool["problemSolving"] = terms_by_hints(all_terms, CATEGORY_HINTS["problemSolving"])[:14]
    # 옵션형 상품은 OCR 치수 눈금/이미지 배경 숫자를 상품 규격으로 오인하기 쉽다.
    # 숫자 규격은 옵션/원본명에서만 보고, OCR에서는 재질명 중심으로만 가져온다.
    pool["materialSpec"] = unique_terms(
        find_material_spec_terms(row.get("name", ""), row.get("opt", ""), " ".join(option_terms), include_free_numeric=True)
        + find_material_spec_terms(ocr_text, " ".join(all_terms), include_free_numeric=not bool(option_terms))
    )[:18]
    pool["userSituation"] = terms_by_hints(all_terms, CATEGORY_HINTS["userSituation"])[:16]
    pool["synonyms"] = unique_terms(generated_longtails + synonym_terms(all_terms))[:24]
    return pool


def read_pipeline_ocr_summary(output_file: str, output_root: str = "") -> dict:
    path = Path(output_file or "")
    summary = {
        "available": False,
        "workbook": str(path) if output_file else "",
        "sheets": [],
        "ocrByGs": {},
    }
    if not output_file or not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return summary

    def read_ocr_workbook(workbook_path: Path) -> bool:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            summary["available"] = True
            summary["workbook"] = str(workbook_path)
            summary["sheets"] = list(workbook.sheetnames)
            if "OCR결과" not in workbook.sheetnames:
                return False

            sheet = workbook["OCR결과"]
            rows = sheet.iter_rows(values_only=True)
            headers = [text_value(value) for value in next(rows, [])]
            gs_col = next((i for i, header in enumerate(headers) if "GS" in header.upper()), 0)
            for row in rows:
                values = {headers[i] or f"col{i + 1}": text_value(row[i] if i < len(row) else "") for i in range(len(headers))}
                gs = extract_gs(row[gs_col] if gs_col < len(row) else "", values)
                if not gs:
                    continue
                base_gs, _ = split_gs(gs)
                raw_text_parts = [
                    value for key, value in values.items()
                    if value and not re.search(r"(이미지|image|path|경로|url)", key, re.IGNORECASE)
                ]
                summary["ocrByGs"][base_gs] = {
                    "gs": gs,
                    "rawText": "\n".join(raw_text_parts)[:8000],
                    "fields": values,
                }
            return bool(summary["ocrByGs"])
        finally:
            workbook.close()

    try:
        if read_ocr_workbook(path):
            return summary

        search_roots = []
        if output_root:
            search_roots.append(Path(output_root))
        search_roots.append(path.parent)
        candidates: list[Path] = []
        for root in search_roots:
            if root.is_dir():
                candidates.extend(root.glob("OCR결과_*.xlsx"))
        candidates = sorted(set(candidates), key=lambda item: item.stat().st_mtime, reverse=True)
        for candidate in candidates:
            if read_ocr_workbook(candidate):
                summary["linkedFrom"] = str(path)
                return summary
    except Exception as exc:
        summary["available"] = False
        summary["error"] = str(exc)
    return summary


def append_keyword_record(target: dict, base_gs: str, field: str, values: object) -> None:
    if not base_gs:
        return
    record = target.setdefault(base_gs, {
        "productNames": [],
        "searchTerms": [],
        "longtails": [],
        "debugTerms": [],
    })
    record[field] = unique_terms(record.get(field, []) + split_candidate_terms(values))


def read_pipeline_keyword_summary(output_file: str) -> dict:
    path = Path(output_file or "")
    summary = {
        "available": False,
        "workbook": str(path) if output_file else "",
        "keywordByGs": {},
    }
    if not output_file or not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return summary
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            pending: list[tuple[str, object]] = []
            for sheet_name in workbook.sheetnames:
                if sheet_name not in {"분리추출후", "B마켓", "디버그"}:
                    continue
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                headers = [text_value(value) for value in next(rows, [])]
                if not headers:
                    continue
                for row in rows:
                    values = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]}
                    gs = extract_gs(
                        values.get("자체 상품코드", ""),
                        values.get("적용 GS코드", ""),
                        values.get("상품명(원본)", ""),
                        values.get("공급사 상품명", ""),
                    )
                    base_gs = split_gs(gs)[0] if gs else ""
                    product_names = [
                        values.get("상품명", ""),
                        values.get("최종상품명", ""),
                        values.get("기본상품명(옵션제외)", ""),
                    ]
                    search_terms = [
                        values.get("검색어설정", ""),
                        values.get("검색어설정(R)", ""),
                        values.get("검색키워드", ""),
                        values.get("키워드정렬", ""),
                        values.get("로컬빌더", ""),
                    ]
                    longtails = [
                        values.get("롱테일10", ""),
                        values.get("앵커", ""),
                        values.get("베이스라인", ""),
                    ]
                    debug_terms = [
                        values.get("상세요약", ""),
                        values.get("옵션(본문)", ""),
                        values.get("옵션_토큰", ""),
                    ]
                    if base_gs:
                        append_keyword_record(summary["keywordByGs"], base_gs, "productNames", " ".join(text_value(v) for v in product_names))
                        append_keyword_record(summary["keywordByGs"], base_gs, "searchTerms", " ".join(text_value(v) for v in search_terms))
                        append_keyword_record(summary["keywordByGs"], base_gs, "longtails", " ".join(text_value(v) for v in longtails))
                        append_keyword_record(summary["keywordByGs"], base_gs, "debugTerms", " ".join(text_value(v) for v in debug_terms))
                    else:
                        for field, vals in (
                            ("productNames", product_names),
                            ("searchTerms", search_terms),
                            ("longtails", longtails),
                            ("debugTerms", debug_terms),
                        ):
                            joined = " ".join(text_value(v) for v in vals)
                            if joined:
                                pending.append((field, joined))
            if len(summary["keywordByGs"]) == 1 and pending:
                only_base = next(iter(summary["keywordByGs"]))
                for field, value in pending:
                    append_keyword_record(summary["keywordByGs"], only_base, field, value)
            summary["available"] = bool(summary["keywordByGs"])
        finally:
            workbook.close()
    except Exception as exc:
        summary["error"] = str(exc)
    return summary


def find_processed_listing_images(output_root: object, base_gs: str, min_count: int = 3) -> dict:
    base = text_value(base_gs).upper()
    if not base:
        return {"urls": [], "folder": "", "status": "missing"}
    search_roots: list[Path] = []
    root_value = text_value(output_root)
    if root_value:
        search_roots.append(Path(root_value))
    search_roots.append(EXPORT_ROOT)

    listing_roots: list[Path] = []
    for root in search_roots:
        root = root.resolve()
        candidates = [root / "listing_images"]
        if root.name.lower() == "listing_images":
            candidates.insert(0, root)
        for candidate in candidates:
            if candidate.is_dir() and candidate not in listing_roots:
                listing_roots.append(candidate)

    folders: list[Path] = []
    for listing_root in listing_roots:
        direct_candidates = [
            listing_root / base,
            *(path for path in listing_root.glob(f"*/{base}") if path.is_dir()),
        ]
        for folder in direct_candidates:
            if folder.is_dir() and folder not in folders:
                folders.append(folder)

    folders.sort(key=lambda folder: folder.stat().st_mtime, reverse=True)
    best_files: list[Path] = []
    best_folder = ""
    for folder in folders:
        files = sorted(
            [path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS],
            key=image_file_sort_key,
        )
        if len(files) > len(best_files):
            best_files = files
            best_folder = str(folder)
        if len(files) >= min_count:
            return {
                "urls": [runtime_web_url(path) for path in files],
                "folder": str(folder),
                "status": "ready",
            }

    return {
        "urls": [],
        "folder": best_folder,
        "status": "insufficient" if best_files else "missing",
        "foundCount": len(best_files),
        "requiredCount": min_count,
    }


def hydrate_seed_payload(seed_payload: dict) -> dict:
    if not isinstance(seed_payload, dict):
        return seed_payload
    output_root = (
        seed_payload.get("pipelineResult", {}).get("output_root", "")
        if isinstance(seed_payload.get("pipelineResult"), dict)
        else ""
    ) or (
        seed_payload.get("artifacts", {}).get("outputRoot", "")
        if isinstance(seed_payload.get("artifacts"), dict)
        else ""
    )
    products = seed_payload.get("products") if isinstance(seed_payload.get("products"), list) else []
    for product in products:
        if not isinstance(product, dict):
            continue
        base_gs = text_value(product.get("baseGs")) or split_gs(product.get("gs", ""))[0]
        product["categoryBasis"] = category_basis_from_product(product)
        if not isinstance(product.get("naverProvidedNotice"), dict):
            ocr = product.get("ocrAnalysis") if isinstance(product.get("ocrAnalysis"), dict) else {}
            product["naverProvidedNotice"] = build_naver_provided_notice({
                "gs": product.get("gs", ""),
                "baseGs": base_gs,
                "name": product.get("sourceName", ""),
                "opt": product.get("optionSummary", ""),
                "optionItems": product.get("optionItems", []),
            }, ocr)
        processed = find_processed_listing_images(output_root, base_gs)
        processed_images = processed.get("urls") if isinstance(processed.get("urls"), list) else []
        if not processed_images:
            continue
        images = product.get("images") if isinstance(product.get("images"), dict) else {}
        images.update({
            "sourceThumb": processed_images[0],
            "representative": processed_images[0],
            "listingCandidates": processed_images,
            "additional": processed_images[1:20],
            "selectionSource": "processed_listing_images",
            "processedFolder": processed.get("folder", ""),
            "processedStatus": processed.get("status", ""),
            "processedFoundCount": len(processed_images),
            "processedRequiredCount": 3,
        })
        product["images"] = images
    return seed_payload


def build_seed_products(
    source_path: str,
    selected_gs: list[str],
    ocr_summary: dict,
    keyword_summary: dict | None = None,
    output_root: object = "",
) -> list[dict]:
    parsed = parse_source_preview(Path(source_path), max_preview=100000)
    selected = set(selected_gs or [])
    ocr_by_gs = (ocr_summary or {}).get("ocrByGs", {})
    keyword_by_gs = (keyword_summary or {}).get("keywordByGs", {})
    products: list[dict] = []
    for row in parsed.get("preview", []):
        if selected and row.get("gs") not in selected and row.get("baseGs") not in selected:
            continue
        base_gs = row.get("baseGs") or split_gs(row.get("gs", ""))[0]
        ocr_record = ocr_by_gs.get(base_gs) or ocr_by_gs.get(row.get("gs", "")) or {}
        keyword_record = keyword_by_gs.get(base_gs) or keyword_by_gs.get(row.get("gs", "")) or {}
        ocr_text = ocr_record.get("rawText", "")
        option_summary = row.get("opt", "단일")
        option_items = row.get("optionItems", [])
        option_info = option_meta(option_summary, option_items)
        processed = find_processed_listing_images(output_root, base_gs)
        processed_images = processed.get("urls") if isinstance(processed.get("urls"), list) else []
        raw_listing_images = unique_terms(row.get("listingImages") or [])
        raw_source_thumb = row.get("thumb", "") or (raw_listing_images[0] if raw_listing_images else "")
        detail_image_set = set(row.get("detailImages") or [])
        if processed_images:
            listing_images = processed_images
            source_thumb = processed_images[0]
            additional_images = processed_images[1:20]
            selection_source = "processed_listing_images"
        else:
            listing_images = raw_listing_images
            source_thumb = raw_source_thumb
            additional_images = [
                url
                for url in unique_terms((row.get("additionalImages") or []) + [url for url in listing_images if url != source_thumb])
                if url and url != source_thumb and url not in detail_image_set
            ][:20]
            selection_source = "source_listing_columns_only"
        naver_provided_notice = build_naver_provided_notice(row, ocr_record)
        product_record = {
            "gs": row.get("gs", ""),
            "baseGs": base_gs,
            "sourceName": row.get("name", ""),
            "price": row.get("price", 0),
            "supplyPrice": row.get("supplyPrice", 0),
            "salePrice": row.get("salePrice", row.get("price", 0)),
            "consumerPrice": row.get("consumerPrice", 0),
            "optionSummary": option_summary,
            "optionInput": row.get("optionInput", ""),
            "optionAdditionalAmounts": row.get("optionAdditionalAmounts", []),
            "optionType": option_info["optionType"],
            "optionCount": option_info["optionCount"],
            "optionItems": option_items,
            "detailHtml": row.get("detailHtml", ""),
            "images": {
                "sourceThumb": source_thumb,
                "representative": source_thumb,
                "listingCandidates": listing_images,
                "additional": additional_images,
                "detail": row.get("detailImages", []),
                "processedSize": "1000x1000",
                "selectionSource": selection_source,
                "processedFolder": processed.get("folder", ""),
                "processedStatus": processed.get("status", ""),
                "processedFoundCount": processed.get("foundCount", len(processed_images)),
                "processedRequiredCount": processed.get("requiredCount", 3),
            },
            "ocrAnalysis": {
                "status": "loaded" if ocr_record else "pending",
                "rawText": ocr_text,
                "fields": ocr_record.get("fields", {}),
            },
            "naverProvidedNotice": naver_provided_notice,
            "photoAnalysis": {
                "status": "pending",
                "facts": [],
                "notes": "대표/추가 이미지 분석 결과를 여기에 누적한다.",
            },
            "keywordCandidatePool": seed_keyword_pool_from_product(row, ocr_text, keyword_record),
            "generatedKeywordSeed": keyword_record,
            "reviewFields": {
                "productName": "",
                "searchTerms": "",
                "memo": "",
            },
        }
        product_record["categoryBasis"] = category_basis_from_product(product_record)
        products.append(product_record)
    return products


def normalize_keyword_channels(channels: list[str] | None, account_scope: str = "전체") -> list[str]:
    raw_channels = channels or [f"{account}:{market}" for account in ("A", "B") for market in SALES_MARKETS]
    out: list[str] = []
    seen: set[str] = set()
    for raw in raw_channels:
        key = text_value(raw)
        if ":" not in key:
            continue
        account, market = key.split(":", 1)
        account = account.upper()
        if account not in {"A", "B"} or market not in SALES_MARKETS:
            continue
        if account_scope and account_scope != "전체" and account != account_scope:
            continue
        normalized = f"{account}:{market}"
        if normalized in seen:
            continue
        seen.add(normalized)
        out.append(normalized)
    return out


def build_keyword_job_products(seed_payload: dict, selected_gs: list[str]) -> list[dict]:
    selected = set(selected_gs or [])
    products = seed_payload.get("products") if isinstance(seed_payload.get("products"), list) else []
    out: list[dict] = []
    for product in products:
        if not isinstance(product, dict):
            continue
        gs = text_value(product.get("gs"))
        base_gs = text_value(product.get("baseGs"))
        if selected and gs not in selected and base_gs not in selected:
            continue
        ocr = product.get("ocrAnalysis") if isinstance(product.get("ocrAnalysis"), dict) else {}
        photo = product.get("photoAnalysis") if isinstance(product.get("photoAnalysis"), dict) else {}
        out.append({
            "gs": gs,
            "baseGs": base_gs,
            "sourceName": product.get("sourceName", ""),
            "price": product.get("price", 0),
            "supplyPrice": product.get("supplyPrice", 0),
            "salePrice": product.get("salePrice", product.get("price", 0)),
            "consumerPrice": product.get("consumerPrice", 0),
            "optionSummary": product.get("optionSummary", ""),
            "optionInput": product.get("optionInput", ""),
            "optionAdditionalAmounts": product.get("optionAdditionalAmounts", []),
            "optionType": product.get("optionType", ""),
            "optionCount": product.get("optionCount", 0),
            "optionItems": product.get("optionItems", []),
            "detailHtml": product.get("detailHtml", ""),
            "images": product.get("images", {}),
            "ocrAnalysis": {
                "status": ocr.get("status", ""),
                "rawText": text_value(ocr.get("rawText"))[:5000],
                "fields": ocr.get("fields", {}),
            },
            "photoAnalysis": photo,
            "keywordCandidatePool": product.get("keywordCandidatePool", {}),
            "generatedKeywordSeed": product.get("generatedKeywordSeed", {}),
        })
    return out


def build_keyword_prompt(input_file: str, output_file: str, has_images: bool = False) -> str:
    image_block = ""
    if has_images:
        image_block = """
이미지 분석 (중요):
- 이 요청에는 상품 상세페이지 이미지가 첨부되어 있다.
- 첨부된 이미지를 직접 보고 OCR 및 시각 분석을 수행한다.
- 이미지에서 파악한 상품 특성(소재, 규격, 브랜드, 용도, 기능, 구성품 등)을 키워드 생성에 적극 반영한다.
- 이미지 OCR로 읽은 텍스트가 입력 JSON의 ocrAnalysis.rawText보다 정확하면 이미지 기반 정보를 우선한다.
"""
    return f"""
너는 한국 오픈마켓 상품명/검색어 생성 담당자다.
현재 작업 폴더의 `{input_file}`을 읽고, 반드시 `{output_file}` 하나만 UTF-8 JSON으로 작성해라.
코드 파일이나 다른 파일은 수정하지 마라.
{image_block}
목표:
- Cafe24는 실제 판매 채널이 아니다. 공통 후보 풀로만 참고한다.
- 실제 생성 채널은 네이버/쿠팡/롯데ON/11번가/ESM의 A/B 계정이다.
- 입력 상품마다 요청된 모든 채널의 상품명과 검색어를 실제로 새로 작성한다.
- 단순 후보 나열이 아니라 상품 정체성, 기능, 사용처, 문제해결, 재질/규격, 현장명/동의어를 판단해서 조합한다.

출력 JSON 스키마:
{{
  "schema": "webocr.keyword.v1",
  "provider": "codex-cli",
  "products": [
    {{
      "gs": "GS코드",
      "channels": {{
        "A:네이버": {{
          "title": "상품명",
          "searchTerms": "검색어 또는 태그 문자열",
          "tags": ["태그1", "태그2"],
          "candidateCount": 12,
          "notes": "짧은 판단 근거"
        }}
      }}
    }}
  ]
}}

생성 규칙:
- 상품명은 정확하게, 검색어/태그는 넓게 간다.
- 상품명은 후보 단어를 단순히 이어붙이지 말고 아래 공식으로 만든다.
  1) 대표 상품 정체성 또는 표준명
  2) 실제 기능/효과
  3) 주 사용처 또는 구매 상황
  4) 재질/규격은 상품 식별에 중요할 때만 추가
- 상품명이 짧아질 때는 수량을 넣지 말고 기능/사용처/문제해결 단어로 채운다.
- 좋은 상품명 예시 구조: `카라비너 릴고리 와이어 릴홀더 백팩 스트랩 연결고리`, `쿠션 깔창 신발 밑창 보강 패드 운동화 구두 PU`.
- 네이버 상품명은 30자 전후를 우선하고 최대 40자를 넘기지 않는다.
- 네이버는 SEO 기준을 우선한다. 상품명 공식은 `브랜드/제조사 + 모델명/모델코드 + 상품유형/카테고리 + 색상/용도/주요속성`이지만, 브랜드/모델이 없으면 빼고 핵심 카테고리와 용도/속성을 앞에 둔다.
- 네이버에서 참고할 항목은 브랜드/제조사, 시리즈, 모델명/모델코드, 상품 유형/카테고리, 색상, 소재, 구성품/수량, 사이즈, 대상 성별/연령, 용량/규격/주요 속성, 판매 옵션이다.
- 네이버 상품명은 중요한 키워드를 앞쪽에 배치하고, 중복 단어/관련 없는 키워드/할인/세일/무료배송/광고 문구를 제외한다.
- 네이버는 한글 중심으로 작성하고 필요한 경우에만 영문/숫자를 사용한다.
- 쿠팡/롯데ON/11번가/ESM도 30자 전후, 최대 40자 기준으로 만든다.
- 길이를 맞추기 위해 무의미한 단어를 넣지 않는다. 30자보다 짧아도 상품 정체성, 기능, 용도, 규격이 충분하면 그대로 둔다.
- 상품명 보강 우선순위는 `상품유형/카테고리 > 핵심 기능 > 사용처 > 식별 규격/소재`다. 광고어, 상황을 억지로 늘린 말, 반복 단어는 넣지 않는다.
- 11번가/ESM 검색어는 최소 14개 이상을 목표로 하고, 표준어 + 현장명 + 동의어 + 사용처 + 문제해결어를 충분히 넣는다.
- A계정은 표준어/대표어 중심, B계정은 동의어/사용처/현장명 중심으로 다르게 만든다.
- A/B 계정은 같은 단어만 순서 바꾸지 말고, 대표어와 현장명을 실제로 다르게 섞는다.
- 색상 옵션은 상품명에 넣지 않는다. 옵션 컬럼에서 처리한다.
- 옵션형 사이즈는 전체 나열하지 말고 필요하면 `230-270 사이즈 선택형`처럼 압축한다.
- `1개`, `1입`, `1매`, `1p`처럼 단품을 뜻하는 수량은 상품명/검색어에 넣지 않는다. 노출 가치가 낮다.
- `2개`, `5매`, `10개`, `100p`, `3세트`처럼 실제 구성 차이를 만드는 수량만 검색어에 남길 수 있다.
- OCR에서만 나온 숫자, 눈금, 이미지 배경 숫자는 상품명/검색어에서 제외한다.
- 무료배송/할인/추천/인기/베스트/가격/문의/상담/판매자명 같은 판매 문구는 금지한다.
- `발편한`, `발 편한`, `편한발`처럼 감성/광고형 문구는 상품명과 검색어에 넣지 않는다. `쿠션`, `충격 완화`, `착용감 보강`처럼 기능어로 바꾼다.
- 같은 단어를 반복하지 않는다.
- 상품명에서는 붙여쓴 합성어를 자연스럽게 띄어쓴다. 예: 카라비너릴고리 -> 카라비너 릴고리, 와이어릴고리 -> 와이어 릴고리, 쿠션깔창 -> 쿠션 깔창, 보강패드 -> 보강 패드, 충격완화 -> 충격 완화.
- 검색어/태그는 12-20개를 목표로 하며, 상품명에 못 넣은 동의어/현장명/사용처/문제해결어를 우선 보강한다.
- 검색어에는 쉼표로 구분된 실제 검색 가능한 말만 넣고, 무관한 인기어와 일부러 만든 오타는 넣지 않는다.
- 검색어/태그의 각 항목은 공백 없이 붙여쓴다. 예: `쿠션 깔창`은 검색어에서 `쿠션깔창`, `보강 패드`는 `보강패드`, `충격 완화`는 `충격완화`로 쓴다.
- 네이버 태그는 검색어를 많이 넣는 칸이 아니라 상품명/카테고리가 잡은 핵심어를 속성, 용도, 대상, 시즌, 스타일, 문제상황으로 확장하는 보조 필드다.
- 네이버 핵심 검색어와 일체형 키워드는 태그에만 의존하지 말고 반드시 상품명에 넣는다.
- 네이버 태그는 `메인키워드+속성`, `메인키워드+용도`, `메인키워드+대상`, `메인키워드+시즌`, `메인키워드+문제상황` 조합을 우선한다.
- 네이버 tags 배열은 8-10개로 압축한다. 단일 넓은 단어보다 구매자가 검색할 조합어를 쓴다. 예: `선반`보다 `욕실선반`, `욕실수납`, `무타공선반`.
- A:네이버 tags는 표준명, 카테고리 정확도, 대표 조합어 중심으로 만든다.
- B:네이버 tags는 A와 3개 이상 겹치지 않게 사용처, 문제상황, 대상, 소재/형태, 현장명 중심으로 다르게 만든다.
- A/B 네이버는 같은 단어의 순서만 바꾸지 말고 실제 태그 구성 축을 다르게 한다.
- `title`, `searchTerms`, `tags`는 비워두지 않는다.
- 요청 채널이 10개면 상품마다 10개 채널 모두 작성한다.

입력 JSON에는 `channels`, `products`, `policy`가 들어 있다.
최종 응답은 설명하지 말고 파일 작성만 끝내라.
""".strip()


def estimate_command_line_chars(args: list[str]) -> int:
    total = 0
    for arg in args:
        text = str(arg)
        needs_quotes = bool(re.search(r"\s", text))
        total += len(text) + 1
        if needs_quotes:
            total += 2
    return total


def append_codex_images_with_budget(cmd: list[str], images: list[str], instruction: str, budget: int = 5600, max_count: int = 8) -> list[str]:
    selected: list[str] = []
    trial = list(cmd)
    for image in images:
        if len(selected) >= max_count:
            break
        candidate = trial + ["-i", image, "--", instruction]
        if estimate_command_line_chars(candidate) > budget:
            continue
        selected.append(image)
        trial.extend(["-i", image])
    return selected


SINGLE_QUANTITY_RE = re.compile(
    r"(?<![0-9A-Za-z가-힣])1\s*(?:개|입|매|p|P|pcs|PCS|pc|PC)(?![0-9A-Za-z가-힣])"
)


def strip_low_value_single_quantity(value: str) -> str:
    cleaned = SINGLE_QUANTITY_RE.sub(" ", text_value(value))
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    cleaned = re.sub(r"\s*,\s*", ", ", cleaned)
    cleaned = re.sub(r"(?:^|,\s*)$", "", cleaned)
    return cleaned.strip(" ,/")


def split_search_keyword_chunks(value: object) -> list[str]:
    raw = text_value(value)
    if not raw:
        return []
    raw = re.sub(r"<[^>]+>", " ", raw)
    chunks: list[str] = []
    for chunk in re.split(r"[,/|;·\n\r\t]+", raw):
        text = normalize_term(chunk)
        if text:
            chunks.append(text)
    return unique_terms(chunks)


def clean_keyword_terms(value: str) -> str:
    terms = [
        compact_search_keyword(strip_low_value_single_quantity(term))
        for term in split_search_keyword_chunks(value)
    ]
    return ", ".join(unique_terms([term for term in terms if term]))


def seo_tag_terms(value: object, limit: int = 10) -> list[str]:
    source = value if isinstance(value, list) else split_candidate_terms(value)
    return unique_terms([
        compact_search_keyword(strip_low_value_single_quantity(term))
        for term in source
        if compact_search_keyword(strip_low_value_single_quantity(term))
    ])[:limit]


def normalize_market_title(title: object, search_terms: object, tags: object, channel: str) -> str:
    cleaned = clean_product_title(strip_low_value_single_quantity(title))
    max_len = NAVER_TITLE_TARGET_MAX if channel.endswith(":네이버") else MARKET_TITLE_TARGET_MAX
    if not cleaned:
        return ""

    def trim_to_limit(value: str) -> str:
        value = clean_product_title(value)
        if len(value) <= max_len:
            return value
        words = value.split()
        if len(words) <= 1:
            return value[:max_len].strip()
        out: list[str] = []
        for word in words:
            candidate = " ".join(out + [word])
            if len(candidate) > max_len:
                break
            out.append(word)
        return clean_product_title(" ".join(out)) or value[:max_len].strip()

    cleaned = trim_to_limit(cleaned)
    if len(cleaned) >= MARKET_TITLE_TARGET_MIN:
        return cleaned

    compact_title = re.sub(r"\s+", "", cleaned).lower()
    candidates = seo_tag_terms(tags, 20) + seo_tag_terms(search_terms, 30)
    for term in unique_terms(candidates):
        display = clean_product_title(strip_low_value_single_quantity(term))
        if not display or len(display) < 2 or len(display) > 12:
            continue
        compact_display = re.sub(r"\s+", "", display).lower()
        if compact_display in compact_title:
            continue
        candidate = clean_product_title(f"{cleaned} {display}")
        if len(candidate) > max_len:
            continue
        cleaned = candidate
        compact_title = re.sub(r"\s+", "", cleaned).lower()
        if len(cleaned) >= MARKET_TITLE_TARGET_MIN:
            break
    return cleaned


def diversify_naver_ab_tags(channel_map: dict[str, dict]) -> None:
    a = channel_map.get("A:네이버")
    b = channel_map.get("B:네이버")
    if not isinstance(a, dict) or not isinstance(b, dict):
        return

    a_tags = seo_tag_terms(a.get("tags") or a.get("searchTerms"), 10)
    b_tags = seo_tag_terms(b.get("tags") or b.get("searchTerms"), 10)
    if not a_tags or not b_tags:
        return

    a_set = {normalize_term(tag).lower() for tag in a_tags}
    b_overlap = [tag for tag in b_tags if normalize_term(tag).lower() in a_set]
    if len(b_overlap) <= 3:
        a["tags"] = a_tags
        b["tags"] = b_tags
        return

    b_pool = seo_tag_terms(b.get("searchTerms"), 30)
    next_b = [tag for tag in b_tags if normalize_term(tag).lower() not in a_set]
    for tag in b_pool:
        key = normalize_term(tag).lower()
        if key in a_set:
            continue
        if key not in {normalize_term(item).lower() for item in next_b}:
            next_b.append(tag)
        if len(next_b) >= 10:
            break

    if len(next_b) >= 5:
        b["tags"] = next_b[:10]
        b["searchTerms"] = ", ".join(unique_terms(next_b + seo_tag_terms(b.get("searchTerms"), 30)))
        b["candidateCount"] = len(b["tags"])
        b["notes"] = (text_value(b.get("notes")) + " / A/B 네이버 태그 중복 축소").strip(" /")
    a["tags"] = a_tags[:10]


def keyword_title_penalty(value: object) -> int:
    title = text_value(value)
    if not title:
        return 10
    penalty = 0
    compact = re.sub(r"\s+", "", title)
    if len(title) < 14:
        penalty += 2
    if len(title) > 85:
        penalty += 2
    if re.search(r"[{}<>]", title):
        penalty += 2
    if re.search(r"(.)\1{4,}", compact):
        penalty += 2
    for pattern in LOW_QUALITY_KEYWORD_TITLE_PATTERNS:
        if re.search(pattern, title, flags=re.IGNORECASE):
            penalty += 3
    return penalty


def is_low_quality_keyword_entry(value: object) -> bool:
    if not isinstance(value, dict):
        return True
    title = text_value(value.get("title"))
    terms = split_candidate_terms(value.get("searchTerms"))
    tags = value.get("tags") if isinstance(value.get("tags"), list) else []
    return keyword_title_penalty(title) >= 3 or not terms or len(tags) < 3


def repair_b_market_keywords_from_a(channel_map: dict[str, dict]) -> None:
    for market in SALES_MARKETS:
        a_key = f"A:{market}"
        b_key = f"B:{market}"
        a_value = channel_map.get(a_key)
        b_value = channel_map.get(b_key)
        if not isinstance(a_value, dict) or not isinstance(b_value, dict):
            continue
        a_title = text_value(a_value.get("title"))
        b_title = text_value(b_value.get("title"))
        title_too_weak = bool(a_title and b_title and len(b_title) < len(a_title))
        if not is_low_quality_keyword_entry(b_value) and not title_too_weak:
            continue
        repaired = dict(a_value)
        repaired["notes"] = (
            text_value(repaired.get("notes"))
            + " / B 생성 결과 품질 저하로 A 기준 상품명과 검색어를 임시 대체"
        ).strip(" /")
        repaired["repairedFrom"] = a_key
        repaired["generatedAt"] = now_text()
        channel_map[b_key] = repaired


def validate_keyword_result(payload: dict, products: list[dict], channels: list[str]) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("keyword result is not an object")
    raw_products = payload.get("products")
    if not isinstance(raw_products, list):
        raise ValueError("keyword result has no products list")

    product_keys = {text_value(product.get("gs")) or text_value(product.get("baseGs")) for product in products}
    normalized_products: list[dict] = []
    for item in raw_products:
        if not isinstance(item, dict):
            continue
        gs = text_value(item.get("gs"))
        if gs not in product_keys:
            continue
        raw_channels = item.get("channels") if isinstance(item.get("channels"), dict) else {}
        channel_map: dict[str, dict] = {}
        for channel in channels:
            value = raw_channels.get(channel)
            if not isinstance(value, dict):
                continue
            search_terms = clean_keyword_terms(value.get("searchTerms") or value.get("search_terms"))
            tag_limit = 10 if channel.endswith(":네이버") else 30
            tags = seo_tag_terms(value.get("tags") if isinstance(value.get("tags"), list) else search_terms, tag_limit)
            title = normalize_market_title(value.get("title"), search_terms, tags, channel)
            if not title or not search_terms:
                continue
            candidate_count = len(tags) or len(split_candidate_terms(search_terms))
            channel_map[channel] = {
                "title": title,
                "searchTerms": search_terms,
                "tags": tags,
                "candidateCount": candidate_count,
                "notes": text_value(value.get("notes"))[:500],
                "provider": text_value(payload.get("provider")) or "codex-cli",
                "generatedAt": now_text(),
            }
        diversify_naver_ab_tags(channel_map)
        repair_b_market_keywords_from_a(channel_map)
        if channel_map:
            normalized_products.append({"gs": gs, "channels": channel_map})
    if not normalized_products:
        raise ValueError("keyword result has no usable channel data")
    return {
        "schema": "webocr.keyword.v1",
        "provider": text_value(payload.get("provider")) or "codex-cli",
        "products": normalized_products,
    }


def apply_keyword_result_to_seed(seed_payload: dict, keyword_result: dict, channels: list[str]) -> None:
    by_gs = {
        text_value(item.get("gs")): item.get("channels", {})
        for item in keyword_result.get("products", [])
        if isinstance(item, dict)
    }
    for product in seed_payload.get("products", []):
        if not isinstance(product, dict):
            continue
        gs = text_value(product.get("gs"))
        channel_map = by_gs.get(gs)
        if not channel_map:
            continue
        existing = product.get("marketKeywords") if isinstance(product.get("marketKeywords"), dict) else {}
        for channel in channels:
            if channel in channel_map:
                current_value = existing.get(channel) if isinstance(existing.get(channel), dict) else {}
                next_value = channel_map[channel]
                current_title = text_value(current_value.get("title"))
                next_title = text_value(next_value.get("title"))
                current_count = int(current_value.get("candidateCount") or len(current_value.get("tags") or []))
                next_count = int(next_value.get("candidateCount") or len(next_value.get("tags") or []))
                current_score = len(current_title) + current_count * 4
                next_score = len(next_title) + next_count * 4
                if current_title and current_score > next_score + 18:
                    current_value["keptBecause"] = "regeneration_result_was_weaker"
                    existing[channel] = current_value
                else:
                    existing[channel] = next_value
        product["marketKeywords"] = existing
    seed_payload["keywordGeneration"] = {
        "status": "completed",
        "provider": keyword_result.get("provider", "codex-cli"),
        "channels": channels,
        "updatedAt": now_text(),
        "note": "키워드 생성 단계에서 Codex CLI를 호출해 실제 마켓별 상품명/검색어를 생성했다.",
    }


def read_csv_rows(path: Path) -> list[dict]:
    for encoding in ("utf-8-sig", "cp949", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                return list(csv.DictReader(handle, dialect=dialect))
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle))


def read_excel_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [text_value(value) for value in next(rows, [])]
        records = []
        for row in rows:
            records.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]})
        workbook.close()
        return records

    import pandas as pd

    frame = pd.read_excel(path, dtype=object)
    frame = frame.where(pd.notnull(frame), "")
    return frame.to_dict(orient="records")


def read_source_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xls"}:
        return read_excel_rows(path)
    return []


def collect_values(row: dict, headers: dict[str, str], names: list[str]) -> list[str]:
    values: list[str] = []
    for name in names:
        key = headers.get(normalize_header(name))
        if key is not None:
            value = text_value(row.get(key, ""))
            if value:
                values.append(value)
    return values


def extract_image_urls(value: str) -> list[str]:
    text = text_value(value)
    if not text:
        return []
    text = text.replace("&amp;", "&")
    urls = re.findall(r"<img[^>]+src=[\"']?([^\"'>\s]+)", text, flags=re.IGNORECASE)
    urls.extend(re.findall(r"https?://[^\s\"'<>|,;]+", text, flags=re.IGNORECASE))
    urls.extend(re.findall(r"//[^\s\"'<>|,;]+", text, flags=re.IGNORECASE))
    urls.extend(re.findall(r"(?:[A-Za-z]:\\|/)[^\s\"'<>|,;]+?\.(?:jpg|jpeg|png|webp|bmp)", text, flags=re.IGNORECASE))
    for part in re.split(r"[\n\r\t,;|]+", text):
        clean_part = part.strip().strip("\"'")
        if re.search(r"\.(?:jpg|jpeg|png|webp|bmp)(?:\?.*)?$", clean_part, re.IGNORECASE):
            urls.append(clean_part)
    out: list[str] = []
    seen: set[str] = set()
    for url in urls:
        clean = urllib.parse.unquote(text_value(url).strip().strip("\"'"))
        if clean.startswith("//"):
            clean = f"https:{clean}"
        if not clean or clean in seen:
            continue
        seen.add(clean)
        out.append(clean)
    return out


def normalize_detail_html_for_upload(value: object) -> str:
    html = text_value(value)
    if not html:
        return ""

    urls: list[str] = []
    seen_urls: set[str] = set()
    for url in extract_image_urls(html):
        if not public_image_url(url):
            continue
        parsed = urllib.parse.urlparse(url)
        key = f"{parsed.netloc.lower()}{parsed.path}"
        if key in seen_urls:
            continue
        seen_urls.add(key)
        urls.append(url)
    if urls and "<img" in html.lower():
        return "<center>" + "".join(f'<img src="{url}">' for url in urls[:80]) + "</center>"
    return html


def collect_image_urls(row: dict, headers: dict[str, str], names: list[str]) -> list[str]:
    return unique_image_urls([
        url
        for value in collect_values(row, headers, names)
        for url in extract_image_urls(value)
    ])


def image_url_identity(value: object) -> str:
    raw = text_value(value)
    if not raw:
        return ""
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme in {"http", "https"}:
        host = parsed.netloc.lower()
        path = urllib.parse.unquote(parsed.path).lower()
        return f"{host}{path}"
    return raw.replace("\\", "/").lower()


def unique_image_urls(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = text_value(value)
        key = image_url_identity(clean)
        if not clean or not key or key in seen:
            continue
        seen.add(key)
        out.append(clean)
    return out


def parse_source_preview(path: Path, max_preview: int = 500) -> dict:
    records = read_source_records(path)
    if not records:
        return {"rows": 0, "gsCodes": 0, "preview": [], "columns": []}

    columns = list(records[0].keys()) if records else []
    headers = {normalize_header(col): col for col in columns}
    groups: dict[str, list[dict]] = {}
    raw_gs_codes: set[str] = set()
    for index, row in enumerate(records):
        code = first_value(row, headers, ["자체 상품코드", "GS코드", "gs코드", "상품코드", "판매자상품코드"])
        name_raw = first_value(row, headers, ["약식 상품명", "약식상품명", "상품명(관리용)", "상품명", "공급사 상품명", "상품 요약설명"])
        gs = extract_gs(code, name_raw, row)
        if not gs:
            continue
        base_gs, suffix = split_gs(gs)
        raw_gs_codes.add(gs)
        price_info = source_price_info(row, headers)
        opt = first_value(row, headers, ["옵션입력", "옵션", "옵션세트명", "옵션명", "규격"]) or "단일"
        option_additional_text = first_value(row, headers, ["옵션추가금", "옵션 추가금", "추가금", "옵션가격", "옵션 가격"])
        listing_images = collect_image_urls(row, headers, LISTING_IMAGE_COLUMNS)
        additional_images = collect_image_urls(row, headers, ADDITIONAL_IMAGE_COLUMNS)
        detail_images = collect_image_urls(row, headers, DETAIL_IMAGE_COLUMNS)
        detail_html = normalize_detail_html_for_upload(first_value(row, headers, DETAIL_HTML_COLUMNS))
        thumb = listing_images[0] if listing_images else ""

        groups.setdefault(base_gs, []).append({
            "index": index,
            "gs": gs,
            "baseGs": base_gs,
            "suffix": suffix,
            "nameRaw": name_raw,
            "name": short_name(name_raw, gs),
            "price": price_info["salePrice"],
            "supplyPrice": price_info["supplyPrice"],
            "salePrice": price_info["salePrice"],
            "consumerPrice": price_info["consumerPrice"],
            "opt": opt,
            "optionAdditionalText": option_additional_text,
            "thumb": thumb,
            "listingImages": listing_images,
            "additionalImages": additional_images,
            "detailImages": detail_images,
            "detailHtml": detail_html,
        })

    preview = []
    ordered_groups = sorted(groups.items(), key=lambda pair: min(item["index"] for item in pair[1]))
    for base_gs, items in ordered_groups[:max_preview]:
        items = sorted(items, key=lambda item: item["index"])
        representative = next((item for item in items if item["suffix"] == "A"), items[0])
        clean_names = [item["name"] for item in items if item["name"]]
        base_name = common_word_prefix(clean_names) if len(items) > 1 else ""
        if not base_name:
            base_name = representative["name"] or representative["gs"]

        option_items = []
        option_labels = []
        for item in items:
            label = strip_option_label(item["name"], base_name)
            if not label and item["opt"] and item["opt"] != "단일":
                label = item["opt"]
            if not label and len(items) > 1:
                label = item["suffix"] or item["gs"].replace(base_gs, "", 1) or item["name"]
            if label:
                option_labels.append(label)
            option_items.append({
                "gs": item["gs"],
                "suffix": item["suffix"],
                "name": item["name"],
                "option": option_label_with_index(label, len(option_items)) or "단일",
                "price": item["price"],
                "supplyPrice": item.get("supplyPrice", 0),
                "salePrice": item.get("salePrice", item["price"]),
                "consumerPrice": item.get("consumerPrice", 0),
                "thumb": item["thumb"],
            })

        if len(option_items) <= 1:
            pm_option_items = pm_option_items_for_base(base_gs)
            if len(pm_option_items) > 1:
                option_items = pm_option_items
                option_labels = [text_value(item.get("option")) for item in option_items if text_value(item.get("option"))]

        listing_images = unique_image_urls([
            url
            for item in items
            for url in (item.get("listingImages", []) + ([item.get("thumb", "")] if item.get("thumb") else []))
        ])
        detail_images = unique_image_urls([
            url
            for item in items
            for url in item.get("detailImages", [])
        ])
        detail_html = next((text_value(item.get("detailHtml")) for item in items if text_value(item.get("detailHtml"))), "")
        representative_thumb = representative["thumb"] or (listing_images[0] if listing_images else "")
        additional_images = unique_image_urls([
            url
            for item in items
            for url in item.get("additionalImages", [])
        ] + [url for url in listing_images if url != representative_thumb])
        detail_keys = {image_url_identity(url) for url in detail_images}
        representative_key = image_url_identity(representative_thumb)
        additional_images = [
            url for url in additional_images
            if image_url_identity(url) not in detail_keys and image_url_identity(url) != representative_key
        ][:20]
        base_sale = option_items[0]["salePrice"] if option_items else representative["price"]
        explicit_additionals = option_price_list(representative.get("optionAdditionalText"))
        option_additionals = explicit_additionals if explicit_additionals else [
            int((item.get("salePrice") or item.get("price") or 0) - base_sale)
            for item in option_items
        ] if len(option_items) > 1 else []
        option_input = option_input_from_labels([
            text_value(item.get("option"))
            for item in option_items[:26]
            if text_value(item.get("option")) and text_value(item.get("option")) != "단일"
        ]) if len(option_items) > 1 else option_summary_to_input(representative.get("opt"))

        preview.append({
            "id": f"row-{len(preview) + 1}",
            "gs": representative["gs"],
            "baseGs": base_gs,
            "name": base_name,
            "price": representative["price"],
            "supplyPrice": representative.get("supplyPrice", 0),
            "salePrice": representative.get("salePrice", representative["price"]),
            "consumerPrice": representative.get("consumerPrice", 0),
            "opt": compact_option_summary(option_labels, len(items)),
            "optionInput": option_input,
            "optionAdditionalAmounts": option_additionals,
            "thumb": representative_thumb,
            "listingImages": listing_images,
            "detailImages": detail_images,
            "detailHtml": detail_html,
            "additionalImages": additional_images,
            "optionItems": option_items,
        })
        preview[-1].update(option_meta(preview[-1]["opt"], option_items))

    return {
        "rows": len(records),
        "gsCodes": len(groups),
        "rawGsCodes": len(raw_gs_codes),
        "preview": preview,
        "columns": columns,
    }


def create_selected_source_file(source_path: str, selected_gs: list[str]) -> tuple[str, dict]:
    source = Path(source_path)
    selected_codes = {text_value(value).upper() for value in selected_gs or [] if text_value(value)}
    selected_bases = {split_gs(code)[0] for code in selected_codes if code}
    if not selected_bases:
        raise ValueError("selectedGs is empty")

    records = read_source_records(source)
    if not records:
        raise ValueError("source file has no rows")

    columns = list(records[0].keys())
    headers = {normalize_header(col): col for col in columns}
    filtered: list[dict] = []
    matched_codes: set[str] = set()
    for row in records:
        code = first_value(row, headers, ["자체 상품코드", "GS코드", "gs코드", "상품코드", "판매자상품코드"])
        name_raw = first_value(row, headers, ["약식 상품명", "약식상품명", "상품명(관리용)", "상품명", "공급사 상품명", "상품 요약설명"])
        gs = extract_gs(code, name_raw, row)
        if not gs:
            continue
        base_gs, _ = split_gs(gs)
        if gs in selected_codes or base_gs in selected_bases:
            filtered.append(row)
            matched_codes.add(gs)

    if not filtered:
        raise ValueError(f"selected GS not found in source: {', '.join(sorted(selected_codes))}")

    for row in filtered:
        for key in row.keys():
            if key not in columns:
                columns.append(key)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = UPLOAD_ROOT / safe_name(f"{source.stem}_{stamp}_selected_{len(selected_bases)}.csv")
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(filtered)

    return str(target), {
        "sourceFile": str(source),
        "filteredFile": str(target),
        "selectedProducts": len(selected_bases),
        "rawRows": len(records),
        "filteredRows": len(filtered),
        "matchedGs": sorted(matched_codes),
    }


def parse_multipart(body: bytes, content_type: str) -> dict[str, dict]:
    match = re.search(r"boundary=(?P<boundary>[^;]+)", content_type or "")
    if not match:
        raise ValueError("multipart boundary not found")
    boundary = match.group("boundary").strip('"').encode("utf-8")
    out: dict[str, dict] = {}
    for raw in body.split(b"--" + boundary):
        raw = raw.strip()
        if not raw or raw == b"--":
            continue
        if raw.endswith(b"--"):
            raw = raw[:-2].strip()
        if b"\r\n\r\n" not in raw:
            continue
        header_blob, content = raw.split(b"\r\n\r\n", 1)
        if content.endswith(b"\r\n"):
            content = content[:-2]
        headers = header_blob.decode("utf-8", errors="replace")
        disposition = next((line for line in headers.split("\r\n") if line.lower().startswith("content-disposition:")), "")
        name_match = re.search(r'name="([^"]+)"', disposition)
        if not name_match:
            continue
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        field_name = name_match.group(1)
        out[field_name] = {
            "filename": filename_match.group(1) if filename_match else "",
            "content": content,
            "headers": headers,
        }
    return out


def py_command() -> list[str]:
    py = shutil.which("py")
    if py:
        return [py, "-3"]
    return [sys.executable]


def bool_arg(value: object) -> str:
    return "true" if bool(value) else "false"


def bridge_args(source_path: str, settings: dict) -> list[str]:
    phase = "full"
    return [
        *py_command(),
        str(BRIDGE_SCRIPT),
        "--legacy-root", str(BACKEND_ROOT),
        "--source", source_path,
        "--make-listing", "true",
        "--listing-size", "1000",
        "--listing-pad", str(settings.get("ListingPad", 20)),
        "--listing-max", str(settings.get("ListingMax", 20)),
        "--logo-path", str(settings.get("LogoPath", "") or ""),
        "--logo-ratio", str(settings.get("LogoRatio", 14)),
        "--logo-opacity", str(settings.get("LogoOpacity", 65)),
        "--logo-pos", str(settings.get("LogoPosition", "tr") or "tr"),
        "--use-auto-contrast", bool_arg(settings.get("UseAutoContrast", True)),
        "--use-sharpen", bool_arg(settings.get("UseSharpen", True)),
        "--use-small-rotate", bool_arg(settings.get("UseSmallRotate", True)),
        "--rotate-zoom", str(settings.get("RotateZoom", 1.04)),
        "--ultra-angle-deg", str(settings.get("UltraAngleDeg", 0.35)),
        "--ultra-translate-px", str(settings.get("UltraTranslatePx", 0.6)),
        "--ultra-scale-pct", str(settings.get("UltraScalePct", 0.25)),
        "--trim-tol", str(settings.get("TrimTolerance", 8)),
        "--jpeg-q-min", str(settings.get("JpegQualityMin", 88)),
        "--jpeg-q-max", str(settings.get("JpegQualityMax", 92)),
        "--flip-lr", bool_arg(settings.get("FlipLeftRight", False)),
        "--logo-path-b", str(settings.get("LogoPathB", "") or ""),
        "--img-tag", str(settings.get("ImgTag", "") or ""),
        "--img-tag-b", str(settings.get("ImgTagB", "") or ""),
        "--a-name-min", str(settings.get("ANameMin", 80)),
        "--a-name-max", str(settings.get("ANameMax", 100)),
        "--b-name-min", str(settings.get("BNameMin", 63)),
        "--b-name-max", str(settings.get("BNameMax", 98)),
        "--a-tag-count", str(settings.get("ATagCount", 20)),
        "--b-tag-count", str(settings.get("BTagCount", 14)),
        "--phase", phase,
        "--export-root", str(EXPORT_ROOT),
        "--model", str(settings.get("Model", "gpt-4.1-mini")),
        "--chunk-size", str(settings.get("ChunkSize", 10)),
        "--keyword-version", str(settings.get("KeywordVersion", "3.0")),
    ]


def stop_process_tree(process: subprocess.Popen) -> None:
    if process.poll() is not None:
        return
    try:
        if os.name == "nt":
            subprocess.run(
                ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
        else:
            process.terminate()
    except Exception:
        try:
            process.kill()
        except Exception:
            pass


def register_active_process(job_id: str, process: subprocess.Popen) -> None:
    with ACTIVE_PROCESS_LOCK:
        ACTIVE_PROCESSES.setdefault(job_id, []).append(process)


def unregister_active_process(job_id: str, process: subprocess.Popen | None = None) -> None:
    with ACTIVE_PROCESS_LOCK:
        if process is None:
            ACTIVE_PROCESSES.pop(job_id, None)
            return
        processes = ACTIVE_PROCESSES.get(job_id)
        if not processes:
            return
        ACTIVE_PROCESSES[job_id] = [item for item in processes if item is not process]
        if not ACTIVE_PROCESSES[job_id]:
            ACTIVE_PROCESSES.pop(job_id, None)


def stop_active_job(job_id: str) -> bool:
    with ACTIVE_PROCESS_LOCK:
        CANCELLED_JOB_IDS.add(job_id)
        processes = list(ACTIVE_PROCESSES.get(job_id) or [])
    if not processes:
        return False
    for process in processes:
        stop_process_tree(process)
    unregister_active_process(job_id)
    return True


def run_seed_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    job = read_json(job_path, {"jobId": job_id})
    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "progressPercent": 2,
        "currentStage": "작업 준비",
    })
    write_json(job_path, job)

    source_path = payload.get("sourcePath") or payload.get("sourceFilePath") or ""
    selected_gs = payload.get("selectedGs", [])
    settings = payload.get("listingImageSettings") or {}
    if not source_path or not Path(source_path).exists():
        job.update({"status": "failed", "finishedAt": now_text(), "error": "source file not found"})
        write_json(job_path, job)
        return

    try:
        effective_source_path, filter_meta = create_selected_source_file(source_path, selected_gs)
    except Exception as exc:
        job.update({"status": "failed", "finishedAt": now_text(), "error": str(exc)})
        write_json(job_path, job)
        return

    job.update({
        "selectedGs": selected_gs,
        "filter": filter_meta,
        "progressPercent": 8,
        "currentStage": "선택 상품 원본 필터링",
        "updatedAt": now_text(),
    })
    write_json(job_path, job)

    cmd = bridge_args(effective_source_path, settings)
    result_payload: dict | None = None
    try:
        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START sourceToSeed\n")
            log.write(f"selected products: {filter_meta['selectedProducts']} / filtered rows: {filter_meta['filteredRows']}\n")
            log.write(f"filtered source: {effective_source_path}\n")
            log.write(" ".join(f'"{part}"' if " " in part else part for part in cmd) + "\n\n")
            process = subprocess.Popen(
                cmd,
                cwd=str(PROJECT_ROOT),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
            )
            register_active_process(job_id, process)
            assert process.stdout is not None
            tail: list[str] = []
            max_progress = 12
            result_seen = False
            try:
                for line in process.stdout:
                    clean = line.rstrip("\n")
                    log.write(clean + "\n")
                    log.flush()
                    tail = (tail + [clean])[-20:]
                    if clean.startswith("__RESULT__"):
                        result_payload = json.loads(clean[len("__RESULT__"):])
                        result_seen = True
                    line_progress, line_stage, current_gs = infer_seed_progress(clean, max_progress)
                    max_progress = max(max_progress, line_progress)
                    current = read_json(job_path, {"jobId": job_id})
                    current.update({
                        "status": "running",
                        "updatedAt": now_text(),
                        "tail": tail,
                        "progressPercent": max_progress,
                        "currentStage": line_stage,
                        "currentGs": current_gs or current.get("currentGs", ""),
                    })
                    write_json(job_path, current)
                    if result_seen:
                        break
                if result_payload:
                    try:
                        exit_code = process.wait(timeout=3)
                    except subprocess.TimeoutExpired:
                        stop_process_tree(process)
                        exit_code = 0
                else:
                    exit_code = process.wait()
            finally:
                unregister_active_process(job_id, process)

        if exit_code != 0:
            raise RuntimeError(f"pipeline failed with exit code {exit_code}")

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        source_stem = Path(source_path).stem
        seed_name = safe_name(f"{source_stem}_{stamp}.webseed.json")
        seed_path = SEED_ROOT / seed_name
        ocr_summary = read_pipeline_ocr_summary(
            (result_payload or {}).get("output_file", ""),
            (result_payload or {}).get("output_root", ""),
        )
        keyword_summary = read_pipeline_keyword_summary((result_payload or {}).get("output_file", ""))
        products = build_seed_products(
            effective_source_path,
            selected_gs,
            ocr_summary,
            keyword_summary,
            (result_payload or {}).get("output_root", ""),
        )
        job.update({"progressPercent": 92, "currentStage": "시드 파일 작성", "updatedAt": now_text()})
        write_json(job_path, job)
        seed_payload = {
            "schema": "webocr.seed.v2",
            "seedType": "source_to_seed",
            "createdAt": now_text(),
            "sourceFile": source_path,
            "effectiveSourceFile": effective_source_path,
            "sourceFilter": filter_meta,
            "selectedGs": selected_gs,
            "analysisPolicy": SEED_ANALYSIS_POLICY,
            "products": products,
            "summary": {
                "products": len(products),
                "ocrLoaded": sum(1 for product in products if product.get("ocrAnalysis", {}).get("status") == "loaded"),
                "imageRule": "1000x1000",
            },
            "pipelineResult": result_payload or {},
            "listingImageSettings": settings,
            "artifacts": {
                "outputRoot": (result_payload or {}).get("output_root", ""),
                "outputFile": (result_payload or {}).get("output_file", ""),
                "logFile": str(log_path),
            },
            "workbookSummary": {
                "sheets": ocr_summary.get("sheets", []),
                "ocrAvailable": ocr_summary.get("available", False),
                "ocrError": ocr_summary.get("error", ""),
                "keywordAvailable": keyword_summary.get("available", False),
                "keywordError": keyword_summary.get("error", ""),
            },
            "note": "1차 시드: 원본 정리, 이미지 1000x1000 가공 기준, OCR/사진분석, 키워드 후보 풀을 담는 기준 데이터셋",
        }
        write_json(seed_path, seed_payload)

        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "1차 시드 생성 완료",
            "result": {
                "seedFileName": seed_name,
                "seedPath": str(seed_path),
                "seedSize": seed_path.stat().st_size,
                "filter": filter_meta,
                "pipelineResult": result_payload or {},
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        if job_id in CANCELLED_JOB_IDS:
            return
        with log_path.open("a", encoding="utf-8", errors="replace") as log:
            log.write(f"\n[{now_text()}] ERROR {exc}\n")
        job.update({"status": "failed", "finishedAt": now_text(), "error": str(exc), "currentStage": "실패"})
        write_json(job_path, job)


def run_keyword_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    work_dir = JOBS_ROOT / f"{job_id}_keyword"
    work_dir.mkdir(parents=True, exist_ok=True)
    job = read_json(job_path, {"jobId": job_id})
    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "workDir": str(work_dir),
        "progressPercent": 3,
        "currentStage": "키워드 생성 준비",
    })
    write_json(job_path, job)

    try:
        seed_path = resolve_seed_path(payload.get("seedPath") or payload.get("sourcePath") or payload.get("path") or "")
        if not seed_path.exists():
            raise FileNotFoundError("seed file not found")
        seed_payload = read_json(seed_path, {})
        channels = normalize_keyword_channels(payload.get("channels"), text_value(payload.get("accountScope") or "전체"))
        selected_gs = payload.get("selectedGs") if isinstance(payload.get("selectedGs"), list) else []
        products = build_keyword_job_products(seed_payload, selected_gs)
        if not products:
            raise ValueError("selected seed products not found")
        if not channels:
            raise ValueError("selected market channels not found")

        a_channels = [ch for ch in channels if ch.startswith("A:")]
        b_channels = [ch for ch in channels if ch.startswith("B:")]
        channel_groups: list[tuple[str, list[str]]] = []
        if a_channels:
            channel_groups.append(("A", a_channels))
        if b_channels:
            channel_groups.append(("B", b_channels))
        if not channel_groups:
            channel_groups.append(("all", channels))

        codex_bin = shutil.which("codex.cmd") or shutil.which("codex") or "codex"

        output_root = ""
        if isinstance(seed_payload.get("pipelineResult"), dict):
            output_root = seed_payload["pipelineResult"].get("output_root", "")
        if not output_root and isinstance(seed_payload.get("artifacts"), dict):
            output_root = seed_payload["artifacts"].get("outputRoot", "")

        product_image_map: dict[str, list[str]] = {}
        if output_root:
            ocr_tmp = Path(output_root) / "_ocr_tmp"
            if ocr_tmp.is_dir():
                for p in products:
                    gs = text_value(p.get("gs")).upper()
                    base = text_value(p.get("baseGs")).upper() or base_gs_code(gs)
                    found: list[str] = []
                    seen_paths: set[str] = set()
                    for prefix in {gs, base, f"{base}A", f"{gs}A"}:
                        d = ocr_tmp / prefix
                        if not d.is_dir():
                            continue
                        for f in sorted(d.iterdir(), key=lambda x: x.name):
                            if f.is_file() and f.suffix.lower() in IMAGE_EXTENSIONS and str(f) not in seen_paths:
                                if re.match(r"^\d+$", f.stem):
                                    seen_paths.add(str(f))
                                    found.append(str(f.resolve()))
                    product_image_map[gs] = found[:5]

        def images_for_products(chunk_products: list[dict]) -> list[str]:
            chunk_images: list[str] = []
            for image_idx in range(5):
                for product in chunk_products:
                    gs = text_value(product.get("gs")).upper()
                    images = product_image_map.get(gs, [])
                    if image_idx < len(images):
                        chunk_images.append(images[image_idx])
            return chunk_images

        raw_chunk_size = payload.get("keywordProductChunkSize") or payload.get("productChunkSize") or KEYWORD_PRODUCT_CHUNK_SIZE
        try:
            chunk_size = int(raw_chunk_size)
        except Exception:
            chunk_size = KEYWORD_PRODUCT_CHUNK_SIZE
        chunk_size = max(1, min(20, chunk_size))
        try:
            stall_seconds = int(payload.get("keywordStallSeconds") or KEYWORD_CODEX_STALL_SECONDS)
        except Exception:
            stall_seconds = KEYWORD_CODEX_STALL_SECONDS
        stall_seconds = max(120, stall_seconds)
        product_chunks = [products[i:i + chunk_size] for i in range(0, len(products), chunk_size)]

        work_items: list[dict] = []
        total_codex_images = sum(len(images_for_products(chunk)) for chunk in product_chunks)
        for group_label, group_channels in channel_groups:
            for chunk_index, chunk_products in enumerate(product_chunks, start=1):
                work_label = group_label if len(product_chunks) == 1 else f"{group_label}_{chunk_index:02d}"
                gdir = work_dir / f"group_{work_label}"
                gdir.mkdir(parents=True, exist_ok=True)
                input_name = "keyword_input.json"
                output_name = "keyword_result.json"
                chunk_images = images_for_products(chunk_products)
                input_payload = {
                    "schema": "webocr.keyword.input.v1",
                    "createdAt": now_text(),
                    "seedFile": str(seed_path),
                    "channels": group_channels,
                    "options": payload.get("options", {}),
                    "policy": SEED_ANALYSIS_POLICY,
                    "chunk": {
                        "index": chunk_index,
                        "total": len(product_chunks),
                        "productChunkSize": chunk_size,
                    },
                    "products": chunk_products,
                }
                write_json(gdir / input_name, input_payload)
                (gdir / "prompt.md").write_text(build_keyword_prompt(input_name, output_name, has_images=bool(chunk_images)), encoding="utf-8")
                work_items.append({
                    "workLabel": work_label,
                    "groupLabel": group_label,
                    "channels": group_channels,
                    "products": chunk_products,
                    "images": chunk_images,
                    "workDir": gdir,
                    "inputName": input_name,
                    "outputName": output_name,
                    "chunkIndex": chunk_index,
                    "chunkTotal": len(product_chunks),
                })

        job.update({
            "seedPath": str(seed_path),
            "channels": channels,
            "selectedGs": [product["gs"] for product in products],
            "totalProducts": len(products),
            "totalChannels": len(channels),
            "parallelGroups": len(work_items),
            "keywordProductChunkSize": chunk_size,
            "keywordStallSeconds": stall_seconds,
            "progressPercent": 10,
            "currentStage": f"Codex 입력 작성 완료 · 상품 {len(products)}개 · 채널 {len(channels)}개 · 작업 {len(work_items)}개({chunk_size}개씩 분할) · 이미지 {total_codex_images}장",
            "updatedAt": now_text(),
        })
        write_json(job_path, job)

        job_lock = threading.Lock()
        shared_tail: list[str] = []
        shared_output_lines = [0]

        def run_codex_group(work_item: dict) -> tuple[str, list[str], list[dict], Path]:
            group_label = text_value(work_item.get("groupLabel"))
            work_label = text_value(work_item.get("workLabel"))
            group_channels = work_item.get("channels") or []
            chunk_products = work_item.get("products") or []
            chunk_images = work_item.get("images") or []
            gdir = Path(work_item.get("workDir"))
            input_name = text_value(work_item.get("inputName")) or "keyword_input.json"
            output_name = text_value(work_item.get("outputName")) or "keyword_result.json"
            image_instruction = (
                f"`prompt.md` 지시서를 먼저 읽고 `{input_name}`의 상품/채널 데이터를 분석해서 "
                f"`{output_name}`을 스키마에 맞는 JSON으로 작성해. "
                f"첨부된 이미지들은 상품 상세페이지 캡처이다. 이미지를 직접 보고 OCR/시각 분석하여 "
                f"상품 특성, 소재, 규격, 브랜드 등을 파악하고 키워드 생성에 반영해라. "
                f"다른 파일은 수정하지 마."
            )
            text_instruction = (
                f"`prompt.md` 지시서를 먼저 읽고 `{input_name}`의 상품/채널 데이터를 분석해서 "
                f"`{output_name}`을 스키마에 맞는 JSON으로 작성해. 다른 파일은 수정하지 마."
            )
            cmd = [
                codex_bin, "exec",
                "--skip-git-repo-check",
                "--dangerously-bypass-approvals-and-sandbox",
                "-C", str(gdir),
            ]
            selected_images = []
            if chunk_images:
                selected_images = append_codex_images_with_budget(cmd, chunk_images, image_instruction)
            instruction = image_instruction if selected_images else text_instruction
            for img_path in selected_images:
                cmd.extend(["-i", img_path])
            if selected_images:
                cmd.append("--")
            cmd.append(instruction)
            group_log_path = JOBS_ROOT / f"{job_id}_{work_label}.log"
            stdout_done = object()

            def read_stdout(stream, oq: queue.Queue) -> None:
                try:
                    for line in stream:
                        oq.put(line)
                finally:
                    oq.put(stdout_done)

            with group_log_path.open("w", encoding="utf-8", errors="replace") as glog:
                glog.write(f"[{now_text()}] START keyword group {work_label} channels={group_channels}\n")
                glog.write(f"products: {len(chunk_products)} / chunk {work_item.get('chunkIndex')}/{work_item.get('chunkTotal')}\n")
                glog.write(f"images attached: {len(selected_images)} / {len(chunk_images)} candidates, command chars ~= {estimate_command_line_chars(cmd)}\n")
                glog.write(" ".join(f'"{p}"' if " " in p else p for p in cmd) + "\n\n")
                process = subprocess.Popen(
                    cmd, cwd=str(gdir),
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
                )
                proc_key = f"{job_id}_{work_label}"
                register_active_process(job_id, process)
                register_active_process(proc_key, process)
                assert process.stdout is not None
                oq: queue.Queue = queue.Queue()
                reader = threading.Thread(target=read_stdout, args=(process.stdout, oq), daemon=True)
                reader.start()
                reader_done = False
                started_at = time.time()
                last_activity = started_at
                result_path = gdir / output_name
                last_result_signature: tuple[float, int] | None = None
                last_hb = 0.0
                try:
                    while True:
                        if job_id in CANCELLED_JOB_IDS:
                            stop_process_tree(process)
                            raise RuntimeError("cancelled")
                        try:
                            item = oq.get(timeout=0.5)
                        except queue.Empty:
                            item = None
                        if item is stdout_done:
                            reader_done = True
                        elif item is not None:
                            clean = str(item).rstrip("\n")
                            glog.write(clean + "\n")
                            glog.flush()
                            last_activity = time.time()
                            with job_lock:
                                shared_output_lines[0] += 1
                                shared_tail.append(clean)
                                while len(shared_tail) > 20:
                                    shared_tail.pop(0)
                        now_ts = time.time()
                        if result_path.exists():
                            stat = result_path.stat()
                            signature = (stat.st_mtime, stat.st_size)
                            if signature != last_result_signature:
                                last_result_signature = signature
                                last_activity = now_ts
                        if process.poll() is None and now_ts - last_activity > stall_seconds:
                            glog.write(f"\n[{now_text()}] STALLED no output/result change for {stall_seconds}s\n")
                            glog.flush()
                            stop_process_tree(process)
                            raise RuntimeError(f"codex group {work_label} stalled for {stall_seconds}s")
                        if now_ts - last_hb >= 2.0:
                            elapsed = int(now_ts - started_at)
                            with job_lock:
                                progress = min(88, max(12 + elapsed // 5, 18 + min(shared_output_lines[0], 60)))
                                current = read_json(job_path, {"jobId": job_id})
                                current.update({
                                    "status": "running", "updatedAt": now_text(),
                                    "tail": list(shared_tail),
                                    "progressPercent": progress,
                                    "currentStage": f"Codex AI 생성 중 ({work_label} · {len(chunk_products)}개 · {len(group_channels)}채널)",
                                })
                                write_json(job_path, current)
                            last_hb = now_ts
                        if process.poll() is not None and reader_done and oq.empty():
                            break
                    exit_code = process.wait()
                finally:
                    unregister_active_process(job_id, process)
                    unregister_active_process(proc_key, process)
            if exit_code != 0:
                raise RuntimeError(f"codex group {work_label} failed (exit {exit_code})")
            result_path = gdir / output_name
            if not result_path.exists():
                raise FileNotFoundError(f"{output_name} not written for group {work_label}")
            return group_label, group_channels, chunk_products, result_path

        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START keywordGenerate (work items {len(work_items)}, max workers {KEYWORD_CODEX_MAX_WORKERS})\n")
            log.write(f"seed: {seed_path}\n")
            log.write(f"products: {len(products)} / channels: {len(channels)} / chunkSize: {chunk_size} / stallSeconds: {stall_seconds}\n")
            for item in work_items:
                log.write(f"  work {item['workLabel']}: {len(item['products'])} products, channels={item['channels']}\n")
            log.flush()

        if len(work_items) >= 2:
            with ThreadPoolExecutor(max_workers=min(KEYWORD_CODEX_MAX_WORKERS, len(work_items))) as pool:
                futures = {
                    pool.submit(run_codex_group, item): item["workLabel"]
                    for item in work_items
                }
                group_results: list[tuple[str, list[str], list[dict], Path]] = []
                for future in as_completed(futures):
                    group_results.append(future.result())
        else:
            group_results = [run_codex_group(work_items[0])]

        merged_keyword_result: dict = {"products": []}
        for group_label, group_channels, chunk_products, result_path in group_results:
            raw = read_json(result_path, {})
            validated = validate_keyword_result(raw, chunk_products, group_channels)
            for product_result in validated.get("products", []):
                gs = text_value(product_result.get("gs"))
                existing = next((p for p in merged_keyword_result["products"] if text_value(p.get("gs")) == gs), None)
                if existing:
                    existing.setdefault("channels", {}).update(product_result.get("channels", {}))
                else:
                    merged_keyword_result["products"].append(product_result)
            if "provider" in validated:
                merged_keyword_result["provider"] = validated["provider"]

        job = read_json(job_path, job)
        job.update({"progressPercent": 90, "currentStage": "결과 검증 및 시드 반영", "updatedAt": now_text()})
        write_json(job_path, job)
        apply_keyword_result_to_seed(seed_payload, merged_keyword_result, channels)
        write_json(seed_path, seed_payload)

        generated_products = len(merged_keyword_result.get("products", []))
        generated_channels = sum(len(item.get("channels", {})) for item in merged_keyword_result.get("products", []))
        job = read_json(job_path, job)
        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "키워드 생성 완료",
            "tail": shared_tail[-20:],
            "result": {
                "seedPath": str(seed_path),
                "keywordResultPath": str(group_results[0][3]) if group_results else "",
                "products": generated_products,
                "channels": channels,
                "generatedChannels": generated_channels,
                "parallelGroups": len(group_results),
                "keywordProductChunkSize": chunk_size,
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        if job_id in CANCELLED_JOB_IDS:
            return
        with log_path.open("a", encoding="utf-8", errors="replace") as log:
            log.write(f"\n[{now_text()}] ERROR {exc}\n")
        job.update({
            "status": "failed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "키워드 생성 실패",
            "error": str(exc),
        })
        write_json(job_path, job)


def run_automation_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    job = read_json(job_path, {"jobId": job_id})
    batch_size = max(1, min(100, int(payload.get("batchSize") or 20)))
    run_count = max(1, min(50, int(payload.get("runCount") or 1)))
    suppliers = payload.get("suppliers") if isinstance(payload.get("suppliers"), list) else []
    upload_date = text_value(payload.get("upload_date"))
    sort_order = text_value(payload.get("sort_order")) or "latest"
    filter_mode = text_value(payload.get("filter_mode")) or "available"
    channels = payload.get("channels") if isinstance(payload.get("channels"), list) else []
    settings = payload.get("listingImageSettings") or {}
    batches: list[dict] = []

    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "progressPercent": 2,
        "currentStage": "자동화 대상 수집",
        "batchSize": batch_size,
        "countPerSupplier": batch_size,
        "runCount": run_count,
        "selectedSupplierCount": len(suppliers),
        "expectedPerRun": len(suppliers) * batch_size,
        "expectedTotal": len(suppliers) * batch_size * run_count,
        "updatedAt": now_text(),
    })
    write_json(job_path, job)

    try:
        selected_batches = pm_select_automation_sku_batches(suppliers, upload_date, sort_order, filter_mode, batch_size, run_count)
        skus = [sku for batch in selected_batches for sku in batch]
        if not selected_batches:
            raise ValueError("자동화 대상 SKU가 없습니다.")

        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START automationPrepare\n")
            log.write(f"countPerSupplier={batch_size} suppliers={len(suppliers)} runCount={run_count} selectedSku={len(skus)} channels={channels}\n")

        def run_child_with_parent_status(child_job_id: str, runner, runner_payload: dict, parent_stage: str, progress_start: int, progress_end: int) -> dict:
            child_path = JOBS_ROOT / f"{child_job_id}.json"
            thread = threading.Thread(target=runner, args=(child_job_id, runner_payload), daemon=True)
            thread.start()
            last_tail: list[str] = []
            while thread.is_alive():
                if job_id in CANCELLED_JOB_IDS:
                    with ACTIVE_PROCESS_LOCK:
                        child_processes = list(ACTIVE_PROCESSES.get(child_job_id) or [])
                    for process in child_processes:
                        stop_process_tree(process)
                    CANCELLED_JOB_IDS.add(child_job_id)
                    break
                child = read_json(child_path, {})
                child_progress = int(child.get("progressPercent") or 0)
                mapped = progress_start + int((progress_end - progress_start) * min(max(child_progress, 0), 100) / 100)
                tail = child.get("tail") if isinstance(child.get("tail"), list) else []
                if tail:
                    last_tail = tail[-8:]
                parent = read_json(job_path, job)
                parent.update({
                    "status": "running",
                    "progressPercent": max(3, min(98, mapped)),
                    "currentStage": f"{parent_stage} · {child.get('currentStage') or child.get('status') or '진행 중'}",
                    "childJobId": child_job_id,
                    "childProgressPercent": child_progress,
                    "currentGs": child.get("currentGs", parent.get("currentGs", "")),
                    "tail": last_tail,
                    "updatedAt": now_text(),
                })
                write_json(job_path, parent)
                time.sleep(2)
            thread.join()
            return read_json(child_path, {})

        for index, batch_skus in enumerate(selected_batches, start=1):
            if job_id in CANCELLED_JOB_IDS:
                raise RuntimeError("cancelled")
            progress_base = int(((index - 1) / max(1, len(selected_batches))) * 96)
            job = read_json(job_path, job)
            job.update({
                "progressPercent": max(3, min(95, progress_base + 3)),
                "currentStage": f"{index}/{len(selected_batches)} 배치 소스 CSV 생성",
                "currentBatch": index,
                "updatedAt": now_text(),
            })
            write_json(job_path, job)

            products = pm_get_products_by_skus(batch_skus)
            csv_path = pm_build_source_csv(products)
            preview = parse_source_preview(csv_path)
            selected_gs = [text_value(row.get("gs")) for row in preview.get("preview", []) if text_value(row.get("gs"))]
            if not selected_gs:
                raise ValueError(f"{index}번 배치에서 GS 코드를 찾지 못했습니다.")

            seed_job_id = uuid.uuid4().hex[:12]
            seed_job_path = JOBS_ROOT / f"{seed_job_id}.json"
            write_json(seed_job_path, {
                "ok": True,
                "jobId": seed_job_id,
                "action": "sourceToSeed",
                "status": "queued",
                "createdAt": now_text(),
                "parentJobId": job_id,
            })
            seed_job = run_child_with_parent_status(seed_job_id, run_seed_job, {
                "sourcePath": str(csv_path),
                "sourceFilePath": str(csv_path),
                "selectedGs": selected_gs,
                "listingImageSettings": settings,
            }, f"{index}/{len(selected_batches)} 배치 소스 CSV/전처리", progress_base + 3, progress_base + 34)
            if seed_job.get("status") != "completed":
                raise RuntimeError(f"{index}번 배치 시드 생성 실패: {seed_job.get('error') or 'unknown error'}")
            seed_path = text_value((seed_job.get("result") or {}).get("seedPath"))

            keyword_job_id = ""
            keyword_job = {}
            if channels:
                keyword_job_id = uuid.uuid4().hex[:12]
                keyword_job_path = JOBS_ROOT / f"{keyword_job_id}.json"
                write_json(keyword_job_path, {
                    "ok": True,
                    "jobId": keyword_job_id,
                    "action": "keywordGenerate",
                    "status": "queued",
                    "createdAt": now_text(),
                    "parentJobId": job_id,
                })
                job = read_json(job_path, job)
                job.update({
                    "progressPercent": max(5, min(97, progress_base + 35)),
                    "currentStage": f"{index}/{len(selected_batches)} 배치 키워드 생성",
                    "updatedAt": now_text(),
                })
                write_json(job_path, job)
                keyword_job = run_child_with_parent_status(keyword_job_id, run_keyword_job, {
                    "seedPath": seed_path,
                    "selectedGs": selected_gs,
                    "channels": channels,
                    "listingImageSettings": settings,
                    "concurrency": 50,
                    "keywordProductChunkSize": KEYWORD_PRODUCT_CHUNK_SIZE,
                    "keywordStallSeconds": KEYWORD_CODEX_STALL_SECONDS,
                }, f"{index}/{len(selected_batches)} 배치 키워드 생성", progress_base + 35, progress_base + 92)
                if keyword_job.get("status") != "completed":
                    raise RuntimeError(f"{index}번 배치 키워드 생성 실패: {keyword_job.get('error') or 'unknown error'}")

            batch_result = {
                "index": index,
                "skuCount": len(batch_skus),
                "gsCount": len(selected_gs),
                "sourcePath": str(csv_path),
                "seedJobId": seed_job_id,
                "keywordJobId": keyword_job_id,
                "seedPath": seed_path,
                "seedFileName": Path(seed_path).name if seed_path else "",
                "status": "completed",
                "supplierCount": len(suppliers),
                "countPerSupplier": batch_size,
            }
            batches.append(batch_result)
            job = read_json(job_path, job)
            job.update({
                "batches": batches,
                "progressPercent": min(98, int((index / max(1, len(selected_batches))) * 98)),
                "currentStage": f"{index}/{len(selected_batches)} 배치 완료",
                "updatedAt": now_text(),
            })
            write_json(job_path, job)

        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "자동화 시드 생성 완료",
            "batches": batches,
            "result": {
                "totalBatches": len(batches),
                "totalSkus": sum(item.get("skuCount", 0) for item in batches),
                "totalGs": sum(item.get("gsCount", 0) for item in batches),
                "supplierCount": len(suppliers),
                "countPerSupplier": batch_size,
                "seedPaths": [item.get("seedPath") for item in batches if item.get("seedPath")],
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        if job_id in CANCELLED_JOB_IDS:
            job.update({
                "status": "cancelled",
                "finishedAt": now_text(),
                "progressPercent": 100,
                "currentStage": "사용자 중지",
                "batches": batches,
            })
        else:
            with log_path.open("a", encoding="utf-8", errors="replace") as log:
                log.write(f"\n[{now_text()}] ERROR {exc}\n")
            job.update({
                "status": "failed",
                "finishedAt": now_text(),
                "progressPercent": 100,
                "currentStage": "자동화 실패",
                "error": str(exc),
                "batches": batches,
            })
        write_json(job_path, job)


def normalize_upload_entries(payload: dict) -> list[dict]:
    rows = payload.get("rows")
    if not isinstance(rows, list):
        rows = []
    entries: list[dict] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        account = text_value(row.get("account", "")).upper()
        market = text_value(row.get("market", ""))
        gs = text_value(row.get("gs", "")).upper()
        if not account or not market or not gs:
            continue
        channel = text_value(row.get("channel") or row.get("channelKey") or f"{account}:{market}")
        additional_image_srcs = row.get("additionalImageSrcs")
        if not isinstance(additional_image_srcs, list):
            additional_image_srcs = extract_image_urls(row.get("additionalImageSrcs", ""))
        detail_image_srcs = row.get("detailImageSrcs")
        if not isinstance(detail_image_srcs, list):
            detail_image_srcs = extract_image_urls(row.get("detailImageSrcs", ""))
        option_items = row.get("optionItems")
        if not isinstance(option_items, list):
            option_items = []
        option_additionals = row.get("optionAdditionalAmounts")
        if not isinstance(option_additionals, list):
            option_additionals = []
        tags = row.get("tags")
        if not isinstance(tags, list):
            tags = split_candidate_terms(row.get("tags", ""))
        naver_tags = row.get("naverTags")
        if not isinstance(naver_tags, list):
            naver_tags = split_candidate_terms(row.get("naverTags", ""))
        entries.append({
            "queueKey": text_value(row.get("queueKey") or f"{channel}:{gs}"),
            "account": account,
            "market": market,
            "channel": channel,
            "gs": gs,
            "sourceName": text_value(row.get("sourceName", "")),
            "title": text_value(row.get("title", "")),
            "searchTerms": text_value(row.get("searchTerms", "")),
            "tags": seo_tag_terms(tags, 30),
            "naverTags": seo_tag_terms(naver_tags or tags, 10),
            "mainImage": text_value(row.get("mainImage", "")),
            "mainImageSrc": text_value(row.get("mainImageSrc", "")),
            "additionalImageSrcs": [text_value(url) for url in additional_image_srcs if text_value(url)],
            "detailImageSrcs": [text_value(url) for url in detail_image_srcs if text_value(url)],
            "detailHtml": normalize_detail_html_for_upload(row.get("detailHtml", "")),
            "cafe24Url": text_value(row.get("cafe24Url", "")),
            "price": text_value(row.get("price") or row.get("salePrice") or ""),
            "supplyPrice": text_value(row.get("supplyPrice") or ""),
            "salePrice": text_value(row.get("salePrice") or row.get("price") or ""),
            "consumerPrice": text_value(row.get("consumerPrice") or ""),
            "optionSummary": text_value(row.get("optionSummary") or row.get("opt") or ""),
            "optionInput": text_value(row.get("optionInput") or ""),
            "optionAdditionalAmounts": option_additionals,
            "optionItems": option_items,
            "naverProvidedNotice": row.get("naverProvidedNotice") if isinstance(row.get("naverProvidedNotice"), dict) else {},
            "brand": text_value(row.get("brand", "")),
            "force": bool(row.get("force")),
            "categories": row.get("categories") if isinstance(row.get("categories"), dict) else {},
        })
    return entries


def parse_upload_price(value: object) -> int:
    text = text_value(value)
    if not text:
        return 0
    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned:
        return 0
    try:
        return int(float(cleaned))
    except Exception:
        return 0


def option_summary_to_input(option_summary: object) -> str:
    return option_input_from_labels(option_labels_from_input(option_summary))


def option_input_from_entry(entry: dict) -> str:
    existing = text_value(entry.get("optionInput"))
    if existing:
        return option_summary_to_input(existing) or existing
    items = entry.get("optionItems") if isinstance(entry.get("optionItems"), list) else []
    labels: list[str] = []
    for item in items:
        raw_label = item.get("option") if isinstance(item, dict) else item
        labels.extend(option_labels_from_input(raw_label))
    if labels:
        return option_input_from_labels(labels)
    return option_summary_to_input(entry.get("optionSummary"))


def option_additionals_from_entry(entry: dict) -> str:
    existing = entry.get("optionAdditionalAmounts")
    if isinstance(existing, list) and existing:
        return "|".join(str(int(parse_upload_price(value))) for value in existing)
    items = entry.get("optionItems") if isinstance(entry.get("optionItems"), list) else []
    if len(items) > 1:
        prices = [
            parse_upload_price(item.get("salePrice") or item.get("price"))
            for item in items
            if isinstance(item, dict)
        ]
        if len(prices) > 1:
            base = prices[0]
            return "|".join(str(price - base) for price in prices)

    labels = option_labels_from_input(option_input_from_entry(entry))
    base_value, _suffix = split_gs(text_value(entry.get("baseGs") or entry.get("gs")))
    pm_items = pm_option_items_for_base(base_value)
    if len(pm_items) > 1:
        prices = [parse_upload_price(item.get("salePrice") or item.get("price")) for item in pm_items]
        base = prices[0]
        count = len(labels) if labels else len(prices)
        return "|".join(str(price - base) for price in prices[:count])
    if len(labels) > 1:
        return "|".join("0" for _ in labels)
    return ""


def upload_image_src_to_path(src: object) -> Path | None:
    raw = text_value(src)
    if not raw:
        return None
    parsed = urllib.parse.urlparse(raw)
    path_text = ""
    if parsed.scheme in {"http", "https"}:
        if parsed.path.startswith("/data/"):
            path_text = urllib.parse.unquote(parsed.path.lstrip("/"))
    elif raw.startswith("/data/"):
        path_text = urllib.parse.unquote(raw.lstrip("/"))
    elif raw.startswith("data/"):
        path_text = urllib.parse.unquote(raw)
    elif Path(raw).is_absolute():
        path = Path(raw)
        return path.resolve() if path.exists() else None

    if not path_text:
        return None
    path = (ROOT / path_text).resolve()
    return path if path.exists() and is_within(DATA_ROOT, path) else None


def public_image_url(src: object) -> str:
    raw = text_value(src)
    if raw.startswith("http://") or raw.startswith("https://"):
        parsed = urllib.parse.urlparse(raw)
        file_name = Path(urllib.parse.unquote(parsed.path)).name
        if file_name and not re.match(r"^[A-Za-z0-9._%+\-=()]+$", file_name):
            return ""
        return raw
    return ""


def direct_upload_image_ref(src: object) -> str:
    path = upload_image_src_to_path(src)
    if path is not None:
        return str(path)
    return public_image_url(src)


def infer_additional_upload_images(entry: dict) -> list[str]:
    explicit = entry.get("additionalImageSrcs")
    if isinstance(explicit, list) and explicit:
        return [text_value(url) for url in explicit if text_value(url)]

    main_path = upload_image_src_to_path(entry.get("mainImageSrc"))
    if main_path is None or not main_path.is_file():
        return []

    base_code = base_gs_code(entry.get("baseGs") or entry.get("gs"))
    if not base_code:
        return []

    siblings = sorted(
        main_path.parent.glob(f"{base_code}_*.jpg"),
        key=image_file_sort_key,
    )
    return [str(path) for path in siblings if path.resolve() != main_path.resolve()][:8]


def seed_detail_images_for_gs(gs: object) -> list[str]:
    gs_value = text_value(gs).upper()
    if not gs_value:
        return []
    base_value = base_gs_code(gs_value)
    for seed_path in sorted(SEED_ROOT.glob("*.webseed.json"), key=lambda path: path.stat().st_mtime, reverse=True):
        seed = read_json(seed_path, {})
        products = seed.get("products")
        if not isinstance(products, list):
            continue
        for product in products:
            if not isinstance(product, dict):
                continue
            product_gs = text_value(product.get("gs")).upper()
            product_base = text_value(product.get("baseGs") or base_gs_code(product_gs)).upper()
            if gs_value not in {product_gs, product_base} and base_value not in {product_gs, product_base}:
                continue
            images = product.get("images")
            if not isinstance(images, dict):
                return []
            detail = images.get("detail")
            if isinstance(detail, list):
                return [text_value(url) for url in detail if text_value(url)]
            return extract_image_urls(detail)
    return []


def seed_detail_html_for_gs(gs: object) -> str:
    gs_value = text_value(gs).upper()
    if not gs_value:
        return ""
    base_value = base_gs_code(gs_value)
    for seed_path in sorted(SEED_ROOT.glob("*.webseed.json"), key=lambda path: path.stat().st_mtime, reverse=True):
        seed = read_json(seed_path, {})
        products = seed.get("products")
        if not isinstance(products, list):
            continue
        for product in products:
            if not isinstance(product, dict):
                continue
            product_gs = text_value(product.get("gs")).upper()
            product_base = text_value(product.get("baseGs") or base_gs_code(product_gs)).upper()
            if gs_value not in {product_gs, product_base} and base_value not in {product_gs, product_base}:
                continue
            return text_value(product.get("detailHtml"))
    return ""


def base_gs_code(gs: object) -> str:
    value = text_value(gs).upper()
    match = GS_CODE_RE.search(value)
    return match.group(1).upper() if match else value


def image_selection_for_entry(entry: dict) -> tuple[int | None, list[int]]:
    main_path = upload_image_src_to_path(entry.get("mainImageSrc"))
    add_paths = [upload_image_src_to_path(url) for url in entry.get("additionalImageSrcs", [])]
    paths = [path for path in [main_path, *add_paths] if path is not None]
    if not paths:
        return None, []
    folder = paths[0].parent
    files = [
        path.resolve()
        for path in sorted(folder.iterdir(), key=image_file_sort_key)
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
    ]
    if not files:
        return None, []
    index_by_path = {str(path): index for index, path in enumerate(files)}
    main_index = index_by_path.get(str(main_path.resolve())) if main_path else None
    additional: list[int] = []
    for path in add_paths:
        if path is None:
            continue
        index = index_by_path.get(str(path.resolve()))
        if index is not None and index != main_index and index not in additional:
            additional.append(index)
    return main_index, additional


def update_export_image_selections(entries: list[dict]) -> None:
    if not entries:
        return
    target = EXPORT_ROOT / "image_selections.json"
    payload = read_json(target, {})
    if not isinstance(payload, dict):
        payload = {}
    nested = payload.get("image_selections")
    if not isinstance(nested, dict):
        nested = {}
    changed = False
    for entry in entries:
        gs_key = base_gs_code(entry.get("gs"))
        if not gs_key:
            continue
        main_index, additional = image_selection_for_entry(entry)
        if main_index is None:
            continue
        selection = {"main": main_index, "additional": additional}
        payload[gs_key[:9]] = selection
        nested[gs_key[:9]] = selection
        changed = True
    if changed:
        payload["image_selections"] = nested
        write_json(target, payload)


def infer_direct_upload_categories(entry: dict) -> dict[str, str]:
    text = " ".join(
        text_value(entry.get(key))
        for key in ("title", "sourceName", "searchTerms", "optionSummary", "gs")
    )
    compact = re.sub(r"\s+", "", text)
    out: dict[str, str] = {}
    explicit = entry.get("categories")
    if isinstance(explicit, dict):
        for key in ("naver", "coupang", "lotte_standard", "lotte_display", "lotte_item", "elevenst", "11st", "esm", "auction", "gmarket"):
            val = text_value(explicit.get(key))
            if val:
                out[key] = val

    def apply(**values: str) -> None:
        for key, value in values.items():
            if value and not out.get(key):
                out[key] = value

    if re.search(r"깔창|인솔|신발밑창|밑창보강", compact, re.IGNORECASE):
        apply(
            naver="50000667",
            coupang="64623",
            lotte_standard="BC43071000",
            lotte_display="FC18101001",
            lotte_item="38",
        )
    if re.search(r"카라비너|릴고리|릴홀더|키홀더|키링|연결고리|등산고리|고리", compact, re.IGNORECASE):
        apply(
            naver="50002646",
            coupang="81718",
            lotte_standard="BC20040800",
            lotte_display="EC10400324",
            lotte_item="38",
        )
    if re.search(r"에어컨.*커버|실외기.*커버|커버.*실외기", compact, re.IGNORECASE):
        apply(
            naver="50003518",
            coupang="78137",
            lotte_standard="BC63120300",
            lotte_display="FC11160703",
            lotte_item="38",
        )
    if re.search(r"가구발커버|의자발커버|가구커버|소파커버|커버류", compact, re.IGNORECASE):
        apply(
            naver="50003521",
            coupang="78133",
            lotte_standard="BC63120300",
            lotte_display="FC11160703",
            lotte_item="38",
        )
    if re.search(r"작업장갑|안전장갑|코팅장갑|나일론.*장갑|오픈핑거.*장갑|장갑", compact, re.IGNORECASE):
        apply(
            naver="50003450",
            coupang="64387",
            lotte_standard="BC10040800",
            lotte_display="FC19041003",
            lotte_item="04",
        )
    if re.search(r"나사|스크류|볼트|너트|브라켓|고정핀|철물|부속|부품", compact, re.IGNORECASE):
        apply(
            naver="50003466",
            coupang="64310",
            lotte_standard="BC10080200",
            lotte_display="FC19040401",
            lotte_item="38",
        )
    if not out.get("naver"):
        matches = search_category_reference("naver", text, 1)
        if matches:
            out["naver"] = text_value(matches[0].get("code"))
    if not out.get("coupang"):
        matches = search_category_reference("coupang", text, 1)
        if matches:
            out["coupang"] = text_value(matches[0].get("code"))
    if not out.get("elevenst") and not out.get("11st"):
        matches = search_category_reference("11st", text, 1)
        if matches:
            out["elevenst"] = text_value(matches[0].get("code"))
    if not out.get("esm"):
        matches = search_category_reference("esm", text, 1)
        if matches:
            out["esm"] = text_value(matches[0].get("code"))
    if not out.get("lotte_standard") or not out.get("lotte_display"):
        matches = search_category_reference("lotteon", text, 12)
        if not out.get("lotte_standard"):
            standard = next((item for item in matches if text_value(item.get("type")).lower() == "standard"), None)
            if standard:
                out["lotte_standard"] = text_value(standard.get("code"))
        if not out.get("lotte_display"):
            display = next((item for item in matches if text_value(item.get("type")).lower() == "display"), None)
            if display:
                out["lotte_display"] = text_value(display.get("code"))
    return out


def naver_notice_payload_for_entry(entry: dict) -> dict:
    notice = entry.get("naverProvidedNotice") if isinstance(entry.get("naverProvidedNotice"), dict) else {}
    payload = notice.get("productInfoProvidedNotice") if isinstance(notice.get("productInfoProvidedNotice"), dict) else {}
    current_type = text_value(payload.get("productInfoProvidedNoticeType") or notice.get("productInfoProvidedNoticeType")).upper()
    candidate = build_naver_provided_notice({
        "gs": entry.get("gs", ""),
        "baseGs": split_gs(entry.get("gs", ""))[0],
        "name": entry.get("title") or entry.get("sourceName") or entry.get("gs", ""),
        "sourceName": entry.get("sourceName", ""),
        "opt": entry.get("optionSummary", ""),
        "optionItems": entry.get("optionItems", []),
    }, {
        "rawText": " ".join([
            text_value(entry.get("title")),
            text_value(entry.get("sourceName")),
            text_value(entry.get("searchTerms")),
            text_value(entry.get("detailHtml")),
        ])
    })
    candidate_payload = candidate.get("productInfoProvidedNotice") if isinstance(candidate, dict) else {}
    candidate_type = text_value(candidate_payload.get("productInfoProvidedNoticeType")).upper() if isinstance(candidate_payload, dict) else ""
    if not payload or (current_type == "ETC" and candidate_type and candidate_type != "ETC"):
        return candidate_payload if isinstance(candidate_payload, dict) else {}
    return payload


def write_direct_upload_workbook(job_id: str, entry: dict) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl 로드 실패: {exc}") from exc

    file_name = safe_name(f"api_upload_{job_id}_{entry['channel']}_{entry['gs']}.xlsx")
    target = EXPORT_ROOT / file_name
    title = clean_product_title(entry.get("title")) or clean_product_title(entry.get("sourceName")) or entry["gs"]
    search_terms = text_value(entry.get("searchTerms"))
    naver_tag_terms = seo_tag_terms(entry.get("naverTags") or entry.get("tags") or search_terms, 10)
    naver_tag_text = ", ".join(naver_tag_terms) if naver_tag_terms else search_terms
    price = parse_upload_price(entry.get("salePrice") or entry.get("price")) or 1000
    consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2) or price
    option_input = option_input_from_entry(entry)
    option_additionals = option_additionals_from_entry(entry)
    additional_sources = infer_additional_upload_images(entry)
    upload_images = [
        url for url in [
            direct_upload_image_ref(entry.get("mainImageSrc")),
            *[direct_upload_image_ref(url) for url in additional_sources],
        ]
        if url
    ]
    public_images = [
        url for url in [
            public_image_url(entry.get("mainImageSrc")),
            *[public_image_url(url) for url in additional_sources],
        ]
        if url
    ]
    detail_sources = entry.get("detailImageSrcs") if isinstance(entry.get("detailImageSrcs"), list) else []
    if not detail_sources:
        detail_sources = seed_detail_images_for_gs(entry.get("gs"))
    public_detail_images = [
        public_image_url(url)
        for url in detail_sources
        if public_image_url(url)
    ]
    image_list = "|".join(upload_images[:9])
    detail_html = normalize_detail_html_for_upload(entry.get("detailHtml")) or normalize_detail_html_for_upload(seed_detail_html_for_gs(entry.get("gs")))
    if detail_html and "<img" not in detail_html.lower() and public_detail_images:
        detail_html = ""
    if not detail_html and public_detail_images:
        detail_html = "<center>" + "".join(f'<img src="{url}">' for url in public_detail_images[:80]) + "</center>"
    elif not detail_html and public_images:
        detail_html = "<center>" + "".join(f'<img src="{url}">' for url in public_images[:12]) + "</center>"
    if not detail_html:
        detail_html = "상세페이지 참조"
    categories = infer_direct_upload_categories(entry)
    upload_category_mode = text_value(entry.get("categoryMode")).lower() == "upload"
    naver_category = "" if upload_category_mode and entry["market"] == "네이버" else categories.get("naver", "")
    coupang_category = "" if upload_category_mode and entry["market"] == "쿠팡" else categories.get("coupang", "")
    naver_notice_payload = naver_notice_payload_for_entry(entry)
    naver_notice_json = json.dumps(naver_notice_payload, ensure_ascii=False, separators=(",", ":")) if naver_notice_payload else ""

    headers = [
        "상품코드", "자체 상품코드", "판매자내부상품번호", "상품명", "공급사 상품명",
        "홈런_공통마켓상품명", "홈런_네이버상품명", "네이버상품명", "홈런_네이버태그", "네이버태그",
        "홈런_쿠팡상품명", "쿠팡상품명", "홈런_쿠팡검색태그", "쿠팡검색태그",
        "홈런_롯데ON상품명", "롯데ON상품명", "홈런_롯데ON검색키워드", "롯데ON검색키워드",
        "네이버카테고리코드", "쿠팡카테고리코드", "롯데ON표준카테고리코드", "롯데ON전시카테고리코드", "롯데ON상품품목코드",
        "홈런_공통마켓검색키워드", "공통마켓검색키워드", "검색어설정", "검색키워드",
        "판매가", "상품가", "소비자가", "옵션입력", "옵션추가금",
        "이미지등록(목록)", "이미지등록(추가)", "이미지등록(상세)",
        "상품 상세설명", "상세설명", "상품정보제공고시", "네이버상품정보고시", "브랜드",
    ]
    row = {
        "상품코드": entry["gs"],
        "자체 상품코드": entry["gs"],
        "판매자내부상품번호": entry["gs"],
        "상품명": title,
        "공급사 상품명": title,
        "홈런_공통마켓상품명": title,
        "홈런_네이버상품명": title if entry["market"] == "네이버" else "",
        "네이버상품명": title if entry["market"] == "네이버" else "",
        "홈런_네이버태그": naver_tag_text if entry["market"] == "네이버" else "",
        "네이버태그": naver_tag_text if entry["market"] == "네이버" else "",
        "홈런_쿠팡상품명": title if entry["market"] == "쿠팡" else "",
        "쿠팡상품명": title if entry["market"] == "쿠팡" else "",
        "홈런_쿠팡검색태그": search_terms if entry["market"] == "쿠팡" else "",
        "쿠팡검색태그": search_terms if entry["market"] == "쿠팡" else "",
        "홈런_롯데ON상품명": title if entry["market"] == "롯데ON" else "",
        "롯데ON상품명": title if entry["market"] == "롯데ON" else "",
        "홈런_롯데ON검색키워드": search_terms if entry["market"] == "롯데ON" else "",
        "롯데ON검색키워드": search_terms if entry["market"] == "롯데ON" else "",
        "네이버카테고리코드": naver_category,
        "쿠팡카테고리코드": coupang_category,
        "롯데ON표준카테고리코드": categories.get("lotte_standard", ""),
        "롯데ON전시카테고리코드": categories.get("lotte_display", ""),
        "롯데ON상품품목코드": categories.get("lotte_item", ""),
        "홈런_공통마켓검색키워드": search_terms,
        "공통마켓검색키워드": search_terms,
        "검색어설정": search_terms,
        "검색키워드": search_terms,
        "판매가": price,
        "상품가": price,
        "소비자가": consumer_price,
        "옵션입력": option_input,
        "옵션추가금": option_additionals,
        "이미지등록(목록)": upload_images[0] if upload_images else "",
        "이미지등록(추가)": "|".join(upload_images[1:9]),
        "이미지등록(상세)": "|".join(
            [direct_upload_image_ref(url) for url in detail_sources if direct_upload_image_ref(url)][:20]
        ) if detail_sources else "",
        "상품 상세설명": detail_html,
        "상세설명": detail_html,
        "상품정보제공고시": naver_notice_json if entry["market"] == "네이버" else "",
        "네이버상품정보고시": naver_notice_json if entry["market"] == "네이버" else "",
        "브랜드": text_value(entry.get("brand")) or "샤플라이",
    }

    workbook = Workbook()
    first = True
    for sheet_name in ("분리추출후", "A마켓", "B마켓"):
        sheet = workbook.active if first else workbook.create_sheet(sheet_name)
        first = False
        sheet.title = sheet_name
        sheet.append(headers)
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(target)
    return target


def market_cli_flag(market: str) -> str:
    return {
        "네이버": "--naver",
        "쿠팡": "--coupang",
        "롯데ON": "--lotteon",
    }.get(market, "")


def resolve_market_key_path(settings: dict, account: str, market: str) -> Path:
    item = settings.get(market_key_id(account, market))
    if not isinstance(item, dict):
        raise FileNotFoundError(f"{account}:{market} 키 파일이 없습니다.")
    path = resolve_market_key_item_path(item)
    if path is None:
        raise FileNotFoundError(f"{account}:{market} 키 파일을 찾지 못했습니다.")
    return path


def backup_or_remove_key_file(target: Path, temp_dir: Path) -> tuple[str, Path | None]:
    if target.exists():
        backup = temp_dir / target.name
        shutil.copy2(target, backup)
        return "backup", backup
    return "missing", None


@contextmanager
def market_key_overlay(account: str, market: str, settings: dict, log=None):
    DESKTOP_KEY_ROOT.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(tempfile.mkdtemp(prefix="webocr_key_overlay_", dir=str(EXPORT_ROOT)))
    restored: list[tuple[Path, str, Path | None]] = []

    def overlay_file(source: Path, target_name: str, hide_only: bool = False) -> None:
        target = (DESKTOP_KEY_ROOT / target_name).resolve()
        if source.resolve() == target:
            return
        state, backup = backup_or_remove_key_file(target, temp_dir)
        restored.append((target, state, backup))
        if hide_only:
            if target.exists():
                target.unlink()
            return
        shutil.copy2(source, target)

    with MARKET_KEY_OVERLAY_LOCK:
        try:
            source = resolve_market_key_path(settings, account, market)
            if market == "네이버":
                overlay_file(source, "naver_client_key.txt")
            elif market == "쿠팡":
                overlay_file(source, "coupang_wing_api.txt")
            elif market == "롯데ON":
                if source.suffix.lower() == ".json":
                    overlay_file(source, "lotteon_upload_id.json")
                else:
                    overlay_file(source, "lotteon_api.txt")
                    overlay_file(source, "lotteon_upload_id.json", hide_only=True)
            else:
                raise ValueError(f"{market}은 API 업로드 대상이 아닙니다.")

            cafe24_item = settings.get(market_key_id(account, "Cafe24"))
            if isinstance(cafe24_item, dict):
                cafe24_path = resolve_market_key_item_path(cafe24_item)
                if cafe24_path is not None:
                    refreshed, message = refresh_cafe24_upload_token(cafe24_path)
                    if log:
                        status = "완료" if refreshed else "실패"
                        log.write(f"[{now_text()}] {account}:Cafe24 토큰 리프레쉬 {status}: {scrub_secret(message)}\n")
                    overlay_file(cafe24_path, cafe24_path.name)
                    overlay_file(cafe24_path, "cafe24_token.json")
            yield
        finally:
            for target, state, backup in reversed(restored):
                try:
                    if state == "backup" and backup and backup.exists():
                        shutil.copy2(backup, target)
                    elif state == "missing" and target.exists():
                        target.unlink()
                except Exception:
                    pass
            shutil.rmtree(temp_dir, ignore_errors=True)


@contextmanager
def account_wide_overlay(account: str, markets: set[str], settings: dict, log=None):
    DESKTOP_KEY_ROOT.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(tempfile.mkdtemp(prefix="webocr_acct_overlay_", dir=str(EXPORT_ROOT)))
    restored: list[tuple[Path, str, Path | None]] = []

    def overlay_file(source: Path, target_name: str, hide_only: bool = False) -> None:
        target = (DESKTOP_KEY_ROOT / target_name).resolve()
        if source.resolve() == target:
            return
        if any(t.resolve() == target for t, _, _ in restored):
            return
        state, backup = backup_or_remove_key_file(target, temp_dir)
        restored.append((target, state, backup))
        if hide_only:
            if target.exists():
                target.unlink()
            return
        shutil.copy2(source, target)

    with MARKET_KEY_OVERLAY_LOCK:
        try:
            for market in markets:
                try:
                    source = resolve_market_key_path(settings, account, market)
                except FileNotFoundError:
                    continue
                if market == "네이버":
                    overlay_file(source, "naver_client_key.txt")
                elif market == "쿠팡":
                    overlay_file(source, "coupang_wing_api.txt")
                elif market == "롯데ON":
                    if source.suffix.lower() == ".json":
                        overlay_file(source, "lotteon_upload_id.json")
                    else:
                        overlay_file(source, "lotteon_api.txt")
                        overlay_file(source, "lotteon_upload_id.json", hide_only=True)
                    category_map_source = (Path.home() / "Desktop" / "key" / "lotteon_category_map.json").resolve()
                    if category_map_source.exists():
                        overlay_file(category_map_source, "lotteon_category_map.json")
            cafe24_item = settings.get(market_key_id(account, "Cafe24"))
            if isinstance(cafe24_item, dict):
                cafe24_path = resolve_market_key_item_path(cafe24_item)
                if cafe24_path is not None:
                    refreshed, message = refresh_cafe24_upload_token(cafe24_path)
                    if log:
                        status = "완료" if refreshed else "실패"
                        log.write(f"[{now_text()}] {account}:Cafe24 토큰 리프레쉬 {status}: {scrub_secret(message)}\n")
                    overlay_file(cafe24_path, cafe24_path.name)
                    overlay_file(cafe24_path, "cafe24_token.json")
            yield
        finally:
            for target, state, backup in reversed(restored):
                try:
                    if state == "backup" and backup and backup.exists():
                        shutil.copy2(backup, target)
                    elif state == "missing" and target.exists():
                        target.unlink()
                except Exception:
                    pass
            shutil.rmtree(temp_dir, ignore_errors=True)


@contextmanager
def account_key_root(account: str, markets: set[str], settings: dict, log=None):
    temp_dir = Path(tempfile.mkdtemp(prefix=f"webocr_keys_{safe_name(account, 'account')}_", dir=str(EXPORT_ROOT)))

    def copy_key(source: Path, target_name: str) -> None:
        target = temp_dir / target_name
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)

    try:
        for market in markets:
            try:
                source = resolve_market_key_path(settings, account, market)
            except FileNotFoundError:
                continue
            if market == "네이버":
                copy_key(source, "naver_client_key.txt")
            elif market == "쿠팡":
                copy_key(source, "coupang_wing_api.txt")
            elif market == "롯데ON":
                if source.suffix.lower() == ".json":
                    copy_key(source, "lotteon_upload_id.json")
                else:
                    copy_key(source, "lotteon_api.txt")
                category_map_source = (Path.home() / "Desktop" / "key" / "lotteon_category_map.json").resolve()
                if category_map_source.exists():
                    copy_key(category_map_source, "lotteon_category_map.json")

        cafe24_item = settings.get(market_key_id(account, "Cafe24"))
        if isinstance(cafe24_item, dict):
            cafe24_path = resolve_market_key_item_path(cafe24_item)
            if cafe24_path is not None:
                refreshed, message = refresh_cafe24_upload_token(cafe24_path)
                if log:
                    status = "완료" if refreshed else "실패"
                    log.write(f"[{now_text()}] {account}:Cafe24 토큰 리프레쉬 {status}: {scrub_secret(message)}\n")
                copy_key(cafe24_path, cafe24_path.name)
                copy_key(cafe24_path, "cafe24_token.json")
        yield temp_dir
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def parse_direct_upload_result(entry: dict, exit_code: int, lines: list[str]) -> dict:
    result = {**entry, "status": "failed", "updatedAt": now_text(), "error": "", "productId": ""}
    tail_text = "\n".join(lines[-80:])
    if exit_code != 0:
        result["error"] = scrub_secret(tail_text or f"dotnet exit code {exit_code}", 400)
        return result

    market = entry["market"]
    if market == "네이버":
        pattern = r"\[네이버\]\s+row=\d+\s+status=(\S+)\s+id=(.*?)\s+error=(.*)"
        id_name = "productId"
    elif market == "쿠팡":
        pattern = r"\[쿠팡\]\s+row=\d+\s+status=(\S+)\s+id=(.*?)\s+category=.*?\s+error=(.*)"
        id_name = "sellerProductId"
    else:
        pattern = r"\[롯데ON\]\s+row=\d+\s+status=(\S+)\s+spdNo=(.*?)\s+error=(.*)"
        id_name = "spdNo"

    matches = re.findall(pattern, tail_text)
    if not matches:
        result["error"] = scrub_secret(tail_text or "업로드 결과를 파싱하지 못했습니다.", 400)
        return result

    status, product_id, error = matches[-1]
    status_upper = status.upper()
    result[id_name] = product_id.strip()
    result["productId"] = product_id.strip()
    result["rawStatus"] = status_upper
    if status_upper in {"OK", "SUCCESS", "DRY_RUN_OK", "DRY_RUN"}:
        result["status"] = "uploaded"
    elif status_upper.startswith("SKIP_DUP"):
        result["status"] = "skipped"
        result["duplicate"] = True
        result["error"] = error.strip() or "기존 성공 이력 있음"
    else:
        result["status"] = "failed"
        result["error"] = error.strip() or status_upper
    return result


def parse_dt_from_file_path(path: Path) -> datetime:
    match = re.search(r"(20\d{6})_(\d{6})", str(path))
    if match:
        try:
            return datetime.strptime(match.group(1) + match.group(2), "%Y%m%d%H%M%S")
        except Exception:
            pass
    try:
        return datetime.fromtimestamp(path.stat().st_mtime)
    except Exception:
        return datetime.min


def canonical_upload_market(value: object) -> str:
    text = text_value(value)
    lowered = text.lower()
    if "네이버" in text or "naver" in lowered or "smartstore" in lowered:
        return "네이버"
    if "쿠팡" in text or "coupang" in lowered:
        return "쿠팡"
    if "롯데" in text or "lotte" in lowered:
        return "롯데ON"
    return text


def upload_product_id_from_result(value: dict) -> str:
    for key in ("productId", "sellerProductId", "spdNo", "ProductId", "SellerProductId", "SpdNo"):
        result = text_value(value.get(key))
        if result:
            return result
    return ""


def collect_deleted_upload_product_ids() -> set[tuple[str, str]]:
    deleted: set[tuple[str, str]] = set()
    delete_actions = {"ALREADY_DELETED", "MISSING_OR_INACCESSIBLE"}
    for path in JOBS_ROOT.glob("*duplicate_cleanup_*.json"):
        try:
            rows = read_json(path, [])
        except Exception:
            continue
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            product_id = upload_product_id_from_result(row)
            market = canonical_upload_market(row.get("market") or path.name)
            if not product_id or not market:
                continue
            action = text_value(row.get("action")).upper()
            delete_ok = str(row.get("deleteOk", "")).lower() == "true"
            if action in delete_actions or delete_ok:
                deleted.add((market, product_id))

    for path in JOBS_ROOT.glob("*delete_report.json"):
        try:
            rows = read_json(path, [])
        except Exception:
            continue
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            product_id = upload_product_id_from_result(row)
            market = canonical_upload_market(row.get("market") or path.name)
            delete_ok = str(row.get("deleteOk", "")).lower() == "true"
            if product_id and market and delete_ok:
                deleted.add((market, product_id))
    return deleted


def collect_market_upload_success_index() -> dict[tuple[str, str, str], dict]:
    deleted_ids = collect_deleted_upload_product_ids()
    index: dict[tuple[str, str, str], dict] = {}

    def consider(row: dict, path: Path, dt: datetime, context_gs: str = "") -> None:
        gs = text_value(row.get("gs") or row.get("GsCode") or context_gs).upper()
        market = canonical_upload_market(row.get("market") or row.get("Market") or row.get("channel") or row.get("queueKey"))
        queue_key = text_value(row.get("queueKey"))
        account = text_value(row.get("account"))
        if not account and ":" in queue_key:
            account = queue_key.split(":", 1)[0]
        status = text_value(row.get("status") or row.get("Status")).lower()
        raw_status = text_value(row.get("rawStatus") or row.get("Status")).upper()
        product_id = upload_product_id_from_result(row)
        if not gs.startswith("GS") or market not in {"네이버", "쿠팡", "롯데ON"} or not product_id:
            return
        if (market, product_id) in deleted_ids:
            return
        if status not in {"uploaded", "ok", "success"} and raw_status not in {"OK", "SUCCESS", "SKIP_DUP"}:
            return
        key = (account, market, gs)
        existing = index.get(key)
        if existing and existing.get("_dt", datetime.min) >= dt:
            return
        index[key] = {
            "_dt": dt,
            "account": account,
            "market": market,
            "gs": gs,
            "productId": product_id,
            "name": text_value(row.get("sourceName") or row.get("title") or row.get("name") or row.get("Name")),
            "sourceFile": str(path),
            "uploadedAt": dt.strftime("%Y-%m-%d %H:%M:%S") if dt != datetime.min else "",
            "reason": "기존 업로드 성공 이력 있음",
        }

    def walk(value: object, path: Path, dt: datetime, context_gs: str = "") -> None:
        if isinstance(value, dict):
            next_gs = text_value(value.get("gs") or value.get("GsCode") or context_gs).upper()
            consider(value, path, dt, context_gs)
            for child in value.values():
                walk(child, path, dt, next_gs or context_gs)
        elif isinstance(value, list):
            for child in value:
                walk(child, path, dt, context_gs)

    for path in JOBS_ROOT.glob("*.json"):
        if "duplicate_cleanup_" in path.name:
            continue
        try:
            walk(read_json(path, {}), path, parse_dt_from_file_path(path))
        except Exception:
            continue

    for market_dir in (EXPORT_ROOT / "logs").glob("*_upload"):
        market = canonical_upload_market(market_dir.name.replace("_upload", ""))
        for summary in market_dir.glob("*/summary.json"):
            dt = parse_dt_from_file_path(summary)
            try:
                rows = read_json(summary, [])
            except Exception:
                continue
            if not isinstance(rows, list):
                continue
            request_files = sorted((summary.parent / "requests").glob("*.json"), key=lambda p: ("_retry" in p.stem, p.name))
            gs = request_files[0].stem.replace("_retry", "").upper() if request_files else ""
            for row in rows:
                if isinstance(row, dict):
                    consider({**row, "gs": gs, "market": market}, summary, dt, gs)

    for value in index.values():
        value.pop("_dt", None)
    return index


def write_market_duplicate_skip_report(job_id: str, rows: list[dict]) -> str:
    if not rows:
        return ""
    path = JOBS_ROOT / f"{job_id}_duplicate_skips.csv"
    fields = ["account", "market", "channel", "gs", "sourceName", "existingProductId", "existingUploadedAt", "existingSourceFile", "reason"]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fields})
    return str(path)


def execute_direct_market_upload(job_id: str, entry: dict, workbook_path: Path, log, results: list[dict]) -> tuple[dict, list[str]]:
    flag = market_cli_flag(entry["market"])
    if not flag:
        raise ValueError(f"{entry['market']}은 API 업로드 대상이 아닙니다.")
    upload_args = [
        "--direct-market-upload",
        "--file", str(workbook_path),
        "--gs", entry["gs"],
        flag,
    ]
    if DOTNET_UPLOAD_EXE.exists():
        cmd = [str(DOTNET_UPLOAD_EXE), *upload_args]
    else:
        cmd = ["dotnet", "run", "--project", str(DOTNET_UPLOAD_PROJECT), "--", *upload_args]
    if entry.get("force"):
        cmd.append("--force")
    if entry.get("dryRun"):
        cmd.append("--dry-run")

    log.write(" ".join(f'"{part}"' if " " in part else part for part in cmd) + "\n")
    child_env = os.environ.copy()
    key_root = Path(text_value(entry.get("_keyRoot")) or str(DESKTOP_KEY_ROOT))
    child_env["KEYWORDOCR_KEY_DIR"] = str(key_root)
    child_env["WEBOCR_KEY_ROOT"] = str(key_root)
    child_env["WEBOCR_ORIGINAL_KEY_ROOT"] = str(DESKTOP_KEY_ROOT)
    process = subprocess.Popen(
        cmd,
        cwd=str(PROJECT_ROOT),
        env=child_env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
    )
    register_active_process(job_id, process)
    lines: list[str] = []
    output_queue: queue.Queue[str | None] = queue.Queue()

    def read_stdout() -> None:
        try:
            assert process.stdout is not None
            for raw_line in process.stdout:
                output_queue.put(raw_line)
        finally:
            output_queue.put(None)

    reader = threading.Thread(target=read_stdout, daemon=True)
    reader.start()
    started = time.monotonic()
    exit_code = -1
    try:
        stream_closed = False
        while True:
            if job_id in CANCELLED_JOB_IDS:
                stop_process_tree(process)
                exit_code = -1
                break
            if time.monotonic() - started > MARKET_UPLOAD_TIMEOUT_SECONDS:
                lines.append(f"TIMEOUT: {entry['channel']} {entry['gs']} API 업로드가 {MARKET_UPLOAD_TIMEOUT_SECONDS}초를 초과했습니다.")
                stop_process_tree(process)
                exit_code = -1
                break
            try:
                item = output_queue.get(timeout=0.5)
            except queue.Empty:
                if process.poll() is not None and stream_closed:
                    exit_code = process.returncode
                    break
                continue
            if item is None:
                stream_closed = True
                if process.poll() is not None:
                    exit_code = process.returncode
                    break
                continue
            clean = item.rstrip("\n")
            if clean:
                lines.append(clean)
                log.write(clean + "\n")
                log.flush()
                current = read_json(JOBS_ROOT / f"{job_id}.json", {"jobId": job_id})
                current.update({
                    "status": "running",
                    "updatedAt": now_text(),
                    "currentStage": f"{entry['channel']} API 실행",
                    "tail": lines[-20:],
                    "results": results,
                })
                write_json(JOBS_ROOT / f"{job_id}.json", current)
        if exit_code == -1 and process.poll() is not None and not any(line.startswith("TIMEOUT:") for line in lines):
            exit_code = process.returncode
    finally:
        unregister_active_process(job_id, process)
    return parse_direct_upload_result(entry, exit_code, lines), lines


def run_market_upload_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    entries = normalize_upload_entries(payload)
    dry_run = bool(payload.get("dryRun"))
    if dry_run:
        for entry in entries:
            entry["dryRun"] = True
    job = read_json(job_path, {"jobId": job_id})
    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "progressPercent": 5,
        "currentStage": "업로드 요청 검증",
        "total": len(entries),
    })
    write_json(job_path, job)

    settings = read_market_key_settings().get("items", {})
    results: list[dict] = []
    api_markets = {"네이버", "쿠팡", "롯데ON"}
    duplicate_check_enabled = payload.get("duplicateCheck") is not False
    parallel_accounts = payload.get("parallelAccounts") is not False
    duplicate_index = collect_market_upload_success_index() if duplicate_check_enabled else {}
    duplicate_skip_rows: list[dict] = []
    upload_lock = threading.Lock()
    log_lock = threading.Lock()
    active_entries: dict[str, dict] = {}
    completed_count = [0]

    def upload_result_summary(result: dict) -> str:
        status = text_value(result.get("status")) or "unknown"
        channel = text_value(result.get("channel"))
        gs = text_value(result.get("gs"))
        product_id = text_value(result.get("productId"))
        error = text_value(result.get("error"))
        workbook = text_value(result.get("workbookPath"))
        if error:
            return f"{channel} {gs} {status} | 실패사유: {error}"
        if product_id:
            return f"{channel} {gs} {status} | productId={product_id}"
        if workbook:
            return f"{channel} {gs} {status} | workbook={Path(workbook).name}"
        return f"{channel} {gs} {status}"

    def upload_result_key(result: dict) -> str:
        return text_value(result.get("queueKey") or f"{result.get('channel')}:{result.get('gs')}:{result.get('market')}")

    def update_job_progress(stage: str = "") -> None:
        with upload_lock:
            done = completed_count[0]
            total = max(len(entries), 1)
            job_snap = read_json(job_path, {"jobId": job_id})
            job_snap.update({
                "status": "running",
                "updatedAt": now_text(),
                "progressPercent": min(95, 10 + int(done / total * 80)),
                "currentStage": stage or f"API 업로드 {done}/{total}",
                "results": list(results),
                "activeEntries": list(active_entries.values()),
                "tail": [upload_result_summary(r) for r in results[-20:]],
            })
            write_json(job_path, job_snap)

    def upload_single_entry(entry: dict, log_lines: list[str]) -> dict:
        market = entry["market"]
        key_id = market_key_id(entry["account"], market)
        result = {**entry, "status": "failed", "updatedAt": now_text()}
        if market not in api_markets:
            result["status"] = "skipped"
            result["error"] = "API 업로드 대상이 아닙니다. 11번가/ESM은 엑셀 export를 사용합니다."
            return result
        if key_id not in settings:
            result["error"] = f"{entry['channel']} 키 파일이 없습니다."
            return result
        result["keyFile"] = settings[key_id].get("fileName", "")
        workbook_path = write_direct_upload_workbook(job_id, entry)
        result["workbookPath"] = str(workbook_path)
        log_lines.append(f"[{now_text()}] {entry['channel']} {entry['gs']} API 업로드 시작")
        dummy_log = type("Log", (), {"write": lambda self, s: log_lines.append(s.rstrip()), "flush": lambda self: None})()
        active_key = upload_result_key(entry)
        with upload_lock:
            active_entries[active_key] = {**entry, "status": "running", "updatedAt": now_text()}
        update_job_progress(f"{entry['channel']} API 실행")
        try:
            result, _lines = execute_direct_market_upload(job_id, entry, workbook_path, dummy_log, results)
        finally:
            with upload_lock:
                active_entries.pop(active_key, None)
        result["keyFile"] = settings[key_id].get("fileName", "")
        result["workbookPath"] = str(workbook_path)
        return result

    def upload_market_group(market: str, market_entries: list[dict]) -> list[dict]:
        group_results: list[dict] = []
        group_log: list[str] = []
        for entry in market_entries:
            if job_id in CANCELLED_JOB_IDS:
                group_results.append({**entry, "status": "failed", "error": "업로드 작업이 중지되었습니다."})
                break
            result = upload_single_entry(entry, group_log)
            group_results.append(result)
            group_log.append(f"{entry['channel']} {entry['gs']} -> {result['status']} {result.get('error', '')}")
            with upload_lock:
                results.append(result)
                completed_count[0] += 1
            update_job_progress(f"{market} 업로드 {entry['gs']}")
        return group_results

    try:
        update_export_image_selections(entries)
        skipped: list[dict] = []
        api_entries: list[dict] = []
        for entry in entries:
            market = entry["market"]
            key_id = market_key_id(entry["account"], market)
            if market not in api_markets:
                skip_result = {**entry, "status": "skipped", "updatedAt": now_text(),
                               "error": "API 업로드 대상이 아닙니다. 11번가/ESM은 엑셀 export를 사용합니다."}
                skipped.append(skip_result)
            elif not entry.get("force") and (entry.get("account", ""), market, text_value(entry.get("gs")).upper()) in duplicate_index:
                duplicate = duplicate_index[(entry.get("account", ""), market, text_value(entry.get("gs")).upper())]
                skip_result = {
                    **entry,
                    "status": "skipped",
                    "updatedAt": now_text(),
                    "error": "기존 업로드 성공 이력 있음",
                    "duplicate": True,
                    "existingProductId": duplicate.get("productId", ""),
                    "existingUploadedAt": duplicate.get("uploadedAt", ""),
                    "existingSourceFile": duplicate.get("sourceFile", ""),
                }
                skipped.append(skip_result)
                duplicate_skip_rows.append({
                    "account": entry.get("account", ""),
                    "market": market,
                    "channel": entry.get("channel", ""),
                    "gs": text_value(entry.get("gs")).upper(),
                    "sourceName": entry.get("sourceName", ""),
                    "existingProductId": duplicate.get("productId", ""),
                    "existingUploadedAt": duplicate.get("uploadedAt", ""),
                    "existingSourceFile": duplicate.get("sourceFile", ""),
                    "reason": duplicate.get("reason", "기존 업로드 성공 이력 있음"),
                })
            elif key_id not in settings:
                skip_result = {**entry, "status": "failed", "updatedAt": now_text(),
                               "error": f"{entry['channel']} 키 파일이 없습니다."}
                skipped.append(skip_result)
            else:
                api_entries.append(entry)

        duplicate_skip_report = write_market_duplicate_skip_report(job_id, duplicate_skip_rows)

        with upload_lock:
            results.extend(skipped)
            completed_count[0] += len(skipped)

        by_account: dict[str, list[dict]] = {}
        for entry in api_entries:
            by_account.setdefault(entry["account"], []).append(entry)

        def run_account(account: str, account_entries: list[dict], log) -> None:
            by_market: dict[str, list[dict]] = {}
            for entry in account_entries:
                by_market.setdefault(entry["market"], []).append(entry)
            market_names = list(by_market.keys())
            with log_lock:
                log.write(f"\n[{now_text()}] account {account}: markets {market_names} ({len(account_entries)} entries)\n")
                log.flush()
            update_job_progress(f"{account}계정 {', '.join(market_names)} 병렬 업로드")

            with account_key_root(account, set(market_names), settings, log) as key_root:
                for entry in account_entries:
                    entry["_keyRoot"] = str(key_root)
                if len(by_market) >= 2:
                    with ThreadPoolExecutor(max_workers=min(len(by_market), 4)) as pool:
                        futures = {
                            pool.submit(upload_market_group, mkt, mkt_entries): mkt
                            for mkt, mkt_entries in by_market.items()
                        }
                        for future in as_completed(futures):
                            mkt = futures[future]
                            try:
                                future.result()
                            except Exception as exc:
                                with log_lock:
                                    log.write(f"[{now_text()}] {account}:{mkt} 그룹 오류: {exc}\n")
                elif by_market:
                    mkt, mkt_entries = next(iter(by_market.items()))
                    upload_market_group(mkt, mkt_entries)

            with log_lock:
                log.write(f"[{now_text()}] account {account} 완료\n")

        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START marketUpload (parallel)\n")
            log.write(f"entries: {len(entries)} / api: {len(api_entries)} / skipped: {len(skipped)} / duplicateCheck={duplicate_check_enabled} / parallelAccounts={parallel_accounts}\n")
            if duplicate_skip_report:
                log.write(f"duplicate skip report: {duplicate_skip_report}\n")
            accounts = list(by_account.keys())
            log.write(f"accounts: {accounts}\n")
            logged_results: set[str] = set()
            if skipped:
                log.write(f"\n[{now_text()}] 사전 실패/스킵\n")
                for result in skipped:
                    logged_results.add(upload_result_key(result))
                    log.write(f"[{now_text()}] {upload_result_summary(result)}\n")
            if not entries:
                raise ValueError("upload entries empty")

            if parallel_accounts and len(accounts) >= 2:
                with ThreadPoolExecutor(max_workers=min(len(accounts), 2)) as pool:
                    futures = {
                        pool.submit(run_account, account, by_account[account], log): account
                        for account in accounts
                    }
                    for future in as_completed(futures):
                        account = futures[future]
                        try:
                            future.result()
                        except Exception as exc:
                            with log_lock:
                                log.write(f"[{now_text()}] account {account} 오류: {exc}\n")
            else:
                for account in accounts:
                    run_account(account, by_account[account], log)

            for result in list(results):
                key = upload_result_key(result)
                if key in logged_results:
                    continue
                logged_results.add(key)
                log.write(f"[{now_text()}] {upload_result_summary(result)}\n")
            log.flush()

            failed_results = [result for result in results if result.get("status") == "failed"]
            if failed_results:
                log.write(f"\n[{now_text()}] 실패 상세 {len(failed_results)}건\n")
                for result in failed_results:
                    log.write(f"- {upload_result_summary(result)}\n")
            summary_success = sum(1 for item in results if item.get("status") == "uploaded")
            summary_failed = sum(1 for item in results if item.get("status") == "failed")
            summary_skipped = sum(1 for item in results if item.get("status") == "skipped")
            log.write(f"\n[{now_text()}] SUMMARY success={summary_success} failed={summary_failed} skipped={summary_skipped}\n")
            log.write(f"\n[{now_text()}] DONE\n")

        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "API 업로드 완료",
            "result": {
                "total": len(results),
                "success": sum(1 for item in results if item.get("status") == "uploaded"),
                "failed": sum(1 for item in results if item.get("status") == "failed"),
                "skipped": sum(1 for item in results if item.get("status") == "skipped"),
                "duplicateSkipped": sum(1 for item in results if item.get("duplicate")),
                "duplicateSkipReport": duplicate_skip_report,
                "results": results,
                "duplicateCheck": duplicate_check_enabled,
                "parallelAccounts": parallel_accounts,
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        with log_path.open("a", encoding="utf-8", errors="replace") as log:
            log.write(f"\n[{now_text()}] ERROR {exc}\n")
        job.update({"status": "failed", "finishedAt": now_text(), "error": str(exc), "currentStage": "업로드 실패"})
        write_json(job_path, job)


def find_active_market_upload_job() -> str:
    for path in sorted(JOBS_ROOT.glob("*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
        try:
            age_seconds = time.time() - path.stat().st_mtime
            if age_seconds > 7200:
                continue
            job = read_json(path, {})
        except Exception:
            continue
        if job.get("action") == "marketUpload" and job.get("status") in {"queued", "running"}:
            job_id = text_value(job.get("jobId") or path.stem)
            with ACTIVE_PROCESS_LOCK:
                has_process = bool(ACTIVE_PROCESSES.get(job_id))
            if age_seconds > 600 and not has_process:
                job.update({
                    "status": "failed",
                    "finishedAt": now_text(),
                    "currentStage": "업로드 중단 감지",
                    "error": "서버 재시작 또는 프로세스 종료로 작업이 멈췄습니다.",
                })
                write_json(path, job)
                continue
            return job_id
    return ""


def latest_market_upload_job() -> dict:
    for path in sorted(JOBS_ROOT.glob("*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
        try:
            job = read_json(path, {})
        except Exception:
            continue
        if job.get("action") == "marketUpload":
            return job
    return {}


def write_emergency_codex_context(payload: dict) -> dict:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    emergency_id = f"upload_{stamp}_{uuid.uuid4().hex[:6]}"
    target = EMERGENCY_ROOT / emergency_id
    target.mkdir(parents=True, exist_ok=True)

    upload_payload = payload.get("uploadPayload") if isinstance(payload.get("uploadPayload"), dict) else {}
    rows = upload_payload.get("rows") if isinstance(upload_payload.get("rows"), list) else []
    history = payload.get("history") if isinstance(payload.get("history"), dict) else {}
    status = payload.get("uploadStatus") if isinstance(payload.get("uploadStatus"), dict) else {}
    latest_job = latest_market_upload_job() if not rows else {}
    job_id = text_value(payload.get("jobId") or latest_job.get("jobId"))
    job = read_json(JOBS_ROOT / f"{safe_name(job_id, '')}.json", latest_job) if job_id else latest_job
    results = []
    if isinstance(job.get("results"), list):
        results.extend(job.get("results") or [])
    if isinstance(job.get("result"), dict) and isinstance(job["result"].get("results"), list):
        results.extend(job["result"].get("results") or [])

    markets: dict[str, dict] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        channel = text_value(row.get("channel") or row.get("channelKey"))
        bucket = markets.setdefault(channel, {"requested": 0, "uploaded": 0, "failed": 0, "exported": 0, "queued": 0})
        bucket["requested"] += 1
        key = text_value(row.get("queueKey") or f"{channel}:{row.get('gs')}")
        row_status = text_value(status.get(key) or (history.get(key) or {}).get("status") or "queued")
        if row_status in bucket:
            bucket[row_status] += 1
        else:
            bucket["queued"] += 1

    for result in results:
        if not isinstance(result, dict):
            continue
        channel = text_value(result.get("channel") or result.get("channelKey"))
        if not channel:
            continue
        bucket = markets.setdefault(channel, {"requested": 0, "uploaded": 0, "failed": 0, "exported": 0, "queued": 0})
        result_status = text_value(result.get("status"))
        if result_status in bucket:
            bucket[result_status] += 1

    context = {
        "id": emergency_id,
        "createdAt": now_text(),
        "seedName": text_value(payload.get("seedName")),
        "seedPath": text_value(payload.get("seedPath")),
        "jobId": job_id,
        "job": job,
        "uploadPayload": upload_payload,
        "uploadStatus": status,
        "history": history,
        "results": results,
        "markets": markets,
    }
    write_json(target / "upload_context.json", context)

    lines = [
        "# 긴급 업로드 복구 컨텍스트",
        "",
        f"- 생성시각: {context['createdAt']}",
        f"- 시드: {context['seedName'] or '-'}",
        f"- 시드 경로: {context['seedPath'] or '-'}",
        f"- 업로드 Job: {job_id or '-'}",
        f"- 대상 행: {len(rows)}",
        "",
        "## 마켓별 상태",
    ]
    if markets:
        for channel, item in sorted(markets.items()):
            lines.append(f"- {channel}: 요청 {item.get('requested', 0)} / 완료 {item.get('uploaded', 0)} / 실패 {item.get('failed', 0)} / 엑셀 {item.get('exported', 0)} / 대기 {item.get('queued', 0)}")
    else:
        lines.append("- 현재 업로드 대기열 정보가 없습니다.")
    lines.extend([
        "",
        "## 복구 선택지",
        "1. API 업로드 성공 상품만 삭제/판매중지 계획 생성",
        "2. 실패 상품만 재시도 계획 생성",
        "3. 특정 마켓만 롤백 계획 생성",
        "4. 업로드 파일/로그 보고서 생성",
        "기타. 사용자가 직접 입력한 지시를 우선한다.",
        "",
        "삭제/판매중지 작업은 바로 실행하지 말고 반드시 계획을 먼저 보여주고 확인을 받은 뒤 실행한다.",
    ])
    (target / "context.md").write_text("\n".join(lines), encoding="utf-8")
    (target / "rollback_plan.md").write_text(
        "1. upload_context.json의 results에서 productId/sellerProductId/spdNo를 확인한다.\n"
        "2. 사용자가 지정한 마켓/상품 범위만 필터링한다.\n"
        "3. 삭제 또는 판매중지 가능 API와 엑셀 수동 처리 대상을 분리한다.\n"
        "4. 실행 전 대상 목록을 사용자에게 확인받는다.\n",
        encoding="utf-8",
    )
    return {
        "ok": True,
        "id": emergency_id,
        "contextPath": str(target / "context.md"),
        "contextJsonPath": str(target / "upload_context.json"),
        "summary": {
            "seedName": context["seedName"],
            "jobId": job_id,
            "rowCount": len(rows),
            "markets": markets,
        },
        "options": [
            "API 업로드 성공 상품만 삭제/판매중지 계획 생성",
            "실패 상품만 재시도 계획 생성",
            "특정 마켓만 롤백 계획 생성",
            "업로드 파일/로그 보고서 생성",
        ],
    }


def resolve_emergency_dir(emergency_id: str) -> Path:
    clean_id = safe_name(text_value(emergency_id), "")
    if not clean_id:
        raise ValueError("emergency id required")
    target = (EMERGENCY_ROOT / clean_id).resolve()
    if not is_within(EMERGENCY_ROOT, target) or not target.exists() or not target.is_dir():
        raise ValueError("emergency context not found")
    return target


def option_label_for_emergency(value: str) -> str:
    text = text_value(value)
    labels = {
        "1": "API 업로드 성공 상품만 삭제/판매중지 계획 생성",
        "2": "실패 상품만 재시도 계획 생성",
        "3": "특정 마켓만 롤백 계획 생성",
        "4": "업로드 파일/로그 보고서 생성",
    }
    return labels.get(text.strip(), "")


def open_emergency_codex_session(payload: dict) -> dict:
    target = resolve_emergency_dir(text_value(payload.get("id")))
    context_path = target / "context.md"
    context_json_path = target / "upload_context.json"
    request_path = target / "emergency_request.md"
    launcher_path = target / "open_codex_recovery.ps1"
    instruction = text_value(payload.get("instruction") or payload.get("note")).strip()
    selected_label = option_label_for_emergency(instruction)
    if not instruction:
        instruction = "4"
        selected_label = option_label_for_emergency(instruction)

    request_lines = [
        "# 긴급 복구 Codex 요청",
        "",
        "이 세션은 WebOcrClude 로컬 업로드 화면에서 열린 긴급 복구 세션입니다.",
        "아래 컨텍스트를 먼저 읽고, 현재 업로드 상황을 요약한 뒤 사용자와 대화하면서 복구 계획을 잡으세요.",
        "",
        "## 사용자가 선택/입력한 요청",
        f"- 입력: {instruction}",
    ]
    if selected_label:
        request_lines.append(f"- 해석: {selected_label}")
    request_lines.extend([
        "",
        "## 반드시 지킬 것",
        "- 삭제, 판매중지, 롤백, 재업로드 같은 파괴적 작업은 바로 실행하지 말고 대상 목록과 계획을 먼저 보여주세요.",
        "- 사용자가 확인하기 전에는 API 삭제/수정 명령을 실행하지 마세요.",
        "- 업로드 성공/실패/대기/엑셀 대상을 분리해서 말하세요.",
        "- 필요한 경우 이 디렉터리의 JSON과 로그를 읽고 근거를 확인하세요.",
        "",
        "## 읽어야 할 파일",
        f"- 컨텍스트 요약: {context_path}",
        f"- 원본 JSON: {context_json_path}",
        f"- 롤백 초안: {target / 'rollback_plan.md'}",
        "",
        "## 첫 응답 형식",
        "1. 현재 작업이 무엇인지 요약",
        "2. 성공/실패/대기/엑셀 대상 개수 요약",
        "3. 사용자의 요청에 맞는 복구 선택지",
        "4. 실행 전 확인해야 할 질문",
    ])
    request_path.write_text("\n".join(request_lines), encoding="utf-8")

    ps = [
        "$ErrorActionPreference = 'Stop'",
        f"Set-Location -LiteralPath {json.dumps(str(PROJECT_ROOT), ensure_ascii=False)}",
        f"$requestPath = {json.dumps(str(request_path), ensure_ascii=False)}",
        "$prompt = Get-Content -LiteralPath $requestPath -Raw -Encoding UTF8",
        "Write-Host ''",
        "Write-Host '=== WebOcrClude 긴급 복구 Codex ===' -ForegroundColor Red",
        "Write-Host ('요청 파일: ' + $requestPath) -ForegroundColor DarkGray",
        "Write-Host 'Codex가 컨텍스트를 읽은 상태로 시작됩니다. 이후 이 창에서 직접 대화하세요.' -ForegroundColor Yellow",
        "Write-Host ''",
        f"codex --cd {json.dumps(str(PROJECT_ROOT), ensure_ascii=False)} --sandbox danger-full-access --ask-for-approval on-request --no-alt-screen $prompt",
    ]
    launcher_path.write_text("\n".join(ps), encoding="utf-8")

    if payload.get("dryRun"):
        return {
            "ok": True,
            "id": target.name,
            "pid": 0,
            "requestPath": str(request_path),
            "launcherPath": str(launcher_path),
            "message": "PowerShell Codex 복구 세션 실행 준비가 완료되었습니다.",
            "dryRun": True,
        }

    args = [
        "powershell.exe",
        "-NoExit",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(launcher_path),
    ]
    creationflags = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
    process = subprocess.Popen(args, cwd=str(PROJECT_ROOT), creationflags=creationflags)
    return {
        "ok": True,
        "id": target.name,
        "pid": process.pid,
        "requestPath": str(request_path),
        "launcherPath": str(launcher_path),
        "message": "PowerShell Codex 복구 세션을 열었습니다.",
    }


def write_market_export(payload: dict) -> dict:
    entries = normalize_upload_entries(payload)
    market = safe_name(text_value(payload.get("market") or "market"), "market")
    if not entries:
        raise ValueError("export entries empty")
    if market == "11번가":
        return write_elevenst_template_export(entries)
    if market.upper() == "ESM":
        return write_esm_template_export(entries)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    headers = [
        "계정", "마켓", "GS코드", "원본상품명", "업로드상품명", "검색어설정",
        "판매가", "소비자가", "옵션입력", "옵션추가금",
        "대표이미지", "이미지등록(목록)", "이미지등록(추가)", "이미지등록(상세)",
        "상품 상세설명", "상품정보제공고시", "Cafe24 URL", "상태",
    ]
    try:
        from openpyxl import Workbook

        file_name = safe_name(f"{market}_upload_queue_{stamp}.xlsx")
        target = EXPORT_ROOT / file_name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = market[:31]
        sheet.append(headers)
        for entry in entries:
            price = parse_upload_price(entry.get("salePrice") or entry.get("price"))
            consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2)
            main_image = direct_upload_image_ref(entry.get("mainImageSrc"))
            add_images = [direct_upload_image_ref(url) for url in entry.get("additionalImageSrcs", [])]
            add_images = [url for url in add_images if url]
            detail_refs = [direct_upload_image_ref(url) for url in entry.get("detailImageSrcs", [])]
            detail_refs = [url for url in detail_refs if url]
            detail_html = normalize_detail_html_for_upload(entry.get("detailHtml")) or normalize_detail_html_for_upload(seed_detail_html_for_gs(entry.get("gs")))
            if not detail_html:
                detail_urls = [public_image_url(url) for url in entry.get("detailImageSrcs", []) if public_image_url(url)]
                if detail_urls:
                    detail_html = "<center>" + "".join(f'<img src="{url}">' for url in detail_urls[:80]) + "</center>"
            naver_notice_payload = naver_notice_payload_for_entry(entry)
            naver_notice_json = json.dumps(naver_notice_payload, ensure_ascii=False, separators=(",", ":")) if naver_notice_payload else ""
            sheet.append([
                entry["account"],
                entry["market"],
                entry["gs"],
                entry["sourceName"],
                entry["title"],
                entry["searchTerms"],
                price,
                consumer_price,
                option_input_from_entry(entry),
                option_additionals_from_entry(entry),
                main_image,
                main_image,
                "|".join(add_images),
                "|".join(detail_refs) if detail_refs else "",
                detail_html,
                naver_notice_json,
                entry["cafe24Url"],
                "대기",
            ])
        workbook.save(target)
        file_format = "xlsx"
    except Exception:
        file_name = safe_name(f"{market}_upload_queue_{stamp}.csv")
        target = EXPORT_ROOT / file_name
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for entry in entries:
                price = parse_upload_price(entry.get("salePrice") or entry.get("price"))
                consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2)
                main_image = direct_upload_image_ref(entry.get("mainImageSrc"))
                add_images = [direct_upload_image_ref(url) for url in entry.get("additionalImageSrcs", [])]
                add_images = [url for url in add_images if url]
                detail_refs = [direct_upload_image_ref(url) for url in entry.get("detailImageSrcs", [])]
                detail_refs = [url for url in detail_refs if url]
                detail_html = normalize_detail_html_for_upload(entry.get("detailHtml")) or normalize_detail_html_for_upload(seed_detail_html_for_gs(entry.get("gs")))
                naver_notice_payload = naver_notice_payload_for_entry(entry)
                naver_notice_json = json.dumps(naver_notice_payload, ensure_ascii=False, separators=(",", ":")) if naver_notice_payload else ""
                writer.writerow([
                    entry["account"],
                    entry["market"],
                    entry["gs"],
                    entry["sourceName"],
                    entry["title"],
                    entry["searchTerms"],
                    price,
                    consumer_price,
                    option_input_from_entry(entry),
                    option_additionals_from_entry(entry),
                    main_image,
                    main_image,
                    "|".join(add_images),
                    "|".join(detail_refs) if detail_refs else "",
                    detail_html,
                    naver_notice_json,
                    entry["cafe24Url"],
                    "대기",
                ])
        file_format = "csv"
    return {
        "fileName": file_name,
        "path": str(target),
        "url": f"/data/exports/{urllib.parse.quote(file_name)}",
        "count": len(entries),
        "format": file_format,
    }


ELEVENST_TEMPLATE_PATH = Path(r"C:\Users\rkghr\Downloads\ExcelUnitProductList-Ver2.50.xlsx")
ESM_TEMPLATE_PATH = Path(r"C:\Users\rkghr\Downloads\new_basic_bulk (1).xlsx")
ESM_TEMPLATE_SHEET = "NEW 일반상품"


def excel_set(sheet, row: int, col: int, value: object) -> None:
    if isinstance(value, str):
        value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", value)
        value = value[:32767]
    sheet.cell(row=row, column=col).value = value


def trim_at_word(value: object, limit: int) -> str:
    text = clean_product_title(value) or text_value(value)
    if len(text) <= limit:
        return text
    cut = text[:limit].rstrip()
    space = cut.rfind(" ")
    if space >= max(10, int(limit * 0.58)):
        cut = cut[:space]
    return cut.rstrip(" ,./|-_")


def export_detail_html_for_entry(entry: dict) -> str:
    detail_sources = entry.get("detailImageSrcs") if isinstance(entry.get("detailImageSrcs"), list) else []
    if not detail_sources:
        detail_sources = seed_detail_images_for_gs(entry.get("gs"))
    detail_html = normalize_detail_html_for_upload(entry.get("detailHtml")) or normalize_detail_html_for_upload(seed_detail_html_for_gs(entry.get("gs")))
    public_detail_images = [public_image_url(url) for url in detail_sources if public_image_url(url)]
    if detail_html and "<img" not in detail_html.lower() and public_detail_images:
        detail_html = ""
    if not detail_html and public_detail_images:
        detail_html = "<center>" + "".join(f'<img src="{url}">' for url in public_detail_images[:80]) + "</center>"
    return detail_html or "상세페이지 참조"


def export_main_image_for_entry(entry: dict) -> str:
    return public_image_url(entry.get("mainImageSrc")) or direct_upload_image_ref(entry.get("mainImageSrc"))


def export_additional_images_for_entry(entry: dict) -> list[str]:
    sources = infer_additional_upload_images(entry)
    return [public_image_url(url) or direct_upload_image_ref(url) for url in sources if public_image_url(url) or direct_upload_image_ref(url)]


def option_labels_for_export(entry: dict) -> list[str]:
    labels = option_labels_from_input(option_input_from_entry(entry))
    return [re.sub(r"^[A-Z]\s+", "", label).strip() or label for label in labels[:50]]


def option_prices_for_export(entry: dict, count: int) -> list[int]:
    raw = option_additionals_from_entry(entry)
    prices = [parse_upload_price(item) for item in raw.split("|") if text_value(item)]
    if len(prices) < count:
        prices.extend([0] * (count - len(prices)))
    return prices[:count]


def category_value(entry: dict, *keys: str) -> str:
    categories = entry.get("categories") if isinstance(entry.get("categories"), dict) else {}
    for key in keys:
        value = text_value(categories.get(key))
        if value:
            return value
    inferred = infer_direct_upload_categories(entry)
    for key in keys:
        value = text_value(inferred.get(key))
        if value:
            return value
    return ""


def convert_xlsx_to_xls_if_possible(source: Path, target: Path) -> Path:
    try:
        import win32com.client  # type: ignore
    except Exception:
        return source
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(source))
        workbook.SaveAs(str(target), FileFormat=56)
        workbook.Close(False)
        source.unlink(missing_ok=True)
        return target
    except Exception:
        try:
            if workbook is not None:
                workbook.Close(False)
        except Exception:
            pass
        return source
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass


def write_elevenst_template_export(entries: list[dict]) -> dict:
    if not ELEVENST_TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"11번가 공식 양식 파일을 찾지 못했습니다: {ELEVENST_TEMPLATE_PATH}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl 로드 실패: {exc}") from exc

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    work_name = safe_name(f"11번가_업로드_{stamp}.xlsx")
    work_path = EXPORT_ROOT / work_name
    xls_name = safe_name(f"11번가_업로드_{stamp}.xls")
    xls_path = EXPORT_ROOT / xls_name
    workbook = load_workbook(ELEVENST_TEMPLATE_PATH)
    sheet = workbook["대량등록 양식"] if "대량등록 양식" in workbook.sheetnames else workbook.worksheets[0]
    if sheet.max_row >= 6:
        sheet.delete_rows(6, max(1, sheet.max_row - 5))
    row_number = 6
    for entry in entries:
        price = parse_upload_price(entry.get("salePrice") or entry.get("price")) or 1000
        consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2) or price
        title = trim_at_word(entry.get("title") or entry.get("sourceName") or entry.get("gs"), 100)
        main_image = export_main_image_for_entry(entry)
        add_images = export_additional_images_for_entry(entry)
        labels = option_labels_for_export(entry)
        option_prices = option_prices_for_export(entry, len(labels))
        category_code = category_value(entry, "elevenst", "11st", "eleven")
        excel_set(sheet, row_number, 2, category_code)
        excel_set(sheet, row_number, 3, entry["gs"])
        excel_set(sheet, row_number, 4, entry["gs"])
        excel_set(sheet, row_number, 5, title)
        excel_set(sheet, row_number, 6, "샤플라이")
        excel_set(sheet, row_number, 8, main_image)
        for index, image in enumerate(add_images[:3], start=9):
            excel_set(sheet, row_number, index, image)
        excel_set(sheet, row_number, 13, export_detail_html_for_entry(entry))
        excel_set(sheet, row_number, 14, "Y")
        excel_set(sheet, row_number, 15, "01")
        excel_set(sheet, row_number, 16, "N")
        excel_set(sheet, row_number, 17, "01")
        excel_set(sheet, row_number, 18, "01")
        excel_set(sheet, row_number, 19, "108")
        excel_set(sheet, row_number, 29, price)
        if labels:
            excel_set(sheet, row_number, 31, "01")
            excel_set(sheet, row_number, 32, "|".join(labels))
            excel_set(sheet, row_number, 33, "|".join(str(value) for value in option_prices))
            excel_set(sheet, row_number, 34, "|".join("999" for _ in labels))
            excel_set(sheet, row_number, 37, len(labels) * 999)
        else:
            excel_set(sheet, row_number, 37, 999)
        if consumer_price:
            excel_set(sheet, row_number, 41, consumer_price)
        excel_set(sheet, row_number, 42, "홈런market")
        excel_set(sheet, row_number, 43, "Y")
        excel_set(sheet, row_number, 44, "01")
        excel_set(sheet, row_number, 45, entry["gs"])
        excel_set(sheet, row_number, 46, "02")
        excel_set(sheet, row_number, 47, "1287")
        excel_set(sheet, row_number, 50, "01|03\n02|03\n03|03\n04|05")
        excel_set(sheet, row_number, 51, "01")
        excel_set(sheet, row_number, 54, "891045")
        excel_set(sheet, row_number, 55, "11800")
        excel_set(sheet, row_number, 56, "상세페이지 참조")
        excel_set(sheet, row_number, 57, "11905")
        excel_set(sheet, row_number, 58, "상세페이지 참조")
        excel_set(sheet, row_number, 59, "23760413")
        excel_set(sheet, row_number, 60, "판매자 고객센터 문의")
        excel_set(sheet, row_number, 61, "23759100")
        excel_set(sheet, row_number, 62, "중국")
        excel_set(sheet, row_number, 63, "23756033")
        excel_set(sheet, row_number, 64, "해당사항 없음")
        excel_set(sheet, row_number, 100, "01")
        excel_set(sheet, row_number, 101, "01")
        excel_set(sheet, row_number, 102, "00034")
        excel_set(sheet, row_number, 103, "1228104")
        excel_set(sheet, row_number, 105, "01")
        excel_set(sheet, row_number, 106, 0)
        excel_set(sheet, row_number, 108, "Y")
        excel_set(sheet, row_number, 109, "03")
        excel_set(sheet, row_number, 111, 3000)
        excel_set(sheet, row_number, 112, "01")
        excel_set(sheet, row_number, 113, 6000)
        excel_set(sheet, row_number, 114, "상품 상세설명을 참고해 주세요.")
        excel_set(sheet, row_number, 115, "상품 상세설명 및 판매자 반품/교환 정책을 참고해 주세요.")
        row_number += 1
    if sheet.max_row >= 5:
        sheet.delete_rows(4, 2)
    workbook.save(work_path)
    final_path = convert_xlsx_to_xls_if_possible(work_path, xls_path)
    return {
        "fileName": final_path.name,
        "path": str(final_path),
        "url": f"/data/exports/{urllib.parse.quote(final_path.name)}",
        "count": len(entries),
        "format": final_path.suffix.lstrip(".").lower(),
        "template": "11st-official",
    }


def write_esm_template_export(entries: list[dict]) -> dict:
    if not ESM_TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"ESM 공식 양식 파일을 찾지 못했습니다: {ESM_TEMPLATE_PATH}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl 로드 실패: {exc}") from exc

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = safe_name(f"ESM_옥션지마켓_업로드_{stamp}.xlsx")
    target = EXPORT_ROOT / file_name
    workbook = load_workbook(ESM_TEMPLATE_PATH)
    sheet = workbook[ESM_TEMPLATE_SHEET] if ESM_TEMPLATE_SHEET in workbook.sheetnames else workbook.worksheets[0]
    if sheet.max_row >= 8:
        sheet.delete_rows(8, max(1, sheet.max_row - 7))
    row_number = 8
    sequence = 1
    for entry in entries:
        price = parse_upload_price(entry.get("salePrice") or entry.get("price")) or 1000
        title = trim_at_word(entry.get("title") or entry.get("sourceName") or entry.get("gs"), 45)
        labels = option_labels_for_export(entry)
        main_image = export_main_image_for_entry(entry)
        add_images = export_additional_images_for_entry(entry)
        categories = entry.get("categories") if isinstance(entry.get("categories"), dict) else {}
        esm_code = category_value(entry, "esm")
        auction_code = category_value(entry, "auction")
        gmarket_code = category_value(entry, "gmarket")
        excel_set(sheet, row_number, 1, sequence)
        excel_set(sheet, row_number, 2, "옥션/G마켓")
        excel_set(sheet, row_number, 3, "rkghrud")
        excel_set(sheet, row_number, 4, "rkghrud")
        excel_set(sheet, row_number, 5, title)
        excel_set(sheet, row_number, 10, text_value(categories.get("esm_template")))
        excel_set(sheet, row_number, 11, esm_code)
        excel_set(sheet, row_number, 12, auction_code)
        excel_set(sheet, row_number, 13, gmarket_code)
        excel_set(sheet, row_number, 14, "90")
        excel_set(sheet, row_number, 15, price)
        excel_set(sheet, row_number, 16, price)
        excel_set(sheet, row_number, 21, 99999)
        excel_set(sheet, row_number, 22, 99999)
        if labels:
            excel_set(sheet, row_number, 23, "단독형")
            excel_set(sheet, row_number, 24, "옵션")
            excel_set(sheet, row_number, 25, "\n".join(f"{label},정상,노출,99999,99999" for label in labels))
        else:
            excel_set(sheet, row_number, 23, "미사용")
        excel_set(sheet, row_number, 26, main_image)
        excel_set(sheet, row_number, 27, ",".join(add_images[:9]))
        excel_set(sheet, row_number, 28, export_detail_html_for_entry(entry))
        excel_set(sheet, row_number, 30, "일반택배")
        excel_set(sheet, row_number, 31, "20223695")
        excel_set(sheet, row_number, 32, "46262933")
        excel_set(sheet, row_number, 33, "4886443")
        excel_set(sheet, row_number, 34, "-13")
        excel_set(sheet, row_number, 35, "-13")
        excel_set(sheet, row_number, 36, "10013")
        excel_set(sheet, row_number, 37, 3000)
        excel_set(sheet, row_number, 38, "36")
        excel_set(sheet, row_number, 39, "235804")
        excel_set(sheet, row_number, 40, "인증대상아님")
        excel_set(sheet, row_number, 43, "인증대상아님")
        excel_set(sheet, row_number, 46, "해당사항없음")
        excel_set(sheet, row_number, 47, "인증대상아님")
        excel_set(sheet, row_number, 50, "해당사항없음")
        excel_set(sheet, row_number, 51, "인증대상아님")
        excel_set(sheet, row_number, 53, "해당없음")
        excel_set(sheet, row_number, 54, "해외수입")
        excel_set(sheet, row_number, 55, "174")
        excel_set(sheet, row_number, 56, "단일원산지")
        excel_set(sheet, row_number, 62, "구매가능")
        excel_set(sheet, row_number, 63, "과세상품")
        row_number += 1
        sequence += 1
    workbook.save(target)
    return {
        "fileName": file_name,
        "path": str(target),
        "url": f"/data/exports/{urllib.parse.quote(file_name)}",
        "count": len(entries),
        "format": "xlsx",
        "template": "esm-official",
    }


def remove_runtime_path(raw_path: object) -> dict:
    path_text = text_value(raw_path)
    if not path_text:
        return {"path": "", "status": "skipped", "reason": "empty"}
    path = Path(path_text)
    if not path.is_absolute():
        path = (ROOT / path).resolve()
    else:
        path = path.resolve()
    allowed_roots = [UPLOAD_ROOT, JOBS_ROOT, SEED_ROOT, EXPORT_ROOT]
    if not any(is_within(root, path) for root in allowed_roots):
        return {"path": str(path), "status": "skipped", "reason": "outside runtime data"}
    if not path.exists():
        return {"path": str(path), "status": "missing"}
    if path.is_dir():
        shutil.rmtree(path)
    else:
        path.unlink()
    return {"path": str(path), "status": "deleted"}


def cleanup_workspace_artifacts(payload: dict) -> dict:
    paths = payload.get("paths") if isinstance(payload.get("paths"), list) else []
    job_ids = payload.get("jobIds") if isinstance(payload.get("jobIds"), list) else []
    removed: list[dict] = []
    stopped: list[str] = []
    seen_paths: set[str] = set()

    def remember_path(path_value: object) -> None:
        clean = text_value(path_value)
        if clean and clean not in seen_paths:
            seen_paths.add(clean)
            removed.append(remove_runtime_path(clean))

    for raw_job_id in job_ids:
        job_id = safe_name(text_value(raw_job_id), "")
        if not job_id:
            continue
        if stop_active_job(job_id):
            stopped.append(job_id)
        job_path = JOBS_ROOT / f"{job_id}.json"
        job_payload = read_json(job_path, {})
        remember_path(job_path)
        remember_path(JOBS_ROOT / f"{job_id}.log")
        remember_path(JOBS_ROOT / f"{job_id}_keyword")
        result = job_payload.get("result") if isinstance(job_payload.get("result"), dict) else {}
        remember_path(result.get("keywordResultPath", ""))

    for path_value in paths:
        remember_path(path_value)

    return {
        "stoppedJobs": stopped,
        "removed": removed,
        "deleted": sum(1 for item in removed if item.get("status") == "deleted"),
        "skipped": sum(1 for item in removed if item.get("status") == "skipped"),
        "missing": sum(1 for item in removed if item.get("status") == "missing"),
    }


def pm_db():
    if not PRODUCT_MANAGER_DB.is_file():
        return None
    conn = sqlite3.connect(str(PRODUCT_MANAGER_DB), timeout=5)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def pm_option_sort_key(p):
    oc = (p.get("option_code") or "").upper()
    pc = p.get("product_code") or ""
    return (0, oc, pc) if oc else (1, pc)


def pm_option_label(p):
    oc = p.get("option_code") or p.get("product_code") or ""
    pc = p.get("product_code") or ""
    pn = p.get("product_name") or ""
    label = pn.split(pc, 1)[1].strip() if pc and pc in pn else pn.strip()
    label = re.sub(r"^[\s\-_/,:]+", "", label)
    label = re.sub(r"\s+", " ", label)
    if not label:
        return oc
    if oc and label.upper().startswith(oc.upper()):
        return label
    return f"{oc} {label}".strip()


def pm_build_group_row(products):
    if len(products) == 1:
        return json.loads(products[0]["raw_data"])
    sorted_prods = sorted(products, key=pm_option_sort_key)
    base = None
    for p in sorted_prods:
        if (p.get("option_code") or "").upper().startswith("A"):
            base = p
            break
    if not base:
        base = sorted_prods[0]
    raw = json.loads(base["raw_data"])
    labels = [pm_option_label(p) for p in sorted_prods]
    raw["자체 상품코드"] = base["product_code"]
    raw["GS상품코드"] = base["product_code"]
    raw["옵션사용"] = "Y"
    raw["품목 구성방식"] = "T"
    raw["옵션 표시방식"] = "C"
    raw["옵션세트명"] = ""
    raw["옵션입력"] = f"옵션{{{'|'.join(labels)}}}"
    prices: list[int] = []
    for product in sorted_prods:
        try:
            product_raw = json.loads(product.get("raw_data") or "{}")
        except Exception:
            product_raw = {}
        headers = {normalize_header(col): col for col in product_raw.keys()}
        price_info = source_price_info(product_raw, headers) if product_raw else {}
        prices.append(price_info.get("salePrice") or price_value(product.get("price")))
    if len(prices) > 1:
        base_price = prices[0]
        raw["옵션추가금"] = "|".join(str(price - base_price) for price in prices)
    raw["필수여부"] = "F"
    return raw


def pm_get_upload_dates() -> list[dict]:
    conn = pm_db()
    if not conn:
        return []
    try:
        rows = conn.execute("""
            SELECT
                substr(upload_date, 1, 10) AS upload_date,
                MAX(upload_date) AS latest_upload_date,
                COUNT(*) AS upload_count,
                SUM(total_count) AS total_count,
                GROUP_CONCAT(file_name, ', ') AS file_names
            FROM upload_history
            GROUP BY substr(upload_date, 1, 10)
            ORDER BY latest_upload_date DESC
        """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def pm_get_suppliers(upload_date: str = "") -> list[dict]:
    conn = pm_db()
    if not conn:
        return []
    try:
        if upload_date:
            rows = conn.execute("""
                WITH scoped AS (
                    SELECT DISTINCT p.product_code, p.supplier_code, p.sku_group,
                        p.is_listed, p.naver_status, p.is_naver_duplicate, p.sale_status
                    FROM products p
                    JOIN upload_product_items upi ON upi.product_code = p.product_code
                    JOIN upload_history uh ON uh.id = upi.upload_history_id
                    WHERE substr(uh.upload_date, 1, 10)=?
                )
                SELECT supplier_code,
                    COUNT(DISTINCT sku_group) as total_skus,
                    COUNT(DISTINCT CASE WHEN is_listed=0 AND naver_status='신규'
                        AND is_naver_duplicate=0 THEN sku_group END) as available_skus,
                    COUNT(DISTINCT CASE WHEN is_listed=1 THEN sku_group END) as listed_skus
                FROM scoped WHERE sale_status='Y'
                GROUP BY supplier_code ORDER BY supplier_code
            """, (upload_date,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT supplier_code,
                    COUNT(DISTINCT sku_group) as total_skus,
                    COUNT(DISTINCT CASE WHEN is_listed=0 AND naver_status='신규'
                        AND is_naver_duplicate=0 THEN sku_group END) as available_skus,
                    COUNT(DISTINCT CASE WHEN is_listed=1 THEN sku_group END) as listed_skus
                FROM products WHERE sale_status='Y'
                GROUP BY supplier_code ORDER BY supplier_code
            """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def pm_get_products_by_skus(sku_list: list[str]) -> list[dict]:
    conn = pm_db()
    if not conn:
        return []
    try:
        ph = ",".join(["?"] * len(sku_list))
        products = conn.execute(f"""
            SELECT product_code, sku_group, supplier_code, option_code,
                   product_name, price, image_url, raw_data
            FROM products WHERE sku_group IN ({ph})
            ORDER BY sku_group, option_code
        """, sku_list).fetchall()
        return [dict(p) for p in products]
    finally:
        conn.close()


def pm_select_automation_skus(
    suppliers: list[str],
    upload_date: str = "",
    sort_order: str = "latest",
    filter_mode: str = "available",
    total_limit: int = 100,
) -> list[str]:
    result = pm_get_products_paginated(
        suppliers,
        upload_date,
        sort_order,
        filter_mode,
        "",
        1,
        max(1, int(total_limit or 100)),
        999999,
    )
    return [text_value(item.get("sku_group")) for item in result.get("products", []) if text_value(item.get("sku_group"))]


def pm_select_automation_sku_batches(
    suppliers: list[str],
    upload_date: str = "",
    sort_order: str = "latest",
    filter_mode: str = "available",
    count_per_supplier: int = 20,
    run_count: int = 1,
) -> list[list[str]]:
    result = pm_get_products_paginated(
        suppliers,
        upload_date,
        sort_order,
        filter_mode,
        "",
        1,
        999999,
        max(1, int(count_per_supplier or 20)) * max(1, int(run_count or 1)),
    )
    by_supplier: dict[str, list[str]] = {}
    for item in result.get("products", []):
        supplier = text_value(item.get("supplier_code"))
        sku = text_value(item.get("sku_group"))
        if supplier and sku:
            by_supplier.setdefault(supplier, []).append(sku)

    batches: list[list[str]] = []
    per_supplier = max(1, int(count_per_supplier or 20))
    for run_index in range(max(1, int(run_count or 1))):
        batch: list[str] = []
        start = run_index * per_supplier
        end = start + per_supplier
        for supplier in suppliers:
            batch.extend(by_supplier.get(text_value(supplier), [])[start:end])
        if batch:
            batches.append(batch)
    return batches


def pm_get_available_products(suppliers: list[str], sort_order: str = "latest",
                               count_per_supplier: int = 999, upload_date: str = "") -> list[dict]:
    conn = pm_db()
    if not conn:
        return []
    try:
        order = {"latest": "p.product_seq DESC", "oldest": "p.product_seq ASC",
                 "random": "RANDOM()"}.get(sort_order, "p.product_seq DESC")
        all_skus: list[str] = []
        for supplier in suppliers:
            if upload_date:
                rows = conn.execute(f"""
                    SELECT DISTINCT p.sku_group
                    FROM products p
                    JOIN upload_product_items upi ON upi.product_code = p.product_code
                    JOIN upload_history uh ON uh.id = upi.upload_history_id
                    WHERE p.supplier_code=? AND substr(uh.upload_date, 1, 10)=?
                        AND p.is_listed=0 AND p.naver_status='신규'
                        AND p.is_naver_duplicate=0 AND p.sale_status='Y'
                    ORDER BY {order} LIMIT ?
                """, (supplier, upload_date, count_per_supplier)).fetchall()
            else:
                rows = conn.execute(f"""
                    SELECT DISTINCT p.sku_group FROM products p
                    WHERE p.supplier_code=? AND p.is_listed=0 AND p.naver_status='신규'
                        AND p.is_naver_duplicate=0 AND p.sale_status='Y'
                    ORDER BY {order} LIMIT ?
                """, (supplier, count_per_supplier)).fetchall()
            all_skus.extend(r["sku_group"] for r in rows)
        if not all_skus:
            return []
        ph = ",".join(["?"] * len(all_skus))
        products = conn.execute(f"""
            SELECT product_code, sku_group, supplier_code, option_code,
                   product_name, price, image_url, raw_data
            FROM products WHERE sku_group IN ({ph})
            ORDER BY sku_group, option_code
        """, all_skus).fetchall()
        return [dict(p) for p in products]
    finally:
        conn.close()


def pm_get_products_paginated(suppliers: list[str], upload_date: str = "",
                                sort_order: str = "latest", filter_mode: str = "available",
                                search: str = "", page: int = 1, per_page: int = 50,
                                count_per_supplier: int = 9999) -> dict:
    conn = pm_db()
    if not conn:
        return {"products": [], "total": 0, "page": page, "per_page": per_page, "total_pages": 0}
    try:
        params: list = []
        # Build the scoped CTE
        if upload_date:
            base_from = (
                "products p "
                "JOIN upload_product_items upi ON upi.product_code = p.product_code "
                "JOIN upload_history uh ON uh.id = upi.upload_history_id"
            )
            where_clauses = ["substr(uh.upload_date, 1, 10) = ?"]
            params.append(upload_date)
        else:
            base_from = "products p"
            where_clauses = []

        if suppliers:
            ph = ",".join(["?"] * len(suppliers))
            where_clauses.append(f"p.supplier_code IN ({ph})")
            params.extend(suppliers)

        if filter_mode == "available":
            where_clauses.append("p.is_listed=0 AND p.naver_status='신규' AND p.is_naver_duplicate=0 AND p.sale_status='Y'")
        elif filter_mode == "listed":
            where_clauses.append("p.is_listed=1 AND p.sale_status='Y'")
        else:
            where_clauses.append("p.sale_status='Y'")

        search_params: list = []
        if search:
            where_clauses.append("(p.product_name LIKE ? OR p.sku_group LIKE ? OR p.product_code LIKE ?)")
            like_val = f"%{search}%"
            search_params = [like_val, like_val, like_val]

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
        all_params = params + search_params

        # Sort expression
        sort_map = {
            "latest": "MAX(p.product_seq) DESC",
            "oldest": "MIN(p.product_seq) ASC",
            "random": "RANDOM()",
        }
        order_expr = sort_map.get(sort_order, "MAX(p.product_seq) DESC")

        grouped_sql = f"""
            SELECT p.sku_group, p.supplier_code,
                   MIN(p.product_name) as product_name,
                   MIN(p.price) as price,
                   MIN(p.image_url) as image_url,
                   COUNT(*) as option_count,
                   MAX(p.is_listed) as is_listed,
                   MIN(p.naver_status) as naver_status,
                   MAX(p.is_naver_duplicate) as is_naver_duplicate
            FROM {base_from}
            WHERE {where_sql}
            GROUP BY p.sku_group
            ORDER BY {order_expr}
        """

        all_rows = [dict(r) for r in conn.execute(grouped_sql, all_params).fetchall()]
        limited_rows: list[dict] = []
        supplier_counts: dict[str, int] = {}
        limit = max(1, int(count_per_supplier or 9999))
        for row in all_rows:
            supplier = row.get("supplier_code") or ""
            current = supplier_counts.get(supplier, 0)
            if current >= limit:
                continue
            supplier_counts[supplier] = current + 1
            limited_rows.append(row)

        total = len(limited_rows)
        total_pages = math.ceil(total / per_page) if per_page > 0 else 0

        # Paginated results
        offset = (page - 1) * per_page
        rows = limited_rows[offset:offset + per_page]

        # Fetch option details for each sku_group
        products_list: list[dict] = []
        sku_groups = [r["sku_group"] for r in rows]
        options_map: dict[str, list[dict]] = {}
        if sku_groups:
            oph = ",".join(["?"] * len(sku_groups))
            option_rows = conn.execute(f"""
                SELECT product_code, option_code, sku_group, product_name
                FROM products WHERE sku_group IN ({oph})
                ORDER BY sku_group, option_code
            """, sku_groups).fetchall()
            for orow in option_rows:
                sg = orow["sku_group"]
                options_map.setdefault(sg, []).append({
                    "code": orow["option_code"] or orow["product_code"],
                    "name": pm_option_label(dict(orow)),
                })

        for r in rows:
            sg = r["sku_group"]
            products_list.append({
                "sku_group": sg,
                "supplier_code": r["supplier_code"],
                "product_name": r["product_name"],
                "price": r["price"],
                "image_url": r["image_url"],
                "option_count": r["option_count"],
                "options": options_map.get(sg, []),
                "is_listed": r["is_listed"],
                "naver_status": r["naver_status"],
                "is_naver_duplicate": r["is_naver_duplicate"],
            })

        return {
            "products": products_list,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
        }
    finally:
        conn.close()


def pm_build_source_csv(products: list[dict]) -> Path:
    grouped: dict[str, list[dict]] = {}
    for p in products:
        grouped.setdefault(p["sku_group"], []).append(p)
    export_rows: list[dict] = []
    for sku in grouped:
        row = pm_build_group_row(grouped[sku])
        export_rows.append(row)
    if not export_rows:
        raise ValueError("선택된 상품이 없습니다.")
    headers = list(export_rows[0].keys())
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = safe_name(f"pm_source_{stamp}.csv")
    target = UPLOAD_ROOT / file_name
    with target.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in export_rows:
            writer.writerow(row)
    return target


class WebOcrHandler(SimpleHTTPRequestHandler):
    server_version = "WebOcrClude/0.1"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def log_message(self, format: str, *args) -> None:
        sys.stderr.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), format % args))

    def send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def read_body(self) -> bytes:
        length = int(self.headers.get("Content-Length", "0") or "0")
        return self.rfile.read(length)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/api/health":
            self.send_json({"ok": True, "time": now_text()})
            return
        if path == "/api/seeds":
            self.send_json({"ok": True, "seeds": list_seed_summaries()})
            return
        if path == "/api/market-keys":
            self.send_json({"ok": True, "items": market_key_summaries()})
            return
        if path == "/api/upload-history":
            self.send_json({"ok": True, "items": collect_upload_history()})
            return
        if path == "/api/categories":
            params = parse_qs(parsed.query)
            market = (params.get("market") or [""])[0]
            query = (params.get("q") or [""])[0]
            parent_path = (params.get("parentPath") or [""])[0]
            try:
                limit = int((params.get("limit") or ["80"])[0])
            except Exception:
                limit = 80
            items = browse_category_reference(market, parent_path, limit) if parent_path else search_category_reference(market, query, limit)
            self.send_json({
                "ok": True,
                "market": normalise_category_market(market),
                "query": query,
                "parentPath": parent_path,
                "items": items,
            })
            return
        if path == "/api/category-bundle":
            params = parse_qs(parsed.query)
            query = (params.get("q") or [""])[0]
            try:
                limit = int((params.get("limit") or ["8"])[0])
            except Exception:
                limit = 8
            bundle = category_bundle_reference(query, limit)
            self.send_json({"ok": True, **bundle})
            return
        if path == "/api/seed":
            params = parse_qs(parsed.query)
            seed_path = resolve_seed_path((params.get("path") or params.get("name") or [""])[0])
            if not seed_path.exists():
                self.send_json({"ok": False, "error": "seed not found"}, 404)
                return
            self.send_json({"ok": True, "seed": hydrate_seed_payload(read_json(seed_path, {})), "summary": seed_summary(seed_path)})
            return
        if path.startswith("/api/jobs/"):
            job_id = safe_name(path.rsplit("/", 1)[-1], "")
            job = read_json(JOBS_ROOT / f"{job_id}.json", {})
            if not job:
                self.send_json({"ok": False, "error": "job not found"}, 404)
                return
            self.send_json({"ok": True, **job})
            return
        if path == "/api/pm/dates":
            self.send_json({"ok": True, "dates": pm_get_upload_dates()})
            return
        if path == "/api/pm/suppliers":
            params = parse_qs(parsed.query)
            upload_date = (params.get("upload_date") or [""])[0]
            self.send_json({"ok": True, "suppliers": pm_get_suppliers(upload_date)})
            return
        if path == "/api/pm/status":
            self.send_json({"ok": True, "available": PRODUCT_MANAGER_DB.is_file()})
            return
        return super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        try:
            if path == "/api/import-source":
                self.handle_import_file(UPLOAD_ROOT, "file")
                return
            if path == "/api/pm/preview":
                self.handle_pm_preview()
                return
            if path == "/api/pm/import":
                self.handle_pm_import()
                return
            if path == "/api/pm/products":
                self.handle_pm_products()
                return
            if path == "/api/pm/upload-csv":
                self.handle_pm_upload_csv()
                return
            if path == "/api/import-logo":
                self.handle_import_file(LOGO_ROOT, "file")
                return
            if path == "/api/market-key":
                self.handle_market_key_upload()
                return
            if path == "/api/market-key-test":
                self.handle_market_key_test()
                return
            if path == "/api/source-to-seed":
                self.handle_source_to_seed()
                return
            if path == "/api/keyword-generate":
                self.handle_keyword_generate()
                return
            if path == "/api/automation-prepare":
                self.handle_automation_prepare()
                return
            if path == "/api/seed-update-images":
                self.handle_seed_update_images()
                return
            if path == "/api/job-stop":
                self.handle_job_stop()
                return
            if path == "/api/market-upload":
                self.handle_market_upload()
                return
            if path == "/api/emergency-codex-context":
                payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
                self.send_json(write_emergency_codex_context(payload))
                return
            if path == "/api/emergency-codex-open":
                payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
                self.send_json(open_emergency_codex_session(payload))
                return
            if path == "/api/excel-export":
                self.handle_excel_export()
                return
            if path == "/api/workspace-reset":
                self.handle_workspace_reset()
                return
            if path == "/api/seed-action":
                self.handle_seed_action()
                return
            self.send_json({"ok": False, "error": "unknown endpoint"}, 404)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, 500)

    def handle_seed_update_images(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        seed_path = resolve_seed_path(text_value(payload.get("path") or ""))
        gs = text_value(payload.get("gs") or "")
        images = payload.get("images") or {}
        if not seed_path.exists():
            self.send_json({"ok": False, "error": "seed not found"}, 404)
            return
        if not gs:
            self.send_json({"ok": False, "error": "gs required"}, 400)
            return
        if not isinstance(images, dict):
            self.send_json({"ok": False, "error": "images must be object"}, 400)
            return
        seed_payload = read_json(seed_path, {})
        products = seed_payload.get("products") if isinstance(seed_payload, dict) else []
        if not isinstance(products, list):
            self.send_json({"ok": False, "error": "seed products missing"}, 400)
            return
        for product in products:
            if not isinstance(product, dict):
                continue
            if text_value(product.get("gs") or "") != gs:
                continue
            product["images"] = images
            product["updatedAt"] = datetime.now().isoformat(timespec="seconds")
            write_json(seed_path, seed_payload)
            self.send_json({"ok": True, "path": str(seed_path), "gs": gs, "summary": seed_summary(seed_path)})
            return
        self.send_json({"ok": False, "error": "product not found in seed"}, 404)

    def handle_import_file(self, target_root: Path, field_name: str) -> None:
        body = self.read_body()
        parts = parse_multipart(body, self.headers.get("Content-Type", ""))
        file_part = parts.get(field_name)
        if not file_part or not file_part.get("filename"):
            self.send_json({"ok": False, "error": "file field missing"}, 400)
            return
        original = safe_name(file_part["filename"])
        upload_id = uuid.uuid4().hex[:12]
        target_path = target_root / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{upload_id}_{original}"
        target_path.write_bytes(file_part["content"])
        payload = {
            "ok": True,
            "uploadId": upload_id,
            "originalName": original,
            "path": str(target_path),
            "size": target_path.stat().st_size,
        }
        if target_root == UPLOAD_ROOT:
            payload["parsed"] = parse_source_preview(target_path)
        self.send_json(payload)

    def handle_market_key_upload(self) -> None:
        body = self.read_body()
        parts = parse_multipart(body, self.headers.get("Content-Type", ""))
        file_part = parts.get("file")
        if not file_part or not file_part.get("filename"):
            self.send_json({"ok": False, "error": "file field missing"}, 400)
            return
        account = text_value((parts.get("account") or {}).get("content", b"").decode("utf-8", errors="replace")).upper() or "A"
        market = text_value((parts.get("market") or {}).get("content", b"").decode("utf-8", errors="replace"))
        mode = text_value((parts.get("mode") or {}).get("content", b"").decode("utf-8", errors="replace")) or "key"
        if not market:
            self.send_json({"ok": False, "error": "market field missing"}, 400)
            return
        key = market_key_id(account, market)
        target_dir = organized_market_key_dir(account, market)
        target_dir.mkdir(parents=True, exist_ok=True)
        original = safe_name(file_part["filename"])
        target_path = target_dir / original
        target_path.write_bytes(file_part["content"])

        settings = read_market_key_settings()
        settings["items"][key] = {
            "account": account,
            "market": market,
            "mode": mode,
            "fileName": original,
            "path": str(target_path),
            "updatedAt": now_text(),
        }
        write_market_key_settings(settings)
        self.send_json({
            "ok": True,
            "item": {
                "key": key,
                "account": account,
                "market": market,
                "mode": mode,
                "fileName": original,
                "size": target_path.stat().st_size,
                "updatedAt": settings["items"][key]["updatedAt"],
                "exists": True,
            },
            "items": market_key_summaries(),
        })

    def handle_market_key_test(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        settings = read_market_key_settings()
        items = settings.get("items", {})
        target_key = market_key_id(payload.get("account", ""), payload.get("market", "")) if payload.get("account") and payload.get("market") else ""
        selected = [item for key, item in items.items() if not target_key or key == target_key]
        results = [test_market_key(item) for item in selected]
        self.send_json({"ok": True, "results": results})

    def handle_source_to_seed(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = uuid.uuid4().hex[:12]
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = {
            "ok": True,
            "jobId": job_id,
            "action": "sourceToSeed",
            "status": "queued",
            "createdAt": now_text(),
        }
        write_json(job_path, job)
        thread = threading.Thread(target=run_seed_job, args=(job_id, payload), daemon=True)
        thread.start()
        self.send_json(job, 202)

    def handle_keyword_generate(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = uuid.uuid4().hex[:12]
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = {
            "ok": True,
            "jobId": job_id,
            "action": "keywordGenerate",
            "status": "queued",
            "createdAt": now_text(),
            "progressPercent": 1,
            "currentStage": "키워드 생성 대기",
        }
        write_json(job_path, job)
        thread = threading.Thread(target=run_keyword_job, args=(job_id, payload), daemon=True)
        thread.start()
        self.send_json(job, 202)

    def handle_automation_prepare(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = uuid.uuid4().hex[:12]
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = {
            "ok": True,
            "jobId": job_id,
            "action": "automationPrepare",
            "status": "queued",
            "createdAt": now_text(),
            "progressPercent": 1,
            "currentStage": "자동화 작업 대기",
        }
        write_json(job_path, job)
        thread = threading.Thread(target=run_automation_job, args=(job_id, payload), daemon=True)
        thread.start()
        self.send_json(job, 202)

    def handle_job_stop(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = text_value(payload.get("jobId"))
        if not job_id:
            self.send_json({"ok": False, "error": "jobId required"}, 400)
            return
        stopped = stop_active_job(job_id)
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = read_json(job_path, {"jobId": job_id})
        job.update({
            "ok": True,
            "jobId": job_id,
            "status": "cancelled",
            "finishedAt": now_text(),
            "currentStage": "사용자 중지",
            "stopped": stopped,
        })
        write_json(job_path, job)
        self.send_json({"ok": True, "jobId": job_id, "stopped": stopped})

    def handle_market_upload(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        with ACTIVE_MARKET_UPLOAD_LOCK:
            active_job_id = find_active_market_upload_job()
            if active_job_id:
                self.send_json({
                    "ok": True,
                    "jobId": active_job_id,
                    "deduped": True,
                    "status": "running",
                    "currentStage": "기존 업로드 작업 진행 중",
                }, 202)
                return
            job_id = uuid.uuid4().hex[:12]
            job_path = JOBS_ROOT / f"{job_id}.json"
            job = {
                "ok": True,
                "jobId": job_id,
                "action": "marketUpload",
                "status": "queued",
                "createdAt": now_text(),
                "progressPercent": 1,
                "currentStage": "업로드 대기",
            }
            write_json(job_path, job)
            thread = threading.Thread(target=run_market_upload_job, args=(job_id, payload), daemon=True)
            thread.start()
        self.send_json(job, 202)

    def handle_excel_export(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        export = write_market_export(payload)
        self.send_json({"ok": True, "export": export})

    def handle_pm_preview(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        suppliers = payload.get("suppliers", [])
        sort_order = text_value(payload.get("sort_order")) or "latest"
        count = int(payload.get("count") or 999)
        upload_date = text_value(payload.get("upload_date"))
        products = pm_get_available_products(suppliers, sort_order, count, upload_date)
        grouped: dict[str, list[dict]] = {}
        for p in products:
            grouped.setdefault(p["sku_group"], []).append(p)
        skus: list[dict] = []
        for sku_group, items in grouped.items():
            base = sorted(items, key=pm_option_sort_key)[0]
            skus.append({
                "sku_group": sku_group,
                "supplier_code": base.get("supplier_code", ""),
                "product_name": base.get("product_name", ""),
                "price": base.get("price", 0),
                "image_url": base.get("image_url", ""),
                "option_count": len(items),
                "product_codes": [i["product_code"] for i in items],
            })
        self.send_json({"ok": True, "skus": skus, "totalProducts": len(products)})

    def handle_pm_import(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        suppliers = payload.get("suppliers", [])
        sort_order = text_value(payload.get("sort_order")) or "latest"
        count = int(payload.get("count") or 999)
        upload_date = text_value(payload.get("upload_date"))
        selected_skus = payload.get("selectedSkus")
        filter_mode = text_value(payload.get("filter_mode")) or "available"
        if isinstance(selected_skus, list) and selected_skus:
            products = pm_get_products_by_skus(selected_skus)
        else:
            products = pm_get_available_products(suppliers, sort_order, count, upload_date)
        csv_path = pm_build_source_csv(products)
        grouped: dict[str, list[dict]] = {}
        for p in products:
            grouped.setdefault(p["sku_group"], []).append(p)
        parsed = parse_source_preview(csv_path)
        self.send_json({
            "ok": True,
            "path": str(csv_path),
            "fileName": csv_path.name,
            "skuCount": len(grouped),
            "productCount": len(products),
            "size": csv_path.stat().st_size,
            "uploadId": uuid.uuid4().hex[:12],
            "parsed": parsed,
        })

    def handle_pm_products(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        suppliers = payload.get("suppliers", [])
        upload_date = text_value(payload.get("upload_date"))
        sort_order = text_value(payload.get("sort_order")) or "latest"
        filter_mode = text_value(payload.get("filter_mode")) or "available"
        search = text_value(payload.get("search")) or ""
        page = int(payload.get("page") or 1)
        per_page = int(payload.get("per_page") or 50)
        count = int(payload.get("count") or 9999)
        result = pm_get_products_paginated(suppliers, upload_date, sort_order, filter_mode, search, page, per_page, count)
        self.send_json({"ok": True, **result})

    def handle_pm_upload_csv(self) -> None:
        body = self.read_body()
        parts = parse_multipart(body, self.headers.get("Content-Type", ""))
        file_part = parts.get("file")
        if not file_part or not file_part.get("filename"):
            self.send_json({"ok": False, "error": "file field missing"}, 400)
            return
        original = safe_name(file_part["filename"])
        tmp = Path(tempfile.gettempdir()) / f"pm_upload_{original}"
        tmp.write_bytes(file_part["content"])
        pm_root = PRODUCT_MANAGER_DB.parent.parent
        try:
            import importlib.util

            def load_pm_module(name: str, path: Path):
                spec = importlib.util.spec_from_file_location(f"product_manager_{name}", path)
                if spec is None or spec.loader is None:
                    raise RuntimeError(f"ProductManager {name} 모듈을 로드할 수 없습니다: {path}")
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                return module

            pm_parser = load_pm_module("parser", pm_root / "parser.py")
            pm_database = load_pm_module("database", pm_root / "database.py")
            pm_database.init_db()
            ext = tmp.suffix.lower()
            if ext == ".csv":
                parsed = pm_parser.parse_csv(str(tmp))
            elif ext in {".xlsx", ".xls"}:
                parsed = pm_parser.parse_excel(str(tmp))
            else:
                raise ValueError("CSV 또는 Excel 파일(.csv, .xlsx, .xls)만 업로드 가능합니다")

            products = parsed.get("products", [])
            headers = parsed.get("headers", [])
            naver_dups = parsed.get("naver_duplicates", [])
            naver_listed = parsed.get("naver_listed_codes", [])
            new_count, updated_count, skipped_count, listed_count = pm_database.upsert_products(
                products, headers, naver_dups, naver_listed
            )
            upload_info = pm_database.save_upload_history(
                original, len(products), new_count, updated_count, skipped_count
            )
            pm_database.save_upload_product_items(upload_info.get("id"), products)
            self.send_json({
                "ok": True,
                "fileName": original,
                "upload_id": upload_info.get("id"),
                "upload_date": upload_info.get("upload_date"),
                "upload_date_key": text_value(upload_info.get("upload_date"))[:10],
                "total": len(products),
                "new": new_count,
                "updated": updated_count,
                "skipped": skipped_count,
                "new_count": new_count,
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "listed_count": listed_count,
                "naver_duplicates": len(naver_dups),
                "sheet_info": parsed.get("sheet_info", {}),
            })
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, 500)
        finally:
            if tmp.exists():
                tmp.unlink()

    def handle_workspace_reset(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        cleanup = cleanup_workspace_artifacts(payload)
        self.send_json({"ok": True, "cleanup": cleanup, "seeds": list_seed_summaries()})

    def handle_seed_action(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        action = text_value(payload.get("action"))
        seed_path = resolve_seed_path(payload.get("path") or payload.get("name") or "")
        if not seed_path.exists():
            self.send_json({"ok": False, "error": "seed not found"}, 404)
            return
        if action == "delete":
            seed_path.unlink()
            self.send_json({"ok": True, "action": "delete", "seeds": list_seed_summaries()})
            return
        if action == "rename":
            new_name = safe_name(text_value(payload.get("newName")))
            if not new_name.lower().endswith(".webseed.json"):
                new_name = re.sub(r"\.json$", "", new_name, flags=re.IGNORECASE) + ".webseed.json"
            target = (SEED_ROOT / new_name).resolve()
            if not is_within(SEED_ROOT, target):
                raise ValueError("target path outside data/seeds")
            if target.exists() and target != seed_path:
                raise ValueError("same seed name already exists")
            seed_path.rename(target)
            self.send_json({"ok": True, "action": "rename", "seed": seed_summary(target), "seeds": list_seed_summaries()})
            return
        self.send_json({"ok": False, "error": "unknown seed action"}, 400)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=6600)
    parser.add_argument("--host", default="127.0.0.1")
    args = parser.parse_args()

    mimetypes.add_type("text/javascript; charset=utf-8", ".js")
    mimetypes.add_type("text/jsx; charset=utf-8", ".jsx")
    httpd = ThreadingHTTPServer((args.host, args.port), WebOcrHandler)
    print(f"WebOcrClude local API server listening on http://{args.host}:{args.port}", flush=True)
    print(f"ROOT={ROOT}", flush=True)
    httpd.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
